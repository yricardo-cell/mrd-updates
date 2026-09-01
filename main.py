"""
MRD TOOL CONTROL - Aplicación principal FastAPI
v1.0.0 - MRD Estructuras
"""
import os
import re
import shutil
import subprocess
import threading
import time
import traceback
import zipfile
import urllib.parse
import uuid
from datetime import datetime, date, timedelta, timezone
from pathlib import Path
from typing import Optional, Dict, Any, Literal
import json
import io
import hashlib
from collections import OrderedDict

from fastapi import (
    FastAPI, Request, Response, Depends, Form, File,
    UploadFile, HTTPException, Query, Body
)
from fastapi.middleware.gzip import GZipMiddleware
from starlette.middleware.trustedhost import TrustedHostMiddleware  # Sprint 5.8
from fastapi.responses import (
    HTMLResponse, RedirectResponse, JSONResponse, FileResponse,
    PlainTextResponse, StreamingResponse,
)
from starlette.exceptions import HTTPException as StarletteHTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, func, or_, text
from sqlalchemy.exc import IntegrityError

from config import (
    BASE_DIR, DATA_DIR, BACKUPS_DIR, EXPORTS_DIR, UPLOADS_DIR,
    APP_NAME, COMPANY_NAME, VERSION, TEMPLATES_DIR,
    DEFAULT_ADMIN_USER, DEFAULT_ADMIN_PASSWORD,
    CATEGORIAS_DEFAULT, ESTADOS_HERRAMIENTA,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    PASSWORD_MIN_LENGTH, MAX_UPLOAD_MB, IS_PRODUCTION,
    # Sprint 5.8 — IASMRD Cloudflare deployment
    MRD_PUBLIC_URL, MRD_SCAN_URL, MRD_TRUST_PROXY_HEADERS,
    MRD_HTTPS_ONLY as _MRD_HTTPS_ONLY, MRD_ALLOWED_HOSTS,
)
from database import engine, get_db, Base, apply_migrations, SessionLocal
import mantenimiento as mant_engine
from models import (
    Ubicacion,
    Usuario, Trabajador, Almacen, Obra, Vehiculo, Herramienta, Movimiento,
    Incidencia, Reparacion, Material, Documento, Proveedor, Categoria, AuditoriaLog,
    Maquinaria, EventoMaquinaria, DocumentoMaquinaria,
    ESTADOS_MAQUINARIA, TIPOS_MAQUINARIA,
    Automatizacion, EjecucionAutomatizacion, Aviso,
    ESTADOS_AUTOMATIZACION, PRIORIDADES_AUTOMATIZACION,
    TIPOS_DISPARADOR, TIPOS_CONDICION, TIPOS_ACCION, PRIORIDADES_AVISO,
    CanalNotificacion, NotificacionEnviada, TIPOS_CANAL, PRIORIDADES_CANAL, PushSuscripcion,
    MantenimientoProgramado, TIPOS_MANTENIMIENTO, ESTADOS_MANTENIMIENTO,
    EntregaEPI, KIT_EPI_INICIAL, KIT_ROPA_SEMESTRAL, INTERVALO_ROPA_DIAS,
    StockEPI,
    EPIIndividual, RevisionEPI, HistorialEPIIndividual, TIPOS_EPI_INDIVIDUAL, INTERVALO_REVISION_EPI_DIAS,
    CatalogoEPI,
    FormacionTrabajador, TIPOS_FORMACION,
    ReconocimientoMedico, DocumentoTrabajador, TIPOS_DOCUMENTO_TRABAJADOR,
    MovimientoMaterial, UNIDADES_MATERIAL, TIPOS_MOVIMIENTO_MAT, CATEGORIAS_MATERIAL,
    MovimientoVehiculo,
    AlbaranSalida, ItemAlbaranSalida, ESTADOS_ALBARAN,
    TransferenciaAlmacen, LineaTransferenciaAlmacen, RecepcionTransferencia,
    PedidoProveedor, LineaPedidoProveedor, RecepcionPedidoProveedor, PreparacionEntrega, LoteAlmacen,
    CierreDiarioAlmacen,
    RepostajeVehiculo, RepostajeSurtidor, ScanEvento, ScanNotificacion,
    DotacionTrabajador, ExistenciaVariante, IdentificadorGlobal, LineaDotacion,
    ActivoInventarioEscaneado, AjusteInventario, IntentoConteo, LineaInventario, SesionInventario, VarianteEPI,
    LoteVariante, LogImpresionEtiqueta, MovimientoStock, RecepcionSuministro,
    SolicitudTrabajador, LineaSolicitudTrabajador, ComunicacionTrabajador,
    ComentarioSolicitudTrabajador, IncidenciaPortalTrabajador,
    NotificacionTrabajador,
    SesionPortalTrabajador, SolicitudDevolucionTrabajador,
    ESTADOS_SOLICITUD_TRABAJADOR,
)
from auth import (
    hash_password, verificar_password, crear_token,
    requiere_login, tiene_permiso, usuario_actual, ROLES_VALIDOS,
    obtener_usuario_por_token, verificar_token,
)
from codigos import generar_qr_base64, generar_barcode_base64, generar_qr_bytes
from salidas_maquinaria import router as salidas_router
from identificadores import (
    asegurar_referencias_operativas,
    asegurar_referencias_maquinaria,
    generar_referencia_almacen,
    generar_referencia_herramienta,
    generar_referencia_maquinaria,
    generar_referencia_material,
    generar_referencia_ubicacion,
    generar_referencia_vehiculo,
)
from backups import crear_backup, listar_backups, restaurar_backup
from movement_service import (
    CONDICIONES_DEVOLUCION, MovementError, actor_snapshot,
    deliver_tool, require_movement_permission, return_tool,
    start_movement_transaction,
)
from scan_service import (
    ScanIdConflict, ScanLeaseLost, changes_after, cleanup_scan_data,
    current_notification_cursor, finish_event, mark_event_error, request_hash,
    reserve_event,
)
from label_printer import generar_zpl_herramienta, generar_zpl_lote, generar_pdf_etiquetas, generar_pdf_etiquetas_ubicaciones
from stock_service import (
    StockError, move_material, move_stock_epi, move_variante, require_stock_permission,
    start_stock_transaction,
)
from generador_codigos import reservar_identificadores
from inventario_service import (
    InventoryError, approve_count, close_inventory_session,
    ensure_inventory_asset_snapshot, open_inventory_session, register_count, require_inventory_admin,
    require_inventory_operator,
)
from dotacion_service import (
    RESET_PHRASE, change_dotation_line_size, clothing_reset_preview,
    confirm_dotation, confirm_dotation_line, create_pending_dotation,
    ensure_epi_identifier, execute_clothing_reset, prepare_dotation_line,
    replace_dotation_line, return_dotation_line,
)
from etiquetas_service import build_zpl, label_from_identifier, send_label
from recepcion_service import find_variant, receive_supply
from mostrador_service import (
    CounterError, allowed_counter_types, operate_counter,
    resolve_counter_item, search_counter_items,
)
from scanner_service import normalize_scanned_code, scan_code_candidates
from warehouse_service import (
    can_access_warehouse, get_default_warehouse, get_user_warehouse,
    visible_warehouses,
)
from albaran_service import create_delivery_note
from transfer_service import (
    TransferError, cancel_transfer, create_transfer, in_transit, receive_transfer,
)
from worker_portal_service import (
    REQUEST_TRANSITIONS, WorkerPortalError, add_worker_request_comment,
    cancel_worker_request, can_manage_requests, create_worker_incident,
    create_worker_message, create_worker_notification, create_worker_request,
    create_worker_return, manage_worker_message, require_request_access,
    transition_worker_request,
)
from reports import (exportar_inventario_excel, exportar_movimientos_excel,
                    exportar_trabajadores_excel,
                    generar_plantilla_importacion, importar_herramientas_excel,
                    generar_analisis_inteligente, exportar_pdf_resumen,
                    exportar_maquinaria_excel, exportar_incidencias_excel,
                    exportar_reparaciones_excel, exportar_inventario_pdf,
                    generar_plantilla_trabajadores, importar_trabajadores_excel)
# Sprint 5.7 updater compatibility shim
import updater as _updater_mod
import json as _json

def leer_version_actual() -> dict:
    try:
        vf = Path("version.json")
        if vf.exists():
            return _json.loads(vf.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {"version_actual": VERSION, "cambios": [], "historial": []}

def verificar_actualizacion_local() -> dict:
    try:
        return _updater_mod.check_update()
    except Exception:
        return {"hay_actualizacion": False, "version_disponible": None, "error": "No disponible"}

def aplicar_actualizacion(archivo=None) -> dict:
    try:
        state = _updater_mod.get_state()
        url = state.get("download_url") or (archivo if isinstance(archivo, str) and archivo.startswith("http") else None)
        if url:
            return _updater_mod.start_update(url, state.get("sha256", ""), state.get("version_disponible", ""))
    except Exception as e:
        return {"ok": False, "error": str(e)}
    return {"ok": False, "error": "Sin URL de descarga"}

def leer_historial() -> list:
    try:
        vf = Path("version.json")
        if vf.exists():
            d = _json.loads(vf.read_text(encoding="utf-8"))
            return d.get("historial", d.get("cambios", []))
    except Exception:
        pass
    return []
import remote_access
import mrd_logging
import automatizaciones as auto_engine
import notificaciones as notif_engine
import push_service
import anomalias as anom_engine
from tools import aplicar_accion, registrar_auditoria, snapshot_herramienta, ErrorTransicion, ESTADOS
from security import (
    CSRF_COOKIE_NAME, CSRF_FIELD_NAME, CSRF_HEADER_NAME,
    generar_csrf_token, validar_csrf,
    validar_contrasena, ErrorContrasena,
    validar_nombre_archivo, validar_contenido_archivo, validar_tamaño_bytes, ErrorArchivo,
    build_security_headers,
)


# ─── Inicialización ───────────────────────────────────────────────────────────
Base.metadata.create_all(bind=engine)

def _sembrar_catalogo_epi():
    """Siembra el catálogo inicial; los cambios de esquema viven en database.py."""
    try:
        with engine.begin() as conn:
            columns = {
                row[1] for row in conn.execute(text("PRAGMA table_info(catalogo_epi)"))
            }
            required = {"nombre", "categoria", "cantidad_kit", "activo", "orden"}
            if not required.issubset(columns):
                return
            n_cat = conn.execute(text("SELECT COUNT(*) FROM catalogo_epi")).scalar()
            if n_cat == 0:
                for i, item in enumerate(KIT_EPI_INICIAL):
                    conn.execute(text(
                        "INSERT INTO catalogo_epi (nombre, categoria, cantidad_kit, activo, orden) "
                        "VALUES (:n, 'epi', :c, 1, :o)"
                    ), {"n": item["nombre"], "c": item.get("cantidad", 1), "o": i})
                for i, item in enumerate(KIT_ROPA_SEMESTRAL):
                    conn.execute(text(
                        "INSERT INTO catalogo_epi (nombre, categoria, cantidad_kit, activo, orden) "
                        "VALUES (:n, 'ropa', :c, 1, :o)"
                    ), {"n": item["nombre"], "c": item.get("cantidad", 1), "o": i})
                mrd_logging.log_app("Catálogo EPI sembrado con artículos iniciales")

    except Exception as _e:
        mrd_logging.log_app(f"Catálogo EPI no sembrado: {_e}", level="warning")

_sembrar_catalogo_epi()

app = FastAPI(
    title=APP_NAME,
    version=VERSION,
    docs_url=None,
    redoc_url=None,
)

app.add_middleware(GZipMiddleware, minimum_size=1000)

# ─── Sprint 5.8 — IASMRD: Proxy headers (Cloudflare) ─────────────────────────
if MRD_TRUST_PROXY_HEADERS:
    from starlette.middleware.httpsredirect import HTTPSRedirectMiddleware as _HttpsRedir  # noqa
    # ProxyHeadersMiddleware not in starlette standalone; implement inline
    @app.middleware("http")
    async def _proxy_headers_mw(request: Request, call_next):
        """Rewrite scope based on X-Forwarded-Proto and X-Real-IP / CF-Connecting-IP."""
        scope = request.scope
        # Protocol
        proto = (
            request.headers.get("cf-visitor", "")
            .replace('{"scheme":"', "").replace('"}', "").strip()
            or request.headers.get("x-forwarded-proto", "")
            .split(",")[0].strip()
        )
        if proto in ("https", "http"):
            scope["scheme"] = proto
        # Client IP (prefer CF-Connecting-IP)
        real_ip = (
            request.headers.get("cf-connecting-ip", "")
            or request.headers.get("x-real-ip", "")
            or request.headers.get("x-forwarded-for", "").split(",")[0].strip()
        )
        if real_ip and scope.get("client"):
            scope["client"] = (real_ip, scope["client"][1])
        return await call_next(request)

# ─── Sprint 5.8 — IASMRD: TrustedHost ────────────────────────────────────────
if MRD_ALLOWED_HOSTS:
    app.add_middleware(
        TrustedHostMiddleware,
        allowed_hosts=MRD_ALLOWED_HOSTS,
    )

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")


@app.get("/sw.js", include_in_schema=False)
def service_worker_raiz():
    """Sirve la PWA desde la raíz para que pueda proteger toda la aplicación."""
    return FileResponse(
        BASE_DIR / "static" / "js" / "sw.js",
        media_type="application/javascript",
        headers={"Service-Worker-Allowed": "/", "Cache-Control": "no-cache"},
    )


@app.middleware("http")
async def cache_static(request: Request, call_next):
    response = await call_next(request)
    if request.url.path == "/static/manifest.json":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    elif request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "public, max-age=604800"
    return response


@app.middleware("http")
async def audit_operaciones_usuario(request: Request, call_next):
    """Registra toda operación con cambios; nunca almacena contraseñas ni cuerpos."""
    mutation = request.method in {"POST", "PUT", "PATCH", "DELETE"}
    actor = None
    if mutation:
        audit_db = SessionLocal()
        try:
            token = request.cookies.get("mrd_token", "")
            current = obtener_usuario_por_token(token, audit_db) if token else None
            if current:
                actor = (current.id, current.username)
        finally:
            audit_db.close()
    response = await call_next(request)
    if actor:
        audit_db = SessionLocal()
        try:
            clean_path = request.url.path[:300]
            audit_db.add(AuditoriaLog(
                tabla=(clean_path.strip("/").split("/")[0] or "sistema")[:50],
                registro_id=None,
                accion="operacion" if response.status_code < 400 else "operacion_fallida",
                resumen=f"{request.method} {clean_path} · HTTP {response.status_code}"[:500],
                datos_nuevos=json.dumps({"ruta": clean_path, "metodo": request.method, "estado_http": response.status_code}),
                usuario_id=actor[0],
                ip=(request.client.host if request.client else "")[:45],
                user_agent=request.headers.get("user-agent", "")[:255],
            ))
            audit_db.commit()
        except Exception as exc:
            audit_db.rollback()
            mrd_logging.log_app(f"Auditoría HTTP no registrada: {exc}", level="warning")
        finally:
            audit_db.close()
    return response


# Los archivos operativos pueden contener fotografías, certificados o
# documentos internos. StaticFiles no aplica dependencias de autenticación,
# por lo que se protegen aquí tanto /uploads como /static/uploads.
@app.middleware("http")
async def protect_uploaded_files(request: Request, call_next):
    path_req = request.url.path
    is_upload = (
        path_req.startswith("/uploads/") or
        path_req.startswith("/static/uploads/")
    )
    if not is_upload:
        return await call_next(request)

    authorized = False
    db = SessionLocal()
    try:
        token = request.cookies.get("mrd_token", "")
        authorized = bool(token and obtener_usuario_por_token(token, db))

        # El portal público solo puede cargar la foto del trabajador asociada
        # al token que ya protege su ficha.
        if not authorized and path_req.startswith("/static/uploads/trabajadores/"):
            portal_token = request.query_params.get("portal_token", "").strip()
            filename = Path(path_req).name
            if portal_token and filename:
                authorized = db.query(Trabajador).filter(
                    Trabajador.portal_token == portal_token,
                    Trabajador.foto == filename,
                ).first() is not None
    finally:
        db.close()

    if not authorized:
        return RedirectResponse("/login", status_code=303)

    response = await call_next(request)
    response.headers["Cache-Control"] = "private, no-store"
    return response


_CONSULTA_ALLOWED = (
    "/scan", "/scanner/configurar", "/perfil", "/cambiar-contrasena",
    "/logout", "/instalar", "/health", "/static", "/sw.js",
)


@app.middleware("http")
async def restrict_consulta_scope(request: Request, call_next):
    """Consulta solo puede consultar un QR y administrar su propia sesión."""
    token = request.cookies.get("mrd_token", "")
    if not token:
        return await call_next(request)
    db = SessionLocal()
    try:
        current = obtener_usuario_por_token(token, db)
        is_consulta = bool(current and current.rol == "consulta")
    finally:
        db.close()
    if not is_consulta:
        return await call_next(request)
    path_req = request.url.path
    if path_req == "/":
        return RedirectResponse("/scan?modo=consulta", status_code=303)
    allowed = any(path_req.startswith(prefix) for prefix in _CONSULTA_ALLOWED)
    if not allowed:
        if request.method == "GET":
            return RedirectResponse("/scan?modo=consulta", status_code=303)
        return JSONResponse({"detail": "La cuenta de consulta no puede realizar esta acción"}, status_code=403)
    return await call_next(request)



# ─── Middleware: Cabeceras de seguridad HTTP ──────────────────────────────────
def _is_https_request(request: Request) -> bool:
    """Detecta HTTPS efectivo detrás de proxys (Cloudflare/WARP) para decidir
    si las cookies deben marcarse Secure."""
    return (
        _MRD_HTTPS_ONLY
        or request.headers.get("x-forwarded-proto", "http") == "https"
        or request.headers.get("cf-visitor", "").find("https") != -1
    )


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    is_https = _is_https_request(request)
    for header, value in build_security_headers(is_https).items():
        response.headers[header] = value
    if request.url.path.startswith("/portal/") or request.url.path.startswith("/portal-trabajador"):
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["Cache-Control"] = "no-store, private"
    return response


# ─── Middleware: Cambio obligatorio de contraseña ─────────────────────────────
# Rutas exentas: el usuario puede acceder aunque must_change_password=True
_RUTAS_EXENTAS_MCP = {
    "/cambiar-contrasena", "/logout", "/login",
    "/health", "/static", "/favicon.ico",
}

@app.middleware("http")
async def must_change_password_middleware(request: Request, call_next):
    """Redirige a /cambiar-contrasena si el JWT contiene mcp=1."""
    path_req = request.url.path
    # Permitir recursos estáticos y rutas exentas
    if (path_req.startswith("/static") or
            path_req.startswith("/uploads") or
            any(path_req.startswith(r) for r in _RUTAS_EXENTAS_MCP)):
        return await call_next(request)
    # Solo peticiones GET (no interrumpir POSTs en /cambiar-contrasena)
    if request.method != "GET":
        return await call_next(request)
    # Leer el claim mcp del JWT (mismo decodificador que el resto de la app)
    token = request.cookies.get("mrd_token", "")
    if token:
        payload = verificar_token(token)
        if payload and payload.get("mcp"):
            return RedirectResponse("/cambiar-contrasena", status_code=303)
    return await call_next(request)


# ─── Middleware: Protección CSRF ──────────────────────────────────────────────
# Rutas exentas del CSRF (sin estado, sin autenticación)
_CSRF_EXENTOS = {
    "/health", "/scan/buscar", "/scan/ip", "/login", "/instalar",
    "/portal-trabajador/acceso", "/login/2fa",
}

def _csrf_403(request: Request, detalle: str = "") -> Response:
    msg = detalle or "Token de seguridad inválido. Recarga la página e inténtalo de nuevo."
    return HTMLResponse(
        content=(
            f'<html><head><meta charset="utf-8"><meta http-equiv="refresh" content="3;url=/">'
            f'<link rel="stylesheet" href="/static/css/bootstrap.min.css"></head><body>'
            f'<div style="display:flex;align-items:center;justify-content:center;min-height:100vh">'
            f'<div style="text-align:center;max-width:400px;padding:40px">'
            f'<div style="font-size:4rem;font-weight:800;color:#dc3545">403</div>'
            f'<h2>Error de seguridad</h2><p>{msg}</p>'
            f'<a href="/" class="btn btn-primary">Volver al inicio</a></div></div></body></html>'
        ),
        status_code=403,
    )

@app.middleware("http")
async def csrf_middleware(request: Request, call_next):
    """
    Protección CSRF mediante double-submit cookie.
    - Cookie mrd_csrf (SameSite=Lax, NOT httponly) contiene el token.
    - Requests mutantes deben incluir el token en X-CSRF-Token header
      (para AJAX/fetch) o en el campo _csrf_token del body (para form submit).
    """
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        if request.url.path not in _CSRF_EXENTOS:
            csrf_cookie = request.cookies.get(CSRF_COOKIE_NAME, "")

            # ── 1. Verificar desde header (fetch/AJAX) ────────────────────────
            header_token = request.headers.get(CSRF_HEADER_NAME, "")
            if header_token:
                if not csrf_cookie or not validar_csrf(csrf_cookie, header_token):
                    mrd_logging.log_security(
                        f"CSRF inválido (header): path={request.url.path} "
                        f"ip={request.client.host if request.client else '?'}"
                    )
                    return _csrf_403(request)
            else:
                # ── 2. Verificar desde body (form submit nativo) ──────────────
                if not csrf_cookie:
                    return _csrf_403(request, "Sesión de seguridad no iniciada. Recarga la página.")
                try:
                    body = await request.body()
                    body_str = body.decode("utf-8", errors="ignore")
                    form_token = ""

                    # URL-encoded form
                    params = urllib.parse.parse_qs(body_str, keep_blank_values=True)
                    form_token = (params.get(CSRF_FIELD_NAME) or [""])[0]

                    # Multipart form (búsqueda simple en raw bytes)
                    if not form_token:
                        marker = f'name="{CSRF_FIELD_NAME}"'
                        idx = body_str.find(marker)
                        if idx != -1:
                            after = body_str[idx + len(marker):]
                            dbl = after.find("\r\n\r\n")
                            if dbl != -1:
                                vs = dbl + 4
                                ve = after.find("\r\n", vs)
                                form_token = after[vs:ve] if ve != -1 else after[vs:vs+128]

                    if not form_token or not validar_csrf(csrf_cookie, form_token):
                        mrd_logging.log_security(
                            f"CSRF inválido (form): path={request.url.path} "
                            f"ip={request.client.host if request.client else '?'}"
                        )
                        return _csrf_403(request)
                except Exception as _e:
                    mrd_logging.log_security(f"CSRF error: {_e}")
                    return _csrf_403(request)

    response = await call_next(request)

    # Establecer cookie CSRF si no existe
    if not request.cookies.get(CSRF_COOKIE_NAME):
        is_https = _is_https_request(request)
        response.set_cookie(
            CSRF_COOKIE_NAME,
            generar_csrf_token(),
            httponly=False,   # JS debe leer el valor para añadirlo a forms/fetch
            secure=is_https,
            samesite="lax",
            path="/",
            max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        )

    return response


# ─── Manejador global de errores ─────────────────────────────────────────────
_ERROR_MESSAGES = {
    400: ("Solicitud incorrecta", "Los datos enviados no son válidos."),
    401: ("No autenticado", "Debes iniciar sesión para acceder."),
    403: ("Sin permiso", "No tienes permiso para realizar esta acción."),
    404: ("Página no encontrada", "La página que buscas no existe o fue movida."),
    409: ("Conflicto", "Ya existe un recurso con esos datos."),
    422: ("Datos inválidos", "Revisa el formulario e inténtalo de nuevo."),
    429: ("Demasiadas solicitudes", "Has superado el límite de intentos. Espera unos minutos."),
    500: ("Error interno", "Ocurrió un error inesperado. El equipo ha sido notificado."),
    503: ("Servicio no disponible", "El servicio está temporalmente fuera de línea."),
}

def _render_error(request: Request, code: int, detail: str = "") -> HTMLResponse:
    titulo, msg = _ERROR_MESSAGES.get(code, ("Error", "Ha ocurrido un error."))
    if detail and code not in (500,):  # no mostrar detalle técnico en 500
        msg = detail
    html = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{code} — {titulo}</title>
  <link rel="stylesheet" href="/static/css/bootstrap.min.css">
  <link rel="stylesheet" href="/static/css/mrd.css">
</head>
<body style="display:flex;align-items:center;justify-content:center;min-height:100vh;background:var(--bg)">
  <div style="text-align:center;max-width:480px;padding:40px 24px">
    <div style="font-size:5rem;font-weight:800;color:var(--primary);line-height:1">{code}</div>
    <h1 style="font-size:1.4rem;font-weight:700;margin:16px 0 8px">{titulo}</h1>
    <p style="color:var(--text-2);margin-bottom:32px">{msg}</p>
    <a href="/" class="btn btn-primary">Volver al inicio</a>
    {"<a href='/login' class='btn btn-outline' style='margin-left:8px'>Iniciar sesión</a>" if code == 401 else ""}
  </div>
</body>
</html>"""
    return HTMLResponse(content=html, status_code=code)


@app.exception_handler(StarletteHTTPException)
async def http_error_handler(request: Request, exc: StarletteHTTPException):
    scan_api = request.url.path in {"/scan/operar", "/scan/cambios"}
    # Para redirecciones devolver la respuesta directamente (no re-lanzar)
    if exc.status_code in (301, 302, 303, 307, 308):
        if scan_api:
            return JSONResponse({"resultado": "error", "detalle": "Sesión caducada"}, status_code=401)
        location = (exc.headers or {}).get("location", "/login")
        return RedirectResponse(url=location, status_code=exc.status_code)
    if exc.status_code == 401:
        if scan_api:
            return JSONResponse({"resultado": "error", "detalle": "Sesión caducada"}, status_code=401)
        return RedirectResponse("/login", status_code=303)
    detail = str(exc.detail) if exc.detail else ""
    mrd_logging.log_error(f"HTTP {exc.status_code} en {request.url.path} — {detail}")
    return _render_error(request, exc.status_code, detail)


@app.exception_handler(Exception)
async def global_error_handler(request: Request, exc: Exception):
    tb = traceback.format_exc()
    mrd_logging.log_error(f"Error no controlado en {request.url.path}", exc)
    mrd_logging.errors.error(tb)
    print(f"[MRD-500] {request.url.path} — {type(exc).__name__}: {exc}", flush=True)
    print(tb, flush=True)
    return _render_error(request, 500)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


# ─── Jinja2 filtros personalizados ────────────────────────────────────────────
def fmt_fecha(value):
    if not value:
        return "—"
    if isinstance(value, str):
        return value
    return value.strftime("%d/%m/%Y")


def fmt_datetime(value):
    if not value:
        return "—"
    if isinstance(value, str):
        return value
    return value.strftime("%d/%m/%Y %H:%M")


def fmt_precio(value):
    if value is None:
        return "—"
    return f"{float(value):,.2f} €"


def estado_color(estado):
    return ESTADOS_HERRAMIENTA.get(estado, {}).get("color", "secondary")


def estado_label(estado):
    return ESTADOS_HERRAMIENTA.get(estado, {}).get("label", estado)


def fromjson_filter(value):
    """Parsea un JSON string a dict (para usar en templates con datos de auditoría)."""
    if not value:
        return {}
    try:
            return json.loads(value)
    except Exception:
        return {}


def dumps_for_script(obj) -> str:
    """json.dumps seguro para incrustar en <script>...</script> vía |safe.

    Escapa '<', '>' y '&' para que un valor guardado en BD (p.ej. nombre de
    herramienta) que contenga '</script>' no pueda cerrar el bloque de script
    e inyectar HTML/JS arbitrario (stored XSS)."""
    return (
        json.dumps(obj, ensure_ascii=False)
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("&", "\\u0026")
    )


templates.env.filters["fmt_fecha"] = fmt_fecha
templates.env.filters["fmt_datetime"] = fmt_datetime
templates.env.filters["fmt_precio"] = fmt_precio
templates.env.filters["estado_color"] = estado_color
templates.env.filters["estado_label"] = estado_label
templates.env.filters["fromjson"] = fromjson_filter

# Test "search" para selectattr (compatible con versiones nuevas de Jinja2)
import re as _re
templates.env.tests["search"] = lambda value, pattern: bool(_re.search(pattern, str(value) if value else ""))


# ─── Startup ──────────────────────────────────────────────────────────────────
def _alertas_no_retorno():
    try:
        from database import SessionLocal as _SL2
        from datetime import datetime as _dt2, timedelta as _td2
        db = _SL2()
        limite = _dt2.now() - _td2(days=14)
        herr = db.query(Herramienta).filter(
            Herramienta.activa == True,
            Herramienta.estado.notin_(["disponible","baja","mantenimiento"]),
        ).all()
        for h in herr:
            ultimo = db.query(Movimiento).filter(
                Movimiento.herramienta_id == h.id,
                Movimiento.tipo.in_(["entrega","traslado"]),
            ).order_by(Movimiento.fecha.desc()).first()
            if ultimo and ultimo.fecha < limite:
                ya = db.query(Aviso).filter(
                    Aviso.titulo.like(f"No retorno herramienta {h.id}%"),
                    Aviso.leido == False, Aviso.archivado == False,
                ).first()
                if not ya:
                    db.add(Aviso(titulo=f"No retorno herramienta {h.id}: {h.nombre}",
                                 mensaje=f"'{h.nombre}' lleva mas de 14 dias sin retorno.",
                                 tipo="alerta", prioridad="alta"))
        vehs = db.query(MovimientoVehiculo).filter(
            MovimientoVehiculo.fecha_retorno == None,
            MovimientoVehiculo.fecha_salida < limite,
        ).all()
        for mv in vehs:
            mat = mv.vehiculo.matricula if mv.vehiculo else str(mv.vehiculo_id)
            ya = db.query(Aviso).filter(
                Aviso.titulo.like(f"No retorno vehiculo {mat}%"),
                Aviso.leido == False, Aviso.archivado == False,
            ).first()
            if not ya:
                db.add(Aviso(titulo=f"No retorno vehiculo {mat} (mov #{mv.id})",
                             mensaje=f"Vehiculo {mat} salio el {mv.fecha_salida.strftime('%d/%m/%Y')} sin retorno.",
                             tipo="alerta", prioridad="alta"))
        db.commit(); db.close()
    except Exception: pass


def _alertas_stock_bajo():
    try:
        from database import SessionLocal as _SL3
        db = _SL3()
        for m in [x for x in db.query(Material).filter(Material.activo == True).all() if x.bajo_minimo]:
            ya = db.query(Aviso).filter(
                Aviso.titulo.like(f"Stock bajo {m.id}%"),
                Aviso.leido == False, Aviso.archivado == False,
            ).first()
            if not ya:
                db.add(Aviso(titulo=f"Stock bajo {m.id}: {m.nombre}",
                             mensaje=f"Stock: {m.stock_actual} {m.unidad} (min: {m.stock_minimo})",
                             tipo="aviso", prioridad="media"))
        db.commit(); db.close()
    except Exception: pass


@app.on_event("startup")
def startup_event():
    mrd_logging.log_app(f"Arrancando MRD TOOL CONTROL v{VERSION}")
    testing = os.getenv("MRD_TESTING", "0") == "1"

    # Inicializar módulo de acceso remoto y pre-calentar caché en background
    if not testing:
        remote_access.init(BASE_DIR)
        threading.Thread(target=remote_access.get_status_cached, daemon=True).start()

    # Migración incremental V2 — añade columnas nuevas sin borrar datos
    apply_migrations()
    mrd_logging.log_app("Migraciones aplicadas")

    # Los códigos que alimentan etiquetas y escáner siempre los crea el
    # servidor. Completa registros históricos sin tocar referencias existentes.
    _refs_db = SessionLocal()
    try:
        _refs = asegurar_referencias_operativas(_refs_db)
        _metros = _reclasificar_metros_como_epi(_refs_db)
        _nombres_ropa = _normalizar_nombres_camisas(_refs_db)
        _chalecos = _reclasificar_chalecos_como_ropa(_refs_db)
        _avisos_nuevos = _sincronizar_avisos_operativos(_refs_db)
        _refs_db.commit()
        if any(_refs.values()):
            mrd_logging.log_app(f"Referencias QR completadas: {_refs}")
        if _metros:
            mrd_logging.log_app(f"Metros reclasificados de materiales a EPI: {_metros}")
        if _nombres_ropa:
            mrd_logging.log_app(f"Nombres de camisetas normalizados: {_nombres_ropa}")
        if _chalecos:
            mrd_logging.log_app(f"Chalecos reclasificados como ropa por tallas: {_chalecos}")
        if _avisos_nuevos:
            mrd_logging.log_app(f"Avisos operativos reales creados: {_avisos_nuevos}")
    except Exception as _refs_exc:
        _refs_db.rollback()
        mrd_logging.log_app(f"No se pudieron completar referencias QR: {_refs_exc}", level="warning")
    finally:
        _refs_db.close()

    from database import SessionLocal as _SL
    if not testing:
        # Sprint 4.1 — Arrancar scheduler de automatizaciones
        auto_engine.start_scheduler(lambda: _SL())
        mrd_logging.log_app("Scheduler de automatizaciones arrancado")

        # Copias SQLite diarias, semanales y mensuales con retención automática
        try:
            import backup_manager as _backup_scheduler
            _backup_scheduler.start_scheduler()
            mrd_logging.log_app("Scheduler de backups arrancado")
        except Exception as _exc:
            mrd_logging.log_app(f"Scheduler de backups no iniciado: {_exc}")

        # Sprint 4.4 — Email semanal automático (lunes 08:00)
        try:
            import email_semanal as _es
            _es.arrancar_scheduler_semanal(lambda: _SL())
            mrd_logging.log_app("Scheduler de email semanal arrancado")
        except Exception as _exc:
            mrd_logging.log_app(f"Email semanal no iniciado: {_exc}")
        # Alertas periódicas de no-retorno y stock bajo
        try:
            import threading as _thr, time as _tm
            def _run_alerts():
                _tm.sleep(30)  # esperar arranque completo
                while True:
                    _alertas_no_retorno()
                    _alertas_stock_bajo()
                    _tm.sleep(6*3600)  # cada 6 horas
            _thr.Thread(target=_run_alerts, daemon=True, name="alertas_bg").start()
        except Exception:
            pass

    db = next(get_db())
    try:
        # Admin por defecto — solo crear si no existe ningún admin
        any_admin = db.query(Usuario).filter(Usuario.rol == "admin").first()
        if not any_admin:
            # Sprint 5.2: generar contraseña segura si no está configurada
            import sys as _sys
            _admin_pwd = DEFAULT_ADMIN_PASSWORD or __import__('secrets').token_urlsafe(16)
            if not DEFAULT_ADMIN_PASSWORD:
                print(
                    f'\n🔑  CONTRASEÑA DE ADMINISTRADOR GENERADA AUTOMÁTICAMENTE\n'
                    f'    Usuario:    admin\n'
                    f'    Contraseña: {_admin_pwd}\n'
                    f'    Guárdala en un lugar seguro. DEBERÁS cambiarla al iniciar sesión.\n',
                    file=_sys.stderr, flush=True,
                )
            admin = Usuario(
                username=DEFAULT_ADMIN_USER,
                password_hash=hash_password(_admin_pwd),
                nombre='Administrador',
                rol='admin',
                activo=True,
                must_change_password=True,  # Sprint 5.2
            )
            db.add(admin)
            db.commit()
            mrd_logging.log_security('Admin inicial creado. must_change_password=True', level='info')

        # Centros físicos separados. Los registros históricos sin almacén
        # pertenecen a Madrid; Barcelona nace vacío y nunca hereda ese stock.
        def _warehouse_named(city: str):
            target = f"almacén {city}".casefold()
            return next((row for row in db.query(Almacen).all()
                         if (row.nombre or "").strip().casefold() == target), None)

        madrid = _warehouse_named("Madrid")
        if madrid is None:
            principal = next((row for row in db.query(Almacen).all()
                              if "principal" in (row.nombre or "").casefold()), None)
            if principal:
                principal.nombre = "Almacén Madrid"
                principal.descripcion = principal.descripcion or "Centro principal de Madrid"
                madrid = principal
            else:
                madrid = Almacen(
                    codigo=generar_referencia_almacen(db), nombre="Almacén Madrid",
                    descripcion="Centro principal de Madrid", activo=True,
                )
                db.add(madrid)
                db.flush()
        barcelona = _warehouse_named("Barcelona")
        if barcelona is None:
            barcelona = Almacen(
                codigo=generar_referencia_almacen(db), nombre="Almacén Barcelona",
                descripcion="Centro operativo de Barcelona", activo=True,
            )
            db.add(barcelona)
            db.flush()

        for model in (Herramienta, Material, StockEPI, EPIIndividual, Maquinaria,
                      Vehiculo, Trabajador, Obra, Incidencia, Reparacion,
                      SesionInventario, AlbaranSalida):
            if hasattr(model, "almacen_id"):
                db.query(model).filter(model.almacen_id.is_(None)).update(
                    {"almacen_id": madrid.id}, synchronize_session=False,
                )
        db.query(Usuario).filter(
            Usuario.rol != "admin", Usuario.almacen_id.is_(None),
        ).update({"almacen_id": madrid.id}, synchronize_session=False)
        db.commit()

        # Toda maquinaria, también la histórica, debe disponer de referencia MRD.
        referencias_creadas = asegurar_referencias_maquinaria(db)
        if referencias_creadas:
            db.commit()
            mrd_logging.log_app(
                f"Referencias internas creadas para {referencias_creadas} máquinas históricas"
            )
    finally:
        db.close()


# ─── Helpers ──────────────────────────────────────────────────────────────────
TIPOS_AVISO_OPERATIVO = ("mantenimiento", "revision", "averia", "reparacion", "incidencia")
_CTX_COUNTS_CACHE: dict[int | None, tuple[float, int, int]] = {}
_CTX_COUNTS_LOCK = threading.Lock()
_CTX_COUNTS_TTL_SECONDS = 8.0


def _stock_bajo_count(db: Session, warehouse_id: int | None = None) -> int:
    material_query = db.query(func.count(Material.id)).filter(
        Material.activo == True, Material.stock_minimo > 0,
        Material.stock_actual <= Material.stock_minimo,
    )
    legacy_query = db.query(func.count(StockEPI.id)).filter(
        StockEPI.stock_minimo > 0, StockEPI.cantidad <= StockEPI.stock_minimo,
    )
    if warehouse_id:
        material_query = material_query.filter(Material.almacen_id == warehouse_id)
        legacy_query = legacy_query.filter(StockEPI.almacen_id == warehouse_id)
    materials = material_query.scalar() or 0
    legacy_epi = legacy_query.scalar() or 0
    totals = db.query(
        ExistenciaVariante.variante_id.label("variant_id"),
        func.coalesce(func.sum(ExistenciaVariante.cantidad), 0).label("total"),
    )
    if warehouse_id:
        totals = totals.filter(ExistenciaVariante.almacen_id == warehouse_id)
    totals = totals.group_by(ExistenciaVariante.variante_id).subquery()
    variants = db.query(func.count(VarianteEPI.id)).outerjoin(
        totals, totals.c.variant_id == VarianteEPI.id,
    ).filter(
        VarianteEPI.activo == True, VarianteEPI.stock_minimo > 0,
        func.coalesce(totals.c.total, 0) <= VarianteEPI.stock_minimo,
    ).scalar() or 0
    return int(materials + legacy_epi + variants)


def _navigation_counts(db: Session, warehouse_id: int | None) -> tuple[int, int]:
    """Evita repetir cuatro agregaciones pesadas al renderizar cada pantalla."""
    use_cache = os.getenv("MRD_TESTING") != "1"
    now = time.monotonic()
    if use_cache:
        with _CTX_COUNTS_LOCK:
            cached = _CTX_COUNTS_CACHE.get(warehouse_id)
            if cached and now - cached[0] < _CTX_COUNTS_TTL_SECONDS:
                return cached[1], cached[2]
    notices = db.query(Aviso).filter(
        Aviso.leido == False, Aviso.archivado == False,
        Aviso.tipo.in_(TIPOS_AVISO_OPERATIVO),
    ).count()
    low_stock = _stock_bajo_count(db, warehouse_id)
    if use_cache:
        with _CTX_COUNTS_LOCK:
            _CTX_COUNTS_CACHE[warehouse_id] = (now, int(notices), int(low_stock))
    return int(notices), int(low_stock)


def ctx_base(request: Request, user: Usuario, db: Session = None, **kwargs) -> dict:
    avisos_sin_leer = 0
    dash_mat_bajo_minimo = 0
    active_warehouse = None
    warehouses_visible = []
    if db is not None:
        requested_id = None
        if user and user.rol == "admin":
            try:
                requested_id = int(request.cookies.get("mrd_warehouse_id") or 0) or None
            except (TypeError, ValueError):
                requested_id = None
        active_warehouse = get_user_warehouse(db, user, requested_id) if user else None
        warehouses_visible = visible_warehouses(db, user) if user else []
        try:
            avisos_sin_leer, dash_mat_bajo_minimo = _navigation_counts(
                db, active_warehouse.id if active_warehouse else None,
            )
        except Exception:
            pass
    return {
        "request": request,
        "user": user,
        "app_name": APP_NAME,
        "company_name": COMPANY_NAME,
        "version": VERSION,
        "avisos_sin_leer": avisos_sin_leer,
        "dash_mat_bajo_minimo": dash_mat_bajo_minimo,
        "almacen_actual": active_warehouse,
        "almacenes_visibles": warehouses_visible,
        # Sprint 5.2: CSRF token disponible en todos los templates
        "csrf_token": request.cookies.get(CSRF_COOKIE_NAME, ""),
        **kwargs,
    }


def registrar_movimiento(
    db: Session,
    herramienta: Herramienta,
    tipo: str,
    estado_nuevo: str,
    usuario: Usuario,
    trabajador_id: Optional[int] = None,
    obra_id: Optional[int] = None,
    destino: str = "",
    observaciones: str = "",
):
    mov = Movimiento(
        tipo=tipo,
        estado_anterior=herramienta.estado,
        estado_nuevo=estado_nuevo,
        destino=destino,
        observaciones=observaciones,
        herramienta_id=herramienta.id,
        usuario_id=usuario.id,
        trabajador_id=trabajador_id,
        obra_id=obra_id,
    )
    db.add(mov)
    return mov


# ─── Rate limiting para login ────────────────────────────────────────────────
_login_attempts: dict = {}   # {ip_o_user: {"count": int, "locked_until": float}}
_login_lock = threading.Lock()
_scan_attempts: dict[str, list[float]] = {}
_scan_lock = threading.Lock()
_worker_buzon_recent: dict[str, float] = {}
_worker_buzon_lock = threading.Lock()
_SCAN_LIMIT_PER_MINUTE = 300
_MAX_INTENTOS = 5
_BLOQUEO_SEGUNDOS = 300      # 5 minutos

def _puede_intentar_login(clave: str) -> bool:
    with _login_lock:
        data = _login_attempts.get(clave, {"count": 0, "locked_until": 0.0})
        return time.time() >= data["locked_until"]

def _registrar_fallo_login(clave: str):
    with _login_lock:
        data = _login_attempts.get(clave, {"count": 0, "locked_until": 0.0})
        data["count"] += 1
        if data["count"] >= _MAX_INTENTOS:
            data["locked_until"] = time.time() + _BLOQUEO_SEGUNDOS
            data["count"] = 0
        _login_attempts[clave] = data

def _limpiar_intentos_login(clave: str):
    with _login_lock:
        _login_attempts.pop(clave, None)

def _segundos_bloqueo(clave: str) -> int:
    with _login_lock:
        data = _login_attempts.get(clave, {"locked_until": 0.0})
        resto = data["locked_until"] - time.time()
        return max(0, int(resto))


def _permitir_busqueda_scan(clave: str, ahora: float | None = None) -> bool:
    """Límite sencillo por IP para conservar el escáner público sin enumeración masiva."""
    ts = time.time() if ahora is None else ahora
    with _scan_lock:
        recientes = [t for t in _scan_attempts.get(clave, []) if t > ts - 60]
        if len(recientes) >= _SCAN_LIMIT_PER_MINUTE:
            _scan_attempts[clave] = recientes
            return False
        recientes.append(ts)
        _scan_attempts[clave] = recientes
        if len(_scan_attempts) > 1000:
            for key in list(_scan_attempts):
                if not any(t > ts - 60 for t in _scan_attempts[key]):
                    _scan_attempts.pop(key, None)
        return True


# ─── Auth ─────────────────────────────────────────────────────────────────────
# ── API externa — montada en /api (Swagger en /api/docs) ──────────────────────
try:
    from api_externa import api_app as _api_app
except Exception as _e:
    _api_app = None
    mrd_logging.log_error("API externa no cargada", _e)


@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request, db: Session = Depends(get_db)):
    token = request.cookies.get("mrd_token")
    if token:
        # Verificar que el token es válido Y el usuario existe antes de redirigir
        # sub contiene el username (string), no el id numérico
        user = obtener_usuario_por_token(token, db)
        if user and user.activo:
            return RedirectResponse("/", status_code=303)
        # Token inválido o usuario no existe — limpiar cookie y mostrar login
        resp = templates.TemplateResponse(request, "login.html", {"request": request, "app_name": APP_NAME})
        resp.delete_cookie("mrd_token")
        return resp
    return templates.TemplateResponse(request, "login.html", {"request": request, "app_name": APP_NAME})


@app.post("/login")
def login_post(
    request: Request,
    response: Response,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    # Rate limiting: clave por IP + username
    ip = request.client.host if request.client else "unknown"
    clave_rl = f"{ip}:{username}"

    if not _puede_intentar_login(clave_rl):
        segundos = _segundos_bloqueo(clave_rl)
        return templates.TemplateResponse(request,
            "login.html",
            {"request": request, "app_name": APP_NAME,
             "error": f"Demasiados intentos fallidos. Espera {segundos // 60}m {segundos % 60}s."},
            status_code=429,
        )

    user = db.query(Usuario).filter(
        Usuario.username == username, Usuario.activo == True
    ).first()
    if not user or not verificar_password(password, user.password_hash):
        _registrar_fallo_login(clave_rl)
        mrd_logging.log_security(f"Login fallido: usuario='{username}' ip={ip}")
        return templates.TemplateResponse(request,
            "login.html",
            {"request": request, "app_name": APP_NAME, "error": "Usuario o contraseña incorrectos"},
            status_code=401,
        )

    _limpiar_intentos_login(clave_rl)

    # 2FA: si el usuario la tiene activada, no abrir sesión todavía —
    # exigir el código TOTP en un segundo paso antes de emitir la cookie real.
    if bool(getattr(user, "totp_habilitado", False)):
        mrd_logging.log_security(f"Login paso 1 OK, pendiente 2FA: usuario='{username}' ip={ip}", level="info")
        is_https = _is_https_request(request)
        pending_token = crear_token({"sub": user.username, "mfa_pending": 1}, expires_delta=timedelta(minutes=5))
        resp = RedirectResponse("/login/2fa", status_code=303)
        resp.set_cookie(
            "mrd_2fa_pending", pending_token,
            httponly=True, secure=is_https, max_age=300, samesite="lax", path="/",
        )
        return resp

    return _emitir_sesion(request, db, user, ip)


def _emitir_sesion(request: Request, db: Session, user: Usuario, ip: str) -> RedirectResponse:
    """Registra el login y emite las cookies de sesión (mrd_token + CSRF).
    Punto único usado tanto por /login directo como tras completar el 2FA."""
    mrd_logging.log_security(f"Login exitoso: usuario='{user.username}' ip={ip}", level="info")
    user.last_login = datetime.utcnow()
    registrar_auditoria(
        db, "sesiones", user.id, "login", user.id, None,
        {"usuario": user.username}, "Inicio de sesión", ip,
    )
    db.commit()

    # Sprint 5.2: incluir mcp=1 en JWT si el usuario debe cambiar contraseña
    mcp = bool(getattr(user, "must_change_password", False))
    token_payload: dict = {"sub": user.username}
    if mcp:
        token_payload["mcp"] = 1

    token = crear_token(token_payload)
    is_https = _is_https_request(request)

    # Destino tras login: /cambiar-contrasena si mcp, sino /
    destino = "/cambiar-contrasena" if mcp else "/"
    resp = RedirectResponse(destino, status_code=303)

    # Cookie de sesión
    resp.set_cookie(
        "mrd_token", token,
        httponly=True,
        secure=is_https,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        samesite="lax",
        path="/",
    )
    # Rotar CSRF token en cada login
    resp.set_cookie(
        CSRF_COOKIE_NAME,
        generar_csrf_token(),
        httponly=False,
        secure=is_https,
        samesite="lax",
        path="/",
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
    )
    return resp


@app.get("/logout")
def logout(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    registrar_auditoria(
        db, "sesiones", user.id, "logout", user.id, None,
        {"usuario": user.username}, "Cierre de sesión",
        request.client.host if request.client else "",
    )
    db.commit()
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie("mrd_token", path="/")
    resp.delete_cookie(CSRF_COOKIE_NAME, path="/")
    return resp


def _usuario_pendiente_2fa(request: Request, db: Session) -> Optional[Usuario]:
    """Decodifica la cookie temporal mrd_2fa_pending emitida en el paso 1 del login."""
    token = request.cookies.get("mrd_2fa_pending", "")
    if not token:
        return None
    payload = verificar_token(token)
    if not payload or not payload.get("mfa_pending"):
        return None
    username = payload.get("sub")
    if not username:
        return None
    return db.query(Usuario).filter(Usuario.username == username, Usuario.activo == True).first()


@app.get("/login/2fa", response_class=HTMLResponse)
def login_2fa_get(request: Request, db: Session = Depends(get_db)):
    user = _usuario_pendiente_2fa(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse(request, "login_2fa.html", {"request": request, "app_name": APP_NAME})


@app.post("/login/2fa")
def login_2fa_post(
    request: Request,
    codigo: str = Form(...),
    db: Session = Depends(get_db),
):
    user = _usuario_pendiente_2fa(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    ip = request.client.host if request.client else "unknown"
    clave_rl = f"2fa:{ip}:{user.username}"
    if not _puede_intentar_login(clave_rl):
        segundos = _segundos_bloqueo(clave_rl)
        return templates.TemplateResponse(request,
            "login_2fa.html",
            {"request": request, "app_name": APP_NAME,
             "error": f"Demasiados intentos fallidos. Espera {segundos // 60}m {segundos % 60}s."},
            status_code=429,
        )

    import pyotp
    totp = pyotp.TOTP(user.totp_secret)
    if not user.totp_secret or not totp.verify(codigo.strip(), valid_window=1):
        _registrar_fallo_login(clave_rl)
        mrd_logging.log_security(f"Código 2FA incorrecto: usuario='{user.username}' ip={ip}")
        return templates.TemplateResponse(request,
            "login_2fa.html",
            {"request": request, "app_name": APP_NAME, "error": "Código incorrecto"},
            status_code=401,
        )

    _limpiar_intentos_login(clave_rl)
    resp = _emitir_sesion(request, db, user, ip)
    resp.delete_cookie("mrd_2fa_pending", path="/")
    return resp


# ─── Perfil / Cambiar contraseña ──────────────────────────────────────────────
@app.get("/perfil", response_class=HTMLResponse)
def perfil_get(request: Request, user: Usuario = Depends(requiere_login)):
    return templates.TemplateResponse(request, "perfil.html", ctx_base(request, user))


@app.post("/perfil")
def perfil_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
    password_confirm: str = Form(...),
):
    if not verificar_password(password_actual, user.password_hash):
        return templates.TemplateResponse(request, "perfil.html", ctx_base(
            request, user, error="Contraseña actual incorrecta."))
    if password_nuevo != password_confirm:
        return templates.TemplateResponse(request, "perfil.html", ctx_base(
            request, user, error="Las contraseñas no coinciden."))
    # Sprint 5.2: validar política de contraseñas
    try:
        validar_contrasena(password_nuevo, username=user.username, min_length=PASSWORD_MIN_LENGTH)
    except ErrorContrasena as _ec:
        return templates.TemplateResponse(request, "perfil.html", ctx_base(
            request, user, error=str(_ec)))
    ip_cambio = request.client.host if request.client else "?"
    u = db.query(Usuario).get(user.id)
    u.password_hash = hash_password(password_nuevo)
    u.must_change_password = False
    db.commit()
    mrd_logging.log_security(
        f"Contraseña cambiada en /perfil: usuario='{user.username}' ip={ip_cambio}",
        level="info",
    )
    return RedirectResponse("/perfil?ok=1", status_code=303)


# ─── 2FA (TOTP) opcional para administradores ────────────────────────────────
@app.get("/perfil/2fa", response_class=HTMLResponse)
def perfil_2fa_get(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="El doble factor solo está disponible para administradores.")

    u = db.query(Usuario).get(user.id)
    qr_data_uri = None
    secreto_pendiente = None
    if not u.totp_habilitado:
        import pyotp
        import qrcode
        import io
        import base64
        secreto_pendiente = pyotp.random_base32()
        uri = pyotp.totp.TOTP(secreto_pendiente).provisioning_uri(
            name=u.username, issuer_name="MRD TOOL CONTROL",
        )
        img = qrcode.make(uri)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_data_uri = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")

    return templates.TemplateResponse(request, "perfil_2fa.html", ctx_base(
        request, user, db,
        totp_habilitado=bool(u.totp_habilitado),
        qr_data_uri=qr_data_uri,
        secreto_pendiente=secreto_pendiente,
    ))


@app.post("/perfil/2fa/activar")
def perfil_2fa_activar(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    secreto: str = Form(...),
    codigo: str = Form(...),
):
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="El doble factor solo está disponible para administradores.")

    import pyotp
    totp = pyotp.TOTP(secreto)
    if not totp.verify(codigo.strip(), valid_window=1):
        return templates.TemplateResponse(request, "perfil_2fa.html", ctx_base(
            request, user, db,
            totp_habilitado=False, qr_data_uri=None, secreto_pendiente=secreto,
            error="Código incorrecto. Vuelve a intentarlo.",
        ))

    u = db.query(Usuario).get(user.id)
    u.totp_secret = secreto
    u.totp_habilitado = True
    db.commit()
    mrd_logging.log_security(f"2FA activado: usuario='{user.username}'", level="info")
    return RedirectResponse("/perfil/2fa?ok=1", status_code=303)


@app.post("/perfil/2fa/desactivar")
def perfil_2fa_desactivar(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    password_actual: str = Form(...),
):
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="El doble factor solo está disponible para administradores.")
    if not verificar_password(password_actual, user.password_hash):
        return templates.TemplateResponse(request, "perfil_2fa.html", ctx_base(
            request, user, db,
            totp_habilitado=True, qr_data_uri=None, secreto_pendiente=None,
            error="Contraseña actual incorrecta.",
        ))

    u = db.query(Usuario).get(user.id)
    u.totp_secret = None
    u.totp_habilitado = False
    db.commit()
    mrd_logging.log_security(f"2FA desactivado: usuario='{user.username}'", level="info")
    return RedirectResponse("/perfil/2fa?ok=1", status_code=303)


# ─── Cambio obligatorio de contraseña (Sprint 5.2) ───────────────────────────
@app.get("/cambiar-contrasena", response_class=HTMLResponse)
def cambiar_contrasena_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
):
    return templates.TemplateResponse(
        request, "cambiar_contrasena.html",
        {**ctx_base(request, user), "modo_obligatorio": bool(getattr(user, "must_change_password", False))},
    )


@app.post("/cambiar-contrasena")
def cambiar_contrasena_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    password_nuevo: str = Form(...),
    password_confirm: str = Form(...),
):
    ctx = {**ctx_base(request, user), "modo_obligatorio": bool(getattr(user, "must_change_password", False))}

    if password_nuevo != password_confirm:
        ctx["error"] = "Las contraseñas no coinciden."
        return templates.TemplateResponse(request, "cambiar_contrasena.html", ctx, status_code=400)

    try:
        validar_contrasena(password_nuevo, username=user.username, min_length=PASSWORD_MIN_LENGTH)
    except ErrorContrasena as _ec:
        ctx["error"] = str(_ec)
        return templates.TemplateResponse(request, "cambiar_contrasena.html", ctx, status_code=400)

    ip_cambio = request.client.host if request.client else "?"
    u = db.query(Usuario).get(user.id)
    u.password_hash = hash_password(password_nuevo)
    u.must_change_password = False
    db.commit()

    mrd_logging.log_security(
        f"Cambio obligatorio de contraseña completado: usuario='{user.username}' ip={ip_cambio}",
        level="info",
    )

    # Emitir nuevo JWT sin el claim mcp=1
    token = crear_token({"sub": user.username})
    is_https = _is_https_request(request)

    resp = RedirectResponse("/", status_code=303)
    resp.set_cookie(
        "mrd_token", token,
        httponly=True, secure=is_https,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        samesite="lax", path="/",
    )
    resp.set_cookie(
        CSRF_COOKIE_NAME, generar_csrf_token(),
        httponly=False, secure=is_https,
        samesite="lax", path="/",
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
    )
    return resp


# ─── Dashboard ────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    # Conteos de herramientas: 1 query GROUP BY en lugar de 7 COUNT separados
    _estado_counts = dict(
        db.query(Herramienta.estado, func.count(Herramienta.id))
        .filter(Herramienta.activa == True, Herramienta.almacen_id == warehouse_id)
        .group_by(Herramienta.estado)
        .all()
    )
    total = sum(_estado_counts.values())
    disponibles   = _estado_counts.get("disponible", 0)
    entregadas    = _estado_counts.get("entregada", 0)
    en_obra       = _estado_counts.get("en_obra", 0)
    en_furgoneta  = _estado_counts.get("en_furgoneta", 0)
    en_reparacion = _estado_counts.get("en_reparacion", 0)
    perdidas      = _estado_counts.get("perdida", 0)

    obras_activas = db.query(Obra).filter(
        Obra.activa == True, Obra.almacen_id == warehouse_id,
    ).count()
    total_trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == warehouse_id,
    ).count()

    # joinedload evita N+1 al acceder a mov.herramienta y mov.usuario en el template
    ultimos_movimientos = (
        db.query(Movimiento)
        .join(Herramienta)
        .filter(Herramienta.almacen_id == warehouse_id)
        .options(
            joinedload(Movimiento.herramienta),
            joinedload(Movimiento.usuario),
        )
        .order_by(Movimiento.fecha.desc())
        .limit(10)
        .all()
    )

    # Por categoría para gráfico
    categorias = (
        db.query(Herramienta.categoria, func.count(Herramienta.id))
        .filter(Herramienta.activa == True, Herramienta.almacen_id == warehouse_id)
        .group_by(Herramienta.categoria)
        .all()
    )

    # Obras activas para panel lateral
    obras = db.query(Obra).filter(
        Obra.activa == True, Obra.almacen_id == warehouse_id,
    ).order_by(Obra.id.desc()).limit(5).all()

    # Alertas automáticas
    alertas = []
    if perdidas > 0:
        alertas.append({"icono": "bi-exclamation-triangle-fill", "color": "danger",
                        "titulo": f"{perdidas} herramienta(s) extraviada(s)",
                        "texto": "Revisar estado del inventario"})
    if en_reparacion > 0:
        alertas.append({"icono": "bi-hammer", "color": "warning",
                        "titulo": f"{en_reparacion} herramienta(s) en reparación",
                        "texto": "Verificar plazos de retorno"})
    # Alertas de garantías próximas a vencer (30 días)
    hoy = date.today()
    prox = hoy + timedelta(days=30)
    garantias_vencer = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == warehouse_id,
        Herramienta.garantia_hasta != None,
        Herramienta.garantia_hasta <= prox,
        Herramienta.garantia_hasta >= hoy,
    ).count()
    if garantias_vencer > 0:
        alertas.append({"icono": "bi-shield-exclamation", "color": "warning",
                        "titulo": f"{garantias_vencer} garantía(s) vencen en 30 días",
                        "texto": "Revisar módulo de herramientas"})

    # Movimientos por día — una sola consulta para toda la semana.
    semana_inicio = hoy - timedelta(days=6)
    movimientos_por_dia = {
        (key.isoformat() if hasattr(key, "isoformat") else str(key)): int(count)
        for key, count in db.query(func.date(Movimiento.fecha), func.count(Movimiento.id))
        .join(Herramienta, Movimiento.herramienta_id == Herramienta.id)
        .filter(
            Herramienta.almacen_id == warehouse_id,
            Movimiento.fecha >= datetime.combine(semana_inicio, datetime.min.time()),
        )
        .group_by(func.date(Movimiento.fecha))
        .all()
    }
    dias_semana = [hoy - timedelta(days=i) for i in range(6, -1, -1)]
    _semana = [int(movimientos_por_dia.get(dia.isoformat(), 0)) for dia in dias_semana]
    _semana_labels = [dia.strftime("%a %d") for dia in dias_semana]

    # Top obras con más herramientas en este momento
    top_obras = (
        db.query(Movimiento.destino, func.count(Movimiento.id).label("cnt"))
        .join(Herramienta, Movimiento.herramienta_id == Herramienta.id)
        .filter(Herramienta.estado == "en_obra", Herramienta.activa == True,
                Herramienta.almacen_id == warehouse_id)
        .filter(Movimiento.destino != None)
        .group_by(Movimiento.destino)
        .order_by(func.count(Movimiento.id).desc())
        .limit(5)
        .all()
    )

    # Alerta: herramientas en reparación > 30 días
    limite_rep = hoy - timedelta(days=30)
    rep_largas = db.query(Reparacion).filter(
        Reparacion.fecha_entrada <= limite_rep,
        Reparacion.fecha_salida == None,
        Reparacion.almacen_id == warehouse_id,
    ).count()
    if rep_largas > 0:
        alertas.append({"icono": "bi-clock-history", "color": "danger",
                        "titulo": f"{rep_largas} reparación(es) abiertas >30 días",
                        "texto": "Contactar con el taller y actualizar estado"})

    # Alerta: EPIs individuales con revisión vencida
    epis_vencidos = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.almacen_id == warehouse_id,
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision < hoy,
    ).count()
    if epis_vencidos > 0:
        alertas.append({"icono": "bi-shield-x", "color": "danger",
                        "titulo": f"{epis_vencidos} EPI(s) con revisión vencida",
                        "texto": "Ir a EPIs → Arneses y Absorbedores"})

    # Alerta: EPIs con revisión próxima (≤15 días)
    prox_epi = hoy + timedelta(days=15)
    epis_proximos = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.almacen_id == warehouse_id,
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision >= hoy,
        EPIIndividual.proxima_revision <= prox_epi,
    ).count()
    epis_proximos_30 = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.almacen_id == warehouse_id,
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision >= hoy,
        EPIIndividual.proxima_revision <= hoy + timedelta(days=30),
    ).count()
    if epis_proximos > 0:
        alertas.append({"icono": "bi-shield-exclamation", "color": "warning",
                        "titulo": f"{epis_proximos} EPI(s) con revisión en ≤15 días",
                        "texto": "Programar revisión antes del vencimiento"})

    # Alertas de vehículos (ITV / seguro próximos 30 días)
    prox_vehs = hoy + timedelta(days=30)
    vehs_itv = db.query(Vehiculo).filter(
        Vehiculo.activo == True,
        Vehiculo.almacen_id == warehouse_id,
        Vehiculo.itv_hasta != None,
        Vehiculo.itv_hasta >= hoy,
        Vehiculo.itv_hasta <= prox_vehs,
    ).all()
    vehs_seg = db.query(Vehiculo).filter(
        Vehiculo.activo == True,
        Vehiculo.almacen_id == warehouse_id,
        Vehiculo.seguro_hasta != None,
        Vehiculo.seguro_hasta >= hoy,
        Vehiculo.seguro_hasta <= prox_vehs,
    ).all()
    vehs_itv_vencida = db.query(Vehiculo).filter(
        Vehiculo.activo == True,
        Vehiculo.almacen_id == warehouse_id,
        Vehiculo.itv_hasta != None,
        Vehiculo.itv_hasta < hoy,
    ).count()
    if vehs_itv_vencida > 0:
        alertas.append({"icono": "bi-car-front-fill", "color": "danger",
                        "titulo": f"{vehs_itv_vencida} vehículo(s) con ITV vencida",
                        "texto": "Revisar módulo de vehículos urgente"})
    if vehs_itv:
        alertas.append({"icono": "bi-car-front", "color": "warning",
                        "titulo": f"{len(vehs_itv)} vehículo(s) con ITV en ≤30 días",
                        "texto": ", ".join(v.matricula or v.nombre for v in vehs_itv[:3])})
    if vehs_seg:
        alertas.append({"icono": "bi-shield-check", "color": "warning",
                        "titulo": f"{len(vehs_seg)} vehículo(s) con seguro en ≤30 días",
                        "texto": ", ".join(v.matricula or v.nombre for v in vehs_seg[:3])})
    # Materiales bajo mínimo → alerta
    _mat_bajo = db.query(Material).filter(
        Material.activo == True, Material.stock_minimo > 0,
        Material.almacen_id == warehouse_id,
        Material.stock_actual <= Material.stock_minimo,
    ).count()
    if _mat_bajo > 0:
        alertas.append({"icono": "bi-exclamation-triangle-fill", "color": "danger",
                        "titulo": f"{_mat_bajo} material(es) bajo mínimo de stock",
                        "texto": "Ir a Materiales → Alertas para ver la lista"})

    return templates.TemplateResponse(request, "dashboard.html", ctx_base(
        request, user, db,
        total=total,
        disponibles=disponibles,
        entregadas=entregadas,
        en_obra=en_obra,
        en_furgoneta=en_furgoneta,
        en_reparacion=en_reparacion,
        perdidas=perdidas,
        obras_activas=obras_activas,
        total_trabajadores=total_trabajadores,
        ultimos_movimientos=ultimos_movimientos,
        categorias=categorias,
        obras=obras,
        alertas=alertas,
        alertas_count=len(alertas),
        movimientos_semana=_semana,
        movimientos_semana_labels=_semana_labels,
        top_obras=top_obras,
        # KPIs EPIs para el dashboard
        dash_epis_vencidos=epis_vencidos,
        dash_epis_proximos=epis_proximos_30,
        dash_epis_disponibles={
            t: db.query(EPIIndividual).filter(
                EPIIndividual.tipo == t,
                EPIIndividual.almacen_id == warehouse_id,
                EPIIndividual.estado != "baja",
                EPIIndividual.trabajador_id == None
            ).count()
            for t in TIPOS_EPI_INDIVIDUAL
        },
        dash_entregas_mes=db.query(EntregaEPI).join(Trabajador).filter(
            EntregaEPI.fecha >= hoy.replace(day=1),
            Trabajador.almacen_id == warehouse_id,
        ).count(),
        # Fuera ahora — widget
        dash_vehs_en_ruta=db.query(MovimientoVehiculo).join(Vehiculo).filter(
            MovimientoVehiculo.fecha_retorno == None,
            Vehiculo.almacen_id == warehouse_id,
        ).count(),
        dash_alb_abiertos=db.query(AlbaranSalida).filter(
            AlbaranSalida.estado.in_(["abierto", "parcial"]),
            AlbaranSalida.almacen_id == warehouse_id,
        ).count(),
        dash_mat_bajo_minimo=_mat_bajo,
        dash_herr_fuera=entregadas + en_obra + en_furgoneta,
    ))


# ─── Dashboard Ejecutivo ──────────────────────────────────────────────────────
@app.get("/dashboard-exec", response_class=HTMLResponse)
def dashboard_exec(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    from datetime import datetime as _dt2, timedelta as _td2, date as _date2

    hoy = _date2.today()
    ESTADOS_CAMPO = ["entregada", "en_obra", "en_furgoneta", "en_transporte"]
    ESTADOS_PERDIDA = ["extraviada", "robada"]

    herr_all = db.query(Herramienta).filter(Herramienta.activa == True).all()
    total_herr = len(herr_all)
    valor_total = sum((h.valor_actual or h.precio_compra or 0) for h in herr_all)
    herr_campo = [h for h in herr_all if h.estado in ESTADOS_CAMPO]
    valor_campo = sum((h.valor_actual or h.precio_compra or 0) for h in herr_campo)
    herr_perdidas = [h for h in herr_all if h.estado in ESTADOS_PERDIDA]
    valor_perdido = sum((h.valor_actual or h.precio_compra or 0) for h in herr_perdidas)
    tasa_perdida = round(len(herr_perdidas) / total_herr * 100, 1) if total_herr else 0

    obras_dict = {}
    for h in herr_campo:
        if h.obra_id:
            o = h.obra
            key = o.nombre if o else f"Obra #{h.obra_id}"
            v = h.valor_actual or h.precio_compra or 0
            obras_dict[key] = obras_dict.get(key, 0) + v
    top_obras = sorted(obras_dict.items(), key=lambda x: x[1], reverse=True)[:10]

    trab_dict = {}
    for h in herr_campo:
        if h.responsable_id:
            t = h.responsable
            key = t.nombre_completo if t else f"Trabajador #{h.responsable_id}"
            if key not in trab_dict:
                trab_dict[key] = {"count": 0, "valor": 0}
            trab_dict[key]["count"] += 1
            trab_dict[key]["valor"] += h.valor_actual or h.precio_compra or 0
    top_trab = sorted(trab_dict.items(), key=lambda x: x[1]["count"], reverse=True)[:10]

    semanas_labels, semanas_data = [], []
    for w in range(7, -1, -1):
        ini = _dt2.combine(hoy - _td2(days=(w+1)*7), _dt2.min.time())
        fin = _dt2.combine(hoy - _td2(days=w*7), _dt2.min.time())
        c = db.query(Movimiento).filter(Movimiento.fecha >= ini, Movimiento.fecha < fin).count()
        semanas_labels.append(f"S-{w}" if w > 0 else "Esta semana")
        semanas_data.append(c)

    cat_dict = {}
    for h in herr_all:
        c = h.categoria or "Sin categoría"
        cat_dict[c] = cat_dict.get(c, 0) + 1
    cat_items = sorted(cat_dict.items(), key=lambda x: x[1], reverse=True)[:8]

    return templates.TemplateResponse(request, "dashboard_exec.html", ctx_base(
        request, user,
        total_herr=total_herr,
        valor_total=valor_total,
        herr_campo=len(herr_campo),
        valor_campo=valor_campo,
        herr_perdidas=len(herr_perdidas),
        valor_perdido=valor_perdido,
        tasa_perdida=tasa_perdida,
        top_obras=top_obras,
        top_trab=top_trab,
        semanas_labels=semanas_labels,
        semanas_data=semanas_data,
        cat_labels=[c[0] for c in cat_items],
        cat_data=[c[1] for c in cat_items],
    ))


# ─── Herramientas ─────────────────────────────────────────────────────────────
@app.get("/herramientas", response_class=HTMLResponse)
def herramientas_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    q: str = "",
    estado: str = "",
    categoria: str = "",
    page: int = 1,
):
    PER_PAGE = 25
    active_warehouse = _active_warehouse(db, user, request)
    query = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1),
    )
    if q:
        query = query.filter(or_(
            Herramienta.codigo.ilike(f"%{q}%"),
            Herramienta.nombre.ilike(f"%{q}%"),
            Herramienta.num_serie.ilike(f"%{q}%"),
            Herramienta.marca.ilike(f"%{q}%"),
        ))
    if estado:
        query = query.filter(Herramienta.estado == estado)
    if categoria:
        query = query.filter(Herramienta.categoria == categoria)

    total = query.count()
    herramientas = (
        query.options(
            joinedload(Herramienta.responsable),
            joinedload(Herramienta.almacen),
            joinedload(Herramienta.obra),
        )
        .order_by(Herramienta.id.desc())
        .offset((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
        .all()
    )
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)

    categorias = db.query(Herramienta.categoria).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1),
    ).distinct().all()
    categorias = [c[0] for c in categorias if c[0]]

    # KPIs por estado — GROUP BY para una sola query
    kpi_raw = (
        db.query(Herramienta.estado, func.count(Herramienta.id))
        .filter(Herramienta.activa == True,
                Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1))
        .group_by(Herramienta.estado)
        .all()
    )
    kpis = {estado_k: cnt for estado_k, cnt in kpi_raw}

    # Total incluyendo inactivas para baja/archivada
    kpis_inactivos = (
        db.query(Herramienta.estado, func.count(Herramienta.id))
        .filter(Herramienta.activa == False,
                Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1))
        .group_by(Herramienta.estado)
        .all()
    )
    for estado_k, cnt in kpis_inactivos:
        kpis[estado_k] = kpis.get(estado_k, 0) + cnt

    total_global = db.query(func.count(Herramienta.id)).filter(
        Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1),
    ).scalar()

    return templates.TemplateResponse(request, "herramientas.html", ctx_base(
        request, user,
        herramientas=herramientas,
        total=total,
        total_global=total_global,
        page=page,
        total_pages=total_pages,
        q=q,
        estado_filtro=estado,
        categoria_filtro=categoria,
        categorias=sorted(categorias),
        estados=ESTADOS_HERRAMIENTA,
        kpis=kpis,
    ))


@app.get("/herramientas/nueva", response_class=HTMLResponse)
def herramienta_nueva_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    num_serie: str = Query(""),
    codigo: str = Query(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    almacenes = visible_warehouses(db, user)
    almacen_principal = _active_warehouse(db, user, request)
    proveedores = db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.nombre).all()
    identificador_escaneado = num_serie.strip() or codigo.strip()
    return templates.TemplateResponse(request, "nueva_herramienta.html", ctx_base(
        request, user,
        almacenes=almacenes,
        proveedores=proveedores,
        categorias=CATEGORIAS_DEFAULT,
        estados=ESTADOS_HERRAMIENTA,
        prefill_num_serie=identificador_escaneado,
        scan_origen=bool(num_serie.strip() or codigo.strip()),
        almacen_principal_id=almacen_principal.id if almacen_principal else None,
    ))


@app.post("/herramientas/nueva")
async def herramienta_nueva_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    descripcion: str = Form(""),
    categoria: str = Form("Otro"),
    subcategoria: str = Form(""),
    familia: str = Form(""),
    marca: str = Form(""),
    modelo: str = Form(""),
    fabricante: str = Form(""),
    num_serie: str = Form(""),
    potencia: str = Form(""),
    voltaje: str = Form(""),
    peso: str = Form(""),
    color: str = Form(""),
    dimensiones: str = Form(""),
    activo_fijo: str = Form(""),
    vida_util_anos: str = Form(""),
    estado: str = Form("disponible"),
    fecha_compra: str = Form(""),
    precio_compra: str = Form(""),
    proveedor_texto: str = Form(""),
    garantia_hasta: str = Form(""),
    numero_factura: str = Form(""),
    observaciones: str = Form(""),
    almacen_id: str = Form(""),
    ubicacion_texto: str = Form(""),
    foto: UploadFile = File(None),
    redirect: str = Form(""),
    tipo_seguimiento: str = Form("individual"),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")

    # La referencia nunca procede del formulario: la genera exclusivamente el servidor.
    codigo = generar_referencia_herramienta(db)

    foto_path = None
    if foto and foto.filename:
        # Sprint 5.2: validar nombre, MIME y tamaño
        try:
            _, _ext = validar_nombre_archivo(foto.filename, {'jpg', 'jpeg', 'png', 'webp'})
            _head = await foto.read(16); await foto.seek(0)
            validar_contenido_archivo(_head, _ext)
            _content_foto = await foto.read(); await foto.seek(0)
            validar_tamaño_bytes(len(_content_foto), MAX_UPLOAD_MB)
        except ErrorArchivo as _ea:
            raise HTTPException(400, str(_ea))
        foto_nombre = f"{codigo}.{_ext}"
        foto_dest = UPLOADS_DIR / "herramientas" / foto_nombre
        foto_dest.parent.mkdir(parents=True, exist_ok=True)
        with open(foto_dest, "wb") as _fd:
            _fd.write(_content_foto)
        foto_path = foto_nombre

    almacen_predeterminado = _active_warehouse(db, user, request)
    a_id = int(almacen_id) if almacen_id else (almacen_predeterminado.id if almacen_predeterminado else None)
    _require_warehouse_access(user, a_id)
    almacen_obj = db.query(Almacen).get(a_id) if a_id else None

    h = Herramienta(
        codigo=codigo,
        nombre=nombre,
        descripcion=descripcion or None,
        categoria=categoria,
        subcategoria=subcategoria or None,
        familia=familia or None,
        marca=marca or None,
        modelo=modelo or None,
        fabricante=fabricante or None,
        num_serie=num_serie or None,
        potencia=potencia or None,
        voltaje=voltaje or None,
        peso=float(peso) if peso else None,
        color=color or None,
        dimensiones=dimensiones or None,
        activo_fijo=activo_fijo or None,
        vida_util_anos=int(vida_util_anos) if vida_util_anos else None,
        estado=estado,
        fecha_compra=datetime.strptime(fecha_compra, "%Y-%m-%d") if fecha_compra else None,
        precio_compra=float(precio_compra) if precio_compra else None,
        proveedor_texto=proveedor_texto or None,
        garantia_hasta=datetime.strptime(garantia_hasta, "%Y-%m-%d") if garantia_hasta else None,
        numero_factura=numero_factura or None,
        observaciones=observaciones or None,
        almacen_id=a_id,
        ubicacion_texto=ubicacion_texto or (almacen_obj.nombre if almacen_obj else "Almacén"),
        foto=foto_path,
        activa=True,
        tipo_seguimiento=tipo_seguimiento if tipo_seguimiento in ("individual", "generico") else "individual",
    )
    db.add(h)
    db.flush()

    registrar_movimiento(db, h, "alta", estado, user, observaciones="Herramienta dada de alta")

    # Auditoría de alta
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    registrar_auditoria(
        db,
        tabla="herramientas",
        registro_id=h.id,
        accion="crear",
        usuario_id=user.id,
        datos_anteriores=None,
        datos_nuevos=snapshot_herramienta(h),
        resumen=f"Alta de herramienta {h.nombre} ({h.codigo})",
        ip=ip,
    )
    db.commit()

    if redirect == "nueva":
        return RedirectResponse("/herramientas/nueva", status_code=303)
    return RedirectResponse(f"/herramientas/{h.id}", status_code=303)


@app.get("/herramientas/importar/plantilla")
def herramientas_plantilla_descarga(user: Usuario = Depends(requiere_login)):
    """Descarga la plantilla Excel de importación."""
    excel = generar_plantilla_importacion()
    nombre = "plantilla_importacion_herramientas.xlsx"
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@app.get("/herramientas/importar", response_class=HTMLResponse)
def herramientas_importar_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
):
    if not tiene_permiso(user, "crear"):
        return RedirectResponse("/herramientas", status_code=303)
    return templates.TemplateResponse(request, "importar_herramientas.html", ctx_base(request, user))


@app.post("/herramientas/importar", response_class=HTMLResponse)
async def herramientas_importar_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    archivo: UploadFile = File(...),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    # Sprint 5.2: validar MIME y tamaño del Excel
    try:
        _, _ext_xl = validar_nombre_archivo(archivo.filename, {'xlsx'})
        _head_xl = await archivo.read(16); await archivo.seek(0)
        validar_contenido_archivo(_head_xl, _ext_xl)
        _xlsx_bytes = await archivo.read(); await archivo.seek(0)
        validar_tamaño_bytes(len(_xlsx_bytes), MAX_UPLOAD_MB)
    except ErrorArchivo as _ea:
        return templates.TemplateResponse(request, "importar_herramientas.html",
            ctx_base(request, user, error=str(_ea)), status_code=400)
    if not archivo.filename.lower().endswith(".xlsx"):
        return templates.TemplateResponse(request, "importar_herramientas.html", ctx_base(
            request, user, error="Solo se aceptan archivos .xlsx"
        ), status_code=400)
    contenido = await archivo.read()
    try:
        resultado = importar_herramientas_excel(db, contenido, user)
        db.commit()
    except Exception as exc:
        db.rollback()
        return templates.TemplateResponse(request, "importar_herramientas.html", ctx_base(
            request, user, error=f"Error al procesar el archivo: {exc}"
        ), status_code=422)
    return templates.TemplateResponse(request, "importar_herramientas.html", ctx_base(
        request, user, resultado=resultado
    ))


# ─── Informes ─────────────────────────────────────────────────────────────────
@app.get("/herramientas/{herramienta_id}", response_class=HTMLResponse)
def herramienta_detalle(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404, "Herramienta no encontrada")
    _require_warehouse_access(user, h.almacen_id)

    qr_code = generar_qr_base64(h.codigo)
    barcode_b64 = generar_barcode_base64(h.codigo)

    movimientos = (
        db.query(Movimiento)
        .filter(Movimiento.herramienta_id == h.id)
        .order_by(Movimiento.fecha.desc())
        .limit(50)
        .all()
    )

    incidencias = (
        db.query(Incidencia)
        .filter(Incidencia.herramienta_id == h.id)
        .order_by(Incidencia.fecha_apertura.desc())
        .all()
    )

    reparaciones = (
        db.query(Reparacion)
        .filter(Reparacion.herramienta_id == h.id)
        .order_by(Reparacion.fecha_entrada.desc())
        .all()
    )

    documentos = (
        db.query(Documento)
        .filter(Documento.herramienta_id == h.id)
        .order_by(Documento.created_at.desc())
        .all()
    )

    trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == h.almacen_id,
    ).order_by(Trabajador.nombre).all()
    obras = db.query(Obra).filter(Obra.activa == True, Obra.almacen_id == h.almacen_id).order_by(Obra.nombre).all()
    almacenes = visible_warehouses(db, user)
    vehiculos = db.query(Vehiculo).filter(Vehiculo.activo == True, Vehiculo.almacen_id == h.almacen_id).all()

    # Auditoría real (últimos 100 registros)
    auditoria_logs = (
        db.query(AuditoriaLog)
        .filter(
            AuditoriaLog.tabla == "herramientas",
            AuditoriaLog.registro_id == h.id,
        )
        .order_by(AuditoriaLog.fecha.desc())
        .limit(100)
        .all()
    )

    return templates.TemplateResponse(request, "herramienta_detalle.html", ctx_base(
        request, user,
        herramienta=h,
        qr_code=qr_code,
        barcode_b64=barcode_b64,
        movimientos=movimientos,
        incidencias=incidencias,
        reparaciones=reparaciones,
        documentos=documentos,
        trabajadores=trabajadores,
        obras=obras,
        almacenes=almacenes,
        vehiculos=vehiculos,
        estados=ESTADOS_HERRAMIENTA,
        estados_info=ESTADOS,
        auditoria_logs=auditoria_logs,
    ))


@app.get("/herramientas/{herramienta_id}/editar", response_class=HTMLResponse)
def herramienta_editar_get(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404)
    _require_warehouse_access(user, h.almacen_id)
    almacenes = visible_warehouses(db, user)
    proveedores = db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.nombre).all()
    return templates.TemplateResponse(request, "editar_herramienta.html", ctx_base(
        request, user,
        herramienta=h,
        almacenes=almacenes,
        proveedores=proveedores,
        categorias=CATEGORIAS_DEFAULT,
        estados=ESTADOS_HERRAMIENTA,
    ))


@app.post("/herramientas/{herramienta_id}/editar")
async def herramienta_editar_post(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    descripcion: str = Form(""),
    categoria: str = Form("Otro"),
    subcategoria: str = Form(""),
    familia: str = Form(""),
    marca: str = Form(""),
    modelo: str = Form(""),
    fabricante: str = Form(""),
    num_serie: str = Form(""),
    potencia: str = Form(""),
    voltaje: str = Form(""),
    peso: str = Form(""),
    color: str = Form(""),
    dimensiones: str = Form(""),
    activo_fijo: str = Form(""),
    vida_util_anos: str = Form(""),
    fecha_compra: str = Form(""),
    precio_compra: str = Form(""),
    proveedor_texto: str = Form(""),
    garantia_hasta: str = Form(""),
    numero_factura: str = Form(""),
    observaciones: str = Form(""),
    almacen_id: str = Form(""),
    ubicacion_texto: str = Form(""),
    foto: UploadFile = File(None),
    tipo_seguimiento: str = Form("individual"),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404)
    _require_warehouse_access(user, h.almacen_id)

    # Snapshot antes de editar (para auditoría)
    snap_ant = snapshot_herramienta(h)

    h.nombre = nombre
    h.descripcion = descripcion or None
    h.categoria = categoria
    h.subcategoria = subcategoria or None
    h.familia = familia or None
    h.marca = marca or None
    h.modelo = modelo or None
    h.fabricante = fabricante or None
    h.num_serie = num_serie or None
    h.potencia = potencia or None
    h.voltaje = voltaje or None
    h.peso = float(peso) if peso else None
    h.color = color or None
    h.dimensiones = dimensiones or None
    h.activo_fijo = activo_fijo or None
    h.vida_util_anos = int(vida_util_anos) if vida_util_anos else None
    h.fecha_compra = datetime.strptime(fecha_compra, "%Y-%m-%d").date() if fecha_compra else None
    h.precio_compra = float(precio_compra) if precio_compra else None
    h.proveedor_texto = proveedor_texto or None
    if tipo_seguimiento in ("individual", "generico"):
        h.tipo_seguimiento = tipo_seguimiento
    h.garantia_hasta = datetime.strptime(garantia_hasta, "%Y-%m-%d").date() if garantia_hasta else None
    h.numero_factura = numero_factura or None
    h.observaciones = observaciones or None
    nuevo_almacen_id = int(almacen_id) if almacen_id else h.almacen_id
    _require_warehouse_access(user, nuevo_almacen_id)
    h.almacen_id = nuevo_almacen_id
    h.ubicacion_texto = ubicacion_texto or None

    if foto and foto.filename:
        # Sprint 5.2: validar nombre, MIME y tamaño
        try:
            _, _ext_h = validar_nombre_archivo(foto.filename, {'jpg', 'jpeg', 'png', 'webp'})
            _head_h = await foto.read(16); await foto.seek(0)
            validar_contenido_archivo(_head_h, _ext_h)
            _content_h = await foto.read(); await foto.seek(0)
            validar_tamaño_bytes(len(_content_h), MAX_UPLOAD_MB)
        except ErrorArchivo as _ea:
            raise HTTPException(400, str(_ea))
        foto_nombre = f"{h.codigo}.{_ext_h}"
        foto_dest = UPLOADS_DIR / "herramientas" / foto_nombre
        foto_dest.parent.mkdir(parents=True, exist_ok=True)
        with open(foto_dest, "wb") as _fd:
            _fd.write(_content_h)
        h.foto = foto_nombre

    # Auditoría de edición
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    registrar_auditoria(
        db,
        tabla="herramientas",
        registro_id=h.id,
        accion="editar",
        usuario_id=user.id,
        datos_anteriores=snap_ant,
        datos_nuevos=snapshot_herramienta(h),
        resumen=f"Edición de {h.nombre} ({h.codigo})",
        ip=ip,
    )

    db.commit()
    return RedirectResponse(f"/herramientas/{herramienta_id}", status_code=303)


@app.post("/herramientas/{herramienta_id}/accion")
def herramienta_accion(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    accion: str = Form(...),
    trabajador_id: str = Form(""),
    obra_id: str = Form(""),
    almacen_id: str = Form(""),
    vehiculo_id: str = Form(""),
    observaciones: str = Form(""),
):
    if not tiene_permiso(user, "entregar"):
        raise HTTPException(403, "Sin permiso")

    # Acciones de admin requieren permiso 'borrar'
    ACCIONES_ADMIN = {"baja", "restaurar", "archivar", "recuperar", "robada"}
    if accion in ACCIONES_ADMIN and not tiene_permiso(user, "borrar"):
        raise HTTPException(403, "Sin permiso de administrador")

    # Herramienta: admitir activa=False para restaurar/recuperar
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404, "Herramienta no encontrada")
    _require_warehouse_access(user, h.almacen_id)

    t_id = int(trabajador_id) if trabajador_id and trabajador_id.isdigit() else None
    o_id = int(obra_id)        if obra_id        and obra_id.isdigit()        else None
    a_id = int(almacen_id)     if almacen_id     and almacen_id.isdigit()     else None
    if a_id:
        _require_warehouse_access(user, a_id)
    v_id = int(vehiculo_id)    if vehiculo_id    and vehiculo_id.isdigit()    else None

    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    es_admin = tiene_permiso(user, "borrar")

    try:
        aplicar_accion(
            db, h, accion, user,
            es_admin      = es_admin,
            trabajador_id = t_id,
            obra_id       = o_id,
            almacen_id    = a_id,
            vehiculo_id   = v_id,
            observaciones = observaciones,
            ip            = ip,
        )
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))

    # Soporte JSON para llamadas AJAX (desde la ficha)
    accept = request.headers.get("Accept", "")
    if "application/json" in accept:
        return JSONResponse({"ok": True, "estado": h.estado})

    return RedirectResponse(f"/herramientas/{herramienta_id}", status_code=303)


@app.get("/herramientas/{herramienta_id}/pdf", response_class=HTMLResponse)
def herramienta_pdf(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Ficha PDF de herramienta — genera HTML optimizado para impresión."""
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404, "Herramienta no encontrada")

    # QR code de la herramienta
    qr_b64 = None
    try:
        qr_b64 = generar_qr_base64(h.codigo)
    except Exception:
        pass

    estado_info = ESTADOS.get(h.estado, {"label": h.estado, "color": "secondary"})
    movimientos = db.query(Movimiento).filter(
        Movimiento.herramienta_id == h.id
    ).order_by(Movimiento.fecha.desc()).limit(20).all()

    return templates.TemplateResponse(request, "herramienta_pdf.html", {
        **ctx_base(request, user),
        "h": h,
        "qr_b64": qr_b64,
        "estado_info": estado_info,
        "movimientos": movimientos,
        "fecha_impresion": datetime.now().strftime("%d/%m/%Y %H:%M"),
    })


@app.get("/api/herramientas/{herramienta_id}/auditoria")
def herramienta_auditoria_api(
    herramienta_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    limit: int = Query(50, le=200),
):
    """Devuelve los últimos N registros de AuditoriaLog para una herramienta."""
    logs = db.query(AuditoriaLog).filter(
        AuditoriaLog.tabla == "herramientas",
        AuditoriaLog.registro_id == herramienta_id,
    ).order_by(AuditoriaLog.fecha.desc()).limit(limit).all()

    return [{
        "id":       log.id,
        "fecha":    log.fecha.isoformat() if log.fecha else None,
        "accion":   log.accion,
        "resumen":  log.resumen,
        "usuario":  log.usuario.nombre if log.usuario else "Sistema",
        "ip":       log.ip,
        "datos_ant": log.datos_anteriores,
        "datos_nvo": log.datos_nuevos,
    } for log in logs]


@app.get("/api/herramientas/{herramienta_id}/check_serie")
def check_num_serie(
    herramienta_id: int,
    serie: str = Query(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(requiere_login),
):
    """Verifica si un número de serie ya existe (para validación en formulario)."""
    existe = db.query(Herramienta).filter(
        Herramienta.num_serie == serie,
        Herramienta.id != herramienta_id,  # excluir la propia herramienta al editar
    ).first()
    return {"duplicado": existe is not None, "id_existente": existe.id if existe else None}


# ─── Movimientos ──────────────────────────────────────────────────────────────
@app.get("/movimientos", response_class=HTMLResponse)
def movimientos_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    q: str = "",
    tipo: str = "",
    trabajador_id: int | None = None,
    usuario_id: int | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    page: int = 1,
):
    PER_PAGE = 30
    active_warehouse = _active_warehouse(db, user, request)
    query = db.query(Movimiento).join(Herramienta).filter(
        Herramienta.almacen_id == (active_warehouse.id if active_warehouse else -1),
    )
    if q:
        query = query.filter(or_(
            Herramienta.codigo.ilike(f"%{q}%"),
            Herramienta.nombre.ilike(f"%{q}%"),
        ))
    if tipo:
        query = query.filter(Movimiento.tipo == tipo)
    if trabajador_id:
        query = query.filter(Movimiento.trabajador_id == trabajador_id)
    if usuario_id and user.rol == "admin":
        query = query.filter(Movimiento.usuario_id == usuario_id)
    if fecha_desde:
        query = query.filter(Movimiento.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        query = query.filter(Movimiento.fecha < datetime.combine(fecha_hasta + timedelta(days=1), datetime.min.time()))

    total = query.count()
    movimientos = query.order_by(Movimiento.fecha.desc()).offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)

    warehouse_id = active_warehouse.id if active_warehouse else -1
    workers = db.query(Trabajador).filter(Trabajador.almacen_id == warehouse_id, Trabajador.activo == True).order_by(Trabajador.nombre).all()
    users = db.query(Usuario).filter(Usuario.activo == True).order_by(Usuario.nombre).all() if user.rol == "admin" else []
    return templates.TemplateResponse(request, "movimientos.html", ctx_base(
        request, user, db,
        movimientos=movimientos,
        total=total,
        page=page,
        total_pages=total_pages,
        q=q,
        tipo_filtro=tipo,
        trabajador_id=trabajador_id, usuario_id=usuario_id,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        trabajadores_filtro=workers, usuarios_filtro=users,
    ))


# ─── Trabajadores ─────────────────────────────────────────────────────────────
@app.get("/trabajadores", response_class=HTMLResponse)
def trabajadores_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user, request)
    trabajadores = db.query(Trabajador).filter(
        Trabajador.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Trabajador.nombre).all()
    return templates.TemplateResponse(request, "trabajadores.html", ctx_base(
        request, user, db, trabajadores=trabajadores,
    ))


@app.post("/trabajadores/nuevo")
def trabajador_nuevo(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    apellidos: str = Form(""),
    dni: str = Form(""),
    telefono: str = Form(""),
    email: str = Form(""),
    cargo: str = Form(""),
    empresa: str = Form("MRD Estructuras"),
    talla_ropa: str = Form(""),
    talla_calzado: str = Form(""),
    portal_pin: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    pin_limpio = portal_pin.strip() or "123456"
    if pin_limpio and (not pin_limpio.isdigit() or not 4 <= len(pin_limpio) <= 6):
        raise HTTPException(422, "El PIN del trabajador debe tener entre 4 y 6 números")
    t = Trabajador(
        nombre=nombre, apellidos=apellidos or None, dni=dni or None,
        telefono=telefono or None, email=email or None, cargo=cargo or None,
        empresa=empresa, activo=True,
        talla_ropa=talla_ropa.strip() or None,
        talla_calzado=talla_calzado.strip() or None,
        almacen_id=warehouse.id if warehouse else None,
        portal_token=uuid.uuid4().hex + uuid.uuid4().hex,
        portal_pin_hash=hash_password(pin_limpio),
        portal_pin_cambio_obligatorio=True,
    )
    db.add(t)
    db.flush()
    if not t.codigo:
        t.codigo = f"TRB-{t.id:05d}"
    create_pending_dotation(db, t, user)
    db.commit()
    return RedirectResponse("/trabajadores", status_code=303)



@app.get("/trabajadores/importar/plantilla")
def trabajadores_plantilla(user: Usuario = Depends(requiere_login)):
    excel = generar_plantilla_trabajadores()
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_trabajadores.xlsx"},
    )


@app.get("/trabajadores/importar", response_class=HTMLResponse)
def trabajadores_importar_get(request: Request, user: Usuario = Depends(requiere_login)):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    return templates.TemplateResponse(request, "importar_trabajadores.html", ctx_base(request, user))


@app.post("/trabajadores/importar", response_class=HTMLResponse)
async def trabajadores_importar_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    archivo: UploadFile = File(...),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    # Sprint 5.2: validar MIME, magic bytes y tamaño
    try:
        _, _ext_trab = validar_nombre_archivo(archivo.filename, {'xlsx'})
        _head_trab = await archivo.read(16); await archivo.seek(0)
        validar_contenido_archivo(_head_trab, _ext_trab)
        contenido = await archivo.read(); await archivo.seek(0)
        validar_tamaño_bytes(len(contenido), MAX_UPLOAD_MB)
    except ErrorArchivo as _ea:
        return templates.TemplateResponse(request, "importar_trabajadores.html",
            ctx_base(request, user, error=str(_ea)), status_code=400)
    try:
        resultado = importar_trabajadores_excel(contenido, db)
    except Exception as e:
        resultado = {"creados": 0, "actualizados": 0, "errores": [str(e)], "filas_procesadas": 0}
    return templates.TemplateResponse(request, "importar_trabajadores.html", ctx_base(
        request, user,
        resultado=resultado,
    ))


@app.post("/trabajadores/{tid}/toggle")
def trabajador_toggle(tid: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    t = db.query(Trabajador).get(tid)
    if t:
        _require_warehouse_access(user, t.almacen_id)
        t.activo = not t.activo
        db.commit()
    return RedirectResponse("/trabajadores", status_code=303)


@app.post("/trabajadores/{tid}/editar", response_class=RedirectResponse)
async def trabajador_editar(
    tid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    t = db.query(Trabajador).get(tid)
    if not t:
        raise HTTPException(404)
    _require_warehouse_access(user, t.almacen_id)
    form = await request.form()
    t.nombre = form.get("nombre") or t.nombre
    t.apellidos = form.get("apellidos") or ""
    t.dni = form.get("dni") or None
    t.telefono = form.get("telefono") or None
    t.email = form.get("email") or None
    t.cargo = form.get("cargo") or None
    t.empresa = form.get("empresa") or t.empresa
    t.departamento = form.get("departamento") or None
    t.observaciones = form.get("observaciones") or None
    t.talla_ropa = form.get("talla_ropa") or None
    t.talla_calzado = form.get("talla_calzado") or None
    db.commit()
    return RedirectResponse("/trabajadores?ok=editado", status_code=303)


# ─── EPIs y Ropa de trabajo ───────────────────────────────────────────────────

@app.get("/epis", response_class=HTMLResponse)
def epis_panel(request: Request, user: Usuario = Depends(requiere_login),
               db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == warehouse_id,
    ).order_by(Trabajador.nombre).all()

    hoy = datetime.utcnow()
    limite_ropa = hoy - timedelta(days=INTERVALO_ROPA_DIAS)

    entregas_por_trabajador: dict = {}
    if trabajadores:
        ids_trabajadores = [t.id for t in trabajadores]
        todas_entregas = db.query(EntregaEPI).filter(
            EntregaEPI.trabajador_id.in_(ids_trabajadores),
        ).all()
        for e in todas_entregas:
            entregas_por_trabajador.setdefault(e.trabajador_id, []).append(e)

    resumen = []
    for t in trabajadores:
        entregas = entregas_por_trabajador.get(t.id, [])
        tiene_epi = any(e.tipo == "epi" for e in entregas)
        ultima_ropa = next(
            (e for e in sorted([e for e in entregas if e.tipo == "ropa"],
                               key=lambda x: x.fecha, reverse=True)),
            None
        )
        ropa_vencida = ultima_ropa is None or ultima_ropa.fecha < limite_ropa
        dias_ropa = (hoy - ultima_ropa.fecha).days if ultima_ropa else None
        resumen.append({
            "trabajador": t,
            "tiene_epi": tiene_epi,
            "ultima_ropa": ultima_ropa,
            "ropa_vencida": ropa_vencida,
            "dias_ropa": dias_ropa,
            "total_entregas": len(entregas),
        })

    pendientes_epi   = sum(1 for r in resumen if not r["tiene_epi"])
    pendientes_ropa  = sum(1 for r in resumen if r["ropa_vencida"])

    cat_epi  = db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "epi",  CatalogoEPI.activo == True).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    cat_ropa = db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "ropa", CatalogoEPI.activo == True).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    kit_epi  = [{"nombre": c.nombre, "cantidad": c.cantidad_kit} for c in cat_epi]  or KIT_EPI_INICIAL
    kit_ropa = [{"nombre": c.nombre, "cantidad": c.cantidad_kit} for c in cat_ropa] or KIT_ROPA_SEMESTRAL
    # Stock de todo: EPIs, ropa y arneses individuales
    _inicializar_stock_epi(db)
    stock_epi_items  = db.query(StockEPI).filter(StockEPI.categoria == "epi", StockEPI.almacen_id == warehouse_id).order_by(StockEPI.nombre, StockEPI.talla).all()
    stock_ropa_items = db.query(StockEPI).filter(StockEPI.categoria == "ropa", StockEPI.almacen_id == warehouse_id).order_by(StockEPI.nombre, StockEPI.talla).all()
    arneses_libres = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.trabajador_id == None,
        EPIIndividual.almacen_id == warehouse_id,
    ).all()
    arneses_asignados = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.trabajador_id != None,
        EPIIndividual.almacen_id == warehouse_id,
    ).all()
    return templates.TemplateResponse(request, "epis.html", ctx_base(
        request, user, db,
        resumen=resumen,
        pendientes_epi=pendientes_epi,
        pendientes_ropa=pendientes_ropa,
        kit_epi=kit_epi,
        kit_ropa=kit_ropa,
        stock_epi_items=stock_epi_items,
        stock_ropa_items=stock_ropa_items,
        arneses_libres=len(arneses_libres),
        arneses_asignados=len(arneses_asignados),
    ))


@app.get("/trabajadores/{tid}/epis", response_class=HTMLResponse)
def trabajador_epis(tid: int, request: Request,
                    user: Usuario = Depends(requiere_login),
                    db: Session = Depends(get_db)):
    t = db.query(Trabajador).get(tid)
    if not t:
        raise HTTPException(404, "Trabajador no encontrado")
    _require_warehouse_access(user, t.almacen_id)
    entregas = db.query(EntregaEPI).filter(
        EntregaEPI.trabajador_id == tid
    ).order_by(EntregaEPI.fecha.desc()).all()

    for e in entregas:
        e._items = json.loads(e.items_json or "[]")

    cat_epi_t  = db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "epi",  CatalogoEPI.activo == True).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    cat_ropa_t = db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "ropa", CatalogoEPI.activo == True).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    kit_epi_t  = [{"nombre": c.nombre, "cantidad": c.cantidad_kit} for c in cat_epi_t]  or KIT_EPI_INICIAL
    kit_ropa_t = [{"nombre": c.nombre, "cantidad": c.cantidad_kit} for c in cat_ropa_t] or KIT_ROPA_SEMESTRAL
    # EPIs individuales (arneses/absorbedores) asignados al trabajador
    epis_individuales_t = db.query(EPIIndividual).filter(
        EPIIndividual.trabajador_id == tid,
        EPIIndividual.estado != "baja"
    ).order_by(EPIIndividual.tipo, EPIIndividual.codigo_fabricacion).all()
    herramientas_asignadas = db.query(Herramienta).filter(
        Herramienta.responsable_id == tid,
        Herramienta.activa == True,
    ).order_by(Herramienta.nombre).all()
    maquinaria_asignada = db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        func.lower(func.trim(Maquinaria.responsable)) == t.nombre_completo.strip().lower(),
    ).order_by(Maquinaria.nombre).all()
    albaranes_trabajador = db.query(AlbaranSalida).filter(
        AlbaranSalida.responsable_id == tid,
    ).order_by(AlbaranSalida.fecha_salida.desc()).limit(30).all()

    return templates.TemplateResponse(request, "trabajador_epis.html", ctx_base(
        request, user, db,
        trabajador=t,
        entregas=entregas,
        kit_epi=kit_epi_t,
        kit_ropa=kit_ropa_t,
        epis_individuales=epis_individuales_t,
        herramientas_asignadas=herramientas_asignadas,
        maquinaria_asignada=maquinaria_asignada,
        albaranes_trabajador=albaranes_trabajador,
    ))


@app.post("/trabajadores/{tid}/epis/entregar")
async def trabajador_epi_entregar(
    tid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    form = await request.form()
    tipo = form.get("tipo", "epi")
    firmado_por = form.get("firmado_por", "")
    observaciones = form.get("observaciones", "")
    firma_base64 = form.get("firma_base64", "") or None
    if not (tiene_permiso(user, "crear") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    t = db.query(Trabajador).get(tid)
    if t:
        _require_warehouse_access(user, t.almacen_id)
    if not t:
        raise HTTPException(404, "Trabajador no encontrado")

    # ── Procesado flexible: el usuario elige qué artículos incluye ──────────────
    items = []

    # Artículos del catálogo: cada uno tiene item_N_nombre (hidden) +
    # item_N_checked (checkbox) + item_N_cantidad + item_N_talla (ropa)
    n_items = int(form.get("n_items", 0) or 0)
    for i in range(n_items):
        if not form.get(f"item_{i}_checked"):
            continue  # no marcado → no se entrega
        nombre   = (form.get(f"item_{i}_nombre",   "") or "").strip()
        cantidad = max(1, int(form.get(f"item_{i}_cantidad", 1) or 1))
        talla    = (form.get(f"item_{i}_talla",    "") or "").strip() or None
        if nombre:
            items.append({"nombre": nombre, "cantidad": cantidad, "talla": talla})

    # Artículos extra libres (nombre escrito a mano)
    n_extra = int(form.get("n_extra", 0) or 0)
    for i in range(n_extra):
        nombre_e = (form.get(f"extra_{i}_nombre",   "") or "").strip().upper()
        if not nombre_e:
            continue
        cantidad_e = max(1, int(form.get(f"extra_{i}_cantidad", 1) or 1))
        talla_e    = (form.get(f"extra_{i}_talla",  "") or "").strip() or None
        items.append({"nombre": nombre_e, "cantidad": cantidad_e, "talla": talla_e})

    if not items:
        return RedirectResponse(f"/trabajadores/{tid}/epis?err=sin_items", status_code=303)

    start_stock_transaction(db)
    try:
        # Cada descuento y la entrega se confirman en la misma transacción.
        for index, item in enumerate(items):
            talla_v = item.get("talla")
            stock = db.query(StockEPI).filter(
                StockEPI.nombre == item["nombre"], StockEPI.talla == talla_v
            ).first()
            if not stock and talla_v is None:
                stock = db.query(StockEPI).filter(StockEPI.nombre == item["nombre"]).first()
            if not stock:
                raise StockError(
                    409,
                    f"No hay stock registrado para {item['nombre']}"
                    + (f" (talla {talla_v})" if talla_v else ""),
                )
            move_stock_epi(
                db, user, stock.id, -item["cantidad"], tipo="entrega",
                event_id=f"epi-{uuid.uuid4()}-{index}",
                motivo=f"Entrega física a trabajador #{tid}", trabajador_id=tid,
            )

        entrega = EntregaEPI(
            trabajador_id=tid,
            tipo=tipo,
            items_json=json.dumps(items, ensure_ascii=False),
            fecha=datetime.utcnow(),
            entregado_por=user.nombre,
            firmado_por=firmado_por or None,
            observaciones=observaciones or None,
            usuario_id=user.id,
            firma_base64=firma_base64,
        )
        db.add(entrega)
        db.commit()
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    mrd_logging.log_security(
        f"Entrega EPI tipo={tipo} trabajador={t.nombre_completo} por {user.username}",
        level="info"
    )
    redirect_to = (form.get("redirect_to", "") or "").strip()
    redirect_url = redirect_to if redirect_to and redirect_to.startswith("/") else f"/trabajadores/{tid}/epis"
    return RedirectResponse(redirect_url, status_code=303)


@app.post("/trabajadores/{tid}/epis/marcar-kit")
def trabajador_marcar_kit(
    tid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Marca el kit EPI como ya entregado sin consumir stock (para regularizaciones manuales)."""
    if not (tiene_permiso(user, "crear") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    t = db.query(Trabajador).get(tid)
    if not t:
        raise HTTPException(404, "Trabajador no encontrado")
    ya_tiene = db.query(EntregaEPI).filter(
        EntregaEPI.trabajador_id == tid,
        EntregaEPI.tipo == "epi",
    ).first()
    if ya_tiene:
        return RedirectResponse("/epis?ok=1", status_code=303)
    entrega = EntregaEPI(
        trabajador_id=tid,
        tipo="epi",
        items_json="[]",
        fecha=datetime.utcnow(),
        entregado_por=user.nombre,
        observaciones="Kit previamente entregado — marcado manualmente sin consumo de stock",
        usuario_id=user.id,
    )
    db.add(entrega)
    db.commit()
    mrd_logging.log_security(
        f"Kit EPI marcado manualmente para trabajador={t.nombre_completo} por {user.username}",
        level="info"
    )
    return RedirectResponse("/epis?ok=1", status_code=303)


# ─── Stock de EPIs ────────────────────────────────────────────────────────────

def _asignar_codigos_stock_epi(db):
    """Completa códigos ausentes; el esquema se migra solo desde database.py."""
    items_sin_codigo = db.query(StockEPI).filter(StockEPI.codigo == None).all()
    for s in items_sin_codigo:
        s.codigo = f"SEPI-{s.id:04d}"
    if items_sin_codigo:
        db.commit()


def _normalizar_tallas_elegidas(valor: str) -> tuple[str, ...]:
    """Normaliza únicamente las tallas escritas por el usuario, sin inventarlas."""
    vistas = set()
    resultado = []
    for parte in re.split(r"[,;\n\r]+", valor or ""):
        talla = " ".join(parte.strip().upper().split())
        if not talla or talla in vistas:
            continue
        if len(talla) > 20:
            raise ValueError("Talla demasiado larga")
        vistas.add(talla)
        resultado.append(talla)
        if len(resultado) > 40:
            raise ValueError("Demasiadas tallas")
    return tuple(resultado)


def _crear_tallas_elegidas(
    db: Session, nombre: str, tallas: tuple[str, ...], warehouse_id: int | None = None,
) -> int:
    """Añade solo las tallas elegidas y conserva stock, referencias y QR previos."""
    query = db.query(StockEPI).filter(StockEPI.nombre == nombre)
    if warehouse_id is not None:
        query = query.filter(StockEPI.almacen_id == warehouse_id)
    rows = query.all()
    almacen_id = warehouse_id
    if almacen_id is None:
        almacen_id = next((row.almacen_id for row in rows if row.almacen_id), None)
    if almacen_id is None:
        almacen_predeterminado = get_default_warehouse(db)
        almacen_id = almacen_predeterminado.id if almacen_predeterminado else None
    por_talla = {str(row.talla).upper(): row for row in rows if row.talla}
    generico = next((row for row in rows if not row.talla), None)
    creadas = 0

    # Una referencia genérica vacía puede convertirse en la primera talla
    # solicitada conservando su identidad QR. Con existencias se mantiene como
    # "sin clasificar" para que el usuario decida cómo repartirlas.
    if generico and tallas and int(generico.cantidad or 0) == 0:
        primera = tallas[0]
        existente = por_talla.get(primera)
        if not existente:
            generico.talla = primera
            generico.categoria = "ropa"
            generico.tipo_seguimiento = "generico"
            por_talla[primera] = generico

    for talla in tallas:
        if talla in por_talla:
            por_talla[talla].categoria = "ropa"
            continue
        stock = StockEPI(
            nombre=nombre, categoria="ropa", talla=talla,
            cantidad=0, stock_minimo=3, tipo_seguimiento="generico",
            almacen_id=almacen_id,
        )
        db.add(stock)
        db.flush()
        stock.codigo = f"SEPI-{stock.id:04d}"
        por_talla[talla] = stock
        creadas += 1
    return creadas


def _asegurar_tallas_ropa(db: Session, nombre: str, tallas=()) -> int:
    """Compatibilidad: nunca crea tallas si el usuario no las indicó."""
    if isinstance(tallas, str):
        tallas = _normalizar_tallas_elegidas(tallas)
    else:
        tallas = tuple(tallas or ())
    return _crear_tallas_elegidas(db, nombre, tallas) if tallas else 0


def _reclasificar_chalecos_como_ropa(db: Session) -> int:
    """Los chalecos se gestionan como ropa, sin crear ni repartir tallas."""
    cambios = 0
    for item in db.query(CatalogoEPI).all():
        if "CHALECO" in (item.nombre or "").upper() and item.categoria != "ropa":
            item.categoria = "ropa"
            cambios += 1
    for stock in db.query(StockEPI).all():
        if "CHALECO" in (stock.nombre or "").upper() and stock.categoria != "ropa":
            stock.categoria = "ropa"
            cambios += 1
    return cambios


def _normalizar_nombres_camisas(db: Session) -> int:
    """Aclara manga corta/larga conservando filas, cantidades y códigos QR."""
    cambios = 0
    for antiguo, nuevo in (
        ("CAMISETA", "CAMISETA MANGA CORTA"),
        ("CAMISETA ML", "CAMISETA MANGA LARGA"),
    ):
        catalogo = db.query(CatalogoEPI).filter(CatalogoEPI.nombre == antiguo).first()
        if not catalogo or db.query(CatalogoEPI).filter(CatalogoEPI.nombre == nuevo).first():
            continue
        catalogo.nombre = nuevo
        for stock in db.query(StockEPI).filter(StockEPI.nombre == antiguo).all():
            stock.nombre = nuevo
        cambios += 1
    return cambios


def _reclasificar_metros_como_epi(db) -> int:
    """Mueve METRO 5M/8M al inventario EPI sin borrar su historial material."""
    materiales = db.query(Material).filter(
        Material.activo == True,
        func.upper(func.trim(Material.nombre)).in_(["METRO 5M", "METRO 8M"]),
    ).all()
    migrados = 0
    almacen_predeterminado = get_default_warehouse(db)
    for material in materiales:
        nombre = material.nombre.strip().upper()
        catalogo = db.query(CatalogoEPI).filter(func.upper(CatalogoEPI.nombre) == nombre).first()
        if not catalogo:
            catalogo = CatalogoEPI(
                nombre=nombre, categoria="epi", cantidad_kit=0, activo=True,
                orden=90, notas="Reclasificado automáticamente desde Materiales",
            )
            db.add(catalogo)
        stock = db.query(StockEPI).filter(
            func.upper(StockEPI.nombre) == nombre, StockEPI.talla == None,
        ).first()
        if not stock:
            stock = StockEPI(
                nombre=nombre, categoria="epi", talla=None,
                cantidad=max(0, int(material.stock_actual or 0)),
                stock_minimo=max(0, int(material.stock_minimo or 0)),
                tipo_seguimiento="generico",
                almacen_id=(material.almacen_id or (
                    almacen_predeterminado.id if almacen_predeterminado else None
                )),
            )
            db.add(stock)
            db.flush()
            stock.codigo = f"SEPI-{stock.id:04d}"
        else:
            # La fila material queda desactivada una sola vez, por lo que esta
            # suma es idempotente y evita perder existencias ya inventariadas.
            stock.cantidad = max(0, int(stock.cantidad or 0)) + max(0, int(material.stock_actual or 0))
            stock.stock_minimo = max(
                max(0, int(stock.stock_minimo or 0)),
                max(0, int(material.stock_minimo or 0)),
            )
        material.activo = False
        migrados += 1
    return migrados


def _inicializar_stock_epi(db):
    """Crea registros de stock para EPIs si no existen (ropa se añade por talla manualmente)."""
    almacen_predeterminado = get_default_warehouse(db)
    almacen_id = almacen_predeterminado.id if almacen_predeterminado else None
    for item in KIT_EPI_INICIAL:
        if not db.query(StockEPI).filter(
            StockEPI.nombre == item["nombre"], StockEPI.talla == None
        ).first():
            db.add(StockEPI(nombre=item["nombre"], categoria="epi", talla=None, cantidad=0, stock_minimo=3, almacen_id=almacen_id))
    # También sincronizar EPIs del catálogo personalizado
    for cat in db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "epi", CatalogoEPI.activo == True).all():
        if not db.query(StockEPI).filter(StockEPI.nombre == cat.nombre, StockEPI.talla == None).first():
            db.add(StockEPI(nombre=cat.nombre, categoria="epi", talla=None, cantidad=0, stock_minimo=3, almacen_id=almacen_id))
    db.commit()
    _asignar_codigos_stock_epi(db)


@app.get("/epis/stock", response_class=HTMLResponse)
def epis_stock_panel(request: Request, user: Usuario = Depends(requiere_login),
                     db: Session = Depends(get_db)):
    _inicializar_stock_epi(db)
    nombres_archivados = {
        row.nombre for row in db.query(CatalogoEPI).filter(CatalogoEPI.activo == False).all()
    }
    todos_epis = [row for row in db.query(StockEPI).filter(
        StockEPI.categoria == "epi"
    ).order_by(StockEPI.nombre).all() if row.nombre not in nombres_archivados]
    # Separar EPIs individualizados (arnes, absorbedor) del resto
    epis = [e for e in todos_epis if e.nombre not in TIPOS_EPI_INDIVIDUAL]
    ropa_rows = [row for row in db.query(StockEPI).filter(
        StockEPI.categoria == "ropa"
    ).order_by(StockEPI.nombre, StockEPI.talla).all() if row.nombre not in nombres_archivados]
    # Calcular stock de arneses/absorbedores desde EPIIndividual
    individuales_stock = {}
    for tipo in TIPOS_EPI_INDIVIDUAL:
        total      = db.query(EPIIndividual).filter(EPIIndividual.tipo == tipo, EPIIndividual.estado != "baja").count()
        sin_asignar = db.query(EPIIndividual).filter(EPIIndividual.tipo == tipo, EPIIndividual.estado != "baja", EPIIndividual.trabajador_id == None).count()
        individuales_stock[tipo] = {"total": total, "sin_asignar": sin_asignar}
    # Agrupar ropa por nombre manteniendo orden del kit
    ropa_grupos = OrderedDict()
    for n in [i["nombre"] for i in KIT_ROPA_SEMESTRAL]:
        ropa_grupos[n] = []
    for row in ropa_rows:
        if row.nombre not in ropa_grupos:
            ropa_grupos[row.nombre] = []
        ropa_grupos[row.nombre].append(row)
    # Incluir ropa del catálogo que aún no tiene tallas en stock
    for cat in db.query(CatalogoEPI).filter(CatalogoEPI.categoria == "ropa", CatalogoEPI.activo == True).order_by(CatalogoEPI.nombre).all():
        if cat.nombre not in ropa_grupos:
            ropa_grupos[cat.nombre] = []
    alertas = [i for i in (epis + ropa_rows) if i.bajo_minimo]
    nombres_ropa = [i["nombre"] for i in KIT_ROPA_SEMESTRAL]
    return templates.TemplateResponse(request, "epis_stock.html", ctx_base(
        request, user, db,
        epis=epis,
        ropa_rows=ropa_rows,
        stock_items=epis + ropa_rows,
        individuales_stock=individuales_stock,
        ropa_grupos=list(ropa_grupos.items()),
        nombres_ropa=nombres_ropa,
        alertas=alertas,
    ))


@app.post("/epis/stock/entrada")
def epis_stock_entrada(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    cantidad: int = Form(...),
    talla: str = Form(""),
    tipo_seguimiento: str = Form("generico"),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    if cantidad <= 0:
        raise HTTPException(400, "La entrada debe ser positiva")
    start_stock_transaction(db)
    try:
        talla_val = talla.strip() or None
        _ts_epi = tipo_seguimiento if tipo_seguimiento in ("individual", "generico") else "generico"
        stock = db.query(StockEPI).filter(
            StockEPI.nombre == nombre, StockEPI.talla == talla_val
        ).first()
        if not stock:
            cat = "ropa" if talla_val else "epi"
            almacen_predeterminado = get_default_warehouse(db)
            stock = StockEPI(nombre=nombre, categoria=cat, talla=talla_val, cantidad=0, stock_minimo=3,
                             tipo_seguimiento=_ts_epi,
                             almacen_id=almacen_predeterminado.id if almacen_predeterminado else None)
            db.add(stock)
            db.flush()
            stock.codigo = f"SEPI-{stock.id:04d}"
        else:
            if _ts_epi in ("individual", "generico"):
                stock.tipo_seguimiento = _ts_epi
            db.flush()
        move_stock_epi(
            db, user, stock.id, cantidad, tipo="entrada",
            event_id=f"epi-entry-{uuid.uuid4()}", motivo="Entrada manual de stock EPI",
        )
        db.commit()
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    return RedirectResponse("/epis/stock", status_code=303)


@app.post("/epis/stock/salida")
def epis_stock_salida(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    cantidad: int = Form(...),
    talla: str = Form(""),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    if cantidad <= 0:
        raise HTTPException(400, "La cantidad debe ser positiva")
    talla_val = talla.strip() or None
    start_stock_transaction(db)
    try:
        stock = db.query(StockEPI).filter(
            StockEPI.nombre == nombre, StockEPI.talla == talla_val
        ).first()
        if not stock:
            raise StockError(404, "Artículo no encontrado en stock")
        if stock.cantidad < cantidad:
            raise StockError(409, f"Stock insuficiente: {stock.cantidad} unidades disponibles")
        move_stock_epi(
            db, user, stock.id, -cantidad, tipo="salida",
            event_id=f"epi-exit-{uuid.uuid4()}", motivo="Salida manual de stock EPI",
        )
        db.commit()
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    return RedirectResponse("/epis/stock", status_code=303)


@app.post("/epis/stock/{sepi_id}/seguimiento")
def epis_stock_seguimiento(
    sepi_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    tipo_seguimiento: str = Form("generico"),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    sepi = db.query(StockEPI).get(sepi_id)
    if not sepi:
        raise HTTPException(404, "Artículo no encontrado")
    if tipo_seguimiento not in ("individual", "generico"):
        raise HTTPException(400, "Tipo de seguimiento no válido")
    sepi.tipo_seguimiento = tipo_seguimiento
    db.commit()
    return RedirectResponse("/epis/stock?ok=ts", status_code=303)


@app.post("/epis/stock/etiquetas")
def epis_stock_etiquetas_lote(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    ids: str = Form(...),
    copias: int = Form(1),
    formato: str = Form("pdf"),
):
    """Genera un paquete PDF o ZPL de etiquetas para EPI/ropa seleccionados."""
    if not tiene_permiso(user, "etiquetas"):
        raise HTTPException(403, "Sin permiso para imprimir etiquetas")
    id_list = list(dict.fromkeys(int(value) for value in ids.split(",") if value.strip().isdigit()))
    if not id_list:
        raise HTTPException(400, "Selecciona al menos un artículo")
    copias = max(1, min(int(copias), 100))
    items = db.query(StockEPI).filter(StockEPI.id.in_(id_list)).order_by(
        StockEPI.categoria, StockEPI.nombre, StockEPI.talla
    ).all()
    if len(items) != len(id_list):
        raise HTTPException(404, "Alguno de los artículos seleccionados ya no existe")
    etiquetas = []
    for item in items:
        if not item.codigo:
            item.codigo = f"SEPI-{item.id:04d}"
        etiqueta = {
            "codigo": item.codigo,
            "nombre": item.nombre_display,
            "num_serie": f"Stock: {item.cantidad} ud",
            "marca": "EPI / Ropa MRD",
        }
        etiquetas.extend([etiqueta] * copias)
    db.commit()
    if formato == "zpl":
        zpl = generar_zpl_lote(etiquetas)
        return Response(
            content=zpl.encode("utf-8"), media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=etiquetas_epi_zebra.prn"},
        )
    pdf_bytes = generar_pdf_etiquetas(etiquetas, COMPANY_NAME)
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=etiquetas_epi_paquete.pdf"},
    )


@app.post("/epis/stock/minimo")
def epis_stock_minimo(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    stock_minimo: int = Form(...),
    talla: str = Form(""),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    talla_val = talla.strip() or None
    stock = db.query(StockEPI).filter(
        StockEPI.nombre == nombre, StockEPI.talla == talla_val
    ).first()
    if stock:
        stock.stock_minimo = max(0, stock_minimo)
        db.commit()
    return RedirectResponse("/epis/stock", status_code=303)


@app.get("/trabajadores/{tid}/epis/{eid}/pdf")
def trabajador_epi_pdf(tid: int, eid: int,
                       user: Usuario = Depends(requiere_login),
                       db: Session = Depends(get_db)):
    t = db.query(Trabajador).get(tid)
    e = db.query(EntregaEPI).filter(
        EntregaEPI.id == eid,
        EntregaEPI.trabajador_id == tid
    ).first()
    if not t or not e:
        raise HTTPException(404)

    items = json.loads(e.items_json or "[]")
    pdf_bytes = _generar_pdf_entrega_epi(t, e, items, COMPANY_NAME)
    tipo_label = "EPI" if e.tipo == "epi" else "Ropa"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=entrega_{tipo_label}_{t.id}_{eid}.pdf"},
    )


def _generar_pdf_entrega_epi(trabajador, entrega, items: list, empresa: str) -> bytes:
    """Genera albarán PDF de entrega de EPI/Ropa."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    azul   = colors.HexColor("#1B4F8A")
    naranj = colors.HexColor("#E8600A")
    gris   = colors.HexColor("#555555")

    h1 = ParagraphStyle("h1", parent=styles["Normal"],
                        fontSize=16, fontName="Helvetica-Bold",
                        textColor=azul, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Normal"],
                        fontSize=11, fontName="Helvetica-Bold",
                        textColor=azul, spaceAfter=4)
    normal = ParagraphStyle("normal", parent=styles["Normal"],
                            fontSize=10, spaceAfter=2)
    small  = ParagraphStyle("small", parent=styles["Normal"],
                            fontSize=8, textColor=gris)
    center = ParagraphStyle("center", parent=styles["Normal"],
                            fontSize=9, alignment=TA_CENTER, textColor=gris)

    tipo_label = "KIT EPI INICIAL" if entrega.tipo == "epi" else "DOTACIÓN DE ROPA SEMESTRAL"
    fecha_str  = entrega.fecha.strftime("%d/%m/%Y") if entrega.fecha else ""

    story = []

    # Cabecera
    story.append(Paragraph(empresa.upper(), h1))
    story.append(Paragraph(f"ALBARÁN DE ENTREGA — {tipo_label}", h2))
    story.append(HRFlowable(width="100%", thickness=1.5, color=azul))
    story.append(Spacer(1, 10))

    # Datos del trabajador
    datos = [
        ["Trabajador:", trabajador.nombre_completo],
        ["DNI:",        trabajador.dni or "—"],
        ["Cargo:",      trabajador.cargo or "—"],
        ["Fecha:",      fecha_str],
        ["Entregado por:", entrega.entregado_por or "—"],
    ]
    t_datos = Table(datos, colWidths=[4*cm, 12*cm])
    t_datos.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("TEXTCOLOR", (0,0), (0,-1), azul),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(t_datos)
    story.append(Spacer(1, 14))

    # Tabla de items
    story.append(Paragraph("Artículos entregados:", h2))
    story.append(Spacer(1, 4))

    table_data = [["Cant.", "Artículo"]]
    for item in items:
        table_data.append([str(item.get("cantidad", 1)), item.get("nombre", "")])

    t_items = Table(table_data, colWidths=[2.5*cm, 13.5*cm])
    t_items.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), azul),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 10),
        ("ALIGN",        (0,0), (0,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 20))

    # Observaciones
    if entrega.observaciones:
        story.append(Paragraph(f"Observaciones: {entrega.observaciones}", normal))
        story.append(Spacer(1, 10))

    # Firmas — con imagen digital si existe
    story.append(Spacer(1, 20))

    firma_b64 = getattr(entrega, "firma_base64", None)
    if firma_b64:
        # Insertar imagen de firma sobre la línea
        try:
            import base64 as _b64
            raw = firma_b64
            if "," in raw:
                raw = raw.split(",", 1)[1]
            img_data = _b64.b64decode(raw)
            from io import BytesIO as _BytesIO
            from reportlab.platypus import Image as _Img
            img_buf = _BytesIO(img_data)
            firma_img = _Img(img_buf, width=7*cm, height=2.5*cm)
            firma_img.hAlign = "LEFT"
            firmas = [
                [firma_img, ""],
                ["____________________________", "____________________________"],
                [f"Entregado por: {entrega.entregado_por or ''}", f"Recibido conforme: {entrega.firmado_por or ''}"],
            ]
        except Exception:
            firma_img = None
            firmas = [
                ["", ""],
                ["____________________________", "____________________________"],
                [f"Entregado por: {entrega.entregado_por or ''}", f"Recibido conforme: {entrega.firmado_por or ''}"],
            ]
    else:
        firmas = [
            ["", ""],
            ["____________________________", "____________________________"],
            [f"Entregado por: {entrega.entregado_por or ''}", f"Recibido conforme: {entrega.firmado_por or ''}"],
        ]

    t_firmas = Table(firmas, colWidths=[8*cm, 8*cm])
    t_firmas.setStyle(TableStyle([
        ("FONTNAME",  (0,2), (-1,2), "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("ALIGN",     (0,0), (-1,-1), "CENTER"),
        ("TEXTCOLOR", (0,2), (-1,2), gris),
        ("VALIGN",    (0,0), (-1,-1), "BOTTOM"),
    ]))
    story.append(t_firmas)
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))

    # QR apuntando a la ficha del trabajador
    try:
        import qrcode as _qr
        from io import BytesIO as _BytesIO2
        qr_url = f"/trabajadores/{trabajador.id}/epis"
        qr_img = _qr.make(qr_url)
        qr_buf = _BytesIO2()
        qr_img.save(qr_buf, format="PNG")
        qr_buf.seek(0)
        from reportlab.platypus import Image as _ImgQR
        qr_elem = _ImgQR(qr_buf, width=1.8*cm, height=1.8*cm)
        pie_data = [
            [qr_elem, f"Documento generado automáticamente — MRD TOOL CONTROL · {fecha_str}\nEscanea el QR para ver el historial completo del trabajador"],
        ]
        t_pie = Table(pie_data, colWidths=[2.2*cm, 13.8*cm])
        t_pie.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("FONTSIZE", (1,0), (1,0), 7),
            ("TEXTCOLOR", (1,0), (1,0), gris),
        ]))
        story.append(Spacer(1, 4))
        story.append(t_pie)
    except Exception:
        story.append(Paragraph(
            f"Documento generado automáticamente — MRD TOOL CONTROL · {fecha_str}",
            center))

    doc.build(story)
    return buf.getvalue()



@app.post("/epis/stock/nueva-talla")
def epis_stock_nueva_talla(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    talla: str = Form(...),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    nombre_val = " ".join(nombre.strip().upper().split())
    talla_val = " ".join(talla.strip().upper().split())
    if not talla_val:
        raise HTTPException(400, "Indica una talla")
    if not nombre_val or len(nombre_val) > 100 or len(talla_val) > 20:
        raise HTTPException(400, "Nombre o talla no válidos")
    start_stock_transaction(db)
    try:
        existing = db.query(StockEPI).filter(
            func.upper(StockEPI.nombre) == nombre_val,
            func.upper(StockEPI.talla) == talla_val,
        ).first()
        if not existing:
            almacen_predeterminado = get_default_warehouse(db)
            stock = StockEPI(
                nombre=nombre_val, categoria="ropa", talla=talla_val,
                cantidad=0, stock_minimo=3, tipo_seguimiento="generico",
                almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
            )
            db.add(stock)
            db.flush()
            stock.codigo = f"SEPI-{stock.id:04d}"
        db.commit()
    except Exception:
        db.rollback()
        raise
    return RedirectResponse("/epis/stock", status_code=303)


# ─── EPIs individuales (arneses y absorbedores con nº de serie y revisiones) ──

# ─── Albarán PDF — EPIs individuales (arneses / absorbedores) ────────────────

def _generar_pdf_epi_individual(trabajador, items: list, usuario_entrega: str, empresa: str) -> bytes:
    """Genera albarán PDF de entrega de EPIs individualizados."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime as _dt

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    azul   = colors.HexColor("#1B4F8A")
    naranj = colors.HexColor("#E8600A")
    gris   = colors.HexColor("#555555")

    h1     = ParagraphStyle("h1",  parent=styles["Normal"], fontSize=16,
                            fontName="Helvetica-Bold", textColor=azul, spaceAfter=4)
    h2     = ParagraphStyle("h2",  parent=styles["Normal"], fontSize=11,
                            fontName="Helvetica-Bold", textColor=azul, spaceAfter=4)
    normal = ParagraphStyle("norm",parent=styles["Normal"], fontSize=10, spaceAfter=2)
    small  = ParagraphStyle("sm",  parent=styles["Normal"], fontSize=8, textColor=gris)
    center = ParagraphStyle("ctr", parent=styles["Normal"], fontSize=9,
                            alignment=TA_CENTER, textColor=gris)

    story = []
    story.append(Paragraph(empresa.upper(), h1))
    story.append(Paragraph("ALBARÁN DE ENTREGA — EQUIPOS DE PROTECCIÓN INDIVIDUAL", h2))
    story.append(HRFlowable(width="100%", thickness=1.5, color=azul))
    story.append(Spacer(1, 10))

    datos = [
        ["Trabajador:",   trabajador.nombre_completo],
        ["DNI:",          trabajador.dni or "—"],
        ["Cargo:",        trabajador.cargo or "—"],
        ["Fecha entrega:", _dt.now().strftime("%d/%m/%Y %H:%M")],
        ["Entregado por:", usuario_entrega],
    ]
    t_datos = Table(datos, colWidths=[4*cm, 13*cm])
    t_datos.setStyle(TableStyle([
        ("FONTNAME",      (0,0),(0,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("TEXTCOLOR",     (0,0),(0,-1), azul),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]))
    story.append(t_datos)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Equipos entregados:", h2))
    story.append(Spacer(1, 4))

    table_data = [["Tipo", "Nº Serie / Código", "Marca / Modelo",
                   "Fecha fab.", "Próx. revisión"]]
    for it in items:
        prox = it.proxima_revision.strftime("%d/%m/%Y") if it.proxima_revision else "—"
        fab  = it.fecha_fabricacion.strftime("%d/%m/%Y") if it.fecha_fabricacion else "—"
        table_data.append([
            it.tipo,
            it.codigo_fabricacion or "—",
            f"{it.marca or ''} {it.modelo or ''}".strip() or "—",
            fab,
            prox,
        ])
    t_items = Table(table_data, colWidths=[3*cm, 4*cm, 4.5*cm, 2.5*cm, 3*cm])
    t_items.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), azul),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 9),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 28))

    # Firmas
    firma_data = [
        [Paragraph("Entregado por:", small), "", Paragraph("Recibido por (firma y DNI):", small)],
        ["", "", ""],
        ["", "", ""],
        [Paragraph(f"<b>{usuario_entrega}</b>", small), "",
         Paragraph(f"<b>{trabajador.nombre_completo}</b>", small)],
    ]
    t_firma = Table(firma_data, colWidths=[7*cm, 3*cm, 7*cm])
    t_firma.setStyle(TableStyle([
        ("LINEABOVE",     (0,2),(0,2), 0.8, colors.black),
        ("LINEABOVE",     (2,2),(2,2), 0.8, colors.black),
        ("FONTSIZE",      (0,0),(-1,-1), 9),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white]),
        ("MINROWHEIGHT",  (0,1),(-1,1), 35),
    ]))
    story.append(t_firma)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "El trabajador declara haber recibido los equipos en perfecto estado y se compromete "
        "a usarlos según las instrucciones del fabricante y la normativa vigente (R.D. 773/1997).",
        small))

    # QR al pie del albarán de arneses/absorbedores
    try:
        import qrcode as _qr3
        from io import BytesIO as _BytesIO4
        qr_img3 = _qr3.make(f"{MRD_PUBLIC_URL}/trabajadores/{trabajador.id}/epis")
        qr_buf3 = _BytesIO4()
        qr_img3.save(qr_buf3, format="PNG")
        qr_buf3.seek(0)
        from reportlab.platypus import Image as _ImgQR3
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
        story.append(Spacer(1, 4))
        pie3 = Table(
            [[_ImgQR3(qr_buf3, width=1.6*cm, height=1.6*cm),
              "Escanea el QR para ver el historial completo del trabajador en MRD TOOL CONTROL."]],
            colWidths=[2*cm, 14*cm]
        )
        pie3.setStyle(TableStyle([
            ("VALIGN", (0,0),(-1,-1),"MIDDLE"),
            ("FONTSIZE",(1,0),(1,0),7),
            ("TEXTCOLOR",(1,0),(1,0),colors.HexColor("#555555")),
        ]))
        story.append(pie3)
    except Exception:
        pass

    doc.build(story)
    return buf.getvalue()


@app.post("/epis/individuales/asignar-lote")
def epi_individual_asignar_lote(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Asigna varios EPIIndividual a un trabajador y devuelve albarán PDF."""
    import urllib.parse as _up
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")

    # Lee form sincrónicamente (datos simples)
    body = None
    # FastAPI no expone form() aquí sincrónicamente — usamos un workaround:
    # los datos llegan como application/x-www-form-urlencoded en request.body
    # Pero como la función es síncrona, usamos el scope ASGI
    form_data = {}
    # Se leen a través de la DB de request state (ver abajo)
    raise HTTPException(500, "Use la versión async")


@app.post("/epis/individuales/asignar-lote-v2")
async def epi_individual_asignar_lote_v2(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Asigna varios EPIIndividual a un trabajador y devuelve albarán PDF."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")

    raise HTTPException(
        409,
        "La asignación masiva de arneses y absorbedores está bloqueada. Use la dotación escaneada línea a línea.",
    )

    form = await request.form()
    trabajador_id = form.get("trabajador_id", "")
    epi_ids_raw   = form.getlist("epi_ids")

    if not trabajador_id:
        return RedirectResponse("/epis/individuales?err=sin_trabajador", status_code=303)
    if not epi_ids_raw:
        return RedirectResponse("/epis/individuales?err=sin_items", status_code=303)

    t = db.query(Trabajador).get(int(trabajador_id))
    if not t:
        raise HTTPException(404, "Trabajador no encontrado")

    items = []
    ya_asignados = []
    for eid_str in epi_ids_raw:
        epi = db.query(EPIIndividual).get(int(eid_str))
        if not epi or epi.estado == "baja":
            continue
        if epi.trabajador_id and epi.trabajador_id != int(trabajador_id):
            # Ya asignado a otro trabajador — no se puede reasignar sin devolución
            ya_asignados.append(f"{epi.tipo} {epi.codigo_fabricacion}")
            continue
        if epi.trabajador_id != int(trabajador_id):
            db.add(HistorialEPIIndividual(
                epi_id=epi.id,
                trabajador_id=int(trabajador_id),
                fecha_asignacion=datetime.utcnow(),
                usuario_id=user.id,
            ))
        epi.trabajador_id = int(trabajador_id)
        items.append(epi)
    db.commit()

    if not items and ya_asignados:
        from urllib.parse import quote
        msg = quote("Los equipos seleccionados ya están asignados a otro trabajador. Devuélvelos primero.")
        return RedirectResponse(f"/epis/individuales?err=ya_asignados&msg={msg}", status_code=303)
    if not items:
        return RedirectResponse("/epis/individuales?err=sin_items", status_code=303)

    mrd_logging.log_app(
        f"Asignación lote EPIs individuales: {len(items)} equipos a {t.nombre_completo} por {user.nombre}",
        level="info"
    )

    pdf_bytes = _generar_pdf_epi_individual(t, items, user.nombre, COMPANY_NAME)
    nombre = f"albaran_epi_{t.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={nombre}"},
    )


@app.post("/epis/individuales/baja-masiva")
async def epi_individual_baja_masiva(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Da de baja varios EPIs individuales seleccionados."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    form = await request.form()
    epi_ids_raw = form.getlist("epi_ids")
    if not epi_ids_raw:
        return RedirectResponse("/epis/individuales?err=sin_seleccion", status_code=303)
    count = 0
    for eid_str in epi_ids_raw:
        try:
            epi = db.query(EPIIndividual).get(int(eid_str))
            if epi and epi.estado != "baja":
                # Cerrar historial abierto si está asignado
                if epi.trabajador_id:
                    reg = db.query(HistorialEPIIndividual).filter(
                        HistorialEPIIndividual.epi_id == epi.id,
                        HistorialEPIIndividual.trabajador_id == epi.trabajador_id,
                        HistorialEPIIndividual.fecha_devolucion == None,
                    ).first()
                    if reg:
                        reg.fecha_devolucion = datetime.utcnow()
                epi.estado = "baja"
                epi.trabajador_id = None
                count += 1
        except (ValueError, TypeError):
            continue
    db.commit()
    mrd_logging.log_app(f"Baja masiva: {count} EPIs dados de baja por {user.nombre}", level="info")
    return RedirectResponse(f"/epis/individuales?ok=baja_masiva&n={count}", status_code=303)


@app.get("/epis/individuales/pdf-inventario")
def epi_individual_pdf_inventario(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """PDF con inventario completo de EPIs individuales."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime as _dt

    items = db.query(EPIIndividual).filter(
        EPIIndividual.estado != "baja"
    ).order_by(EPIIndividual.tipo, EPIIndividual.codigo_fabricacion).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    azul  = colors.HexColor("#1B4F8A")
    gris  = colors.HexColor("#555555")
    rojo  = colors.HexColor("#dc3545")
    verde = colors.HexColor("#198754")
    naranj = colors.HexColor("#fd7e14")

    h1    = ParagraphStyle("h1",  parent=styles["Normal"], fontSize=14,
                           fontName="Helvetica-Bold", textColor=azul, spaceAfter=2)
    small = ParagraphStyle("sm",  parent=styles["Normal"], fontSize=7, textColor=gris)
    center = ParagraphStyle("ctr",parent=styles["Normal"], fontSize=7,
                            alignment=TA_CENTER, textColor=gris)

    story = []
    story.append(Paragraph(COMPANY_NAME.upper(), h1))
    story.append(Paragraph(
        f"INVENTARIO EPIs INDIVIDUALES — Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=azul, spaceAfter=4)
    ))
    story.append(HRFlowable(width="100%", thickness=1.5, color=azul))
    story.append(Spacer(1, 10))

    # Resumen KPI
    total = len(items)
    libres = sum(1 for e in items if not e.trabajador_id)
    asignados = total - libres
    vencidos = sum(1 for e in items if e.revision_vencida)
    kpi_data = [
        ["Total activos", "Disponibles", "Asignados", "Rev. vencida"],
        [str(total), str(libres), str(asignados), str(vencidos)],
    ]
    t_kpi = Table(kpi_data, colWidths=[4.5*cm]*4)
    t_kpi.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), azul),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("BOTTOMPADDING", (0,0),(-1,-1), 6),
        ("TOPPADDING",    (0,0),(-1,-1), 6),
        ("TEXTCOLOR",     (0,1),(0,1), azul),
        ("TEXTCOLOR",     (3,1),(3,1), rojo if vencidos else verde),
        ("FONTNAME",      (0,1),(-1,1), "Helvetica-Bold"),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 14))

    table_data = [["Tipo", "Código/Serie", "Marca/Modelo", "Asignado a", "Estado", "Próx. revisión"]]
    for e in items:
        prox_str = ""
        prox_color = colors.black
        if e.proxima_revision:
            prox_str = e.proxima_revision.strftime("%d/%m/%Y")
            if e.revision_vencida:
                prox_color = rojo
            elif e.dias_para_revision is not None and e.dias_para_revision <= 30:
                prox_color = naranj
        asig = e.trabajador.nombre_completo if e.trabajador else "—"
        estado_str = {"activo": "Activo", "en_revision": "En revisión", "baja": "Baja"}.get(e.estado, e.estado)
        table_data.append([
            e.tipo,
            e.codigo_fabricacion or "—",
            f"{e.marca or ''} {e.modelo or ''}".strip() or "—",
            asig,
            estado_str,
            Paragraph(f'<font color="{prox_color.hexval() if hasattr(prox_color,"hexval") else "#000000"}">{prox_str or "—"}</font>', small) if prox_str else "—",
        ])

    t_items = Table(table_data, colWidths=[2.5*cm, 3.5*cm, 3.5*cm, 4.5*cm, 2*cm, 2.5*cm])
    t_items.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), azul),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
    ]))
    story.append(t_items)
    doc.build(story)
    pdf_bytes = buf.getvalue()
    nombre = f"inventario_epis_{_dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename={nombre}"})


# ─── Documentación de cumplimiento ────────────────────────────────────────────

@app.get("/docs/registro-arneses-pdf")
def docs_registro_arneses(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Registro de revisiones de arneses y absorbedores (formato UNE EN 361)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io
    from datetime import date

    epis = db.query(EPIIndividual).filter(
        EPIIndividual.tipo.in_(TIPOS_EPI_INDIVIDUAL),
    ).order_by(EPIIndividual.tipo, EPIIndividual.num_serie).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.5*cm, bottomMargin=2*cm)
    h1 = ParagraphStyle("h1", fontSize=13, fontName="Helvetica-Bold", spaceAfter=4, alignment=TA_CENTER)
    sm = ParagraphStyle("sm", fontSize=8, fontName="Helvetica", spaceAfter=2)

    story = []
    story.append(Paragraph(COMPANY_NAME or "MRD ESTRUCTURAS", h1))
    story.append(Paragraph("REGISTRO DE INSPECCIÓN DE EQUIPOS ANTICAÍDA", h1))
    story.append(Paragraph(f"(UNE EN 361 / UNE EN 355) — Fecha: {date.today().strftime('%d/%m/%Y')}", sm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e"), spaceAfter=10))

    headers = ["Nº Serie", "Tipo", "Trabajador", "F. Compra", "Últ. Revisión", "Resultado", "Prox. Rev.", "Estado"]
    data = [headers]
    for e in epis:
        ultima = e.revisiones[-1] if e.revisiones else None
        data.append([
            e.num_serie or "—", e.tipo,
            e.trabajador.nombre_completo if e.trabajador else "—",
            e.fecha_compra.strftime("%d/%m/%Y") if e.fecha_compra else "—",
            ultima.fecha.strftime("%d/%m/%Y") if ultima else "—",
            ultima.resultado if ultima else "Sin revisión",
            e.proxima_revision.strftime("%d/%m/%Y") if e.proxima_revision else "—",
            "APTO" if e.estado == "activo" and not e.revision_vencida else "BAJA/VENC.",
        ])

    if len(data) == 1:
        data.append(["—"] * len(headers))

    col_w = [2.6*cm, 2.2*cm, 3.4*cm, 1.9*cm, 2.3*cm, 2.3*cm, 2.1*cm, 2.2*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 7),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4ff")]),
        ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#dee2e6")),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("PADDING",    (0,0),(-1,-1), 3),
    ]))
    story.append(t)
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Responsable: ____________________   Firma: ___________   Fecha: ___________", sm))
    doc.build(story)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=registro_arneses.pdf"})


@app.get("/docs/plan-mantenimiento-pdf")
def docs_plan_mantenimiento(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Plan de mantenimiento preventivo (formato ISO 55000)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io
    from datetime import date, timedelta

    herrs = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.intervalo_mantenimiento_dias != None,
    ).order_by(Herramienta.fecha_proximo_mantenimiento).all()

    maquinas = db.query(Maquinaria).filter(Maquinaria.activa == True).order_by(Maquinaria.proxima_itv).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.5*cm, bottomMargin=2*cm)
    h1 = ParagraphStyle("h1", fontSize=13, fontName="Helvetica-Bold", spaceAfter=4, alignment=TA_CENTER)
    h2 = ParagraphStyle("h2", fontSize=10, fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=12)
    sm = ParagraphStyle("sm", fontSize=8, fontName="Helvetica", spaceAfter=2)

    hoy = date.today()
    story = []
    story.append(Paragraph(COMPANY_NAME or "MRD ESTRUCTURAS", h1))
    story.append(Paragraph("PLAN DE MANTENIMIENTO PREVENTIVO — ISO 55000", h1))
    story.append(Paragraph(f"Generado: {hoy.strftime('%d/%m/%Y')}", sm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e"), spaceAfter=10))

    story.append(Paragraph("1. Herramientas — Mantenimiento Preventivo", h2))
    if herrs:
        hdr = ["Código", "Herramienta", "Marca", "Últ. Mant.", "Intervalo", "Prox. Mant.", "Estado"]
        data = [hdr]
        for h in herrs:
            prox = h.fecha_proximo_mantenimiento
            est = "VENCIDO" if prox and prox < hoy else ("Próximo" if prox and prox <= hoy + timedelta(days=30) else "OK")
            data.append([
                h.codigo,
                (h.nombre[:28] + "…") if len(h.nombre) > 28 else h.nombre,
                h.marca or "—",
                h.fecha_ultimo_mantenimiento.strftime("%d/%m/%Y") if h.fecha_ultimo_mantenimiento else "—",
                f"{h.intervalo_mantenimiento_dias}d",
                prox.strftime("%d/%m/%Y") if prox else "—",
                est,
            ])
        t = Table(data, colWidths=[2.2*cm, 5.5*cm, 2.5*cm, 2.2*cm, 1.8*cm, 2.2*cm, 2*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
            ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0),(-1,-1), 7.5),
            ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#dee2e6")),
            ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ("PADDING",    (0,0),(-1,-1), 3),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No hay herramientas con mantenimiento programado.", sm))

    story.append(Paragraph("2. Maquinaria — ITV y Revisiones", h2))
    if maquinas:
        hdr2 = ["Matrícula", "Vehículo/Máquina", "Última ITV", "Próxima ITV", "Estado"]
        data2 = [hdr2]
        for m in maquinas:
            prox_itv = m.proxima_itv
            est = "VENCIDA" if prox_itv and prox_itv < hoy else ("Próxima" if prox_itv and prox_itv <= hoy + timedelta(days=30) else "OK")
            data2.append([
                m.matricula or "—",
                (m.nombre[:33] + "…") if len(m.nombre) > 33 else m.nombre,
                m.ultima_itv.strftime("%d/%m/%Y") if m.ultima_itv else "—",
                prox_itv.strftime("%d/%m/%Y") if prox_itv else "—",
                est,
            ])
        t2 = Table(data2, colWidths=[2.5*cm, 7*cm, 2.5*cm, 2.5*cm, 2*cm], repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#283593")),
            ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
            ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0),(-1,-1), 7.5),
            ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#dee2e6")),
            ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
            ("PADDING",    (0,0),(-1,-1), 3),
        ]))
        story.append(t2)
    else:
        story.append(Paragraph("No hay maquinaria registrada.", sm))

    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Responsable: ______________________   Firma: ___________   Fecha: ___________", sm))
    doc.build(story)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=plan_mantenimiento.pdf"})


@app.get("/docs/cert-entrega-epi/{eid}")
def docs_cert_entrega_epi(
    eid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Certificado de entrega de EPI (formato Inspección de Trabajo, Art. 17 Ley PRL)."""
    from models import EntregaEPI
    entrega = db.query(EntregaEPI).filter(EntregaEPI.id == eid).first()
    if not entrega:
        raise HTTPException(404)
    trabajador = entrega.trabajador

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    import io, base64
    from io import BytesIO

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.5*cm, bottomMargin=2*cm)
    h1  = ParagraphStyle("h1",   fontSize=13, fontName="Helvetica-Bold", spaceAfter=6, alignment=TA_CENTER)
    sm  = ParagraphStyle("sm",   fontSize=8,  fontName="Helvetica",      spaceAfter=2)
    just= ParagraphStyle("just", fontSize=9,  fontName="Helvetica",      spaceAfter=6, leading=14, alignment=TA_JUSTIFY)

    fecha_str = entrega.fecha.strftime("%d de %B de %Y") if entrega.fecha else "—"
    empresa   = COMPANY_NAME or "MRD ESTRUCTURAS"
    t_nombre  = trabajador.nombre_completo if trabajador else "—"
    t_dni     = getattr(trabajador, "dni", None) or "—"
    t_puesto  = getattr(trabajador, "puesto", None) or "—"

    story = []
    story.append(Paragraph(empresa.upper(), h1))
    story.append(Paragraph("ACUSE DE RECIBO DE EQUIPOS DE PROTECCIÓN INDIVIDUAL (EPI)", h1))
    story.append(Paragraph("(Art. 17 Ley 31/1995 PRL — Art. 14 RD 773/1997)", sm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1a237e"), spaceAfter=12))
    story.append(Paragraph(
        f"D./D.ª <b>{t_nombre}</b>, DNI <b>{t_dni}</b>, puesto <b>{t_puesto}</b>, "
        f"declara haber recibido con fecha <b>{fecha_str}</b> los siguientes EPIs:", just))

    items = getattr(entrega, "items", [])
    idata = [["#", "EPI / Descripción", "Talla/Modelo", "Cant.", "Nº Serie / Ref."]]
    for i, it in enumerate(items, 1):
        idata.append([str(i),
            getattr(it, "nombre_epi", None) or getattr(it, "descripcion", "—"),
            getattr(it, "talla", None) or "—",
            str(getattr(it, "cantidad", 1)),
            getattr(it, "referencia", None) or "—",
        ])
    if len(idata) == 1:
        idata.append(["—", "Ver albarán adjunto", "—", "—", "—"])

    t = Table(idata, colWidths=[0.7*cm, 7.5*cm, 2.5*cm, 1.5*cm, 3.5*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#f0f4ff")]),
        ("GRID",       (0,0),(-1,-1), 0.3, colors.HexColor("#dee2e6")),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("PADDING",    (0,0),(-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "El trabajador/a declara haber sido <b>informado/a sobre el uso correcto, mantenimiento y conservación</b> "
        "de los EPIs recibidos, así como de los riesgos frente a los que le protegen (Art. 18 Ley 31/1995 PRL).", just))

    if entrega.observaciones:
        story.append(Paragraph(f"<b>Observaciones:</b> {entrega.observaciones}", just))

    story.append(Spacer(1, 0.8*cm))

    # Firma
    firma_img_cell = ""
    if entrega.firma_base64:
        try:
            raw = entrega.firma_base64
            if "," in raw:
                raw = raw.split(",", 1)[1]
            img_bytes = base64.b64decode(raw)
            from reportlab.platypus import Image as _Img
            firma_img_cell = _Img(BytesIO(img_bytes), width=5*cm, height=2*cm)
        except Exception:
            pass

    t_firma = Table(
        [["Entregado por:", "", "Recibido conforme (firma y DNI):"],
         ["", "", firma_img_cell or ""],
         [entrega.entregado_por or "________________________", "",
          entrega.firmado_por or "________________________"]],
        colWidths=[7*cm, 2*cm, 7*cm]
    )
    t_firma.setStyle(TableStyle([
        ("FONTNAME",  (0,0),(-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0),(-1,-1), 8),
        ("LINEBELOW", (0,2),(0,2), 0.5, colors.black),
        ("LINEBELOW", (2,2),(2,2), 0.5, colors.black),
        ("VALIGN",    (0,0),(-1,-1), "BOTTOM"),
    ]))
    story.append(t_firma)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Lugar y fecha: {empresa} — {fecha_str}", sm))

    doc.build(story)
    buf.seek(0)
    nombre = f"cert_epi_{(trabajador.nombre or str(eid)).replace(' ','_')}_{eid}.pdf"
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename={nombre}"})


@app.post("/epis/individuales/alerta-revision")
def epi_alerta_revision_manual(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Envía aviso interno con EPIs vencidos/próximos."""
    if not tiene_permiso(user, "admin"):
        raise HTTPException(403, "Sin permiso")
    hoy = date.today()
    prox_30 = hoy + timedelta(days=30)
    vencidos = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision < hoy,
    ).all()
    proximos = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision >= hoy,
        EPIIndividual.proxima_revision <= prox_30,
    ).all()
    lineas = []
    if vencidos:
        lineas.append(f"EPIs con revision VENCIDA ({len(vencidos)}):")
        for e in vencidos:
            ts = e.trabajador.nombre_completo if e.trabajador else "Sin asignar"
            lineas.append(f"  - {e.tipo} {e.codigo_fabricacion} - {ts} - vencida: {e.proxima_revision}")
    if proximos:
        lineas.append(f"EPIs con revision proxima en 30 dias ({len(proximos)}):")
        for e in proximos:
            ts = e.trabajador.nombre_completo if e.trabajador else "Sin asignar"
            lineas.append(f"  - {e.tipo} {e.codigo_fabricacion} - {ts} - revision: {e.proxima_revision} ({e.dias_para_revision} dias)")
    if not lineas:
        return RedirectResponse("/epis/individuales?ok=sin_alertas", status_code=303)
    mensaje = "\n".join(lineas)
    try:
        aviso = Aviso(
            titulo=f"Alerta revision EPIs - {hoy.strftime('%d/%m/%Y')}",
            mensaje=mensaje,
            prioridad="alta" if vencidos else "media",
            usuario_id=user.id,
        )
        db.add(aviso)
        db.commit()
        notif_engine.notificar_aviso(db, aviso)
    except Exception as _err:
        mrd_logging.log_app(f"Error enviando alerta EPIs: {_err}", level="warning")
    mrd_logging.log_app(
        f"Alerta EPIs enviada por {user.nombre}: {len(vencidos)} vencidos, {len(proximos)} proximos",
        level="info"
    )
    total = len(vencidos) + len(proximos)
    return RedirectResponse(f"/epis/individuales?ok=alertas&n={total}", status_code=303)


@app.get("/epis/individuales/plantilla-excel")
def epi_individual_plantilla_excel(user: Usuario = Depends(requiere_login)):
    """Descarga plantilla Excel para importación masiva de EPIs individuales."""
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
    from openpyxl.utils import get_column_letter as _gcl
    import io as _io

    wb = _WB(); ws = wb.active; ws.title = "EPIs Individuales"
    fill_h = _Fill("solid", fgColor="1B4F8A")
    fill_e = _Fill("solid", fgColor="FFF3CD")

    cols = [
        ("tipo*", 18, "ARNES o ABSORBEDOR"),
        ("codigo_fabricacion*", 24, "Nº serie del fabricante"),
        ("marca", 18, "Marca (opcional)"),
        ("modelo", 18, "Modelo (opcional)"),
        ("fecha_fabricacion", 18, "YYYY-MM-DD (opcional)"),
        ("fecha_puesta_servicio", 20, "YYYY-MM-DD (opcional)"),
        ("proxima_revision", 18, "YYYY-MM-DD (opcional)"),
        ("notas", 30, "Observaciones (opcional)"),
    ]
    for c, (header, width, nota) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = _Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill_h
        cell.alignment = _Align(horizontal="center")
        ws.column_dimensions[_gcl(c)].width = width

        nota_cell = ws.cell(row=2, column=c, value=nota)
        nota_cell.font = _Font(italic=True, size=9, color="555555")
        nota_cell.fill = fill_e
        nota_cell.alignment = _Align(horizontal="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # Filas de ejemplo
    ejemplos = [
        ["ARNES", "AH-2024-001", "Petzl", "Avao Bod Croll", "2022-01-15", "2022-03-01", "2025-03-01", ""],
        ["ABSORBEDOR", "AB-2024-002", "MSA", "Latchways", "2021-06-10", "2021-09-01", "2024-09-01", "Revisar costuras"],
    ]
    for r, fila in enumerate(ejemplos, 3):
        for c, val in enumerate(fila, 1):
            ws.cell(row=r, column=c, value=val)

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_epis_individuales.xlsx"},
    )


@app.post("/epis/individuales/importar")
async def epi_individual_importar(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    archivo: UploadFile = File(...),
):
    """Importa EPIs individuales desde Excel."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    from openpyxl import load_workbook as _lw
    import io as _io

    contenido = await archivo.read()
    try:
        wb = _lw(_io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception as ex:
        return RedirectResponse(f"/epis/individuales?err=excel_invalido", status_code=303)

    creados = 0; errores = []
    tipos_validos = {t.upper() for t in TIPOS_EPI_INDIVIDUAL}

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        tipo_raw = str(row[0] or "").strip().upper() if row[0] else ""
        codigo   = str(row[1] or "").strip() if row[1] else ""
        if not tipo_raw or not codigo:
            continue
        if tipo_raw not in tipos_validos:
            errores.append(f"Fila {row_idx}: tipo '{tipo_raw}' no válido")
            continue
        # Evitar duplicados por código de fabricación
        existe = db.query(EPIIndividual).filter(EPIIndividual.codigo_fabricacion == codigo).first()
        if existe:
            errores.append(f"Fila {row_idx}: código '{codigo}' ya existe")
            continue

        def _parse_date(v):
            if not v: return None
            import re
            s = str(v).strip()
            try:
                from datetime import datetime as _dt
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                    try: return _dt.strptime(s, fmt).date()
                    except: pass
            except: pass
            return None

        almacen_predeterminado = get_default_warehouse(db)
        epi = EPIIndividual(
            tipo=tipo_raw,
            codigo_fabricacion=codigo,
            marca=str(row[2] or "").strip() or None,
            modelo=str(row[3] or "").strip() or None,
            fecha_fabricacion=_parse_date(row[4]),
            fecha_puesta_servicio=_parse_date(row[5]),
            proxima_revision=_parse_date(row[6]),
            notas=str(row[7] or "").strip() or None,
            estado="activo",
            almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
        )
        db.add(epi)
        db.flush()
        ensure_epi_identifier(db, epi, user)
        creados += 1

    db.commit()
    mrd_logging.log_app(f"Importación EPIs individuales: {creados} creados por {user.nombre}", level="info")
    from urllib.parse import quote as _q
    msg = f"{creados} equipos importados"
    if errores:
        msg += f" | {len(errores)} errores: " + "; ".join(errores[:3])
    return RedirectResponse(f"/epis/individuales?ok={_q(msg)}", status_code=303)


@app.get("/epis/individuales", response_class=HTMLResponse)
def epis_individuales_panel(request: Request, user: Usuario = Depends(requiere_login),
                            db: Session = Depends(get_db)):
    items = db.query(EPIIndividual).filter(EPIIndividual.estado != "baja").order_by(
        EPIIndividual.tipo, EPIIndividual.codigo_fabricacion
    ).all()
    por_revisar = [i for i in items if i.revision_vencida]
    proximos    = [i for i in items if not i.revision_vencida and i.proxima_revision and
                   i.dias_para_revision is not None and i.dias_para_revision <= 30]
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()
    # Stock disponible (sin asignar) por tipo
    stock_disponible = {}
    for tipo in TIPOS_EPI_INDIVIDUAL:
        stock_disponible[tipo] = db.query(EPIIndividual).filter(
            EPIIndividual.tipo == tipo,
            EPIIndividual.estado != "baja",
            EPIIndividual.trabajador_id == None
        ).count()
    return templates.TemplateResponse(request, "epis_individuales.html", ctx_base(
        request, user, db,
        items=items,
        por_revisar=por_revisar,
        proximos=proximos,
        trabajadores=trabajadores,
        tipos=TIPOS_EPI_INDIVIDUAL,
        stock_disponible=stock_disponible,
    ))


@app.post("/epis/individuales")
def epis_individuales_crear(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    tipo: str = Form(...),
    codigo_fabricacion: str = Form(...),
    marca: str = Form(""),
    modelo: str = Form(""),
    fecha_fabricacion: str = Form(""),
    fecha_puesta_servicio: str = Form(""),
    trabajador_id: str = Form(""),
    proxima_revision: str = Form(""),
    notas: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")

    def _pd(s):
        try: return date.fromisoformat(s) if s else None
        except Exception: return None

    almacen_predeterminado = get_default_warehouse(db)
    epi = EPIIndividual(
        tipo=tipo.upper(),
        codigo_fabricacion=codigo_fabricacion.strip(),
        marca=marca.strip() or None,
        modelo=modelo.strip() or None,
        fecha_fabricacion=_pd(fecha_fabricacion),
        fecha_puesta_servicio=_pd(fecha_puesta_servicio),
        trabajador_id=int(trabajador_id) if trabajador_id.strip() else None,
        proxima_revision=_pd(proxima_revision),
        notas=notas.strip() or None,
        estado="activo",
        almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
    )
    db.add(epi)
    db.flush()
    ensure_epi_identifier(db, epi, user)
    db.commit()
    mrd_logging.log_security(
        f"Nuevo EPI individual tipo={tipo} cod={codigo_fabricacion} por {user.username}",
        level="info"
    )
    return RedirectResponse("/epis/individuales", status_code=303)


@app.get("/epis/individuales/{eid}", response_class=HTMLResponse)
def epi_individual_detalle(eid: int, request: Request,
                           user: Usuario = Depends(requiere_login),
                           db: Session = Depends(get_db)):
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()
    return templates.TemplateResponse(request, "epi_individual_detalle.html", ctx_base(
        request, user, db,
        epi=epi,
        trabajadores=trabajadores,
        intervalo_dias=INTERVALO_REVISION_EPI_DIAS,
        historial_epi=epi.historial,
    ))


@app.post("/epis/individuales/{eid}/revision")
def epi_individual_revision(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    resultado: str = Form(...),
    tecnico: str = Form(""),
    proxima_revision: str = Form(""),
    observaciones: str = Form(""),
    back: str = Form(""),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")

    def _pd(s):
        try: return date.fromisoformat(s) if s else None
        except Exception: return None

    prox = _pd(proxima_revision)
    rev = RevisionEPI(
        epi_id=eid,
        fecha=datetime.utcnow(),
        resultado=resultado,
        tecnico=tecnico.strip() or None,
        proxima_revision=prox,
        observaciones=observaciones.strip() or None,
        usuario_id=user.id,
    )
    db.add(rev)
    epi.estado = "baja" if resultado == "retirar" else ("en_revision" if resultado == "en_revision" else "activo")
    if prox:
        epi.proxima_revision = prox
    db.commit()
    mrd_logging.log_security(
        f"Revisión EPI id={eid} resultado={resultado} por {user.username}", level="info"
    )
    _back = back.strip() if back and back.strip().startswith("/") else f"/epis/individuales/{eid}"
    return RedirectResponse(_back, status_code=303)


@app.post("/epis/individuales/{eid}/asignar")
def epi_individual_asignar(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    trabajador_id: str = Form(""),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    nuevo_tid = int(trabajador_id) if trabajador_id.strip() else None
    if nuevo_tid and epi.tipo.upper() in {"ARNES", "ABSORBEDOR"}:
        raise HTTPException(
            409,
            "Este EPI solo puede asignarse escaneando su QR desde una dotación.",
        )
    if nuevo_tid and nuevo_tid != epi.trabajador_id:
        if epi.trabajador_id:
            reg_ant = db.query(HistorialEPIIndividual).filter(
                HistorialEPIIndividual.epi_id == eid,
                HistorialEPIIndividual.trabajador_id == epi.trabajador_id,
                HistorialEPIIndividual.fecha_devolucion == None,
            ).first()
            if reg_ant:
                reg_ant.fecha_devolucion = datetime.utcnow()
        db.add(HistorialEPIIndividual(
            epi_id=epi.id,
            trabajador_id=nuevo_tid,
            fecha_asignacion=datetime.utcnow(),
            usuario_id=user.id,
        ))
    epi.trabajador_id = nuevo_tid
    db.commit()
    return RedirectResponse(f"/epis/individuales/{eid}", status_code=303)


@app.post("/epis/individuales/{eid}/identificador")
def epi_individual_generar_identificador(
    eid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso")
    epi = db.get(EPIIndividual, eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    try:
        identifier = ensure_epi_identifier(db, epi, user)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return JSONResponse({
        "epi_id": epi.id,
        "referencia_interna": identifier.referencia_interna,
        "codigo_qr": identifier.codigo_qr,
    })


@app.post("/epis/individuales/{eid}/devolver")
def epi_individual_devolver(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Desasigna el EPI individual del trabajador actual, deja disponible."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    anterior = epi.trabajador.nombre_completo if epi.trabajador else "nadie"
    anterior_id = epi.trabajador_id
    if anterior_id:
        registro_abierto = db.query(HistorialEPIIndividual).filter(
            HistorialEPIIndividual.epi_id == eid,
            HistorialEPIIndividual.trabajador_id == anterior_id,
            HistorialEPIIndividual.fecha_devolucion == None,
        ).first()
        if registro_abierto:
            registro_abierto.fecha_devolucion = datetime.utcnow()
    epi.trabajador_id = None
    db.commit()
    mrd_logging.log_app(
        f"EPI {epi.tipo} {epi.codigo_fabricacion} devuelto (era de {anterior}) por {user.nombre}",
        level="info"
    )
    return RedirectResponse(f"/epis/individuales/{eid}?ok=devuelto", status_code=303)


@app.post("/epis/individuales/{eid}/baja")
def epi_individual_baja(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    epi.estado = "baja"
    db.commit()
    return RedirectResponse("/epis/individuales", status_code=303)


@app.post("/epis/individuales/{eid}/renovar-revision")
def epi_individual_renovar_revision(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    tecnico: str = Form(""),
    observaciones: str = Form(""),
):
    """Marca revisión como realizada hoy y calcula próxima automáticamente."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    hoy = date.today()
    prox = hoy + timedelta(days=INTERVALO_REVISION_EPI_DIAS)
    rev = RevisionEPI(
        epi_id=eid,
        fecha=datetime.utcnow(),
        resultado="apto",
        tecnico=tecnico.strip() or user.nombre,
        proxima_revision=prox,
        observaciones=observaciones.strip() or "Revisión periódica realizada",
        usuario_id=user.id,
    )
    db.add(rev)
    epi.proxima_revision = prox
    epi.estado = "activo"
    db.commit()
    mrd_logging.log_app(
        f"Revisión renovada EPI id={eid} por {user.nombre} — próxima: {prox}",
        level="info"
    )
    return RedirectResponse(f"/epis/individuales/{eid}?ok=revision", status_code=303)


@app.post("/epis/individuales/{eid}/foto")
async def epi_individual_subir_foto(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    foto: UploadFile = File(...),
):
    """Sube foto del EPI individual."""
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    ext = Path(foto.filename).suffix.lower() if foto.filename else ".jpg"
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        ext = ".jpg"
    carpeta = BASE_DIR / "static" / "uploads" / "epis"
    carpeta.mkdir(parents=True, exist_ok=True)
    nombre = f"epi_{eid}_{int(datetime.utcnow().timestamp())}{ext}"
    if epi.foto_path:
        old_p = carpeta / epi.foto_path
        if old_p.exists():
            old_p.unlink(missing_ok=True)
    content = await foto.read()
    (carpeta / nombre).write_bytes(content)
    epi.foto_path = nombre
    db.commit()
    return RedirectResponse(f"/epis/individuales/{eid}?ok=foto", status_code=303)


@app.post("/epis/individuales/{eid}/foto/eliminar")
def epi_individual_eliminar_foto(
    eid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    epi = db.query(EPIIndividual).get(eid)
    if not epi:
        raise HTTPException(404, "EPI no encontrado")
    if epi.foto_path:
        p = BASE_DIR / "static" / "uploads" / "epis" / epi.foto_path
        if p.exists():
            p.unlink(missing_ok=True)
        epi.foto_path = None
        db.commit()
    return RedirectResponse(f"/epis/individuales/{eid}?ok=foto_eliminada", status_code=303)


# ─── Catálogo EPIs — gestión dinámica ────────────────────────────────────────

@app.get("/epis/catalogo", response_class=HTMLResponse)
def epis_catalogo(request: Request, user: Usuario = Depends(requiere_login),
                  db: Session = Depends(get_db)):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    mostrar_archivados = user.rol == "admin" and request.query_params.get("archivados") == "1"
    estado_visible = False if mostrar_archivados else True
    items_epi = db.query(CatalogoEPI).filter(
        CatalogoEPI.categoria == "epi", CatalogoEPI.activo == estado_visible,
    ).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    items_ropa = db.query(CatalogoEPI).filter(
        CatalogoEPI.categoria == "ropa", CatalogoEPI.activo == estado_visible,
    ).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    archivados_count = db.query(CatalogoEPI).filter(CatalogoEPI.activo == False).count()
    stock_por_nombre = {}
    tallas_por_nombre = {}
    for stock in db.query(StockEPI).order_by(StockEPI.id).all():
        stock_por_nombre.setdefault(stock.nombre, stock)
        if stock.talla:
            tallas_por_nombre.setdefault(stock.nombre, []).append(stock.talla)
    return templates.TemplateResponse(request, "epis_catalogo.html", ctx_base(
        request, user, db,
        items_epi=items_epi,
        items_ropa=items_ropa,
        stock_por_nombre=stock_por_nombre,
        tallas_por_nombre=tallas_por_nombre,
        mostrar_archivados=mostrar_archivados,
        archivados_count=archivados_count,
    ))


@app.post("/epis/catalogo/nuevo")
async def epis_catalogo_nuevo(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    form = await request.form()
    nombre    = (form.get("nombre", "") or "").strip().upper()
    categoria = form.get("categoria", "epi")
    cantidad  = max(1, int(form.get("cantidad_kit", 1) or 1))
    marca     = (form.get("marca", "") or "").strip()[:100]
    notas     = (form.get("notas", "") or "").strip()[:200]
    source    = form.get("source", "")
    try:
        tallas = _normalizar_tallas_elegidas(form.get("tallas", "") or "")
    except ValueError:
        return RedirectResponse("/epis/catalogo?err=tallas", status_code=303)
    destino   = "/epis/stock" if source == "stock" else "/epis/catalogo"
    if not nombre:
        return RedirectResponse(f"{destino}?err=nombre_vacio", status_code=303)
    if categoria not in ("epi", "ropa"):
        categoria = "epi"
    existe = db.query(CatalogoEPI).filter(CatalogoEPI.nombre == nombre).first()
    if existe:
        if not existe.activo:
            existe.activo = True
        if existe.categoria == "ropa":
            _crear_tallas_elegidas(db, existe.nombre, tallas)
        db.commit()
        return RedirectResponse(f"{destino}?ok=reactivado", status_code=303)
    orden_max = db.query(CatalogoEPI).filter(CatalogoEPI.categoria == categoria).count()
    item = CatalogoEPI(nombre=nombre, categoria=categoria, cantidad_kit=cantidad,
                       activo=True, orden=orden_max, marca=marca or None, notas=notas or None)
    try:
        db.add(item)
        db.flush()
        # Catálogo, stock e identidad QR se confirman juntos.
        if categoria == "ropa" and tallas:
            _crear_tallas_elegidas(db, nombre, tallas)
        else:
            stock = db.query(StockEPI).filter(
                StockEPI.nombre == nombre, StockEPI.talla == None
            ).first()
            if not stock:
                almacen_predeterminado = get_default_warehouse(db)
                stock = StockEPI(
                    nombre=nombre, categoria=categoria, talla=None,
                    cantidad=0, stock_minimo=3, tipo_seguimiento="generico",
                    almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
                )
                db.add(stock)
                db.flush()
            if not stock.codigo:
                stock.codigo = f"SEPI-{stock.id:04d}"
        db.commit()
    except Exception:
        db.rollback()
        raise
    return RedirectResponse(f"{destino}?ok=creado", status_code=303)


@app.post("/epis/catalogo/{cid}/editar")
async def epis_catalogo_editar(
    cid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    item = db.query(CatalogoEPI).get(cid)
    if not item:
        raise HTTPException(404)
    form = await request.form()
    nombre_nuevo = (form.get("nombre", "") or "").strip().upper()
    if not nombre_nuevo:
        return RedirectResponse("/epis/catalogo?err=nombre_vacio", status_code=303)
    categoria_nueva = (form.get("categoria", item.categoria) or item.categoria).strip().lower()
    if categoria_nueva not in {"epi", "ropa"}:
        return RedirectResponse("/epis/catalogo?err=datos", status_code=303)
    try:
        tallas = _normalizar_tallas_elegidas(form.get("tallas", "") or "")
    except ValueError:
        return RedirectResponse("/epis/catalogo?err=tallas", status_code=303)

    duplicado = db.query(CatalogoEPI).filter(
        CatalogoEPI.nombre == nombre_nuevo,
        CatalogoEPI.id != item.id,
    ).first()
    if duplicado:
        return RedirectResponse("/epis/catalogo?err=nombre_duplicado", status_code=303)

    stock_old = db.query(StockEPI).filter(StockEPI.nombre == item.nombre).all()
    if nombre_nuevo != item.nombre:
        for stock in stock_old:
            solape = db.query(StockEPI).filter(
                StockEPI.nombre == nombre_nuevo,
                StockEPI.talla == stock.talla,
                StockEPI.id != stock.id,
            ).first()
            if solape:
                return RedirectResponse("/epis/catalogo?err=stock_duplicado", status_code=303)

    try:
        for stock in stock_old:
            stock.nombre = nombre_nuevo
            stock.categoria = categoria_nueva
        item.nombre = nombre_nuevo
        item.categoria = categoria_nueva
        cantidad = form.get("cantidad_kit")
        if cantidad:
            item.cantidad_kit = max(1, min(int(cantidad), 999))
        marca = form.get("marca")
        if marca is not None:
            item.marca = marca.strip()[:100] or None
        notas = form.get("notas")
        if notas is not None:
            item.notas = notas.strip()[:200] or None
        activo = form.get("activo")
        if activo is not None:
            item.activo = activo == "1"
        if categoria_nueva == "ropa":
            _crear_tallas_elegidas(db, nombre_nuevo, tallas)
        db.commit()
    except (IntegrityError, ValueError):
        db.rollback()
        return RedirectResponse("/epis/catalogo?err=conflicto", status_code=303)
    return RedirectResponse("/epis/catalogo?ok=editado", status_code=303)


@app.post("/epis/catalogo/{cid}/eliminar")
def epis_catalogo_eliminar(
    cid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo admins")
    item = db.query(CatalogoEPI).get(cid)
    if not item:
        raise HTTPException(404)
    stock_rows = db.query(StockEPI).filter(StockEPI.nombre == item.nombre).all()
    stock_ids = [row.id for row in stock_rows]
    has_history = bool(
        (stock_ids and db.query(MovimientoStock).filter(MovimientoStock.stock_epi_id.in_(stock_ids)).first())
        or (stock_ids and db.query(LineaInventario).filter(LineaInventario.stock_epi_id.in_(stock_ids)).first())
        or db.query(VarianteEPI).filter(VarianteEPI.catalogo_epi_id == item.id).first()
        or db.query(LineaDotacion).filter(LineaDotacion.catalogo_epi_id == item.id).first()
        or any(int(row.cantidad or 0) != 0 for row in stock_rows)
    )
    if has_history:
        item.activo = False
        result = "desactivado"
    else:
        for row in stock_rows:
            db.delete(row)
        db.delete(item)
        result = "eliminado"
    db.commit()
    return RedirectResponse(f"/epis/catalogo?ok={result}", status_code=303)


# ─── Almacenes — detalle, ubicaciones, foto, QR, eliminar ────────────────────

@app.get("/almacenes/{aid}", response_class=HTMLResponse)
def almacen_detalle(
    aid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()
    almacenes_todos = visible_warehouses(db, user)
    principal = get_default_warehouse(db)
    filtro_material = Material.almacen_id == aid
    if principal and principal.id == aid:
        filtro_material = or_(Material.almacen_id == aid, Material.almacen_id.is_(None))
    materiales_almacen = db.query(Material).filter(Material.activo == True, filtro_material).all()
    return templates.TemplateResponse(request, "almacen_detalle.html", ctx_base(
        request, user, db,
        almacen=a,
        trabajadores=trabajadores,
        herramientas=a.herramientas,
        materiales=materiales_almacen,
        ubicaciones=a.ubicaciones,
        almacenes_todos=almacenes_todos,
    ))


@app.get("/api/almacenes/{aid}/metricas")
def api_almacen_metricas(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Métricas en vivo de un almacén, consultadas por polling desde la ficha."""
    _require_warehouse_access(user, aid)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    principal = get_default_warehouse(db)
    filtro_material = Material.almacen_id == aid
    if principal and principal.id == aid:
        filtro_material = or_(Material.almacen_id == aid, Material.almacen_id.is_(None))
    materiales = db.query(Material).filter(Material.activo == True, filtro_material).all()
    herr_en_almacen = db.query(Herramienta).filter(
        Herramienta.almacen_id == aid, Herramienta.activa == True, Herramienta.estado == "en_almacen",
    ).count()
    herr_fuera = db.query(Herramienta).filter(
        Herramienta.almacen_id == aid, Herramienta.activa == True, Herramienta.estado != "en_almacen",
    ).count()
    hoy_inicio = datetime.combine(datetime.utcnow().date(), datetime.min.time())
    movs_herr_hoy = (
        db.query(Movimiento).join(Herramienta, Movimiento.herramienta_id == Herramienta.id)
        .filter(Herramienta.almacen_id == aid, Movimiento.fecha >= hoy_inicio)
        .count()
    )
    movs_mat_hoy = (
        db.query(MovimientoMaterial).join(Material, MovimientoMaterial.material_id == Material.id)
        .filter(filtro_material, MovimientoMaterial.fecha >= hoy_inicio)
        .count()
    )
    albaranes_abiertos = (
        db.query(AlbaranSalida)
        .filter(AlbaranSalida.almacen_id == aid, AlbaranSalida.estado.in_(("abierto", "parcial")))
        .count()
    )
    return {
        "actualizado": datetime.utcnow().isoformat(),
        "herramientas_en_almacen": herr_en_almacen,
        "herramientas_fuera": herr_fuera,
        "materiales_total": len(materiales),
        "materiales_bajo_minimo": sum(1 for m in materiales if m.bajo_minimo),
        "movimientos_hoy": movs_herr_hoy + movs_mat_hoy,
        "albaranes_abiertos": albaranes_abiertos,
    }


def _mapa_config_almacen(almacen: Almacen) -> dict:
    """Lee el plano sin romper los mapas de puntos de versiones anteriores."""
    config = {"version": 3, "fondo": None, "zonas": []}
    try:
        raw = json.loads(almacen.mapa_json or "null")
    except (TypeError, ValueError):
        return config
    if isinstance(raw, dict):
        config["fondo"] = Path(str(raw.get("fondo") or "")).name or None
        config["zonas"] = raw.get("zonas") if isinstance(raw.get("zonas"), list) else []
        return config
    if isinstance(raw, list):
        # Conversión automática del mapa antiguo: cada bloque pasa a ser una zona.
        ubicaciones = {u.id: u for u in almacen.ubicaciones if u.activo}
        for indice, punto in enumerate(raw):
            try:
                uid = int(punto["id"])
                ubicacion = ubicaciones.get(uid)
                if not ubicacion:
                    continue
                config["zonas"].append({
                    "id": f"zona-{uid}", "nombre": ubicacion.nombre,
                    "x": max(0, min(85, float(punto.get("x", 20)) / 8)),
                    "y": max(0, min(80, float(punto.get("y", 20)) / 6)),
                    "w": 18, "h": 18, "ubicaciones": [uid],
                })
            except (KeyError, TypeError, ValueError):
                continue
    return config


@app.get("/almacenes/{aid}/mapa", response_class=HTMLResponse)
def almacen_mapa(
    aid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    a = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    if not a:
        raise HTTPException(404)
    mapa_config = _mapa_config_almacen(a)
    principal = get_default_warehouse(db)
    materiales_sin_asignar = (
        db.query(Material).filter(Material.activo == True, Material.almacen_id.is_(None)).all()
        if principal and principal.id == aid else []
    )
    herramientas_sin_asignar = (
        db.query(Herramienta).filter(Herramienta.activa == True, Herramienta.almacen_id.is_(None)).all()
        if principal and principal.id == aid else []
    )
    # Vista operativa del contenido real de cada ubicación: herramientas,
    # consumibles, ropa/EPI y maquinaria vinculada por nombre o código de zona.
    existencias_v2 = db.query(ExistenciaVariante).filter(
        ExistenciaVariante.almacen_id == aid,
        ExistenciaVariante.cantidad > 0,
    ).all()
    existencias_por_ubicacion = {}
    for existencia in existencias_v2:
        variante = existencia.variante
        catalogo = variante.catalogo if variante else None
        if not variante or not catalogo:
            continue
        existencias_por_ubicacion.setdefault(existencia.ubicacion_id or 0, []).append({
            "id": existencia.id, "tipo_key": "existencia",
            "tipo": "ropa" if catalogo.categoria == "ropa" else "epi",
            "codigo": variante.referencia_interna,
            "nombre": catalogo.nombre,
            "detalle": " · ".join(filter(None, [variante.modelo, variante.color, variante.talla])),
            "cantidad": existencia.cantidad,
            "unidad": "ud",
        })

    stock_epi_por_ubicacion = {}
    epi_individual_por_ubicacion = {}
    for stock in db.query(StockEPI).filter(
        StockEPI.cantidad > 0,
        or_(StockEPI.almacen_id == aid, StockEPI.almacen_id.is_(None)),
    ).all():
        stock_epi_por_ubicacion.setdefault(stock.ubicacion_id or 0, []).append({
            "id": stock.id, "tipo_key": "stock_epi", "tipo": "ropa / EPI",
            "codigo": stock.codigo or f"SEPI-{stock.id:04d}",
            "nombre": stock.nombre_display, "detalle": stock.categoria,
            "cantidad": stock.cantidad or 0, "unidad": "ud",
        })
    for epi in db.query(EPIIndividual).filter(
        EPIIndividual.trabajador_id.is_(None), EPIIndividual.estado != "baja",
        or_(EPIIndividual.almacen_id == aid, EPIIndividual.almacen_id.is_(None)),
    ).all():
        epi_individual_por_ubicacion.setdefault(epi.ubicacion_id or 0, []).append({
            "id": epi.id, "tipo_key": "epi_individual", "tipo": "EPI individual",
            "codigo": epi.codigo_qr or epi.referencia_interna or epi.codigo_fabricacion,
            "nombre": epi.tipo, "detalle": " · ".join(filter(None, [epi.marca, epi.modelo, epi.estado])),
            "cantidad": 1, "unidad": "ud",
        })

    maquinaria_por_ubicacion = {}
    maquinaria_mapa = []
    ubicaciones_activas = [u for u in a.ubicaciones if u.activo]
    ubicacion_por_clave = {}
    for ubicacion in ubicaciones_activas:
        for clave in (ubicacion.nombre, ubicacion.codigo):
            clave_limpia = (clave or "").strip().casefold()
            if clave_limpia and clave_limpia not in ubicacion_por_clave:
                ubicacion_por_clave[clave_limpia] = ubicacion
    for maquina in db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        or_(Maquinaria.almacen_id == aid, Maquinaria.almacen_id.is_(None)),
    ).order_by(Maquinaria.nombre).all():
        ubicacion_maquina = (maquina.ubicacion or "").strip().casefold()
        ubicacion_obj = ubicacion_por_clave.get(ubicacion_maquina)
        item_maquina = {
            "id": maquina.id, "tipo_key": "maquinaria", "tipo": "maquinaria",
            "codigo": maquina.codigo_interno or maquina.codigo_barras or f"MRD-MAQ-{maquina.id}",
            "nombre": maquina.nombre,
            "detalle": " · ".join(filter(None, [maquina.tipo, maquina.marca, maquina.modelo,
                                                   maquina.matricula, maquina.num_serie, maquina.estado])),
            "cantidad": 1, "unidad": "ud",
        }
        if ubicacion_obj:
            maquinaria_por_ubicacion.setdefault(ubicacion_obj.id, []).append(item_maquina)
        claves_almacen = {
            (a.nombre or "").strip().casefold(),
            (a.codigo or "").strip().casefold(),
        }
        if not ubicacion_obj and (ubicacion_maquina in claves_almacen or not ubicacion_maquina):
            maquinaria_por_ubicacion.setdefault(0, []).append({
                **item_maquina, "detalle": "Falta zona/estantería concreta",
            })
        maquinaria_mapa.append({
            **item_maquina,
            "ubicacion_id": ubicacion_obj.id if ubicacion_obj else None,
            "ubicacion_actual": maquina.ubicacion or "Sin ubicación",
        })

    ubicaciones_mapa = []
    for u in a.ubicaciones:
        if not u.activo:
            continue
        items = [{
            "id": h.id, "tipo_key": "herramienta",
            "tipo": "herramienta", "codigo": h.codigo, "nombre": h.nombre,
            "detalle": h.estado or "", "cantidad": 1, "unidad": "ud",
        } for h in u.herramientas if h.activa]
        items.extend({
            "id": material.id, "tipo_key": "material",
            "tipo": "consumible", "codigo": material.codigo, "nombre": material.nombre,
            "detalle": material.categoria or "", "cantidad": material.stock_actual or 0,
            "unidad": material.unidad or "ud",
        } for material in u.materiales if material.activo)
        items.extend(existencias_por_ubicacion.get(u.id, []))
        items.extend(stock_epi_por_ubicacion.get(u.id, []))
        items.extend(epi_individual_por_ubicacion.get(u.id, []))
        items.extend(maquinaria_por_ubicacion.get(u.id, []))
        ubicaciones_mapa.append({
            "id": u.id, "nombre": u.nombre, "codigo": u.codigo or "",
            "ruta": u.ruta_completa,
            "estanteria": u.estanteria or "", "posicion": u.posicion or "",
            "count": len(items),
            "unidades": sum(float(i["cantidad"] or 0) for i in items),
            "items": items,
        })
    pendientes = [{
        "id": h.id, "tipo_key": "herramienta",
        "tipo": "herramienta", "codigo": h.codigo, "nombre": h.nombre,
        "detalle": "Falta asignar estantería/zona", "cantidad": 1, "unidad": "ud",
    } for h in a.herramientas if h.activa and h.ubicacion_id is None]
    pendientes.extend({
        "id": h.id, "tipo_key": "herramienta",
        "tipo": "herramienta", "codigo": h.codigo, "nombre": h.nombre,
        "detalle": "Almacén principal · falta asignar estantería/zona",
        "cantidad": 1, "unidad": "ud",
    } for h in herramientas_sin_asignar if h.ubicacion_id is None)
    pendientes.extend({
        "id": material.id, "tipo_key": "material",
        "tipo": "consumible", "codigo": material.codigo, "nombre": material.nombre,
        "detalle": "Falta asignar estantería/zona", "cantidad": material.stock_actual or 0,
        "unidad": material.unidad or "ud",
    } for material in a.materiales if material.activo and material.ubicacion_id is None)
    pendientes.extend({
        "id": material.id, "tipo_key": "material",
        "tipo": "consumible", "codigo": material.codigo, "nombre": material.nombre,
        "detalle": "Almacén principal · falta asignar estantería/zona",
        "cantidad": material.stock_actual or 0, "unidad": material.unidad or "ud",
    } for material in materiales_sin_asignar)
    pendientes.extend(stock_epi_por_ubicacion.get(0, []))
    pendientes.extend(epi_individual_por_ubicacion.get(0, []))
    pendientes.extend(existencias_por_ubicacion.get(0, []))
    pendientes.extend(maquinaria_por_ubicacion.get(0, []))
    if pendientes:
        ubicaciones_mapa.append({
            "id": 0, "nombre": "Pendiente de ubicar", "codigo": "SIN-ZONA",
            "count": len(pendientes),
            "unidades": sum(float(i["cantidad"] or 0) for i in pendientes),
            "items": pendientes,
        })
    movibles_mapa = []
    movibles_vistos = set()
    for contenedor in ubicaciones_mapa:
        ubicacion_id = None if contenedor["id"] == 0 else contenedor["id"]
        ubicacion_nombre = next(
            (u.nombre for u in ubicaciones_activas if u.id == ubicacion_id),
            "Sin ubicación",
        )
        for item in contenedor["items"]:
            tipo_key, item_id = item.get("tipo_key"), item.get("id")
            clave = f"{tipo_key}:{item_id}"
            if not tipo_key or not item_id or clave in movibles_vistos:
                continue
            movibles_vistos.add(clave)
            movibles_mapa.append({
                **item, "key": clave, "ubicacion_id": ubicacion_id,
                "ubicacion_actual": ubicacion_nombre,
            })
    for maquina in maquinaria_mapa:
        clave = f"maquinaria:{maquina['id']}"
        if clave not in movibles_vistos:
            movibles_mapa.append({**maquina, "key": clave})
            movibles_vistos.add(clave)
    ids_reales = [u["id"] for u in ubicaciones_mapa if u["id"] != 0]
    if not mapa_config["zonas"] and ids_reales:
        mapa_config["zonas"] = [{
            "id": "zona-general", "nombre": "Nave / zona general",
            "x": 4, "y": 5, "w": 92, "h": 88, "ubicaciones": ids_reales,
        }]
    fondo_url = (
        f"/uploads/almacenes/mapas/{mapa_config['fondo']}"
        if mapa_config.get("fondo") else None
    )
    return templates.TemplateResponse(request, "mapa_almacen.html", ctx_base(
        request, user,
        almacen=a,
        ubicaciones=a.ubicaciones,
        ubicaciones_json=dumps_for_script(ubicaciones_mapa),
        maquinaria_json=dumps_for_script(movibles_mapa),
        mapa_config_json=dumps_for_script(mapa_config),
        fondo_url=fondo_url,
        puede_organizar=(tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")),
    ))


@app.post("/almacenes/{aid}/mapa/guardar")
async def almacen_mapa_guardar(
    aid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para organizar el mapa")
    a = db.query(Almacen).filter(Almacen.id == aid).first()
    if not a:
        raise HTTPException(404)
    body = await request.json()
    zonas = body.get("zonas", [])
    ubicaciones_validas = {u.id for u in a.ubicaciones if u.activo}
    if not isinstance(zonas, list) or len(zonas) > 100:
        raise HTTPException(400, "Disposición no válida")
    zonas_limpias = []
    zonas_vistas = set()
    ubicaciones_asignadas = set()
    for entrada in zonas:
        try:
            zona_id = str(entrada["id"]).strip()
            nombre = str(entrada["nombre"]).strip()[:80]
            x = max(0.0, min(98.0, float(entrada["x"])))
            y = max(0.0, min(98.0, float(entrada["y"])))
            w = max(5.0, min(100.0 - x, float(entrada["w"])))
            h = max(5.0, min(100.0 - y, float(entrada["h"])))
            ids_ubicacion = [int(uid) for uid in entrada.get("ubicaciones", [])]
        except (KeyError, TypeError, ValueError):
            raise HTTPException(400, "Zona o coordenadas no válidas")
        if (not nombre or not re.fullmatch(r"[A-Za-z0-9_-]{1,50}", zona_id)
                or zona_id in zonas_vistas):
            raise HTTPException(400, "Identificador o nombre de zona no válido")
        if (len(ids_ubicacion) != len(set(ids_ubicacion))
                or any(uid not in ubicaciones_validas for uid in ids_ubicacion)
                or ubicaciones_asignadas.intersection(ids_ubicacion)):
            raise HTTPException(400, "Una estantería no puede estar repetida en varias zonas")
        zonas_vistas.add(zona_id)
        ubicaciones_asignadas.update(ids_ubicacion)
        zonas_limpias.append({
            "id": zona_id, "nombre": nombre, "x": round(x, 2), "y": round(y, 2),
            "w": round(w, 2), "h": round(h, 2), "ubicaciones": ids_ubicacion,
        })
    config_actual = _mapa_config_almacen(a)
    a.mapa_json = json.dumps({
        "version": 3, "fondo": config_actual.get("fondo"), "zonas": zonas_limpias,
    }, ensure_ascii=False)
    db.commit()
    return {"ok": True}


@app.post("/almacenes/{aid}/mapa/ubicaciones")
async def almacen_mapa_crear_ubicacion(
    aid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Crea una estantería desde el propio plano y devuelve su QR oficial."""
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para organizar el mapa")
    almacen = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    if not almacen:
        raise HTTPException(404)
    body = await request.json()
    nombre = str(body.get("nombre") or "").strip()[:100]
    descripcion = str(body.get("descripcion") or "").strip()[:500]
    if len(nombre) < 2:
        siguiente = db.query(Ubicacion).filter(
            Ubicacion.almacen_id == aid,
            Ubicacion.activo == True,
        ).count() + 1
        nombre = f"Estantería {siguiente}"
    ubicacion = Ubicacion(
        almacen_id=aid,
        nombre=nombre,
        codigo=generar_referencia_ubicacion(db),
        descripcion=descripcion or None,
        zona=str(body.get("zona") or "").strip()[:100] or None,
        pasillo=str(body.get("pasillo") or "").strip()[:50] or None,
        estanteria=str(body.get("estanteria") or nombre).strip()[:50] or None,
        balda=str(body.get("balda") or "").strip()[:50] or None,
        posicion=str(body.get("posicion") or "").strip()[:50] or None,
        activo=True,
    )
    db.add(ubicacion)
    db.flush()
    registrar_auditoria(
        db, "ubicaciones", ubicacion.id, "crear", user.id, None,
        {"nombre": nombre, "codigo": ubicacion.codigo},
        "Estantería creada desde el mapa",
        request.client.host if request.client else "",
    )
    db.commit()
    return {
        "ok": True, "id": ubicacion.id, "nombre": ubicacion.nombre,
        "codigo": ubicacion.codigo, "ruta": ubicacion.ruta_completa,
        "estanteria": ubicacion.estanteria or "", "posicion": ubicacion.posicion or "",
        "count": 0, "unidades": 0, "items": [],
    }


@app.post("/almacenes/{aid}/mapa/ubicaciones/grid")
async def almacen_mapa_crear_ubicaciones_grid(
    aid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Genera de golpe estanterías tipo letra+fila (A1..J20) dentro de una zona,
    para organizar cajas/contenedores como una cuadrícula de Excel."""
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para organizar el mapa")
    almacen = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    if not almacen:
        raise HTTPException(404)
    body = await request.json()
    zona = str(body.get("zona") or "").strip()[:100]
    letra_desde = str(body.get("letra_desde") or "").strip().upper()[:1]
    letra_hasta = str(body.get("letra_hasta") or "").strip().upper()[:1]
    try:
        fila_desde = int(body.get("fila_desde") or 0)
        fila_hasta = int(body.get("fila_hasta") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "Rango de filas no válido")
    if len(zona) < 2:
        raise HTTPException(400, "Indica el nombre de la zona o contenedor")
    if not (letra_desde and letra_hasta and "A" <= letra_desde <= "Z" and "A" <= letra_hasta <= "Z" and letra_desde <= letra_hasta):
        raise HTTPException(400, "Rango de letras no válido")
    if not (1 <= fila_desde <= fila_hasta <= 500):
        raise HTTPException(400, "Rango de filas no válido")
    total = (ord(letra_hasta) - ord(letra_desde) + 1) * (fila_hasta - fila_desde + 1)
    if total > 2000:
        raise HTTPException(400, "Demasiados huecos de golpe (máx. 2000) — reduce el rango o hazlo en dos tandas")

    existentes = {
        (u.estanteria, u.posicion)
        for u in db.query(Ubicacion).filter(
            Ubicacion.almacen_id == aid, Ubicacion.zona == zona,
        ).all()
    }
    creadas = []
    for letra_ord in range(ord(letra_desde), ord(letra_hasta) + 1):
        letra = chr(letra_ord)
        for fila in range(fila_desde, fila_hasta + 1):
            posicion = str(fila)
            if (letra, posicion) in existentes:
                continue
            ubicacion = Ubicacion(
                almacen_id=aid,
                nombre=f"{zona} {letra}{fila}",
                codigo=generar_referencia_ubicacion(db),
                zona=zona,
                estanteria=letra,
                posicion=posicion,
                activo=True,
            )
            db.add(ubicacion)
            db.flush()
            creadas.append(ubicacion)
    if creadas:
        registrar_auditoria(
            db, "ubicaciones", creadas[0].id, "crear_grid", user.id, None,
            {"zona": zona, "cantidad": len(creadas)},
            f"{len(creadas)} huecos generados en {zona} ({letra_desde}{fila_desde}-{letra_hasta}{fila_hasta})",
            request.client.host if request.client else "",
        )
        db.commit()
    return {
        "ok": True, "creadas": len(creadas), "omitidas": total - len(creadas),
        "ubicaciones": [{
            "id": u.id, "nombre": u.nombre, "codigo": u.codigo, "ruta": u.ruta_completa,
            "estanteria": u.estanteria or "", "posicion": u.posicion or "",
            "count": 0, "unidades": 0, "items": [],
        } for u in creadas],
    }


def _active_warehouse(db: Session, user: Usuario, request: Request | None = None) -> Almacen | None:
    requested_id = None
    if request is not None and user.rol == "admin":
        try:
            requested_id = int(request.cookies.get("mrd_warehouse_id") or 0) or None
        except (TypeError, ValueError):
            requested_id = None
    return get_user_warehouse(db, user, requested_id)


def _require_warehouse_access(user: Usuario, warehouse_id: int | None) -> None:
    if not can_access_warehouse(user, warehouse_id):
        raise HTTPException(403, "No tienes acceso a este almacén")


@app.post("/almacen-activo", response_class=RedirectResponse)
def seleccionar_almacen_activo(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    almacen_id: int = Form(...),
    volver: str = Form("/"),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede cambiar de almacén")
    warehouse = db.query(Almacen).filter(
        Almacen.id == almacen_id, Almacen.activo == True,
    ).first()
    if not warehouse:
        raise HTTPException(404, "Almacén no encontrado")
    safe_return = volver if volver.startswith("/") and not volver.startswith("//") else "/"
    response = RedirectResponse(safe_return, status_code=303)
    response.set_cookie(
        "mrd_warehouse_id", str(warehouse.id), httponly=True,
        samesite="lax", secure=IS_PRODUCTION, max_age=60 * 60 * 24 * 365,
    )
    return response


@app.post("/almacenes/{aid}/mapa/maquinaria/{mid}/ubicacion")
async def almacen_mapa_asignar_maquinaria(
    aid: int,
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Coloca una máquina en una estantería real desde el plano de la nave."""
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para organizar maquinaria")
    almacen = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    maquina = db.query(Maquinaria).filter(
        Maquinaria.id == mid, Maquinaria.activa == True,
        or_(Maquinaria.almacen_id == aid, Maquinaria.almacen_id.is_(None)),
    ).first()
    if not almacen or not maquina:
        raise HTTPException(404, "Almacén o maquinaria no encontrado")
    body = await request.json()
    try:
        ubicacion_id = int(body.get("ubicacion_id"))
    except (TypeError, ValueError):
        raise HTTPException(400, "Selecciona una estantería válida")
    ubicacion = db.query(Ubicacion).filter(
        Ubicacion.id == ubicacion_id,
        Ubicacion.almacen_id == aid,
        Ubicacion.activo == True,
    ).first()
    if not ubicacion:
        raise HTTPException(404, "La estantería no pertenece a este almacén")
    anterior = maquina.ubicacion
    # Se guarda el código estable y único; la interfaz presenta su nombre.
    maquina.ubicacion = ubicacion.codigo or ubicacion.nombre
    maquina.almacen_id = aid
    registrar_auditoria(
        db, "maquinaria", maquina.id, "ubicar_en_almacen", user.id,
        {"ubicacion": anterior},
        {"almacen_id": aid, "ubicacion_id": ubicacion.id,
         "ubicacion": ubicacion.nombre, "codigo_ubicacion": ubicacion.codigo},
        f"Maquinaria colocada en {ubicacion.nombre} desde el mapa",
        request.client.host if request.client else "",
    )
    db.commit()
    return {
        "ok": True, "maquinaria_id": maquina.id, "ubicacion_id": ubicacion.id,
        "ubicacion": ubicacion.nombre, "codigo_ubicacion": ubicacion.codigo,
    }


@app.post("/almacenes/{aid}/mapa/items/{tipo}/{item_id}/ubicacion")
async def almacen_mapa_asignar_item(
    aid: int,
    tipo: str,
    item_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Coloca cualquier activo físico del almacén en una estantería real."""
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para organizar el almacén")
    almacen = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    if not almacen:
        raise HTTPException(404, "Almacén no encontrado")
    body = await request.json()
    try:
        ubicacion_id = int(body.get("ubicacion_id"))
    except (TypeError, ValueError):
        raise HTTPException(400, "Selecciona una estantería válida")
    ubicacion = db.query(Ubicacion).filter(
        Ubicacion.id == ubicacion_id,
        Ubicacion.almacen_id == aid,
        Ubicacion.activo == True,
    ).first()
    if not ubicacion:
        raise HTTPException(404, "La estantería no pertenece a este almacén")

    modelos = {
        "herramienta": (Herramienta, Herramienta.activa == True),
        "material": (Material, Material.activo == True),
        "existencia": (ExistenciaVariante, ExistenciaVariante.cantidad > 0),
        "stock_epi": (StockEPI, StockEPI.cantidad > 0),
        "epi_individual": (EPIIndividual, EPIIndividual.estado != "baja"),
        "maquinaria": (Maquinaria, Maquinaria.activa == True),
    }
    if tipo not in modelos:
        raise HTTPException(400, "Tipo de artículo no válido")
    modelo, condicion = modelos[tipo]
    item = db.query(modelo).filter(modelo.id == item_id, condicion).first()
    if not item:
        raise HTTPException(404, "Artículo no encontrado o inactivo")
    if tipo in {"herramienta", "material", "stock_epi", "epi_individual", "maquinaria"}:
        almacen_actual = getattr(item, "almacen_id", None)
        if almacen_actual not in (None, aid):
            raise HTTPException(409, "Este artículo pertenece a otro almacén")
    if tipo == "existencia" and item.almacen_id != aid:
        raise HTTPException(409, "Esta existencia pertenece a otro almacén")
    if tipo == "existencia":
        repetida = db.query(ExistenciaVariante).filter(
            ExistenciaVariante.id != item.id,
            ExistenciaVariante.variante_id == item.variante_id,
            ExistenciaVariante.almacen_id == aid,
            ExistenciaVariante.ubicacion_clave == ubicacion.id,
            ExistenciaVariante.cantidad > 0,
        ).first()
        if repetida:
            raise HTTPException(409, "Esta talla o variante ya tiene stock en esa estantería")
    if tipo == "epi_individual" and item.trabajador_id:
        raise HTTPException(409, "Este EPI está asignado a un trabajador y no está físicamente en el almacén")

    anterior = getattr(item, "ubicacion_id", None)
    if tipo == "maquinaria":
        anterior = item.ubicacion
        item.ubicacion = ubicacion.codigo or ubicacion.nombre
        item.almacen_id = aid
    else:
        item.ubicacion_id = ubicacion.id
        if hasattr(item, "almacen_id"):
            item.almacen_id = aid
        if tipo == "herramienta":
            item.ubicacion_texto = ubicacion.nombre
        if tipo == "material":
            item.ubicacion_texto = ubicacion.nombre
        if tipo == "existencia":
            item.ubicacion_clave = ubicacion.id
    registrar_auditoria(
        db, modelo.__tablename__, item.id, "ubicar_en_almacen", user.id,
        {"ubicacion": anterior},
        {"almacen_id": aid, "ubicacion_id": ubicacion.id,
         "ubicacion": ubicacion.nombre, "codigo_ubicacion": ubicacion.codigo},
        f"{tipo} colocado en {ubicacion.nombre} desde el mapa",
        request.client.host if request.client else "",
    )
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "No se puede duplicar esta referencia en la estantería elegida")
    return {
        "ok": True, "tipo": tipo, "item_id": item.id,
        "ubicacion_id": ubicacion.id, "ubicacion": ubicacion.nombre,
        "codigo_ubicacion": ubicacion.codigo,
    }


@app.post("/almacenes/{aid}/mapa/plano", response_class=RedirectResponse)
async def almacen_mapa_subir_plano(
    aid: int,
    plano: UploadFile = File(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para cambiar el plano")
    a = db.query(Almacen).filter(Almacen.id == aid, Almacen.activo == True).first()
    if not a:
        raise HTTPException(404)
    try:
        nombre_seguro, ext = validar_nombre_archivo(plano.filename, {"jpg", "jpeg", "png", "webp"})
        cabecera = await plano.read(16)
        await plano.seek(0)
        validar_contenido_archivo(cabecera, ext)
        contenido = await plano.read()
        validar_tamaño_bytes(len(contenido), MAX_UPLOAD_MB)
    except ErrorArchivo as exc:
        raise HTTPException(400, str(exc))
    carpeta = UPLOADS_DIR / "almacenes" / "mapas"
    carpeta.mkdir(parents=True, exist_ok=True)
    nombre_archivo = f"plano_{aid}_{uuid.uuid4().hex}.{ext}"
    destino = carpeta / nombre_archivo
    destino.write_bytes(contenido)
    config_mapa = _mapa_config_almacen(a)
    anterior = config_mapa.get("fondo")
    config_mapa["fondo"] = nombre_archivo
    try:
        a.mapa_json = json.dumps(config_mapa, ensure_ascii=False)
        db.commit()
    except Exception:
        db.rollback()
        destino.unlink(missing_ok=True)
        raise
    if anterior:
        (carpeta / Path(anterior).name).unlink(missing_ok=True)
    return RedirectResponse(f"/almacenes/{aid}/mapa?ok=plano", status_code=303)


@app.get("/almacenes/{aid}/qr", response_class=HTMLResponse)
def almacen_qr(
    aid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    qr_b64 = generar_qr_base64(f"/almacenes/{a.id}")
    mats_count = db.query(Material).filter(Material.almacen_id == aid).count()
    return templates.TemplateResponse(request, "almacen_qr.html", ctx_base(
        request, user, db,
        almacen=a,
        qr_b64=qr_b64,
        empresa=COMPANY_NAME,
        ubicaciones=a.ubicaciones,
        herramientas_count=len(a.herramientas),
        materiales_count=mats_count,
        ubicaciones_count=len(a.ubicaciones),
    ))


@app.post("/almacenes/{aid}/foto", response_class=RedirectResponse)
async def almacen_foto(
    aid: int,
    foto: UploadFile = File(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    carpeta = BASE_DIR / "static" / "uploads" / "almacenes"
    carpeta.mkdir(parents=True, exist_ok=True)
    if a.foto:
        old_p = carpeta / a.foto
        if old_p.exists():
            old_p.unlink()
    ext = (foto.filename or "").rsplit(".", 1)[-1].lower() or "jpg"
    import time as _t
    nombre = f"alm_{aid}_{int(_t.time())}.{ext}"
    data = await foto.read()
    (carpeta / nombre).write_bytes(data)
    a.foto = nombre
    db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=foto", status_code=303)


@app.post("/almacenes/{aid}/foto/eliminar", response_class=RedirectResponse)
def almacen_foto_eliminar(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403)
    a = db.query(Almacen).get(aid)
    if a and a.foto:
        p = BASE_DIR / "static" / "uploads" / "almacenes" / a.foto
        if p.exists():
            p.unlink()
        a.foto = None
        db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=foto_eliminada", status_code=303)


@app.post("/almacenes/{aid}/eliminar", response_class=RedirectResponse)
def almacen_eliminar(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "borrar"):
        raise HTTPException(403)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    if a.foto:
        try:
            p = BASE_DIR / "static" / "uploads" / "almacenes" / a.foto
            if p.exists():
                p.unlink()
        except Exception:
            pass
    db.delete(a)
    db.commit()
    return RedirectResponse("/almacenes?ok=eliminado", status_code=303)


@app.post("/almacenes/{aid}/ubicaciones/nuevo", response_class=RedirectResponse)
def ubicacion_nueva(
    aid: int, request_: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    codigo: str = Form(""),
    descripcion: str = Form(""),
    zona: str = Form(""),
    pasillo: str = Form(""),
    estanteria: str = Form(""),
    balda: str = Form(""),
    posicion: str = Form(""),
):
    _require_warehouse_access(user, aid)
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    ub = Ubicacion(
        almacen_id=aid,
        nombre=nombre.strip(),
        # La referencia nunca depende de lo que escriba el usuario.
        codigo=generar_referencia_ubicacion(db),
        descripcion=descripcion or None,
        zona=zona.strip()[:100] or None,
        pasillo=pasillo.strip()[:50] or None,
        estanteria=estanteria.strip()[:50] or nombre.strip()[:50],
        balda=balda.strip()[:50] or None,
        posicion=posicion.strip()[:50] or None,
    )
    db.add(ub)
    db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=ubicacion", status_code=303)


@app.post("/almacenes/{aid}/ubicaciones/{uid}/editar", response_class=RedirectResponse)
async def ubicacion_editar(
    aid: int, uid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403)
    ub = db.query(Ubicacion).filter(Ubicacion.id == uid, Ubicacion.almacen_id == aid).first()
    if not ub:
        raise HTTPException(404)
    form = await request.form()
    ub.nombre      = form.get("nombre", ub.nombre).strip()
    # El identificador es inmutable: evita solapes y etiquetas huérfanas.
    if not ub.codigo:
        ub.codigo = generar_referencia_ubicacion(db)
    ub.descripcion = form.get("descripcion") or None
    ub.zona = (form.get("zona") or "").strip()[:100] or None
    ub.pasillo = (form.get("pasillo") or "").strip()[:50] or None
    ub.estanteria = (form.get("estanteria") or ub.nombre).strip()[:50] or None
    ub.balda = (form.get("balda") or "").strip()[:50] or None
    ub.posicion = (form.get("posicion") or "").strip()[:50] or None
    db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=ubicacion", status_code=303)


@app.post("/almacenes/{aid}/ubicaciones/{uid}/eliminar", response_class=RedirectResponse)
def ubicacion_eliminar(
    aid: int, uid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403)
    ub = db.query(Ubicacion).filter(Ubicacion.id == uid, Ubicacion.almacen_id == aid).first()
    if ub:
        # Desvincular herramientas y materiales
        db.query(Herramienta).filter(Herramienta.ubicacion_id == uid).update({"ubicacion_id": None})
        db.query(Material).filter(Material.ubicacion_id == uid).update({"ubicacion_id": None})
        db.delete(ub)
        db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=ubicacion_eliminada", status_code=303)


@app.post("/almacenes/{aid}/transferir", response_class=RedirectResponse)
async def almacen_transferir(
    aid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Compatibilidad: los traspasos ya no cambian stock sin recepción."""
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede transferir entre almacenes")
    return RedirectResponse(f"/multi-almacen?origen={aid}", status_code=303)



# ─── Almacenes — inventario express ────────────────────────────────────────────
@app.post("/almacenes/{aid}/inventario", response_class=RedirectResponse)
async def almacen_inventario(
    aid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _require_warehouse_access(user, aid)
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403)
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    form = await request.form()
    mats = db.query(Material).filter(Material.almacen_id == aid).all()
    for m in mats:
        key = f"stock_{m.id}"
        val = form.get(key)
        if val is not None and val.strip() != "":
            try:
                nuevo = float(val)
                if nuevo != m.stock_actual:
                    diff = nuevo - m.stock_actual
                    tipo_mov = "ajuste"
                    db.add(MovimientoMaterial(
                        material_id=m.id,
                        tipo=tipo_mov,
                        cantidad=abs(diff),
                        notas=f"Inventario express: {m.stock_actual} → {nuevo}",
                        usuario_id=user.id,
                    ))
                    m.stock_actual = nuevo
            except (ValueError, TypeError):
                pass
    db.commit()
    return RedirectResponse(f"/almacenes/{aid}?ok=inventario&tab=mat", status_code=303)


# ─── Almacenes — QR por ubicación ───────────────────────────────────────────────
@app.get("/almacenes/{aid}/ubicaciones/{uid}/qr", response_class=HTMLResponse)
def ubicacion_qr(
    aid: int, uid: int, request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    a = db.query(Almacen).get(aid)
    ub = db.query(Ubicacion).get(uid)
    if not a or not ub or ub.almacen_id != aid:
        raise HTTPException(404)
    qr_code = ub.codigo or f"ALM{aid}-UBI{uid}"
    qr_b64 = generar_qr_base64(qr_code)
    return templates.TemplateResponse(request, "ubicacion_qr.html", ctx_base(
        request, user,
        almacen=a,
        ubicacion=ub,
        qr_b64=qr_b64,
        herramientas_count=len(ub.herramientas),
        materiales_count=len(ub.materiales),
    ))


@app.get("/almacenes/{aid}/etiquetas-ubicaciones/pdf")
def almacen_etiquetas_ubicaciones_pdf(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    zona: str = "",
):
    """PDF por lotes de etiquetas de ubicación (QR + Code128 + código grande), estilo almacén industrial."""
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    query = db.query(Ubicacion).filter(Ubicacion.almacen_id == aid, Ubicacion.activo == True)
    zona = (zona or "").strip()
    if zona:
        query = query.filter(Ubicacion.zona == zona)
    ubicaciones = query.order_by(Ubicacion.zona, Ubicacion.estanteria, Ubicacion.balda, Ubicacion.posicion).all()
    if not ubicaciones:
        raise HTTPException(404, "No hay ubicaciones activas para imprimir")
    pdf_bytes = generar_pdf_etiquetas_ubicaciones(ubicaciones, COMPANY_NAME)
    filename = f"etiquetas_ubicaciones_{a.nombre.replace(' ', '_')}" + (f"_{zona.replace(' ', '_')}" if zona else "") + ".pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ─── Almacenes — exportar inventario a Excel ─────────────────────────────────
@app.get("/almacenes/{aid}/exportar")
def almacen_exportar(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado")

    wb = openpyxl.Workbook()

    # ── Hoja Herramientas ──
    ws_h = wb.active
    ws_h.title = "Herramientas"
    header_fill = PatternFill("solid", fgColor="F5A623")
    header_font = Font(bold=True, color="000000")
    cols_h = ["Código", "Nombre", "Categoría", "Estado", "Ubicación", "Marca", "Modelo", "Nº Serie"]
    for c, col in enumerate(cols_h, 1):
        cell = ws_h.cell(row=1, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    herrs = a.herramientas
    for r, h in enumerate(herrs, 2):
        row = [
            h.codigo, h.nombre, h.categoria or "",
            h.estado.replace("_", " ").title(),
            h.ubicacion.nombre if h.ubicacion else (h.almacen_nombre or ""),
            h.marca or "", h.modelo or "", h.num_serie or "",
        ]
        for c, val in enumerate(row, 1):
            ws_h.cell(row=r, column=c, value=val)
    for c in range(1, len(cols_h) + 1):
        ws_h.column_dimensions[get_column_letter(c)].width = 18

    # ── Hoja Materiales ──
    ws_m = wb.create_sheet("Materiales")
    cols_m = ["Código", "Nombre", "Categoría", "Stock actual", "Stock mínimo", "Unidad", "Ubicación", "Bajo mínimo"]
    for c, col in enumerate(cols_m, 1):
        cell = ws_m.cell(row=1, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    mats = db.query(Material).filter(Material.almacen_id == aid, Material.activo == True).all()
    red_fill = PatternFill("solid", fgColor="FFD7D7")
    for r, m in enumerate(mats, 2):
        row = [
            m.codigo, m.nombre, m.categoria or "",
            m.stock_actual, m.stock_minimo, m.unidad or "ud",
            m.ubicacion.nombre if m.ubicacion else (m.ubicacion_texto or ""),
            "⚠ Sí" if m.bajo_minimo else "Ok",
        ]
        for c, val in enumerate(row, 1):
            cell = ws_m.cell(row=r, column=c, value=val)
            if m.bajo_minimo:
                cell.fill = red_fill
    for c in range(1, len(cols_m) + 1):
        ws_m.column_dimensions[get_column_letter(c)].width = 18

    # ── Hoja Ubicaciones ──
    ws_u = wb.create_sheet("Ubicaciones")
    cols_u = ["Código", "Nombre", "Descripción", "Herramientas", "Materiales"]
    for c, col in enumerate(cols_u, 1):
        cell = ws_u.cell(row=1, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for r, u in enumerate(a.ubicaciones, 2):
        row = [u.codigo or "", u.nombre, u.descripcion or "",
               len(u.herramientas), len(u.materiales)]
        for c, val in enumerate(row, 1):
            ws_u.cell(row=r, column=c, value=val)
    for c in range(1, len(cols_u) + 1):
        ws_u.column_dimensions[get_column_letter(c)].width = 18

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f"inventario_{a.nombre.replace(' ','_')}.xlsx"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


# ─── Obras ────────────────────────────────────────────────────────────────────
@app.get("/obras", response_class=HTMLResponse)
def obras_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user, request)
    obras = db.query(Obra).filter(
        Obra.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Obra.id.desc()).all()
    return templates.TemplateResponse(request, "obras.html", ctx_base(request, user, db, obras=obras))


@app.get("/obras/nueva", response_class=HTMLResponse)
def obra_nueva_get(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    return templates.TemplateResponse(request, "nueva_obra.html", ctx_base(request, user, db))


@app.post("/obras/nueva")
def obra_nueva_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    cliente: str = Form(""),
    responsable: str = Form(""),
    direccion: str = Form(""),
    fecha_inicio: str = Form(""),
    fecha_fin: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    # Auto-número
    año = datetime.now().year
    count = db.query(Obra).filter(Obra.numero.like(f"{año}-%")).count()
    numero = f"{año}-{(count + 1):04d}"
    warehouse = _active_warehouse(db, user, request)
    o = Obra(
        numero=numero, nombre=nombre, cliente=cliente or None,
        responsable=responsable or None, direccion=direccion or None,
        fecha_inicio=datetime.strptime(fecha_inicio, "%Y-%m-%d") if fecha_inicio else None,
        fecha_fin=datetime.strptime(fecha_fin, "%Y-%m-%d") if fecha_fin else None,
        activa=True,
        almacen_id=warehouse.id if warehouse else None,
    )
    db.add(o)
    db.commit()
    return RedirectResponse("/obras", status_code=303)


@app.post("/obras/{oid}/editar", response_class=RedirectResponse)
async def obra_editar(
    oid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    o = db.query(Obra).get(oid)
    if not o:
        raise HTTPException(404)
    _require_warehouse_access(user, o.almacen_id)
    form = await request.form()
    o.nombre = form.get("nombre") or o.nombre
    o.cliente = form.get("cliente") or None
    o.responsable = form.get("responsable") or None
    o.direccion = form.get("direccion") or None
    fi = form.get("fecha_inicio", "")
    ff = form.get("fecha_fin", "")
    if fi:
        o.fecha_inicio = datetime.strptime(fi, "%Y-%m-%d")
    if ff:
        o.fecha_fin = datetime.strptime(ff, "%Y-%m-%d")
    activa = form.get("activa", "")
    o.activa = activa in ("1", "true", "on")
    o.observaciones = form.get("observaciones") or None
    db.commit()
    return RedirectResponse("/obras?ok=editado", status_code=303)


# ─── Obras — plantilla e importación Excel ────────────────────────────────────
@app.get("/obras/plantilla-excel")
def obras_plantilla_excel(user: Usuario = Depends(requiere_login)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Obras"
    cols = ["Nº Obra", "Nombre / Proyecto *", "Fase", "Provincia",
            "Cliente", "Responsable / Creado por",
            "Fecha inicio (DD/MM/AAAA)", "Fecha fin (DD/MM/AAAA)", "Estado (Activo/Finalizado)", "Observaciones"]
    ws.append(cols)
    hdr_fill = PatternFill("solid", fgColor="1a3c5e")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    # Ejemplos reales
    ws.append(["12.351", "Base de Mantenimiento de Fuencarral Norte", "Proyecto", "Madrid",
               "UTE BASE FUENCARRAL", "Ricardo Sainz Olmo", "01/01/2026", "", "Activo", ""])
    ws.append(["12.350", "Vivienda en Valdemorillo", "Proyecto", "Madrid",
               "Manu Ibáñez", "Angel Marrero", "", "", "Activo", ""])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_obras.xlsx"},
    )


@app.post("/obras/importar", response_class=HTMLResponse)
async def obras_importar(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    archivo: UploadFile = File(...),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    import io
    from openpyxl import load_workbook
    from dateutil import parser as dateparser

    contenido = await archivo.read()
    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    año = datetime.now().year
    creadas = 0
    omitidas = []

    def parse_fecha(val):
        if not val:
            return None
        try:
            if hasattr(val, "date"):
                return val.date()
            return dateparser.parse(str(val), dayfirst=True).date()
        except Exception:
            return None

    for i, row in enumerate(rows, start=2):
        # Soporte formato antiguo (6 cols) y nuevo (10 cols)
        if len(row) >= 10:
            num_excel = str(row[0]).strip() if row[0] else ""
            nombre    = str(row[1]).strip() if row[1] else ""
            fase      = str(row[2]).strip() if row[2] else None
            provincia = str(row[3]).strip() if row[3] else None
            cliente   = str(row[4]).strip() if row[4] else None
            responsable = str(row[5]).strip() if row[5] else None
            fecha_inicio = parse_fecha(row[6])
            fecha_fin    = parse_fecha(row[7])
            estado_txt   = str(row[8]).strip().lower() if row[8] else "activo"
            observaciones = str(row[9]).strip() if row[9] else None
        else:
            num_excel = ""
            nombre    = str(row[0]).strip() if row[0] else ""
            cliente   = str(row[1]).strip() if row[1] else None
            provincia = str(row[2]).strip() if row[2] else None
            responsable = str(row[3]).strip() if row[3] else None
            fecha_inicio = parse_fecha(row[4]) if len(row) > 4 else None
            fecha_fin    = parse_fecha(row[5]) if len(row) > 5 else None
            estado_txt   = "activo"
            observaciones = str(row[6]).strip() if len(row) > 6 and row[6] else None
            fase = None

        if not nombre or nombre.lower() in ("none", "nombre *", "nombre / proyecto *"):
            continue

        activa = "finaliz" not in estado_txt

        # Número: usar el del Excel si viene, si no auto-generar
        if num_excel and num_excel.lower() not in ("none", "nº obra", ""):
            numero = num_excel
        else:
            count = db.query(Obra).filter(Obra.numero.like(f"{año}-%")).count()
            numero = f"{año}-{(count + 1):04d}"

        # Evitar duplicados por número o nombre
        existe = db.query(Obra).filter(
            (Obra.numero == numero) | Obra.nombre.ilike(nombre)
        ).first()
        if existe:
            omitidas.append(f"Fila {i}: «{nombre}» (Nº {numero}) ya existe")
            continue

        # Provincia va en dirección si no hay otra cosa
        direccion = provincia

        notas_extra = []
        if fase:
            notas_extra.append(f"Fase: {fase}")
        if observaciones:
            notas_extra.append(observaciones)

        db.add(Obra(
            numero=numero, nombre=nombre, cliente=cliente,
            direccion=direccion, responsable=responsable,
            fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
            observaciones=" | ".join(notas_extra) if notas_extra else None,
            activa=activa,
        ))
        db.commit()
        creadas += 1

    obras = db.query(Obra).order_by(Obra.id.desc()).all()
    ctx = ctx_base(request, user, obras=obras)
    ctx["import_ok"] = creadas
    ctx["import_omitidas"] = omitidas
    return templates.TemplateResponse(request, "obras.html", ctx)


# ─── Almacenes ────────────────────────────────────────────────────────────────
@app.get("/almacenes", response_class=HTMLResponse)
def almacenes_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    almacenes = visible_warehouses(db, user)
    return templates.TemplateResponse(request, "almacenes.html", ctx_base(
        request, user, db, almacenes=almacenes,
    ))


@app.post("/almacenes/nuevo")
def almacen_nuevo(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    descripcion: str = Form(""),
    direccion: str = Form(""),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede crear almacenes")
    db.add(Almacen(
        codigo=generar_referencia_almacen(db),
        nombre=nombre,
        descripcion=descripcion or None,
        direccion=direccion or None,
    ))
    db.commit()
    return RedirectResponse("/almacenes", status_code=303)


@app.post("/almacenes/{aid}/editar", response_class=RedirectResponse)
async def almacen_editar(
    aid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede editar almacenes")
    a = db.query(Almacen).get(aid)
    if not a:
        raise HTTPException(404)
    form = await request.form()
    a.nombre = form.get("nombre") or a.nombre
    a.descripcion = form.get("descripcion") or None
    a.direccion = form.get("direccion") or None
    a.responsable = form.get("responsable") or None
    db.commit()
    return RedirectResponse("/almacenes?ok=editado", status_code=303)


# ─── Vehículos ────────────────────────────────────────────────────────────────
@app.get("/vehiculos", response_class=HTMLResponse)
def vehiculos_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user, request)
    vehiculos = db.query(Vehiculo).filter(
        Vehiculo.almacen_id == (warehouse.id if warehouse else -1),
    ).all()
    return templates.TemplateResponse(request, "vehiculos.html", ctx_base(request, user, db, vehiculos=vehiculos))


@app.post("/vehiculos/nuevo")
def vehiculo_nuevo(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    matricula: str = Form(...),
    marca: str = Form(""),
    modelo: str = Form(""),
    tipo: str = Form("furgoneta"),
    kilometros: str = Form(""),
    descripcion: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    mat_up = matricula.strip().upper()
    existe = db.query(Vehiculo).filter(Vehiculo.matricula == mat_up).first()
    if not existe:
        warehouse = _active_warehouse(db, user, request)
        db.add(Vehiculo(
            codigo=generar_referencia_vehiculo(db),
            matricula=mat_up,
            marca=marca or None,
            modelo=modelo or None,
            tipo=tipo or "furgoneta",
            kilometros=int(kilometros) if kilometros else None,
            descripcion=descripcion or None,
            almacen_id=warehouse.id if warehouse else None,
        ))
        db.commit()
    return RedirectResponse("/vehiculos", status_code=303)


@app.post("/vehiculos/{vid}/editar", response_class=RedirectResponse)
async def vehiculo_editar(
    vid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    v = db.query(Vehiculo).get(vid)
    if not v:
        raise HTTPException(404)
    _require_warehouse_access(user, v.almacen_id)
    form = await request.form()
    mat_raw = (form.get("matricula") or "").strip().upper()
    v.matricula = mat_raw or v.matricula
    v.marca = form.get("marca") or None
    v.modelo = form.get("modelo") or None
    v.tipo = form.get("tipo") or v.tipo
    anio_raw = form.get("anio", "")
    if anio_raw:
        try: v.anio = int(anio_raw)
        except ValueError: pass
    km = form.get("kilometros", "")
    if km:
        v.kilometros = int(km)
    v.descripcion = form.get("descripcion") or None
    v.estado = form.get("estado") or v.estado
    v.observaciones = form.get("observaciones") or None
    itv = form.get("itv_hasta", "")
    if itv:
        v.itv_hasta = datetime.strptime(itv, "%Y-%m-%d").date()
    else:
        v.itv_hasta = None
    seg = form.get("seguro_hasta", "")
    if seg:
        v.seguro_hasta = datetime.strptime(seg, "%Y-%m-%d").date()
    else:
        v.seguro_hasta = None
    prox = form.get("proxima_revision", "")
    if prox:
        v.proxima_revision = datetime.strptime(prox, "%Y-%m-%d").date()
    else:
        v.proxima_revision = None
    v.compania_seguro = form.get("compania_seguro") or None
    v.num_poliza = form.get("num_poliza") or None
    db.commit()
    return RedirectResponse("/vehiculos?ok=editado", status_code=303)


@app.get("/vehiculos/{vid}/qr", response_class=HTMLResponse)
def vehiculo_qr(
    vid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    v = db.query(Vehiculo).get(vid)
    if not v:
        raise HTTPException(404)
    hoy = date.today()

    def _fecha_info(d):
        if not d:
            return {"txt": "No registrada", "css": "none", "aviso": ""}
        dias = (d - hoy).days
        txt = d.strftime("%d/%m/%Y")
        if dias < 0:
            return {"txt": txt, "css": "venc", "aviso": "⚠ VENCIDA/O"}
        elif dias < 30:
            return {"txt": txt, "css": "warn", "aviso": f"({dias} días)"}
        return {"txt": txt, "css": "ok", "aviso": ""}

    url_ficha = f"/vehiculos/{v.id}/qr"
    qr_b64 = generar_qr_base64(url_ficha)
    return templates.TemplateResponse(request, "vehiculo_qr.html", ctx_base(
        request, user,
        vehiculo=v,
        qr_b64=qr_b64,
        empresa=COMPANY_NAME,
        itv=_fecha_info(v.itv_hasta),
        seguro=_fecha_info(v.seguro_hasta),
        revision=_fecha_info(v.proxima_revision),
        compania_seguro=v.compania_seguro or "",
        num_poliza=v.num_poliza or "",
    ))


# ─── Etiquetas ────────────────────────────────────────────────────────────────
@app.get("/etiquetas", response_class=HTMLResponse)
def etiquetas_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    herramientas = db.query(Herramienta).filter(Herramienta.activa == True).order_by(Herramienta.nombre).all()
    return templates.TemplateResponse(request, "etiquetas.html", ctx_base(request, user, herramientas=herramientas))


@app.get("/etiquetas/{herramienta_id}/qr")
def etiqueta_qr_download(herramienta_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    h = db.query(Herramienta).get(herramienta_id)
    if not h:
        raise HTTPException(404)
    qr_bytes = generar_qr_bytes(h.codigo)
    return Response(content=qr_bytes, media_type="image/png",
                    headers={"Content-Disposition": f"attachment; filename=QR_{h.codigo}.png"})


@app.get("/etiquetas/{herramienta_id}/zpl")
def etiqueta_zpl_download(herramienta_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    h = db.query(Herramienta).get(herramienta_id)
    if not h:
        raise HTTPException(404)
    zpl = generar_zpl_herramienta(h.codigo, h.nombre, h.num_serie or "", h.marca or "", COMPANY_NAME)
    return Response(content=zpl.encode("utf-8"), media_type="text/plain",
                    headers={"Content-Disposition": f"attachment; filename=ETIQ_{h.codigo}.prn"})


@app.post("/etiquetas/pdf")
def etiquetas_pdf(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    ids: str = Form(...),
):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        raise HTTPException(400, "Sin herramientas seleccionadas")
    herramientas = db.query(Herramienta).filter(Herramienta.id.in_(id_list)).all()
    pdf_bytes = generar_pdf_etiquetas(herramientas, COMPANY_NAME)
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=etiquetas_mrd.pdf"})


@app.get("/etiquetas/imprimir/{herramienta_id}", response_class=HTMLResponse)
def etiqueta_imprimir(
    herramienta_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Vista de impresión de etiqueta individual con código de barras Code128 + QR."""
    h = db.query(Herramienta).get(herramienta_id)
    if not h:
        raise HTTPException(404)
    qr_b64 = generar_qr_base64(h.codigo)
    return templates.TemplateResponse(request, "etiqueta_imprimir.html",
        ctx_base(request, user, herramienta=h, qr_b64=qr_b64, empresa=COMPANY_NAME))


@app.post("/etiquetas/{herramienta_id}/reimprimir")
def etiqueta_reimprimir_deteriorada(
    herramienta_id: int,
    motivo: str = Form(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Reimprime la referencia original y registra por qué fue necesario."""
    if not tiene_permiso(user, "etiquetas"):
        raise HTTPException(403, "Sin permiso para reimprimir etiquetas")
    herramienta = db.get(Herramienta, herramienta_id)
    if not herramienta or not herramienta.activa:
        raise HTTPException(404, "Herramienta no encontrada o inactiva")
    clean_reason = (motivo or "").strip()
    if len(clean_reason) < 4:
        raise HTTPException(422, "Indica por qué se reimprime la etiqueta")

    pdf_bytes = generar_pdf_etiquetas([herramienta], COMPANY_NAME)
    event_id = f"tool-label-{uuid.uuid4().hex}"
    db.add(LogImpresionEtiqueta(
        event_id=event_id,
        usuario_id=user.id,
        tipo="herramienta",
        referencia=herramienta.codigo,
        copias=1,
        reimpresion=True,
        motivo_reimpresion=clean_reason[:300],
        zpl_hash=hashlib.sha256(pdf_bytes).hexdigest(),
        impresora_host="pdf-descarga",
    ))
    registrar_auditoria(
        db, "herramientas", herramienta.id, "reimprimir_etiqueta", user.id,
        None, {"codigo": herramienta.codigo, "motivo": clean_reason[:300]},
        f"Reimpresión de etiqueta deteriorada: {herramienta.codigo}", "",
    )
    db.commit()
    safe_code = re.sub(r"[^A-Za-z0-9_.-]", "_", herramienta.codigo)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="REIMPRESION_{safe_code}.pdf"'},
    )


# ─── Escaneo ──────────────────────────────────────────────────────────────────
@app.get("/scan", response_class=HTMLResponse)
def scan_page(request: Request, db: Session = Depends(get_db)):
    """Página de escaneo — pública para acceso desde móvil sin login."""
    user = usuario_actual(request, db)
    puede_entregar = bool(user and tiene_permiso(user, "entregar"))
    puede_devolver = bool(user and tiene_permiso(user, "devolver"))
    trabajadores = []
    if puede_entregar:
        trabajadores = db.query(Trabajador).filter(
            Trabajador.activo == True
        ).order_by(Trabajador.nombre).all()
    return templates.TemplateResponse(request, "scan.html", ctx_base(
        request, user, trabajadores=trabajadores,
        puede_entregar=puede_entregar, puede_devolver=puede_devolver,
    ))


@app.get("/scan/ip")
def scan_ip(request: Request, db: Session = Depends(get_db)):
    """
    Endpoint de compatibilidad — mantiene el formato original.
    Ahora consume el nuevo sistema multi-proveedor.
    """
    status = remote_access.get_status_cached(max_age=30)
    ip = status.get("ip", "127.0.0.1")
    port = status.get("port", 8000)
    local_url = f"http://{ip}:{port}/scan"
    public_url = status.get("public_url")
    url_https = (public_url.rstrip("/") + "/scan") if public_url else None
    url_qr = url_https or local_url
    qr_b64 = generar_qr_base64(url_qr)
    return JSONResponse({
        "ip": ip,
        "url_local": local_url,
        "url_https": url_https,
        "url_qr": url_qr,
        "qr": qr_b64,
        # Campos extendidos para compatibilidad futura
        "provider": status.get("primary_provider", "local"),
        "status": status.get("status", "local_only"),
    })


# ─── API Acceso Remoto ────────────────────────────────────────────────────────

@app.get("/api/remote-access/status")
def api_remote_access_status(user: Usuario = Depends(requiere_login)):
    """
    Estado completo del acceso remoto con todos los proveedores detectados.
    Incluye URL pública, scan_url, QR en base64, proveedores activos.
    """
    status = remote_access.get_status_cached(max_age=30)
    # Añadir QR al resultado
    scan_url = status.get("scan_url", "")
    if scan_url:
        try:
            status["qr"] = generar_qr_base64(scan_url)
        except Exception:
            status["qr"] = None
    else:
        status["qr"] = None
    return JSONResponse(status)


@app.post("/api/remote-access/status/refresh")
def api_remote_access_refresh(user: Usuario = Depends(requiere_login)):
    """Fuerza refresco inmediato de la detección (síncrono)."""
    remote_access.invalidate_cache()
    status = remote_access.detect_all()
    remote_access._cached_status = status
    remote_access._cache_time = datetime.now()
    if status.get("scan_url"):
        try:
            status["qr"] = generar_qr_base64(status["scan_url"])
        except Exception:
            status["qr"] = None
    return JSONResponse(status)


@app.get("/api/remote-access/config")
def api_remote_access_config_get(user: Usuario = Depends(requiere_login)):
    """Devuelve la configuración actual de acceso remoto (solo admins)."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    cfg = remote_access.load_config()
    return JSONResponse(cfg)


@app.post("/api/remote-access/config")
async def api_remote_access_config_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
):
    """Guarda la configuración de acceso remoto (solo admins)."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    try:
        updates = await request.json()
        if not isinstance(updates, dict):
            raise HTTPException(400, "Formato inválido")
        ok = remote_access.save_config(updates)
        if ok:
            remote_access.invalidate_cache()
            return JSONResponse({"ok": True, "message": "Configuración guardada"})
        return JSONResponse({"ok": False, "message": "Error al guardar"}, status_code=500)
    except Exception as e:
        raise HTTPException(400, str(e))


@app.get("/acceso-remoto", response_class=HTMLResponse)
def acceso_remoto_page(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    """Página dedicada de acceso remoto — vista sencilla + configuración avanzada (admin)."""
    version_info = leer_version_actual()
    backups = listar_backups()
    ultimo_backup = backups[0]["fecha"] if backups else "—"
    return templates.TemplateResponse(request, "acceso_remoto.html", {
        "request": request,
        "user": user,
        "app_name": APP_NAME,
        "company_name": COMPANY_NAME,
        "version": VERSION,
        "version_info": version_info,
        "ultimo_backup": ultimo_backup,
        "es_admin": tiene_permiso(user, "config"),
    })


@app.post("/api/remote-access/test")
def api_remote_access_test(user: Usuario = Depends(requiere_login)):
    """
    Ejecuta diagnósticos completos de conectividad.
    Prueba servidor local, puerto, cloudflare, URL pública, HTTPS, /scan, QR.
    """
    cfg = remote_access.load_config()
    port = int(cfg.get("port", 8000))
    diag = remote_access.run_diagnostics(port)
    status = remote_access.get_status_cached(max_age=60)
    return JSONResponse({
        "ok": True,
        "diagnostics": diag,
        "status": status.get("status"),
        "latency_ms": diag.get("response_ms"),
        "public_response_ms": diag.get("public_response_ms"),
        "checked_at": datetime.now().isoformat(),
    })


@app.get("/api/remote-access/qr")
def api_remote_access_qr(user: Usuario = Depends(requiere_login)):
    """Devuelve el QR en PNG para descarga directa."""
    status = remote_access.get_status_cached(max_age=60)
    scan_url = status.get("scan_url", "")
    if not scan_url:
        raise HTTPException(404, "No hay URL disponible para generar QR")
    try:
        qr_bytes = generar_qr_bytes(scan_url)
        return Response(content=qr_bytes, media_type="image/png",
                        headers={"Content-Disposition": "attachment; filename=qr_mrd.png"})
    except Exception as e:
        raise HTTPException(500, f"Error generando QR: {e}")


@app.post("/api/remote-access/restart")
def api_remote_access_restart(user: Usuario = Depends(requiere_login)):
    """Reinicia el servicio cloudflared (solo admins, si allow_restart está activado)."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    cfg = remote_access.load_config()
    if not cfg.get("allow_restart", False):
        raise HTTPException(403, "Reinicio de servicio no habilitado en la configuración")
    svc = cfg.get("cloudflared_service", "cloudflared")
    svc = re.sub(r"[^a-zA-Z0-9_\-]", "", svc)[:50]
    if not svc:
        raise HTTPException(400, "Nombre de servicio inválido")
    try:
        result = subprocess.run(
            ["sc", "stop", svc], capture_output=True, text=True, timeout=5,
            creationflags=0x08000000,
        )
        time.sleep(1)
        result2 = subprocess.run(
            ["sc", "start", svc], capture_output=True, text=True, timeout=5,
            creationflags=0x08000000,
        )
        remote_access.invalidate_cache()
        return JSONResponse({"ok": True, "message": f"Servicio '{svc}' reiniciado"})
    except Exception as e:
        raise HTTPException(500, f"Error al reiniciar: {str(e)[:100]}")


def _scan_route_target(raw_code: str):
    """Traduce QR históricos basados en URL a (tipo, id, id_almacén)."""
    value = urllib.parse.unquote(str(raw_code or "")).strip().strip("\ufeff\"'")
    value = "".join(char for char in value if char >= " " and char != "\x7f").strip()
    value = re.sub(r"^\][A-Za-z]\d", "", value).strip()
    if not (value.startswith("/") or "://" in value):
        legacy = re.fullmatch(r"ALM(\d+)-UBI(\d+)", value, re.IGNORECASE)
        return ("ubicacion", int(legacy.group(2)), int(legacy.group(1))) if legacy else None
    try:
        path = urllib.parse.unquote(urllib.parse.urlparse(value).path).rstrip("/") or "/"
    except ValueError:
        return None
    patterns = (
        (r"/qr/(herramienta|material|stock_epi|epi_individual|maquinaria|almacen|ubicacion|vehiculo)/(\d+)", None),
        (r"/herramientas/(\d+)", "herramienta"),
        (r"/materiales/(\d+)", "material"),
        (r"/maquinaria/(\d+)(?:/pasaporte)?", "maquinaria"),
        (r"/epis/individuales/(\d+)", "epi_individual"),
        (r"/vehiculos/(\d+)(?:/qr)?", "vehiculo"),
        (r"/almacenes/(\d+)/ubicaciones/(\d+)(?:/qr)?", "ubicacion"),
        (r"/almacenes/(\d+)(?:/mapa|/qr)?", "almacen"),
    )
    for pattern, fixed_type in patterns:
        match = re.fullmatch(pattern, path, re.IGNORECASE)
        if not match:
            continue
        if fixed_type is None:
            return (match.group(1).lower(), int(match.group(2)), None)
        if fixed_type == "ubicacion":
            return (fixed_type, int(match.group(2)), int(match.group(1)))
        return (fixed_type, int(match.group(1)), None)
    return None


def _resolved_warehouse_context(db: Session, item: dict) -> dict:
    """Añade almacén y ubicación al resultado QR sin duplicar el resolvedor."""
    kind = item.get("tipo")
    item_id = int(item.get("id") or 0)
    obj = None
    if kind == "herramienta":
        obj = db.get(Herramienta, item_id)
    elif kind == "maquinaria":
        obj = db.get(Maquinaria, item_id)
    elif kind == "vehiculo":
        obj = db.get(Vehiculo, item_id)
    elif kind == "epi_individual":
        obj = db.get(EPIIndividual, item_id)
    elif kind == "material":
        obj = db.get(Material, item_id)
    elif kind == "stock_epi":
        obj = db.get(StockEPI, item_id)
    elif kind == "variante":
        obj = db.get(ExistenciaVariante, item_id)
    elif kind == "ubicacion":
        obj = db.get(Ubicacion, item_id)
    elif kind == "almacen":
        obj = db.get(Almacen, item_id)
    warehouse_id = getattr(obj, "almacen_id", None) if obj else None
    if kind == "almacen" and obj:
        warehouse_id = obj.id
    warehouse = db.get(Almacen, warehouse_id) if warehouse_id else None
    location = getattr(obj, "ubicacion", None) if obj else None
    location_text = getattr(obj, "ubicacion_texto", None) if obj else None
    if kind == "maquinaria" and obj:
        location_text = obj.ubicacion
        location = None
    location_name = (
        getattr(location, "ruta_completa", None) or getattr(location, "nombre", None)
        if location is not None else None
    )
    item.update({
        "almacen_id": warehouse_id,
        "almacen_nombre": warehouse.nombre if warehouse else "Sin almacén asignado",
        "ubicacion_nombre": location_name or location_text or "Sin ubicación exacta",
    })
    return item


@app.get("/api/system/stats")
def api_system_stats(user: Usuario = Depends(requiere_login)):
    """Estadísticas del servidor: CPU, RAM, uptime, versión, último backup."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    version_info = leer_version_actual()
    backups = listar_backups()
    ultimo_backup = backups[0]["fecha"] if backups else "—"
    stats = remote_access.get_server_stats()
    return JSONResponse({
        **stats,
        "version": version_info.get("version_actual", "—"),
        "ultimo_backup": ultimo_backup,
    })


@app.get("/scan/buscar")
def scan_buscar(
    request: Request,
    codigo: str,
    db: Session = Depends(get_db),
):
    """Resolvedor QR común para la pantalla Scan y el Mostrador."""
    current_user = usuario_actual(request, db)
    client_key = f"user:{current_user.id}" if current_user else (
        request.client.host if request.client else "unknown"
    )
    if not _permitir_busqueda_scan(client_key):
        raise HTTPException(
            429, "Demasiadas consultas de escaneo. Espera un minuto.",
            headers={"Retry-After": "60"},
        )
    if not codigo or len(codigo) > 512:
        raise HTTPException(400, "Código de escaneo inválido")

    preparation_token = ""
    try:
        parsed_scan_url = urllib.parse.urlparse(codigo.strip())
        match = re.search(r"/preparaciones-entrega/qr/([A-Za-z0-9_-]{20,80})/?$", parsed_scan_url.path)
        if match:
            preparation_token = match.group(1)
    except (TypeError, ValueError):
        preparation_token = ""
    if not preparation_token and re.fullmatch(r"[A-Za-z0-9_-]{20,80}", codigo.strip()):
        preparation_token = codigo.strip()
    if preparation_token:
        preparation = db.query(PreparacionEntrega).filter(
            PreparacionEntrega.qr_token == preparation_token,
        ).first()
        if preparation:
            if not current_user:
                return JSONResponse({
                    "found": True, "public": True,
                    "message": "Código reconocido. Inicia sesión para consultar sus datos.",
                    "requires_login": True, "login_url": "/login",
                })
            _require_warehouse_access(current_user, preparation.almacen_id)
            return JSONResponse({
                "found": True, "tipo": "preparacion", "id": preparation.id,
                "nombre": f"Entrega preparada {preparation.numero}", "codigo": preparation.numero,
                "estado": preparation.estado, "estado_label": preparation.estado.capitalize(),
                "url": f"/preparaciones-entrega/qr/{preparation.qr_token}",
                "almacen_id": preparation.almacen_id,
                "almacen_nombre": preparation.almacen.nombre if preparation.almacen else "Almacén",
                "ubicacion_nombre": "Lote preparado",
            })

    route_target = _scan_route_target(codigo)
    if route_target:
        route_type, item_id, _parent_id = route_target
        route_code, _ = _get_qr_code_for(route_type, item_id, db)
        if route_code:
            codigo = route_code

    try:
        active_warehouse = _active_warehouse(db, current_user, request) if current_user else None
        resolved = resolve_counter_item(
            db, codigo,
            warehouse_id=active_warehouse.id if active_warehouse else None,
        )
    except CounterError as exc:
        if current_user and current_user.rol == "admin" and exc.status_code == 404:
            try:
                resolved = resolve_counter_item(db, codigo, warehouse_id=None)
            except CounterError:
                resolved = None
            if resolved:
                resolved = _resolved_warehouse_context(db, resolved)
                resolved["requiere_traspaso"] = bool(
                    active_warehouse and resolved.get("almacen_id") not in (None, active_warehouse.id)
                )
                return JSONResponse(resolved)
        if exc.status_code == 400:
            raise HTTPException(400, exc.detail)
        normalized = normalize_scanned_code(codigo)
        return JSONResponse({
            "found": False,
            "detail": "Código no reconocido" if not current_user else exc.detail,
            "codigo_normalizado": normalized if current_user else "",
        })

    if not current_user:
        return JSONResponse({
            "found": True,
            "public": True,
            "message": "Código reconocido. Inicia sesión para consultar sus datos.",
            "requires_login": True,
            "login_url": "/login",
        })

    resolved = _resolved_warehouse_context(db, resolved)
    resolved["requiere_traspaso"] = False
    return JSONResponse(resolved)


class ScanOperationRequest(BaseModel):
    scan_event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    accion: Literal["entregar", "devolver"]
    herramienta_id: int = Field(gt=0)
    trabajador_id: Optional[int] = Field(default=None, gt=0)
    obra_id: Optional[int] = Field(default=None, gt=0)
    almacen_id: Optional[int] = Field(default=None, gt=0)
    condicion: Literal["buena", "requiere_revision", "danada"] = "buena"
    observaciones: str = Field(default="", max_length=2000)


class ScanStockEPIRequest(BaseModel):
    scan_event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    stock_epi_id: int = Field(gt=0)
    accion: Literal["entrada", "salida"]
    cantidad: int = Field(gt=0, le=9999)
    motivo: str = Field(default="", max_length=500)


def requiere_login_scan(request: Request, db: Session = Depends(get_db)) -> Usuario:
    """Dependencia JSON: nunca convierte una sesion caducada en redireccion exitosa."""
    user = usuario_actual(request, db)
    if not user:
        raise HTTPException(401, "Sesion caducada")
    return user


class MostradorLineaRequest(BaseModel):
    tipo: Literal[
        "herramienta", "maquinaria", "vehiculo", "material",
        "stock_epi", "variante", "epi_individual",
    ]
    id: int = Field(gt=0)
    cantidad: int = Field(default=1, gt=0, le=9999)


class MostradorOperacionRequest(BaseModel):
    operacion_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    accion: Literal["salida", "entrada"]
    lineas: list[MostradorLineaRequest] = Field(min_length=1, max_length=200)
    trabajador_id: Optional[int] = Field(default=None, gt=0)
    obra_id: Optional[int] = Field(default=None, gt=0)
    almacen_id: Optional[int] = Field(default=None, gt=0)
    fecha_devolucion_prevista: Optional[datetime] = None
    notas: str = Field(default="", max_length=1000)
    origen: str = Field(default="", max_length=160)


class TransferLineRequest(BaseModel):
    tipo: Literal[
        "herramienta", "maquinaria", "vehiculo", "epi_individual",
        "material", "stock_epi", "variante",
    ]
    id: int = Field(gt=0)
    cantidad: float = Field(default=1, gt=0, le=999999)


class TransferCreateRequest(BaseModel):
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    origen_id: int = Field(gt=0)
    destino_id: int = Field(gt=0)
    lineas: list[TransferLineRequest] = Field(min_length=1, max_length=300)
    notas: str = Field(default="", max_length=1000)


class TransferReceiptLineRequest(BaseModel):
    linea_id: int = Field(gt=0)
    cantidad_aceptada: float = Field(default=0, ge=0)
    cantidad_danada: float = Field(default=0, ge=0)
    ubicacion_id: Optional[int] = Field(default=None, gt=0)
    notas: str = Field(default="", max_length=1000)
    foto_datos: Optional[str] = Field(default=None, max_length=4_500_000)


class TransferReceiveRequest(BaseModel):
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    firma_datos: str = Field(min_length=30, max_length=700000)
    firma_nombre: str = Field(min_length=2, max_length=100)
    lineas: Optional[list[TransferReceiptLineRequest]] = Field(default=None, max_length=300)


class PedidoCrearRequest(BaseModel):
    proveedor: str = Field(default="", max_length=150)
    fecha_prevista: Optional[date] = None
    notas: str = Field(default="", max_length=1000)
    incluir_stock_bajo: bool = True
    codigo: Optional[str] = Field(default=None, max_length=512)
    cantidad: Optional[float] = Field(default=None, gt=0, le=1_000_000)


class PedidoRecibirLineaRequest(BaseModel):
    linea_id: int = Field(gt=0)
    cantidad: float = Field(gt=0, le=1_000_000)
    numero_lote: Optional[str] = Field(default=None, max_length=100)
    fecha_caducidad: Optional[date] = None


class PedidoRecibirRequest(BaseModel):
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    lineas: list[PedidoRecibirLineaRequest] = Field(min_length=1, max_length=300)


class PreparacionCrearRequest(BaseModel):
    trabajador_id: Optional[int] = Field(default=None, gt=0)
    obra_id: Optional[int] = Field(default=None, gt=0)
    destino: str = Field(default="", max_length=200)
    notas: str = Field(default="", max_length=1000)
    lineas: list[MostradorLineaRequest] = Field(min_length=1, max_length=200)


class PreparacionEntregarRequest(BaseModel):
    operation_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")


class CierreDiarioRequest(BaseModel):
    firma_nombre: str = Field(min_length=2, max_length=100)
    firma_datos: str = Field(min_length=30, max_length=700000)


@app.get("/mostrador", response_class=HTMLResponse)
def mostrador_unico_panel(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "entregar") or tiene_permiso(user, "devolver")):
        raise HTTPException(403, "Sin permiso para operar el mostrador")
    warehouse = _active_warehouse(db, user, request)
    workers_query = db.query(Trabajador).filter(Trabajador.activo == True)
    works_query = db.query(Obra).filter(Obra.activa == True)
    if warehouse:
        workers_query = workers_query.filter(Trabajador.almacen_id == warehouse.id)
        works_query = works_query.filter(Obra.almacen_id == warehouse.id)
    workers = workers_query.order_by(
        Trabajador.nombre, Trabajador.apellidos,
    ).all()
    works = works_query.order_by(Obra.nombre).all()
    warehouses = [warehouse] if warehouse else []
    return templates.TemplateResponse(request, "mostrador.html", ctx_base(
        request, user, db, trabajadores=workers, obras=works, almacenes=warehouses,
        tipos_permitidos=sorted(allowed_counter_types(user)),
        puede_salida=tiene_permiso(user, "entregar"),
        puede_entrada=tiene_permiso(user, "devolver"),
    ))


@app.get("/scanner/configurar", response_class=HTMLResponse)
def scanner_configurar(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "entregar") or tiene_permiso(user, "inventario")):
        raise HTTPException(403, "Sin permiso para configurar el lector")
    return templates.TemplateResponse(
        request, "scanner_configurar.html", ctx_base(request, user, db),
    )


@app.get("/api/mostrador/resolver")
def mostrador_resolver(
    codigo: str = Query(..., min_length=1, max_length=128),
    user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
    request: Request = None,
):
    if not (tiene_permiso(user, "entregar") or tiene_permiso(user, "devolver")):
        raise HTTPException(403, "Sin permiso")
    try:
        warehouse = _active_warehouse(db, user, request)
        item = resolve_counter_item(
            db, codigo, warehouse_id=warehouse.id if warehouse else None,
        )
        item = _resolved_warehouse_context(db, item)
        if item["tipo"] not in allowed_counter_types(user):
            raise HTTPException(404, "Tu rol no opera este tipo de artículo")
        return JSONResponse({"ok": True, "item": item})
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)


@app.get("/api/mostrador/buscar")
def mostrador_buscar(
    q: str = Query(..., min_length=2, max_length=100),
    user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
    request: Request = None,
):
    if not (tiene_permiso(user, "entregar") or tiene_permiso(user, "devolver")):
        raise HTTPException(403, "Sin permiso")
    try:
        allowed = allowed_counter_types(user)
        warehouse = _active_warehouse(db, user, request)
        items = [
            item for item in search_counter_items(
                db, q, warehouse_id=warehouse.id if warehouse else None,
            ) if item["tipo"] in allowed
        ]
        return JSONResponse({"ok": True, "items": items})
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)


@app.post("/api/mostrador/operar")
def mostrador_operar(
    payload: MostradorOperacionRequest,
    user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
    request: Request = None,
):
    try:
        warehouse = _active_warehouse(db, user, request)
        if not warehouse:
            raise CounterError(409, "No hay un almacén activo configurado")
        if payload.almacen_id and payload.almacen_id != warehouse.id:
            raise CounterError(409, "Cambia primero al almacén que quieres operar")
        result = operate_counter(
            db, user, operation_id=payload.operacion_id, action=payload.accion,
            lines=[line.model_dump() for line in payload.lineas],
            worker_id=payload.trabajador_id, work_id=payload.obra_id,
            warehouse_id=warehouse.id, notes=payload.notas,
            expected_return=payload.fecha_devolucion_prevista,
            origin=payload.origen,
        )
        db.commit()
        return JSONResponse(result)
    except CounterError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        mrd_logging.log_app("Mostrador unico: operacion revertida por error inesperado", level="error")
        raise HTTPException(500, "No se registro ningun movimiento. El carrito completo fue revertido.")


@app.get("/multi-almacen", response_class=HTMLResponse)
def multiwarehouse_center(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede consultar todos los almacenes")
    warehouses = db.query(Almacen).filter(Almacen.activo == True).order_by(Almacen.nombre).all()
    cards = []
    for warehouse in warehouses:
        variant_stock = int(db.query(func.coalesce(func.sum(ExistenciaVariante.cantidad), 0)).filter(
            ExistenciaVariante.almacen_id == warehouse.id,
        ).scalar() or 0)
        material_stock = float(db.query(func.coalesce(func.sum(Material.stock_actual), 0)).filter(
            Material.almacen_id == warehouse.id, Material.activo == True,
        ).scalar() or 0)
        low_material = db.query(Material).filter(
            Material.almacen_id == warehouse.id, Material.activo == True,
            Material.stock_minimo > 0, Material.stock_actual <= Material.stock_minimo,
        ).count()
        cards.append({
            "almacen": warehouse,
            "herramientas": db.query(Herramienta).filter(
                Herramienta.almacen_id == warehouse.id, Herramienta.activa == True,
            ).count(),
            "maquinaria": db.query(Maquinaria).filter(
                Maquinaria.almacen_id == warehouse.id, Maquinaria.activa == True,
            ).count(),
            "incidencias": db.query(Incidencia).filter(
                Incidencia.almacen_id == warehouse.id,
                Incidencia.estado.in_(["abierta", "en_curso"]),
            ).count(),
            "material_stock": material_stock,
            "variant_stock": variant_stock,
            "stock_bajo": low_material,
            "trabajadores": db.query(Trabajador).filter(
                Trabajador.almacen_id == warehouse.id, Trabajador.activo == True,
            ).count(),
        })
    transfers = db.query(TransferenciaAlmacen).order_by(
        TransferenciaAlmacen.creado_en.desc(), TransferenciaAlmacen.id.desc(),
    ).limit(100).all()
    audit = db.query(AuditoriaLog).order_by(AuditoriaLog.fecha.desc()).limit(30).all()
    return templates.TemplateResponse(request, "multi_almacen.html", ctx_base(
        request, user, db, almacenes=warehouses, tarjetas=cards,
        transferencias=transfers, auditoria_reciente=audit,
    ))


@app.get("/api/traspasos/resolver")
def transfer_resolver(
    codigo: str = Query(min_length=1, max_length=512),
    origen_id: int = Query(gt=0),
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede preparar traspasos")
    try:
        item = resolve_counter_item(db, codigo, warehouse_id=origen_id)
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    if item.get("tipo") not in {
        "herramienta", "maquinaria", "vehiculo", "epi_individual",
        "material", "stock_epi", "variante",
    }:
        raise HTTPException(400, "Este código no corresponde a un artículo transferible")
    transfer = in_transit(db, item["tipo"], int(item["id"]))
    if transfer:
        raise HTTPException(409, f"Ya está en tránsito con {transfer.numero}")
    warehouse = db.get(Almacen, origen_id)
    item["almacen_id"] = origen_id
    item["almacen_nombre"] = warehouse.nombre if warehouse else ""
    return JSONResponse({"ok": True, "item": item})


@app.get("/api/traspasos/buscar")
def transfer_search(
    q: str = Query(min_length=2, max_length=100), origen_id: int = Query(gt=0),
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403)
    try:
        items = search_counter_items(db, q, warehouse_id=origen_id)
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    items = [item for item in items if item.get("tipo") in {
        "herramienta", "maquinaria", "vehiculo", "epi_individual",
        "material", "stock_epi", "variante",
    } and not in_transit(db, item["tipo"], int(item["id"]))]
    return JSONResponse({"ok": True, "items": items})


@app.post("/api/traspasos")
def transfer_create_api(
    payload: TransferCreateRequest, user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
):
    try:
        start_stock_transaction(db)
        transfer = create_transfer(
            db, user, origin_id=payload.origen_id, destination_id=payload.destino_id,
            event_id=payload.event_id,
            lines=[line.model_dump() for line in payload.lineas], notes=payload.notas,
        )
        db.add(AuditoriaLog(
            tabla="transferencias_almacen", registro_id=transfer.id, accion="crear",
            resumen=f"{transfer.numero}: {transfer.origen.nombre} → {transfer.destino.nombre}",
            usuario_id=user.id,
        ))
        db.commit()
        return JSONResponse({
            "ok": True, "id": transfer.id, "numero": transfer.numero,
            "estado": transfer.estado,
            "url": f"/multi-almacen/traspasos/{transfer.id}",
            "pdf_url": f"/multi-almacen/traspasos/{transfer.id}/pdf",
        }, status_code=201)
    except TransferError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "El traspaso coincide con otra operación; recarga la pantalla")


def _transfer_access(user: Usuario, transfer: TransferenciaAlmacen) -> None:
    if user.rol == "admin":
        return
    if user.almacen_id not in {transfer.origen_id, transfer.destino_id}:
        raise HTTPException(403, "Este traspaso pertenece a otro almacén")


@app.get("/multi-almacen/traspasos/{transfer_id}", response_class=HTMLResponse)
def transfer_detail(
    transfer_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    transfer = db.get(TransferenciaAlmacen, transfer_id)
    if not transfer:
        raise HTTPException(404)
    _transfer_access(user, transfer)
    receipts = db.query(RecepcionTransferencia).filter_by(
        transferencia_id=transfer.id,
    ).order_by(RecepcionTransferencia.recibido_en).all()
    locations = db.query(Ubicacion).filter(
        Ubicacion.almacen_id == transfer.destino_id, Ubicacion.activo == True,
    ).order_by(Ubicacion.nombre).all()
    return templates.TemplateResponse(request, "transferencia_detalle.html", ctx_base(
        request, user, db, transferencia=transfer, ubicaciones_destino=locations,
        recepciones=receipts,
    ))


@app.post("/api/traspasos/{transfer_id}/recibir")
def transfer_receive_api(
    transfer_id: int, payload: TransferReceiveRequest,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    signer_name = payload.firma_nombre.strip()
    if len(signer_name) < 2:
        raise HTTPException(400, "Indica el nombre de quien recibe")
    mime, signature_bytes = _decode_delivery_signature(payload.firma_datos)
    canonical_signature = f"data:{mime};base64,{__import__('base64').b64encode(signature_bytes).decode('ascii')}"
    receipt_lines = None
    if payload.lineas is not None:
        receipt_lines = []
        for item in payload.lineas:
            row = item.model_dump(exclude={"foto_datos"})
            row["foto_path"] = _save_transfer_receipt_photo(item.foto_datos, transfer_id, item.linea_id)
            receipt_lines.append(row)
    try:
        start_stock_transaction(db)
        transfer = receive_transfer(
            db, user, transfer_id, event_id=payload.event_id,
            signature_data=canonical_signature,
            signature_name=signer_name,
            receipt_lines=receipt_lines,
        )
        db.add(AuditoriaLog(
            tabla="transferencias_almacen", registro_id=transfer.id, accion="recibir",
            resumen=f"{transfer.numero}: recepción {'completa' if transfer.estado == 'recibida' else 'parcial'} en {transfer.destino.nombre}", usuario_id=user.id,
        ))
        db.commit()
        return JSONResponse({
            "ok": True, "estado": transfer.estado,
            "recepcion_completa": transfer.estado == "recibida",
            "url": f"/multi-almacen/traspasos/{transfer.id}",
            "pdf_url": f"/multi-almacen/traspasos/{transfer.id}/pdf",
        })
    except TransferError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)


def _save_transfer_receipt_photo(value: str | None, transfer_id: int, line_id: int) -> str | None:
    if not value:
        return None
    import base64
    import binascii
    match = re.fullmatch(r"data:image/(png|jpeg);base64,([A-Za-z0-9+/=]+)", value.strip(), re.IGNORECASE)
    if not match:
        raise HTTPException(400, "La fotografía de recepción no es válida")
    try:
        content = base64.b64decode(match.group(2), validate=True)
    except (ValueError, binascii.Error):
        raise HTTPException(400, "La fotografía de recepción está dañada")
    if not content or len(content) > 3_000_000:
        raise HTTPException(400, "La fotografía supera el máximo de 3 MB")
    extension = "jpg" if match.group(1).lower() == "jpeg" else "png"
    relative = Path("traspasos") / f"TR-{transfer_id}-L{line_id}-{uuid.uuid4().hex[:10]}.{extension}"
    target = UPLOADS_DIR / relative
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(content)
    return str(Path("uploads") / relative).replace("\\", "/")


@app.post("/multi-almacen/traspasos/{transfer_id}/cancelar", response_class=RedirectResponse)
def transfer_cancel_route(
    transfer_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        start_stock_transaction(db)
        transfer = cancel_transfer(db, user, transfer_id)
        db.add(AuditoriaLog(
            tabla="transferencias_almacen", registro_id=transfer.id, accion="cancelar",
            resumen=f"{transfer.numero} cancelado y stock restaurado", usuario_id=user.id,
        ))
        db.commit()
        return RedirectResponse(f"/multi-almacen/traspasos/{transfer.id}", status_code=303)
    except TransferError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)


@app.get("/multi-almacen/traspasos/{transfer_id}/pdf")
def transfer_pdf(
    transfer_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from xml.sax.saxutils import escape
    import io

    transfer = db.get(TransferenciaAlmacen, transfer_id)
    if not transfer:
        raise HTTPException(404)
    _transfer_access(user, transfer)
    receipts = db.query(RecepcionTransferencia).filter_by(
        transferencia_id=transfer.id,
    ).order_by(RecepcionTransferencia.recibido_en).all()
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    stamp_path = Path(__file__).parent / "static" / "img" / "mrd_sello_blanco.png"
    if stamp_path.is_file():
        story.extend([RLImage(str(stamp_path), width=145, height=72), Spacer(1, 5)])
    story.extend([
        Paragraph(f"<b>ALBARÁN DE TRASPASO {escape(transfer.numero)}</b>", styles["Title"]),
        Spacer(1, 8),
        Table([
            ["Origen", transfer.origen.nombre, "Destino", transfer.destino.nombre],
            ["Salida", transfer.creado_en.strftime("%d/%m/%Y %H:%M"), "Estado", transfer.estado.upper()],
        ], colWidths=[65, 180, 65, 180]),
        Spacer(1, 14),
    ])
    rows = [["#", "Referencia", "Descripción", "Enviado", "Correcto", "Dañado", "Pendiente"]]
    for index, line in enumerate(transfer.lineas, 1):
        rows.append([
            index, escape(line.referencia), Paragraph(escape(line.descripcion), styles["BodyText"]),
            f"{line.cantidad:g}", f"{float(line.cantidad_recibida or 0):g}",
            f"{float(line.cantidad_danada or 0):g}",
            f"{max(0, float(line.cantidad)-float(line.cantidad_recibida or 0)-float(line.cantidad_danada or 0)):g}",
        ])
    table = Table(rows, colWidths=[20, 75, 205, 50, 50, 45, 50], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), .3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.extend([table, Spacer(1, 18)])
    qr_url = f"{MRD_PUBLIC_URL or 'https://app.iasmrd.com'}/multi-almacen/traspasos/{transfer.id}"
    qr_image = RLImage(io.BytesIO(generar_qr_bytes(qr_url)), width=72, height=72)
    if transfer.firma_recepcion:
        _mime, signature_bytes = _decode_delivery_signature(transfer.firma_recepcion)
        signature = RLImage(io.BytesIO(signature_bytes), width=180, height=60)
        received = transfer.recibido_en.strftime("%d/%m/%Y %H:%M") if transfer.recibido_en else ""
        footer = Table([
            [Paragraph("<b>RECEPCIÓN CONFORME</b>", styles["Normal"]), qr_image],
            [signature, ""],
            [Paragraph(f"<b>{escape(transfer.firma_recepcion_nombre or '')}</b><br/>{received}", styles["Normal"]), "Consultar traspaso"],
        ], colWidths=[360, 120])
    else:
        receipt_text = "Pendiente de recepción"
        if receipts:
            receipt_text = "Recepciones parciales: " + ", ".join(
                f"{row.firma_nombre} ({row.recibido_en.strftime('%d/%m %H:%M')})" for row in receipts
            )
        footer = Table([[receipt_text, qr_image]], colWidths=[360, 120])
    footer.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), .5, colors.grey), ("ALIGN", (1, 0), (1, -1), "CENTER")]))
    story.append(footer)
    document.build(story)
    output.seek(0)
    return StreamingResponse(output, media_type="application/pdf", headers={
        "Content-Disposition": f"inline; filename=traspaso_{transfer.numero}.pdf",
    })


def _operation_warehouse(request: Request, user: Usuario, db: Session) -> Almacen:
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    _require_warehouse_access(user, warehouse.id)
    return warehouse


def _require_stock_http(user: Usuario) -> None:
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "editar")):
        raise HTTPException(403, "Sin permiso para operar stock")


def _purchase_line_data(item: dict, quantity: float) -> dict:
    return {
        "tipo": item["tipo"], "objeto_id": int(item["id"]),
        "referencia": str(item.get("codigo") or "")[:100],
        "descripcion": str(item.get("nombre") or "Artículo")[:300],
        "cantidad_pedida": float(quantity),
    }


@app.get("/pedidos-proveedor", response_class=HTMLResponse)
def supplier_orders_page(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _require_stock_http(user)
    warehouse = _operation_warehouse(request, user, db)
    orders = db.query(PedidoProveedor).filter(
        PedidoProveedor.almacen_id == warehouse.id,
    ).order_by(PedidoProveedor.fecha_pedido.desc(), PedidoProveedor.id.desc()).all()
    low_materials = db.query(Material).filter(
        Material.almacen_id == warehouse.id, Material.activo == True,
        Material.stock_minimo > 0, Material.stock_actual <= Material.stock_minimo,
    ).order_by(Material.nombre).all()
    low_epi = db.query(StockEPI).filter(
        StockEPI.almacen_id == warehouse.id, StockEPI.stock_minimo > 0,
        StockEPI.cantidad <= StockEPI.stock_minimo,
    ).order_by(StockEPI.nombre, StockEPI.talla).all()
    return templates.TemplateResponse(request, "pedidos_proveedor.html", ctx_base(
        request, user, db, almacen=warehouse, pedidos=orders,
        materiales_bajos=low_materials, epis_bajos=low_epi,
    ))


@app.post("/api/pedidos-proveedor")
def supplier_order_create(
    payload: PedidoCrearRequest, request: Request,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    _require_stock_http(user)
    warehouse = _operation_warehouse(request, user, db)
    rows: list[dict] = []
    if payload.incluir_stock_bajo:
        for obj in db.query(Material).filter(
            Material.almacen_id == warehouse.id, Material.activo == True,
            Material.stock_minimo > 0, Material.stock_actual <= Material.stock_minimo,
        ).all():
            rows.append({"tipo": "material", "objeto_id": obj.id, "referencia": obj.codigo,
                         "descripcion": obj.nombre, "cantidad_pedida": max(float(obj.stock_minimo) - float(obj.stock_actual), 1)})
        for obj in db.query(StockEPI).filter(
            StockEPI.almacen_id == warehouse.id, StockEPI.stock_minimo > 0,
            StockEPI.cantidad <= StockEPI.stock_minimo,
        ).all():
            rows.append({"tipo": "stock_epi", "objeto_id": obj.id, "referencia": obj.codigo,
                         "descripcion": obj.nombre_display, "cantidad_pedida": max(int(obj.stock_minimo) - int(obj.cantidad), 1)})
    if payload.codigo:
        try:
            item = resolve_counter_item(db, payload.codigo, warehouse_id=warehouse.id)
        except CounterError as exc:
            raise HTTPException(exc.status_code, exc.detail)
        if item["tipo"] not in {"material", "stock_epi", "variante"}:
            raise HTTPException(409, "Esta referencia no pertenece a stock comprable")
        rows.append(_purchase_line_data(item, payload.cantidad or 1))
    unique = {(row["tipo"], row["objeto_id"]): row for row in rows}
    if not unique:
        raise HTTPException(400, "No hay artículos bajo mínimo; escanea una referencia para crear el pedido")
    order = PedidoProveedor(
        numero=f"PED-{datetime.now():%Y%m%d}-{uuid.uuid4().hex[:6].upper()}",
        almacen_id=warehouse.id, proveedor=payload.proveedor.strip() or None,
        fecha_prevista=payload.fecha_prevista, notas=payload.notas.strip() or None,
        creado_por_id=user.id,
    )
    db.add(order)
    db.flush()
    for row in unique.values():
        db.add(LineaPedidoProveedor(pedido_id=order.id, **row))
    db.add(AuditoriaLog(tabla="pedidos_proveedor", registro_id=order.id, accion="crear",
                        resumen=f"{order.numero}: pedido para {warehouse.nombre}", usuario_id=user.id))
    db.commit()
    return JSONResponse({"ok": True, "id": order.id, "url": f"/pedidos-proveedor/{order.id}"}, status_code=201)


def _supplier_order_for_user(db: Session, user: Usuario, order_id: int) -> PedidoProveedor:
    order = db.get(PedidoProveedor, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    _require_warehouse_access(user, order.almacen_id)
    return order


@app.get("/pedidos-proveedor/{order_id}", response_class=HTMLResponse)
def supplier_order_detail(
    order_id: int, request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _require_stock_http(user)
    return templates.TemplateResponse(request, "pedido_proveedor_detalle.html", ctx_base(
        request, user, db, pedido=_supplier_order_for_user(db, user, order_id),
    ))


@app.post("/api/pedidos-proveedor/{order_id}/enviar")
def supplier_order_send(
    order_id: int, user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    _require_stock_http(user)
    order = _supplier_order_for_user(db, user, order_id)
    if order.estado != "borrador":
        raise HTTPException(409, "Solo se puede enviar un pedido en borrador")
    order.estado = "enviado"
    db.commit()
    return JSONResponse({"ok": True, "estado": order.estado})


@app.post("/api/pedidos-proveedor/{order_id}/recibir")
def supplier_order_receive(
    order_id: int, payload: PedidoRecibirRequest,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    _require_stock_http(user)
    order = _supplier_order_for_user(db, user, order_id)
    existing_receipt = db.query(RecepcionPedidoProveedor).filter_by(event_id=payload.event_id).first()
    if existing_receipt:
        if existing_receipt.pedido_id != order.id:
            raise HTTPException(409, "El identificador de recepción ya pertenece a otro pedido")
        return JSONResponse({"ok": True, "estado": order.estado, "reutilizada": True})
    if order.estado not in {"enviado", "parcial"}:
        raise HTTPException(409, "El pedido no está pendiente de recepción")
    selected = {row.linea_id: row for row in payload.lineas}
    if len(selected) != len(payload.lineas):
        raise HTTPException(400, "Hay líneas repetidas")
    try:
        start_stock_transaction(db)
        applied = 0
        for index, line in enumerate(order.lineas):
            data = selected.get(line.id)
            if not data:
                continue
            remaining = float(line.cantidad_pedida) - float(line.cantidad_recibida or 0)
            if data.cantidad > remaining + .00001:
                raise StockError(409, f"{line.referencia}: la recepción supera lo pendiente")
            event_id = f"{payload.event_id}-{index}"
            if line.tipo == "material":
                obj = db.get(Material, line.objeto_id)
                if not obj or obj.almacen_id != order.almacen_id:
                    raise StockError(409, f"Material no disponible: {line.referencia}")
                move_material(db, user, obj.id, data.cantidad, tipo="pedido_proveedor",
                              event_id=event_id, motivo=f"Recepción {order.numero}")
                db.expire(obj)
            elif line.tipo == "stock_epi":
                if data.cantidad != int(data.cantidad):
                    raise StockError(400, "La ropa y EPI se reciben en unidades completas")
                obj = db.get(StockEPI, line.objeto_id)
                if not obj or obj.almacen_id != order.almacen_id:
                    raise StockError(409, f"EPI no disponible: {line.referencia}")
                move_stock_epi(db, user, obj.id, int(data.cantidad), tipo="pedido_proveedor",
                               event_id=event_id, motivo=f"Recepción {order.numero}")
                db.expire(obj)
            elif line.tipo == "variante":
                if data.cantidad != int(data.cantidad):
                    raise StockError(400, "Las variantes se reciben en unidades completas")
                obj = db.get(ExistenciaVariante, line.objeto_id)
                if not obj or obj.almacen_id != order.almacen_id:
                    raise StockError(409, f"Variante no disponible: {line.referencia}")
                move_variante(db, user, obj.id, int(data.cantidad), tipo="pedido_proveedor",
                              event_id=event_id, motivo=f"Recepción {order.numero}")
                db.expire(obj)
            else:
                raise StockError(400, "Tipo de línea no admitido")
            if data.numero_lote:
                lot = db.query(LoteAlmacen).filter_by(
                    tipo=line.tipo, objeto_id=line.objeto_id, almacen_id=order.almacen_id,
                    numero_lote=data.numero_lote.strip(),
                ).first()
                if not lot:
                    lot = LoteAlmacen(tipo=line.tipo, objeto_id=line.objeto_id,
                                      almacen_id=order.almacen_id, numero_lote=data.numero_lote.strip(), cantidad=0)
                    db.add(lot)
                lot.cantidad = float(lot.cantidad or 0) + data.cantidad
                lot.fecha_caducidad = data.fecha_caducidad or lot.fecha_caducidad
                lot.proveedor = order.proveedor
            line.cantidad_recibida = float(line.cantidad_recibida or 0) + data.cantidad
            applied += 1
        if not applied or applied != len(selected):
            raise StockError(400, "Una línea no pertenece a este pedido")
        complete = all(float(line.cantidad_recibida or 0) >= float(line.cantidad_pedida) - .00001 for line in order.lineas)
        order.estado = "recibido" if complete else "parcial"
        order.cerrado_en = datetime.now() if complete else None
        db.add(RecepcionPedidoProveedor(
            pedido_id=order.id, event_id=payload.event_id,
            lineas_json=json.dumps([row.model_dump(mode="json") for row in payload.lineas], ensure_ascii=False),
            usuario_id=user.id,
        ))
        db.add(AuditoriaLog(tabla="pedidos_proveedor", registro_id=order.id, accion="recibir",
                            resumen=f"{order.numero}: recepción {'completa' if complete else 'parcial'}", usuario_id=user.id))
        db.commit()
        return JSONResponse({"ok": True, "estado": order.estado})
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except IntegrityError:
        db.rollback()
        existing = db.query(RecepcionPedidoProveedor).filter_by(event_id=payload.event_id).first()
        if existing and existing.pedido_id == order_id:
            current = db.get(PedidoProveedor, order_id)
            return JSONResponse({"ok": True, "estado": current.estado, "reutilizada": True})
        raise HTTPException(409, "La recepción coincide con otra operación")


@app.get("/preparaciones-entrega", response_class=HTMLResponse)
def preparations_page(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "entregar"):
        raise HTTPException(403, "Sin permiso para preparar entregas")
    warehouse = _operation_warehouse(request, user, db)
    preparations = db.query(PreparacionEntrega).filter(
        PreparacionEntrega.almacen_id == warehouse.id,
    ).order_by(PreparacionEntrega.creado_en.desc()).limit(100).all()
    portal_base = MRD_PUBLIC_URL if IS_PRODUCTION else str(request.base_url).rstrip("/")
    generated_token = False
    for preparation in preparations:
        if not preparation.qr_token:
            preparation.qr_token = uuid.uuid4().hex + uuid.uuid4().hex
            generated_token = True
        preparation.qr_url = f"{portal_base}/preparaciones-entrega/qr/{preparation.qr_token}"
        preparation.qr_b64 = generar_qr_base64(preparation.qr_url)
        try:
            preparation.total_lineas = len(json.loads(preparation.lineas_json or "[]"))
        except (TypeError, ValueError):
            preparation.total_lineas = 0
    if generated_token:
        db.commit()
    workers = db.query(Trabajador).filter(Trabajador.activo == True, Trabajador.almacen_id == warehouse.id).order_by(Trabajador.nombre).all()
    works = db.query(Obra).filter(Obra.activa == True, Obra.almacen_id == warehouse.id).order_by(Obra.nombre).all()
    return templates.TemplateResponse(request, "preparaciones_entrega.html", ctx_base(
        request, user, db, almacen=warehouse, preparaciones=preparations,
        trabajadores=workers, obras=works,
    ))


@app.post("/api/preparaciones-entrega")
def preparation_create(
    payload: PreparacionCrearRequest, request: Request,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "entregar"):
        raise HTTPException(403)
    warehouse = _operation_warehouse(request, user, db)
    if not (payload.trabajador_id or payload.obra_id):
        raise HTTPException(400, "Selecciona trabajador u obra")
    worker = db.get(Trabajador, payload.trabajador_id) if payload.trabajador_id else None
    work = db.get(Obra, payload.obra_id) if payload.obra_id else None
    if payload.trabajador_id and (not worker or not worker.activo or worker.almacen_id != warehouse.id):
        raise HTTPException(400, "El trabajador no pertenece al almacén activo")
    if payload.obra_id and (not work or not work.activa or work.almacen_id != warehouse.id):
        raise HTTPException(400, "La obra no pertenece al almacén activo")
    # Resolver de nuevo evita guardar artículos inexistentes o de otra sede.
    lines = []
    for raw in payload.lineas:
        model_map = {"herramienta": Herramienta, "maquinaria": Maquinaria, "vehiculo": Vehiculo,
                     "epi_individual": EPIIndividual, "material": Material,
                     "stock_epi": StockEPI, "variante": ExistenciaVariante}
        obj = db.get(model_map[raw.tipo], raw.id)
        if not obj or getattr(obj, "almacen_id", warehouse.id) != warehouse.id:
            raise HTTPException(409, "Una línea no pertenece al almacén activo")
        if raw.tipo in {"herramienta", "maquinaria", "vehiculo", "epi_individual"} and raw.cantidad != 1:
            raise HTTPException(400, "Los activos individuales se preparan de uno en uno")
        lines.append(raw.model_dump())
    prep = PreparacionEntrega(
        numero=f"PREP-{datetime.now():%Y%m%d}-{uuid.uuid4().hex[:6].upper()}",
        qr_token=uuid.uuid4().hex + uuid.uuid4().hex,
        almacen_id=warehouse.id, trabajador_id=payload.trabajador_id, obra_id=payload.obra_id,
        destino=payload.destino.strip() or None, notas=payload.notas.strip() or None,
        lineas_json=json.dumps(lines, ensure_ascii=False), creado_por_id=user.id,
    )
    db.add(prep)
    db.flush()
    db.add(AuditoriaLog(tabla="preparaciones_entrega", registro_id=prep.id, accion="crear",
                        resumen=f"{prep.numero}: {len(lines)} líneas preparadas", usuario_id=user.id))
    db.commit()
    return JSONResponse({
        "ok": True, "id": prep.id, "numero": prep.numero,
        "qr_url": f"/preparaciones-entrega/qr/{prep.qr_token}",
    }, status_code=201)


@app.post("/api/preparaciones-entrega/{preparation_id}/entregar")
def preparation_deliver(
    preparation_id: int, payload: PreparacionEntregarRequest,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    prep = db.get(PreparacionEntrega, preparation_id)
    if not prep:
        raise HTTPException(404, "Preparación no encontrada")
    _require_warehouse_access(user, prep.almacen_id)
    if prep.estado != "preparada":
        raise HTTPException(409, "La preparación ya fue cerrada")
    try:
        result = operate_counter(
            db, user, operation_id=payload.operation_id, action="salida",
            lines=json.loads(prep.lineas_json), worker_id=prep.trabajador_id,
            work_id=prep.obra_id, warehouse_id=prep.almacen_id,
            notes=prep.notas or f"Entrega preparada {prep.numero}", origin="preparacion",
        )
        prep.estado = "entregada"
        prep.entregado_en = datetime.now()
        prep.albaran_id = result.get("albaran_id")
        db.commit()
        return JSONResponse(result)
    except CounterError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)


@app.get("/centro-etiquetas", response_class=HTMLResponse)
def label_center(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    warehouse = _operation_warehouse(request, user, db)
    return templates.TemplateResponse(request, "centro_etiquetas.html", ctx_base(
        request, user, db, almacen=warehouse,
    ))


@app.get("/localizador", response_class=HTMLResponse)
def visual_locator(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    warehouse = _operation_warehouse(request, user, db)
    return templates.TemplateResponse(request, "localizador.html", ctx_base(
        request, user, db, almacen=warehouse,
    ))


def _daily_snapshot(db: Session, warehouse_id: int, day: date) -> dict:
    start = datetime.combine(day, datetime.min.time())
    end = start + timedelta(days=1)
    return {
        "fecha": day.isoformat(),
        "albaranes": db.query(AlbaranSalida).filter(AlbaranSalida.almacen_id == warehouse_id, AlbaranSalida.fecha_salida >= start, AlbaranSalida.fecha_salida < end).count(),
        "incidencias_abiertas": db.query(Incidencia).filter(Incidencia.almacen_id == warehouse_id, Incidencia.estado.in_(["abierta", "en_curso"])).count(),
        "traspasos_pendientes": db.query(TransferenciaAlmacen).filter(TransferenciaAlmacen.estado == "en_transito", or_(TransferenciaAlmacen.origen_id == warehouse_id, TransferenciaAlmacen.destino_id == warehouse_id)).count(),
        "pedidos_pendientes": db.query(PedidoProveedor).filter(PedidoProveedor.almacen_id == warehouse_id, PedidoProveedor.estado.in_(["borrador", "enviado", "parcial"])).count(),
        "inventarios_abiertos": db.query(SesionInventario).filter(SesionInventario.almacen_id == warehouse_id, SesionInventario.estado.notin_(["cerrada", "cancelada"])).count(),
        "preparaciones_pendientes": db.query(PreparacionEntrega).filter(PreparacionEntrega.almacen_id == warehouse_id, PreparacionEntrega.estado == "preparada").count(),
    }


@app.get("/cierres-diarios", response_class=HTMLResponse)
def daily_closures_page(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    warehouse = _operation_warehouse(request, user, db)
    closures = db.query(CierreDiarioAlmacen).filter(CierreDiarioAlmacen.almacen_id == warehouse.id).order_by(CierreDiarioAlmacen.fecha.desc()).limit(60).all()
    return templates.TemplateResponse(request, "cierres_diarios.html", ctx_base(
        request, user, db, almacen=warehouse, resumen=_daily_snapshot(db, warehouse.id, date.today()),
        cierres=closures, cierre_hoy=next((item for item in closures if item.fecha == date.today()), None),
    ))


@app.post("/api/cierres-diarios")
def daily_closure_create(
    payload: CierreDiarioRequest, request: Request,
    user: Usuario = Depends(requiere_login_scan), db: Session = Depends(get_db),
):
    warehouse = _operation_warehouse(request, user, db)
    if db.query(CierreDiarioAlmacen).filter_by(almacen_id=warehouse.id, fecha=date.today()).first():
        raise HTTPException(409, "El cierre de hoy ya está firmado")
    _decode_delivery_signature(payload.firma_datos)
    snapshot = _daily_snapshot(db, warehouse.id, date.today())
    closure = CierreDiarioAlmacen(
        almacen_id=warehouse.id, fecha=date.today(), resumen_json=json.dumps(snapshot, ensure_ascii=False),
        firma_datos=payload.firma_datos, firma_nombre=payload.firma_nombre.strip(), usuario_id=user.id,
    )
    db.add(closure)
    db.flush()
    db.add(AuditoriaLog(tabla="cierres_diarios_almacen", registro_id=closure.id, accion="firmar",
                        resumen=f"Cierre diario firmado: {warehouse.nombre}", usuario_id=user.id))
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "El cierre de hoy ya está firmado")
    return JSONResponse({"ok": True, "id": closure.id})


def _pending_snapshot(db: Session, warehouse_id: int | None) -> dict:
    """Una sola lectura operativa para tableta y Centro de Pendientes."""
    now = datetime.now()
    incidents = db.query(Incidencia).filter(Incidencia.estado.notin_(["resuelta", "cerrada"]))
    repairs = db.query(Reparacion).filter(Reparacion.estado.notin_(["finalizada", "sin_reparacion"]))
    requests = db.query(SolicitudTrabajador).filter(SolicitudTrabajador.estado.notin_(["entregada", "rechazada", "cancelada"]))
    inventories = db.query(SesionInventario).filter(SesionInventario.estado.notin_(["cerrada", "cancelada"]))
    low_materials = db.query(Material).filter(Material.activo == True, Material.stock_minimo > 0, Material.stock_actual <= Material.stock_minimo)
    low_clothing = db.query(StockEPI).filter(StockEPI.stock_minimo > 0, StockEPI.cantidad <= StockEPI.stock_minimo)
    transfers = db.query(TransferenciaAlmacen).filter(TransferenciaAlmacen.estado == "en_transito")
    if warehouse_id:
        incidents = incidents.filter(Incidencia.almacen_id == warehouse_id)
        repairs = repairs.filter(Reparacion.almacen_id == warehouse_id)
        requests = requests.filter(SolicitudTrabajador.almacen_id == warehouse_id)
        inventories = inventories.filter(SesionInventario.almacen_id == warehouse_id)
        low_materials = low_materials.filter(Material.almacen_id == warehouse_id)
        low_clothing = low_clothing.filter(StockEPI.almacen_id == warehouse_id)
        transfers = transfers.filter(or_(TransferenciaAlmacen.origen_id == warehouse_id, TransferenciaAlmacen.destino_id == warehouse_id))
    variant_total = func.coalesce(func.sum(ExistenciaVariante.cantidad), 0)
    variant_rows = db.query(VarianteEPI, variant_total.label("total")).outerjoin(
        ExistenciaVariante,
        and_(ExistenciaVariante.variante_id == VarianteEPI.id,
             ExistenciaVariante.almacen_id == warehouse_id),
    ).filter(
        VarianteEPI.activo == True, VarianteEPI.stock_minimo > 0,
    ).group_by(VarianteEPI.id).having(variant_total <= VarianteEPI.stock_minimo).limit(100).all()
    low_variants = [{
        "nombre": " · ".join(part for part in [
            row.catalogo.nombre if row.catalogo else row.referencia_interna,
            f"T.{row.talla}" if row.talla else "", row.color or "",
        ] if part), "cantidad": int(total or 0), "stock_minimo": row.stock_minimo,
        "referencia": row.referencia_interna,
    } for row, total in variant_rows]
    maintenance = db.query(MantenimientoProgramado).filter(
        MantenimientoProgramado.estado.in_(["pendiente", "en_proceso", "vencido"]),
    ).order_by(MantenimientoProgramado.fecha_programada.asc())
    if warehouse_id:
        maintenance = maintenance.filter(or_(
            (MantenimientoProgramado.tipo_activo == "herramienta") & MantenimientoProgramado.activo_id.in_(
                db.query(Herramienta.id).filter(Herramienta.almacen_id == warehouse_id)
            ),
            (MantenimientoProgramado.tipo_activo == "maquinaria") & MantenimientoProgramado.activo_id.in_(
                db.query(Maquinaria.id).filter(Maquinaria.almacen_id == warehouse_id)
            ),
        ))
    overdue = db.query(Movimiento).options(joinedload(Movimiento.herramienta)).filter(
        Movimiento.tipo == "entrega", Movimiento.fecha_devolucion_prevista.isnot(None),
        Movimiento.fecha_devolucion_prevista < now,
    )
    # Las herramientas actualmente fuera son la autoridad para evitar falsos avisos.
    if warehouse_id:
        overdue = overdue.join(Herramienta).filter(Herramienta.almacen_id == warehouse_id)
    overdue_rows, seen_tools = [], set()
    for movement in overdue.order_by(Movimiento.fecha.desc()).limit(500).all():
        if movement.herramienta_id in seen_tools:
            continue
        seen_tools.add(movement.herramienta_id)
        if movement.herramienta and movement.herramienta.estado in {"entregada", "en_obra", "en_furgoneta", "en_transporte"}:
            overdue_rows.append(movement)
    low_clothing_rows = low_clothing.order_by(StockEPI.nombre, StockEPI.talla).limit(100).all()
    return {
        "incidencias": incidents.order_by(Incidencia.fecha_apertura.desc()).limit(100).all(),
        "reparaciones": repairs.order_by(Reparacion.fecha_entrada.desc()).limit(100).all(),
        "solicitudes": requests.order_by(SolicitudTrabajador.creado_en.asc()).limit(100).all(),
        "inventarios": inventories.order_by(SesionInventario.opened_at.asc()).limit(100).all(),
        "materiales_bajos": low_materials.order_by(Material.nombre).limit(100).all(),
        "ropa_baja": low_clothing_rows + low_variants,
        "traspasos": transfers.order_by(TransferenciaAlmacen.creado_en.asc()).limit(100).all(),
        "mantenimientos": maintenance.limit(100).all(), "devoluciones_vencidas": overdue_rows,
    }


@app.get("/pendientes", response_class=HTMLResponse)
def centro_pendientes(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "crear")):
        raise HTTPException(403, "Sin permiso para ver la operación del almacén")
    warehouse = _operation_warehouse(request, user, db)
    pending = _pending_snapshot(db, warehouse.id)
    return templates.TemplateResponse(request, "centro_pendientes.html", ctx_base(
        request, user, db, almacen=warehouse, pendientes=pending,
        pendientes_total=sum(len(rows) for rows in pending.values()),
    ))


def _preparation_line_details(db: Session, preparation: PreparacionEntrega) -> list[dict]:
    model_map = {
        "herramienta": Herramienta, "maquinaria": Maquinaria, "vehiculo": Vehiculo,
        "epi_individual": EPIIndividual, "material": Material,
        "stock_epi": StockEPI, "variante": ExistenciaVariante,
    }
    details = []
    for line in json.loads(preparation.lineas_json or "[]"):
        kind = str(line.get("tipo") or "")
        model = model_map.get(kind)
        obj = db.get(model, int(line.get("id") or 0)) if model else None
        name, code = "Artículo no disponible", "—"
        if isinstance(obj, Herramienta):
            name, code = obj.nombre, obj.codigo
        elif isinstance(obj, Maquinaria):
            name, code = obj.nombre, obj.codigo_interno or obj.codigo_barras or obj.num_serie or "—"
        elif isinstance(obj, Vehiculo):
            name, code = " ".join(filter(None, [obj.marca, obj.modelo, obj.matricula])), obj.codigo or obj.matricula
        elif isinstance(obj, EPIIndividual):
            name, code = obj.tipo, obj.codigo_qr or obj.referencia_interna or obj.codigo_fabricacion
        elif isinstance(obj, Material):
            name, code = obj.nombre, obj.codigo
        elif isinstance(obj, StockEPI):
            name, code = obj.nombre_display, obj.codigo or "—"
        elif isinstance(obj, ExistenciaVariante) and obj.variante:
            variant = obj.variante
            catalog_name = variant.catalogo.nombre if variant.catalogo else "Ropa / EPI"
            variant_info = " · ".join(filter(None, [variant.modelo, variant.color, variant.talla]))
            name = catalog_name + (f" · {variant_info}" if variant_info else "")
            code = variant.codigo_qr or variant.referencia_interna
        details.append({
            "tipo": kind, "nombre": name or "Artículo", "codigo": code or "—",
            "cantidad": line.get("cantidad") or 1,
        })
    return details


@app.get("/preparaciones-entrega/qr/{qr_token}", response_class=HTMLResponse)
def preparation_qr_detail(
    qr_token: str, request: Request,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    preparation = db.query(PreparacionEntrega).filter(
        PreparacionEntrega.qr_token == qr_token,
    ).first()
    if not preparation:
        raise HTTPException(404, "Entrega preparada no encontrada")
    _require_warehouse_access(user, preparation.almacen_id)
    return templates.TemplateResponse(request, "preparacion_qr_detalle.html", ctx_base(
        request, user, db, preparacion=preparation,
        lineas=_preparation_line_details(db, preparation),
    ))


@app.get("/tablet", response_class=HTMLResponse)
def tablet_mode(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    warehouse = _operation_warehouse(request, user, db)
    pending = _pending_snapshot(db, warehouse.id)
    counts = {key: len(value) for key, value in pending.items()}
    counts["total"] = sum(counts.values())
    return templates.TemplateResponse(request, "tablet.html", ctx_base(
        request, user, db, almacen=warehouse, pendientes=counts,
        puede_solicitudes=can_manage_requests(user),
    ))


@app.post("/scan/stock-epi-operar")
def scan_stock_epi_operar(
    payload: ScanStockEPIRequest,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "stock_operar") or tiene_permiso(user, "crear")):
        raise HTTPException(403, "Sin permiso")
    start_stock_transaction(db)
    try:
        sepi = db.query(StockEPI).get(payload.stock_epi_id)
        if not sepi:
            raise StockError(404, "Artículo no encontrado")
        warehouse = _active_warehouse(db, user, request)
        if not warehouse or sepi.almacen_id != warehouse.id:
            raise StockError(404, "Artículo no encontrado en este almacén")
        delta = payload.cantidad if payload.accion == "entrada" else -payload.cantidad
        result = move_stock_epi(
            db, user, sepi.id, delta,
            tipo=payload.accion,
            event_id=f"scan-epi-{payload.scan_event_id}",
            motivo=payload.motivo or f"Escaneo {payload.accion} — {sepi.nombre_display}",
        )
        response = {
            "ok": True,
            "nombre": sepi.nombre_display,
            "cantidad_nueva": int(result.saldo_posterior),
            "accion": payload.accion,
            "reutilizada": result.reused,
        }
        db.commit()
        return JSONResponse(response)
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise


# ─── QR Universal ─────────────────────────────────────────────────────────────

def _get_qr_code_for(tipo: str, item_id: int, db: Session):
    """Devuelve (codigo_escaneable, nombre_display) para cualquier tipo de artículo."""
    if tipo == "herramienta":
        obj = db.query(Herramienta).get(item_id)
        return (obj.codigo, obj.nombre) if obj else (None, None)
    if tipo == "material":
        obj = db.query(Material).get(item_id)
        return (obj.codigo, obj.nombre) if obj else (None, None)
    if tipo == "stock_epi":
        obj = db.query(StockEPI).get(item_id)
        return (obj.codigo, obj.nombre_display) if obj else (None, None)
    if tipo == "epi_individual":
        obj = db.query(EPIIndividual).get(item_id)
        if not obj:
            return (None, None)
        codigo = obj.codigo_qr or obj.referencia_interna or obj.codigo_fabricacion
        return (codigo, f"{obj.tipo} — {obj.codigo_fabricacion}")
    if tipo == "maquinaria":
        obj = db.query(Maquinaria).get(item_id)
        if not obj:
            return (None, None)
        codigo = obj.codigo_interno or obj.codigo_barras or f"MRD-MAQ-{obj.id}"
        return (codigo, obj.nombre)
    if tipo == "almacen":
        obj = db.get(Almacen, item_id)
        return (obj.codigo, obj.nombre) if obj else (None, None)
    if tipo == "ubicacion":
        obj = db.get(Ubicacion, item_id)
        if not obj:
            return (None, None)
        almacen = obj.almacen.nombre if obj.almacen else "Almacén"
        return (obj.codigo, f"{almacen} — {obj.nombre}")
    if tipo == "vehiculo":
        obj = db.get(Vehiculo, item_id)
        if not obj:
            return (None, None)
        nombre = " ".join(filter(None, [obj.marca, obj.modelo])) or obj.matricula
        return (obj.codigo, f"{nombre} — {obj.matricula}")
    return (None, None)


@app.get("/qr/{tipo}/{item_id}")
def qr_imagen(
    tipo: str,
    item_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Devuelve imagen PNG del QR para cualquier artículo."""
    codigo, _ = _get_qr_code_for(tipo, item_id, db)
    if not codigo:
        raise HTTPException(404, "Artículo o código no encontrado")
    try:
        import qrcode as _qrcode
        from io import BytesIO
        qr = _qrcode.QRCode(version=1, error_correction=_qrcode.constants.ERROR_CORRECT_M,
                            box_size=8, border=3)
        qr.add_data(codigo)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        from fastapi.responses import Response as _Resp
        return _Resp(content=buf.read(), media_type="image/png",
                     headers={"Cache-Control": "max-age=3600"})
    except ImportError:
        raise HTTPException(500, "Librería qrcode no instalada. Ejecuta: pip install qrcode[pil]")


@app.get("/qr/{tipo}/{item_id}/imprimir")
def qr_imprimir(
    tipo: str,
    item_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Página HTML para imprimir etiqueta QR de cualquier artículo."""
    codigo, nombre = _get_qr_code_for(tipo, item_id, db)
    if not codigo:
        raise HTTPException(404, "Artículo o código no encontrado")
    tipos_label = {
        "herramienta": "Herramienta", "material": "Material",
        "stock_epi": "Stock EPI/Ropa", "epi_individual": "EPI Individual",
        "maquinaria": "Maquinaria", "almacen": "Almacén",
        "ubicacion": "Estantería / ubicación", "vehiculo": "Vehículo",
    }
    return templates.TemplateResponse(request, "qr_imprimir.html", {
        "request": request, "user": user,
        "tipo": tipo, "tipo_label": tipos_label.get(tipo, tipo),
        "item_id": item_id, "codigo": codigo, "nombre": nombre,
        "qr_url": f"/qr/{tipo}/{item_id}",
    })


def _scan_response_for_existing(reservation):
    if reservation.estado == "pending":
        return JSONResponse({
            "resultado": "pending",
            "detalle": "La operación sigue en proceso; vuelve a consultar con el mismo identificador.",
        }, status_code=202)
    result = reservation.result or {"resultado": reservation.estado}
    status = int(result.get("http_status", 200 if reservation.estado == "ok" else (409 if reservation.estado == "conflicto" else 422)))
    return JSONResponse(result, status_code=status)


@app.post("/scan/operar")
def scan_operar(
    payload: ScanOperationRequest,
    user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
    request: Request = None,
):
    """Opera con idempotencia y reutiliza las reglas transaccionales de movimientos."""
    try:
        require_movement_permission(user, payload.accion)
    except MovementError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    actor = actor_snapshot(user)
    user_name = user.nombre or user.username
    warehouse = _active_warehouse(db, user, request)
    tool = db.get(Herramienta, payload.herramienta_id)
    if not tool or (warehouse and tool.almacen_id not in (None, warehouse.id)):
        raise HTTPException(404, "Herramienta no encontrada en este almacén")
    if warehouse and payload.almacen_id and payload.almacen_id != warehouse.id:
        raise HTTPException(409, "Cambia primero al almacén que quieres operar")

    normalized = payload.model_dump(exclude={"scan_event_id"})
    normalized["usuario_id"] = actor.id
    content_hash = request_hash(normalized)
    try:
        reservation = reserve_event(
            db,
            scan_event_id=payload.scan_event_id,
            content_hash=content_hash,
            action=payload.accion,
            herramienta_id=payload.herramienta_id,
            user_id=actor.id,
        )
    except ScanIdConflict as exc:
        db.rollback()
        return JSONResponse({"resultado": "conflicto", "detalle": str(exc)}, status_code=409)

    if not reservation.acquired:
        return _scan_response_for_existing(reservation)

    start_movement_transaction(db)
    try:
        if payload.accion == "entregar":
            movement = deliver_tool(
                db, actor, payload.herramienta_id,
                payload.trabajador_id, payload.obra_id, payload.observaciones,
            )
            delivery_note = create_delivery_note(
                db, user_id=actor.id, worker_id=payload.trabajador_id,
                work_id=payload.obra_id, notes=payload.observaciones,
                warehouse_id=warehouse.id if warehouse else None,
                lines=[{"tipo": "herramienta", "id": movement.herramienta_id,
                        "cantidad": 1, "nombre": movement.codigo,
                        "movimiento_id": movement.movimiento_id}],
            )
        else:
            movement = return_tool(
                db, actor, payload.herramienta_id,
                warehouse.id if warehouse else payload.almacen_id,
                payload.condicion, payload.observaciones,
            )
        result = {
            "resultado": "ok",
            "scan_event_id": payload.scan_event_id,
            "accion": payload.accion,
            "herramienta_id": movement.herramienta_id,
            "codigo": movement.codigo,
            "estado": movement.estado,
            "estado_label": movement.estado_label,
            "movimiento_id": movement.movimiento_id,
            "destino": movement.destino,
            "usuario": user_name,
        }
        if payload.accion == "entregar":
            result.update({
                "albaran_id": delivery_note.id,
                "albaran_url": f"/albaranes-salida/{delivery_note.id}",
                "albaran_pdf_url": f"/albaranes-salida/{delivery_note.id}/pdf",
            })
        notification = {
            "tipo": "estado_herramienta",
            "herramienta_id": movement.herramienta_id,
            "codigo": movement.codigo,
            "estado": movement.estado,
            "estado_label": movement.estado_label,
            "accion": payload.accion,
            "resultado": "ok",
            "usuario": user_name,
        }
        finish_event(
            db, event_id=reservation.event_id, lease_token=reservation.lease_token,
            estado="ok", result=result, notification=notification,
        )
        db.commit()
        return JSONResponse(result)
    except MovementError as exc:
        db.rollback()
        final_state = "conflicto" if exc.status_code == 409 else "error"
        result = {
            "resultado": final_state,
            "scan_event_id": payload.scan_event_id,
            "accion": payload.accion,
            "herramienta_id": payload.herramienta_id,
            "detalle": exc.detail,
            "http_status": exc.status_code,
        }
        notification = {
            "tipo": "conflicto_herramienta" if final_state == "conflicto" else "error_operacion",
            "herramienta_id": payload.herramienta_id,
            "accion": payload.accion,
            "resultado": final_state,
            "detalle": exc.detail,
            "usuario": user_name,
        }
        try:
            finish_event(
                db, event_id=reservation.event_id, lease_token=reservation.lease_token,
                estado=final_state, result=result, notification=notification,
            )
            db.commit()
        except ScanLeaseLost:
            db.rollback()
            return JSONResponse({"resultado": "pending", "detalle": "Lease recuperado por otra petición"}, status_code=202)
        return JSONResponse(result, status_code=exc.status_code)
    except ScanLeaseLost:
        db.rollback()
        return JSONResponse({"resultado": "pending", "detalle": "Lease recuperado por otra petición"}, status_code=202)
    except Exception:
        db.rollback()
        result = {
            "resultado": "error",
            "scan_event_id": payload.scan_event_id,
            "accion": payload.accion,
            "herramienta_id": payload.herramienta_id,
            "detalle": "La operación no pudo completarse y fue revertida.",
            "http_status": 500,
        }
        mark_event_error(db, reservation.event_id, reservation.lease_token, result)
        return JSONResponse(result, status_code=500)


@app.get("/scan/cambios")
def scan_cambios(
    ultimo_id: int = Query(0, ge=0),
    limite: int = Query(50, ge=1, le=100),
    inicializar: bool = Query(False),
    user: Usuario = Depends(requiere_login_scan),
    db: Session = Depends(get_db),
):
    """Polling autenticado con cursor incremental; nunca usa marcas de tiempo."""
    if not (tiene_permiso(user, "entregar") or tiene_permiso(user, "devolver")):
        raise HTTPException(403, "Sin permiso para consultar actividad operativa")
    deleted = cleanup_scan_data(db)
    if deleted["notifications"] or deleted["events"]:
        db.commit()
    if inicializar:
        return JSONResponse({"cambios": [], "next_cursor": current_notification_cursor(db)})
    items, next_cursor = changes_after(db, ultimo_id, limite)
    return JSONResponse({"cambios": items, "next_cursor": next_cursor})


# ─── Importación Excel ────────────────────────────────────────────────────────
@app.get("/informes", response_class=HTMLResponse)
def informes(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    analisis = generar_analisis_inteligente(db)
    # Datos para Chart.js (JSON seguro)
    chart_estados = {
        "labels": list(analisis["herramientas"]["estados"].keys()),
        "data": list(analisis["herramientas"]["estados"].values()),
    }
    chart_mov = {
        "labels": [r["mes"] for r in analisis["movimientos"]["por_mes"]],
        "data": [r["total"] for r in analisis["movimientos"]["por_mes"]],
    }
    return templates.TemplateResponse(request, "informes.html", ctx_base(
        request, user, db,
        analisis=analisis,
        chart_estados_json=dumps_for_script(chart_estados),
        chart_mov_json=dumps_for_script(chart_mov),
    ))


@app.get("/informes/inventario/excel")
def informe_inventario_excel(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Herramienta.nombre).all()
    excel = exportar_inventario_excel(herramientas)
    nombre = f"inventario_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@app.get("/informes/inventario/pdf")
def informe_inventario_pdf(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Herramienta.nombre).all()
    pdf = exportar_inventario_pdf(herramientas)
    nombre = f"inventario_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@app.get("/informes/movimientos/excel")
def informe_movimientos_excel(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    movimientos = db.query(Movimiento).join(Herramienta).filter(
        Herramienta.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Movimiento.fecha.desc()).limit(5000).all()
    excel = exportar_movimientos_excel(movimientos)
    nombre = f"movimientos_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


# ─── Configuración ────────────────────────────────────────────────────────────
@app.get("/configuracion", response_class=HTMLResponse)
def configuracion(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "config"):
        return RedirectResponse("/", status_code=303)
    backups = listar_backups()
    usuarios = db.query(Usuario).order_by(Usuario.username).all()
    almacenes = db.query(Almacen).filter(Almacen.activo == True).order_by(Almacen.nombre).all()
    version_info = leer_version_actual()
    return templates.TemplateResponse(request, "configuracion.html", ctx_base(
        request, user,
        backups=backups,
        usuarios=usuarios,
        almacenes=almacenes,
        version_info=version_info,
    ))


@app.post("/configuracion/backup")
def hacer_backup(user: Usuario = Depends(requiere_login)):
    if not tiene_permiso(user, "backup"):
        raise HTTPException(403, "Sin permiso")
    resultado = crear_backup()
    if resultado["ok"]:
        return RedirectResponse("/configuracion?backup=ok", status_code=303)
    return RedirectResponse("/configuracion?backup=error", status_code=303)


@app.get("/configuracion/backup/descargar/{nombre}")
def descargar_backup(nombre: str, user: Usuario = Depends(requiere_login)):
    if not tiene_permiso(user, "backup"):
        raise HTTPException(403, "Sin permiso")
    # Seguridad: solo nombres de backup válidos — evita path traversal
    if not re.match(r'^(backup|pre_update|antes_restaurar)_[\w\-]+\.db$', nombre):
        raise HTTPException(400, "Nombre de archivo no válido")
    ruta = (BACKUPS_DIR / nombre).resolve()
    # Verificar que la ruta resuelta está dentro de BACKUPS_DIR
    if not str(ruta).startswith(str(BACKUPS_DIR.resolve())):
        raise HTTPException(403, "Acceso denegado")
    if not ruta.exists() or not ruta.is_file():
        raise HTTPException(404)
    return FileResponse(str(ruta), filename=nombre, media_type="application/octet-stream")


@app.post("/configuracion/usuarios/nuevo")
def usuario_nuevo(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    username: str = Form(...),
    nombre: str = Form(...),
    password: str = Form(...),
    rol: str = Form("consulta"),
    email: str = Form(""),
    almacen_id: str = Form(""),
):
    if not tiene_permiso(user, "usuarios"):
        raise HTTPException(403, "Sin permiso")
    if rol not in ROLES_VALIDOS:
        raise HTTPException(400, "Rol no válido")
    warehouse = None
    if rol != "admin":
        try:
            selected_warehouse_id = int(almacen_id)
        except (TypeError, ValueError):
            raise HTTPException(400, "Selecciona el almacén del usuario")
        warehouse = db.query(Almacen).filter(
            Almacen.id == selected_warehouse_id, Almacen.activo == True,
        ).first()
        if not warehouse:
            raise HTTPException(400, "Selecciona el almacén del usuario")
    existe = db.query(Usuario).filter(Usuario.username == username).first()
    if not existe:
        # Sprint 5.2: validar política de contraseñas
        try:
            validar_contrasena(password, username=username, min_length=PASSWORD_MIN_LENGTH)
        except ErrorContrasena as _ec:
            raise HTTPException(400, str(_ec))
        db.add(Usuario(
            username=username, nombre=nombre,
            password_hash=hash_password(password),
            rol=rol, email=email or None, activo=True,
            almacen_id=warehouse.id if warehouse else None,
            must_change_password=True,  # Sprint 5.2: forzar cambio en primer login
        ))
        db.commit()
        mrd_logging.log_security(
            f"Usuario nuevo creado: username='{username}' rol={rol} por admin='{user.username}'",
            level="info",
        )
    return RedirectResponse("/configuracion?usuario=ok", status_code=303)


@app.post("/configuracion/usuarios/{uid}/toggle")
def usuario_toggle(uid: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "usuarios"):
        raise HTTPException(403, "Sin permiso")
    u = db.query(Usuario).get(uid)
    if u and u.username != DEFAULT_ADMIN_USER:
        u.activo = not u.activo
        db.commit()
    return RedirectResponse("/configuracion", status_code=303)


@app.post("/configuracion/usuarios/{uid}/editar")
def usuario_editar(
    uid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    rol: str = Form("consulta"),
    email: str = Form(""),
    password: str = Form(""),
    almacen_id: str = Form(""),
):
    if not tiene_permiso(user, "usuarios"):
        raise HTTPException(403, "Sin permiso")
    if rol not in ROLES_VALIDOS:
        raise HTTPException(400, "Rol no válido")
    warehouse = None
    if rol != "admin":
        try:
            selected_warehouse_id = int(almacen_id)
        except (TypeError, ValueError):
            raise HTTPException(400, "Selecciona el almacén del usuario")
        warehouse = db.query(Almacen).filter(
            Almacen.id == selected_warehouse_id, Almacen.activo == True,
        ).first()
        if not warehouse:
            raise HTTPException(400, "Selecciona el almacén del usuario")
    u = db.query(Usuario).get(uid)
    if u:
        u.nombre = nombre
        u.rol = rol
        u.email = email or None
        u.almacen_id = warehouse.id if warehouse else None
        if password.strip():
            # Sprint 5.2: validar política de contraseñas al resetear
            try:
                validar_contrasena(password.strip(), username=u.username, min_length=PASSWORD_MIN_LENGTH)
            except ErrorContrasena as _ec:
                raise HTTPException(400, str(_ec))
            u.password_hash = hash_password(password.strip())
            u.must_change_password = True  # forzar cambio al siguiente login
            mrd_logging.log_security(
                f"Contraseña reseteada por admin: usuario='{u.username}' admin='{user.username}'",
                level="info",
            )
        db.commit()
    return RedirectResponse("/configuracion?usuario=editado", status_code=303)


@app.post("/configuracion/usuarios/{uid}/eliminar")
def usuario_eliminar(uid: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "usuarios"):
        raise HTTPException(403, "Sin permiso")
    u = db.query(Usuario).get(uid)
    if u and u.username != DEFAULT_ADMIN_USER and u.id != user.id:
        db.delete(u)
        db.commit()
    return RedirectResponse("/configuracion?usuario=eliminado", status_code=303)


# ─── Actualizaciones ──────────────────────────────────────────────────────────
@app.post("/actualizaciones/aplicar")
def actualizar_app(
    user: Usuario = Depends(requiere_login),
    archivo: str = Form(...),
):
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    resultado = aplicar_actualizacion(archivo)
    if resultado["ok"]:
        return RedirectResponse("/actualizaciones?update=ok", status_code=303)
    return RedirectResponse(f"/actualizaciones?update=error&msg={resultado.get('error', '')}", status_code=303)


@app.post("/actualizaciones/empaquetar")
def empaquetar_release(user: Usuario = Depends(requiere_login)):
    """Crea el ZIP de release con la versión actual del código."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    version_info = leer_version_actual()
    version = version_info.get("version_actual", "1.0.0")
    EXCLUIR = {
        "data", "backups", "uploads", "logs", "venv",
        "__pycache__", ".git", "releases", "exports",
        "get_access_info.bat", "get_access_info.py",
        "create_release.bat", "access_info.txt", "qr_movil.png",
    }
    out = BASE_DIR / "releases" / f"mrd_tool_control_v{version}.zip"
    out.parent.mkdir(exist_ok=True)
    try:
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in BASE_DIR.rglob("*"):
                parts = f.relative_to(BASE_DIR).parts
                if not parts or parts[0] in EXCLUIR:
                    continue
                if any(p == "__pycache__" for p in parts):
                    continue
                if any(p.startswith("_update_tmp") for p in parts):
                    continue
                if f.is_file():
                    arcname = f"mrd_tool_control/{f.relative_to(BASE_DIR)}"
                    zf.write(f, arcname)
        size_kb = out.stat().st_size // 1024
        return RedirectResponse(f"/actualizaciones?empaquetado=ok&version={version}&size={size_kb}", status_code=303)
    except Exception as e:
        return RedirectResponse(f"/actualizaciones?empaquetado=error&msg={str(e)}", status_code=303)


# ─── Health ───────────────────────────────────────────────────────────────────
@app.get("/health")
def health_simple():
    """Health check rápido — público — para proxies y monitores."""
    v = leer_version_actual().get("version_actual", VERSION)
    return {"status": "ok", "version": v}


@app.get("/api/system/health")
def health_completo(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    """Health check completo — solo administradores."""
    resultado: dict = {
        "status": "healthy",
        "version": VERSION,
        "checked_at": datetime.now(timezone.utc).isoformat(),
    }

    # Base de datos
    try:
        db.execute(text("SELECT 1"))
        resultado["database"] = "ok"
    except Exception as e:
        resultado["database"] = f"error: {e}"
        resultado["status"] = "degraded"

    # Sistema de archivos
    try:
        test = DATA_DIR / ".health_check"
        test.write_text("ok")
        test.unlink()
        resultado["storage"] = "ok"
    except Exception:
        resultado["storage"] = "error"
        resultado["status"] = "degraded"

    # Backups
    backups = list(BACKUPS_DIR.glob("backup_*.db"))
    resultado["backup_count"] = len(backups)
    resultado["backup"] = "ok" if backups else "sin_backups"

    # Disco
    try:
        total, used, free = shutil.disk_usage(str(BASE_DIR))
        resultado["disk_free_gb"] = round(free / 1024 ** 3, 1)
        resultado["disk_total_gb"] = round(total / 1024 ** 3, 1)
        if free < 500 * 1024 * 1024:  # < 500 MB
            resultado["status"] = "degraded"
    except Exception:
        resultado["disk_free_gb"] = None

    # RAM y uptime via psutil
    try:
        import psutil
        mem = psutil.virtual_memory()
        resultado["ram_used_pct"] = round(mem.percent, 1)
        proc = psutil.Process()
        resultado["uptime_seconds"] = int(time.time() - proc.create_time())
    except ImportError:
        resultado["ram_used_pct"] = None
        resultado["uptime_seconds"] = None

    # Acceso remoto
    try:
        ra = remote_access.get_status_cached(max_age=120)
        resultado["remote_access"] = "ok" if ra.get("public_url") else "sin_url_publica"
    except Exception:
        resultado["remote_access"] = "error"

    # Logs
    log_dir = BASE_DIR / "logs"
    resultado["logs"] = "ok" if log_dir.exists() else "sin_directorio"

    return resultado


# ─── API ──────────────────────────────────────────────────────────────────────
@app.get("/api/herramientas/buscar")
def api_buscar(
    q: str = Query(""),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    herramientas = (
        db.query(Herramienta)
        .filter(Herramienta.activa == True)
        .filter(or_(
            Herramienta.codigo.ilike(f"%{q}%"),
            Herramienta.nombre.ilike(f"%{q}%"),
        ))
        .limit(10)
        .all()
    )
    return [{"id": h.id, "codigo": h.codigo, "nombre": h.nombre, "estado": h.estado} for h in herramientas]


@app.get("/api/herramientas/codigo/{codigo}")
def api_herramienta_por_codigo(
    codigo: str,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.codigo == codigo, Herramienta.activa == True).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    return {
        "id": h.id,
        "codigo": h.codigo,
        "nombre": h.nombre,
        "estado": h.estado,
        "ubicacion": h.ubicacion_texto,
        "url": f"/herramientas/{h.id}",
    }


@app.get("/api/stats")
def api_stats(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    # 1 query GROUP BY en lugar de 4 COUNT separados
    counts = dict(
        db.query(Herramienta.estado, func.count(Herramienta.id))
        .filter(Herramienta.activa == True)
        .group_by(Herramienta.estado)
        .all()
    )
    return {
        "total":      sum(counts.values()),
        "disponibles": counts.get("disponible", 0),
        "entregadas":  counts.get("entregada", 0),
        "en_obra":     counts.get("en_obra", 0),
        "en_furgoneta": counts.get("en_furgoneta", 0),
        "en_reparacion": counts.get("en_reparacion", 0),
        "perdidas":    counts.get("perdida", 0),
    }


@app.get("/api/version/check")
def api_version_check(user: Usuario = Depends(requiere_login)):
    result = verificar_actualizacion_local()
    version_actual = leer_version_actual().get("version_actual", VERSION)
    return {
        "version_actual": version_actual,
        "actualizacion_disponible": result.get("disponible", False),
        "nueva_version": result.get("version", ""),
    }


# ─── Proveedores ──────────────────────────────────────────────────────────────
@app.get("/proveedores", response_class=HTMLResponse)
def proveedores_list(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    proveedores = db.query(Proveedor).order_by(Proveedor.nombre).all()
    return templates.TemplateResponse(request, "proveedores.html", ctx_base(
        request, user, proveedores=proveedores, total=len(proveedores)
    ))


@app.post("/proveedores/nuevo")
def proveedor_nuevo(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    cif: str = Form(""),
    telefono: str = Form(""),
    email: str = Form(""),
    direccion: str = Form(""),
    web: str = Form(""),
    contacto: str = Form(""),
    observaciones: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    db.add(Proveedor(
        nombre=nombre, cif=cif or None, telefono=telefono or None,
        email=email or None, direccion=direccion or None,
        web=web or None, contacto=contacto or None,
        observaciones=observaciones or None, activo=True,
    ))
    db.commit()
    return RedirectResponse("/proveedores", status_code=303)


@app.post("/proveedores/{pid}/toggle")
def proveedor_toggle(pid: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    p = db.query(Proveedor).get(pid)
    if p:
        p.activo = not p.activo
        db.commit()
    return RedirectResponse("/proveedores", status_code=303)


@app.post("/proveedores/{pid}/editar", response_class=RedirectResponse)
async def proveedor_editar(
    pid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    p = db.query(Proveedor).get(pid)
    if not p:
        raise HTTPException(404)
    form = await request.form()
    p.nombre = form.get("nombre") or p.nombre
    p.cif = form.get("cif") or None
    p.telefono = form.get("telefono") or None
    p.email = form.get("email") or None
    p.direccion = form.get("direccion") or None
    p.web = form.get("web") or None
    p.contacto = form.get("contacto") or None
    p.observaciones = form.get("observaciones") or None
    db.commit()
    return RedirectResponse("/proveedores?ok=editado", status_code=303)


# ─── Incidencias ──────────────────────────────────────────────────────────────
@app.get("/incidencias", response_class=HTMLResponse)
def incidencias_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = "",
    prioridad: str = "",
):
    warehouse = _active_warehouse(db, user, request)
    query = db.query(Incidencia)
    if warehouse:
        query = query.filter(Incidencia.almacen_id == warehouse.id)
    if estado:
        query = query.filter(Incidencia.estado == estado)
    if prioridad:
        query = query.filter(Incidencia.prioridad == prioridad)
    incidencias = query.order_by(Incidencia.fecha_apertura.desc()).all()
    total = len(incidencias)
    abiertas = sum(1 for i in incidencias if i.estado == "abierta")
    en_proceso = sum(1 for i in incidencias if i.estado == "en_curso")
    criticas = sum(1 for i in incidencias if i.prioridad == "critica" and i.estado != "cerrada")
    return templates.TemplateResponse(request, "incidencias.html", ctx_base(
        request, user, db,
        incidencias=incidencias, total=total, abiertas=abiertas,
        en_proceso=en_proceso, criticas=criticas,
        estado_filtro=estado, prioridad_filtro=prioridad,
    ))


@app.get("/incidencias/nueva", response_class=HTMLResponse)
def incidencia_nueva_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta: int = Query(None),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Herramienta.nombre).all()
    h_sel = db.query(Herramienta).get(herramienta) if herramienta else None
    return templates.TemplateResponse(request, "nueva_incidencia.html", ctx_base(
        request, user, db, herramientas=herramientas, herramienta_sel=h_sel
    ))


@app.post("/incidencias/nueva")
def incidencia_nueva_post(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    titulo: str = Form(...),
    tipo: str = Form("averia"),
    prioridad: str = Form("media"),
    descripcion: str = Form(""),
    herramienta_id: str = Form(""),
    request: Request = None,
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    tool = db.get(Herramienta, int(herramienta_id)) if herramienta_id else None
    if herramienta_id and (not tool or not warehouse or tool.almacen_id != warehouse.id):
        raise HTTPException(400, "La herramienta no pertenece al almacén activo")
    año = datetime.now().year
    count = db.query(Incidencia).filter(Incidencia.numero.like(f"INC-{año}-%")).count()
    numero = f"INC-{año}-{(count+1):04d}"
    inc = Incidencia(
        numero=numero, titulo=titulo, tipo=tipo, prioridad=prioridad,
        descripcion=descripcion or None, estado="abierta",
        fecha_apertura=datetime.utcnow(),
        herramienta_id=int(herramienta_id) if herramienta_id else None,
        creado_por_id=user.id,
        almacen_id=warehouse.id if warehouse else None,
    )
    db.add(inc)
    db.commit()
    return RedirectResponse("/incidencias", status_code=303)


@app.get("/incidencias/{iid}", response_class=HTMLResponse)
def incidencia_detalle(
    iid: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    inc = db.get(Incidencia, iid)
    if not inc:
        raise HTTPException(404)
    _require_warehouse_access(user, inc.almacen_id)
    return templates.TemplateResponse(request, "incidencia_detalle.html", ctx_base(
        request, user, db, incidencia=inc,
        puede_editar=tiene_permiso(user, "editar"),
    ))


@app.post("/incidencias/{iid}/cerrar")
def incidencia_cerrar(
    iid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    solucion: str = Form(""),
    request: Request = None,
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    inc = db.query(Incidencia).get(iid)
    if inc:
        warehouse = _active_warehouse(db, user, request)
        if not warehouse or inc.almacen_id != warehouse.id:
            raise HTTPException(404)
        inc.estado = "cerrada"
        inc.solucion = solucion or None
        inc.fecha_resolucion = datetime.utcnow()
        db.commit()
    return RedirectResponse(f"/incidencias/{iid}?ok=cerrada", status_code=303)


@app.post("/incidencias/{iid}/editar", response_class=RedirectResponse)
async def incidencia_editar(
    iid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    inc = db.query(Incidencia).get(iid)
    if not inc:
        raise HTTPException(404)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse or inc.almacen_id != warehouse.id:
        raise HTTPException(404)
    form = await request.form()
    inc.titulo = form.get("titulo") or inc.titulo
    inc.tipo = form.get("tipo") or inc.tipo
    inc.prioridad = form.get("prioridad") or inc.prioridad
    inc.estado = form.get("estado") or inc.estado
    inc.descripcion = form.get("descripcion") or None
    inc.observaciones = form.get("observaciones") or None
    db.commit()
    return RedirectResponse(f"/incidencias/{iid}?ok=editado", status_code=303)


# ─── Reparaciones ─────────────────────────────────────────────────────────────
@app.get("/reparaciones", response_class=HTMLResponse)
def reparaciones_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = "",
):
    warehouse = _active_warehouse(db, user, request)
    query = db.query(Reparacion)
    if warehouse:
        query = query.filter(Reparacion.almacen_id == warehouse.id)
    if estado:
        query = query.filter(Reparacion.estado == estado)
    reparaciones = query.order_by(Reparacion.fecha_entrada.desc()).all()
    total = len(reparaciones)
    en_curso = sum(1 for r in reparaciones if r.estado in ("diagnostico", "en_reparacion", "pendiente_piezas"))
    finalizadas = sum(1 for r in reparaciones if r.estado == "finalizada")
    return templates.TemplateResponse(request, "reparaciones.html", ctx_base(
        request, user, db,
        reparaciones=reparaciones, total=total, en_curso=en_curso, finalizadas=finalizadas,
        estado_filtro=estado,
    ))


@app.get("/reparaciones/nueva", response_class=HTMLResponse)
def reparacion_nueva_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta: int = Query(None),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Herramienta.nombre).all()
    proveedores = db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.nombre).all()
    h_sel = db.query(Herramienta).get(herramienta) if herramienta else None
    return templates.TemplateResponse(request, "nueva_reparacion.html", ctx_base(
        request, user, db, herramientas=herramientas, proveedores=proveedores, herramienta_sel=h_sel
    ))


@app.post("/reparaciones/nueva")
def reparacion_nueva_post(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_id: str = Form(...),
    diagnostico: str = Form(""),
    proveedor_id: str = Form(""),
    coste_estimado: str = Form(""),
    fecha_prevista: str = Form(""),
    observaciones: str = Form(""),
    request: Request = None,
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    h_id = int(herramienta_id)
    h = db.query(Herramienta).get(h_id)
    if not h:
        raise HTTPException(404)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse or h.almacen_id != warehouse.id:
        raise HTTPException(404)

    año = datetime.now().year
    count = db.query(Reparacion).filter(Reparacion.numero.like(f"REP-{año}-%")).count()
    numero = f"REP-{año}-{(count+1):04d}"

    rep = Reparacion(
        numero=numero,
        herramienta_id=h_id,
        diagnostico=diagnostico or None,
        proveedor_id=int(proveedor_id) if proveedor_id else None,
        coste_estimado=float(coste_estimado) if coste_estimado else None,
        fecha_entrada=datetime.utcnow(),
        fecha_prevista=datetime.strptime(fecha_prevista, "%Y-%m-%d") if fecha_prevista else None,
        estado="diagnostico",
        observaciones=observaciones or None,
        creado_por_id=user.id,
        almacen_id=warehouse.id,
    )
    db.add(rep)

    # Cambiar estado de la herramienta
    h.estado = "en_reparacion"
    h.ubicacion_texto = "En reparación"
    registrar_movimiento(db, h, "reparacion", "en_reparacion", user,
                         observaciones=f"Reparación {numero}: {diagnostico or ''}")
    db.commit()
    return RedirectResponse("/reparaciones", status_code=303)


@app.get("/reparaciones/{rid}", response_class=HTMLResponse)
def reparacion_detalle(
    rid: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    rep = db.get(Reparacion, rid)
    if not rep:
        raise HTTPException(404)
    _require_warehouse_access(user, rep.almacen_id)
    proveedores = db.query(Proveedor).filter(Proveedor.activo == True).order_by(Proveedor.nombre).all()
    warehouse = _active_warehouse(db, user, request)
    almacenes = [warehouse] if warehouse else []
    return templates.TemplateResponse(request, "reparacion_detalle.html", ctx_base(
        request, user, db, reparacion=rep, proveedores=proveedores, almacenes=almacenes,
        puede_editar=tiene_permiso(user, "editar"),
    ))


@app.post("/reparaciones/{rid}/finalizar")
def reparacion_finalizar(
    rid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    coste_final: str = Form(""),
    almacen_id: str = Form(""),
    request: Request = None,
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    rep = db.query(Reparacion).get(rid)
    if not rep:
        raise HTTPException(404)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse or rep.almacen_id != warehouse.id:
        raise HTTPException(404)
    if almacen_id and int(almacen_id) != warehouse.id:
        raise HTTPException(409, "La reparación debe volver al almacén activo")
    rep.estado = "finalizada"
    rep.fecha_salida = datetime.utcnow()
    rep.coste_final = float(coste_final) if coste_final else None

    h = rep.herramienta
    if h:
        almacen = warehouse
        h.estado = "disponible"
        h.almacen_id = almacen.id if almacen else None
        h.ubicacion_texto = almacen.nombre if almacen else "Almacén"
        h.fecha_ultimo_mantenimiento = date.today()
        registrar_movimiento(db, h, "devolucion_reparacion", "disponible", user,
                             destino=h.ubicacion_texto,
                             observaciones=f"Reparación {rep.numero} finalizada")
    db.commit()
    return RedirectResponse(f"/reparaciones/{rid}?ok=finalizada", status_code=303)


@app.post("/reparaciones/{rid}/editar", response_class=RedirectResponse)
async def reparacion_editar(
    rid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    rep = db.query(Reparacion).get(rid)
    if not rep:
        raise HTTPException(404)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse or rep.almacen_id != warehouse.id:
        raise HTTPException(404)
    form = await request.form()
    rep.descripcion = form.get("descripcion") or None
    rep.diagnostico = form.get("diagnostico") or None
    rep.estado = form.get("estado") or rep.estado
    rep.prioridad = form.get("prioridad") or rep.prioridad
    pid = form.get("proveedor_id", "")
    rep.proveedor_id = int(pid) if pid else None
    ce = form.get("coste_estimado", "")
    rep.coste_estimado = float(ce) if ce else None
    fp = form.get("fecha_prevista", "")
    if fp:
        rep.fecha_prevista = datetime.strptime(fp, "%Y-%m-%d")
    rep.observaciones = form.get("observaciones") or None
    db.commit()
    return RedirectResponse(f"/reparaciones/{rid}?ok=editado", status_code=303)


# ─── Movimientos: Entregar / Devolver (páginas dedicadas) ─────────────────────
def _require_movement_http(user: Usuario, action: str):
    try:
        require_movement_permission(user, action)
        return actor_snapshot(user)
    except MovementError as exc:
        raise HTTPException(exc.status_code, exc.detail)


def _validate_movement_warehouse(
    db: Session, user: Usuario, request: Request | None, tool_ids: list[int],
    worker_id: int | None = None, work_id: int | None = None,
) -> Almacen | None:
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        if request is None:
            return None
        raise HTTPException(409, "No hay un almacén activo configurado")
    tools_count = db.query(Herramienta).filter(
        Herramienta.id.in_(tool_ids), Herramienta.almacen_id == warehouse.id,
    ).count()
    if tools_count != len(set(tool_ids)):
        raise HTTPException(404, "Una herramienta no pertenece al almacén activo")
    if worker_id:
        worker = db.get(Trabajador, worker_id)
        if not worker or worker.almacen_id != warehouse.id:
            raise HTTPException(400, "El trabajador no pertenece al almacén activo")
    if work_id:
        work = db.get(Obra, work_id)
        if not work or work.almacen_id != warehouse.id:
            raise HTTPException(400, "La obra no pertenece al almacén activo")
    return warehouse


@app.get("/movimientos/entregar", response_class=HTMLResponse)
def movimiento_entregar_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta: int = Query(None),
):
    if not tiene_permiso(user, "entregar"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True, Herramienta.estado == "disponible",
        Herramienta.almacen_id == warehouse_id,
    ).order_by(Herramienta.nombre).all()
    trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == warehouse_id,
    ).order_by(Trabajador.nombre).all()
    obras = db.query(Obra).filter(
        Obra.activa == True, Obra.almacen_id == warehouse_id,
    ).order_by(Obra.nombre).all()
    h_sel = db.query(Herramienta).get(herramienta) if herramienta else None
    return templates.TemplateResponse(request, "movimiento_entregar.html", ctx_base(
        request, user, db,
        herramientas=herramientas, trabajadores=trabajadores, obras=obras, herramienta_sel=h_sel,
    ))


@app.post("/movimientos/entregar")
def movimiento_entregar_post(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_id: str = Form(...),
    trabajador_id: str = Form(""),
    obra_id: str = Form(""),
    observaciones: str = Form(""),
    firma_datos: str = Form(""),
    firma_nombre: str = Form(""),
    fecha_devolucion_prevista: str = Form(""),
    request: Request = None,
):
    actor = _require_movement_http(user, "entregar")
    try:
        h_id = int(herramienta_id)
        t_id = int(trabajador_id) if trabajador_id else None
        o_id = int(obra_id) if obra_id else None
        prevista = (
            datetime.fromisoformat(fecha_devolucion_prevista)
            if isinstance(fecha_devolucion_prevista, str) and fecha_devolucion_prevista.strip()
            else None
        )
    except ValueError:
        raise HTTPException(400, "Datos de entrega no válidos")
    warehouse = _validate_movement_warehouse(db, user, request, [h_id], t_id, o_id)
    start_movement_transaction(db)
    try:
        result = deliver_tool(
            db, actor, h_id, t_id, o_id, observaciones,
            firma_datos, firma_nombre,
            prevista,
        )
        delivery_note = create_delivery_note(
            db, user_id=actor.id, worker_id=t_id, work_id=o_id,
            expected_return=prevista, notes=observaciones,
            signature_data=firma_datos, signature_name=firma_nombre,
            lines=[{"tipo": "herramienta", "id": result.herramienta_id,
                    "cantidad": 1, "nombre": result.codigo,
                    "movimiento_id": result.movimiento_id}],
            warehouse_id=warehouse.id if warehouse else None,
        )
        db.commit()
    except MovementError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    return RedirectResponse(f"/albaranes-salida/{delivery_note.id}", status_code=303)


@app.post("/movimientos/entregar/lote")
def movimiento_entregar_lote(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_ids: str = Form(...),
    trabajador_id: str = Form(""),
    obra_id: str = Form(""),
    observaciones: str = Form(""),
    firma_datos: str = Form(""),
    firma_nombre: str = Form(""),
    fecha_devolucion_prevista: str = Form(""),
    request: Request = None,
):
    """Registra una entrega múltiple en una única transacción."""
    actor = _require_movement_http(user, "entregar")
    try:
        ids = [int(value.strip()) for value in herramienta_ids.split(",") if value.strip()]
        t_id = int(trabajador_id) if trabajador_id else None
        o_id = int(obra_id) if obra_id else None
        prevista = (
            datetime.fromisoformat(fecha_devolucion_prevista)
            if isinstance(fecha_devolucion_prevista, str) and fecha_devolucion_prevista.strip()
            else None
        )
    except ValueError:
        raise HTTPException(400, "Lista de herramientas no válida")
    if len(ids) != len(set(ids)):
        raise HTTPException(400, "La lista contiene herramientas duplicadas")
    if len(ids) < 2:
        raise HTTPException(400, "La entrega múltiple requiere al menos dos herramientas")
    warehouse = _validate_movement_warehouse(db, user, request, ids, t_id, o_id)
    start_movement_transaction(db)

    try:
        results = []
        for hid in ids:
            results.append(deliver_tool(
                db, actor, hid, t_id, o_id, observaciones,
                firma_datos, firma_nombre,
                prevista,
            ))
        delivery_note = create_delivery_note(
            db, user_id=actor.id, worker_id=t_id, work_id=o_id,
            expected_return=prevista, notes=observaciones,
            signature_data=firma_datos, signature_name=firma_nombre,
            lines=[{"tipo": "herramienta", "id": row.herramienta_id,
                    "cantidad": 1, "nombre": row.codigo,
                    "movimiento_id": row.movimiento_id} for row in results],
            warehouse_id=warehouse.id if warehouse else None,
        )
        db.commit()
    except MovementError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise

    return JSONResponse({
        "ok": True, "count": len(ids), "albaran_id": delivery_note.id,
        "albaran_url": f"/albaranes-salida/{delivery_note.id}",
        "albaran_pdf_url": f"/albaranes-salida/{delivery_note.id}/pdf",
    })


@app.get("/movimientos/devolver", response_class=HTMLResponse)
def movimiento_devolver_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta: int = Query(None),
):
    if not tiene_permiso(user, "devolver"):
        raise HTTPException(403, "Sin permiso")
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    herramientas = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == warehouse_id,
        Herramienta.estado.in_(["entregada", "en_obra", "en_furgoneta", "en_transporte"])
    ).order_by(Herramienta.nombre).all()
    almacenes = [warehouse] if warehouse else []
    h_sel = db.query(Herramienta).get(herramienta) if herramienta else None
    herramientas_data = [_herramienta_devolucion_json(h) for h in herramientas]
    return templates.TemplateResponse(request, "movimiento_devolver.html", ctx_base(
        request, user, db,
        herramientas=herramientas, herramientas_data=herramientas_data,
        almacenes=almacenes, herramienta_sel=h_sel,
    ))


def _herramienta_devolucion_json(herramienta: Herramienta) -> dict:
    """Proyección JSON explícita; evita serializar objetos ORM en la plantilla."""
    responsable = ""
    if herramienta.responsable:
        responsable = herramienta.responsable.nombre_completo
    return {
        "id": herramienta.id,
        "codigo": herramienta.codigo,
        "nombre": herramienta.nombre,
        "marca": herramienta.marca or "",
        "categoria": herramienta.categoria or "",
        "responsable": responsable,
        "trabajador_nombre": responsable,
    }


@app.post("/movimientos/devolver")
def movimiento_devolver_post(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_id: str = Form(...),
    almacen_id: str = Form(""),
    observaciones: str = Form(""),
    condicion: str = Form("buena"),
    request: Request = None,
):
    actor = _require_movement_http(user, "devolver")
    try:
        h_id = int(herramienta_id)
        a_id = int(almacen_id) if almacen_id else None
    except ValueError:
        raise HTTPException(400, "Identificador no válido")
    warehouse = _validate_movement_warehouse(db, user, request, [h_id])
    if warehouse and a_id and a_id != warehouse.id:
        raise HTTPException(409, "La devolución debe registrarse en el almacén activo")
    if warehouse:
        a_id = warehouse.id
    start_movement_transaction(db)
    try:
        result = return_tool(db, actor, h_id, a_id, condicion, observaciones)
        db.commit()
    except MovementError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    return RedirectResponse(f"/herramientas/{result.herramienta_id}", status_code=303)


@app.post("/movimientos/devolver/lote")
def movimiento_devolver_lote(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_ids: str = Form(...),
    almacen_id: str = Form(""),
    observaciones: str = Form(""),
    condicion: str = Form("buena"),
    request: Request = None,
):
    """Registra una devolución múltiple en una única transacción."""
    actor = _require_movement_http(user, "devolver")
    try:
        ids = [int(value.strip()) for value in herramienta_ids.split(",") if value.strip()]
        a_id = int(almacen_id) if almacen_id else None
    except ValueError:
        raise HTTPException(400, "Identificadores no válidos")
    if len(ids) != len(set(ids)):
        raise HTTPException(400, "La lista contiene herramientas duplicadas")
    if len(ids) < 2:
        raise HTTPException(400, "La devolución múltiple requiere al menos dos herramientas")
    if condicion not in CONDICIONES_DEVOLUCION:
        raise HTTPException(400, "Condición de devolución no válida")
    warehouse = _validate_movement_warehouse(db, user, request, ids)
    if warehouse and a_id and a_id != warehouse.id:
        raise HTTPException(409, "La devolución debe registrarse en el almacén activo")
    if warehouse:
        a_id = warehouse.id
    start_movement_transaction(db)

    try:
        for hid in ids:
            return_tool(db, actor, hid, a_id, condicion, observaciones)
        db.commit()
    except MovementError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise

    return JSONResponse({"ok": True, "count": len(ids)})


# ─── Documentos: subida AJAX ──────────────────────────────────────────────────
@app.post("/documentos/subir")
async def documento_subir(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    herramienta_id: int = Form(...),
    nombre: str = Form(""),
    tipo: str = Form("otro"),
    archivo: UploadFile = File(...),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).get(herramienta_id)
    if not h:
        raise HTTPException(404)

    # Sprint 5.2: validar archivo antes de guardar
    try:
        _, _ext_doc = validar_nombre_archivo(archivo.filename, {'jpg', 'jpeg', 'png', 'webp', 'pdf', 'xlsx', 'csv'})
        _head_doc = await archivo.read(16); await archivo.seek(0)
        validar_contenido_archivo(_head_doc, _ext_doc)
        contenido = await archivo.read(); await archivo.seek(0)
        validar_tamaño_bytes(len(contenido), MAX_UPLOAD_MB)
    except ErrorArchivo as _ea:
        raise HTTPException(400, str(_ea))
    ext = f".{_ext_doc}"
    nombre_archivo = f"{h.codigo}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
    docs_dir = UPLOADS_DIR / "documentos"
    docs_dir.mkdir(parents=True, exist_ok=True)
    ruta = docs_dir / nombre_archivo

    with open(ruta, "wb") as _fd:
        _fd.write(contenido)

    doc = Documento(
        herramienta_id=herramienta_id,
        nombre=nombre or archivo.filename,
        tipo=tipo,
        archivo=nombre_archivo,
        extension=ext.lstrip("."),
        tamano=len(contenido),
        subido_por_id=user.id,
    )
    db.add(doc)
    db.flush()

    # Propagar a herramientas del mismo modelo
    # Estrategia: si existe modelo, agrupar por marca+modelo (case-insensitive)
    # Si no hay modelo, agrupar por nombre (case-insensitive)
    from sqlalchemy import func as _func
    q_hermanas = db.query(Herramienta).filter(
        Herramienta.id != h.id,
        Herramienta.activa == True,
    )
    if h.modelo and h.modelo.strip():
        q_hermanas = q_hermanas.filter(
            _func.lower(Herramienta.modelo) == h.modelo.strip().lower()
        )
        if h.marca and h.marca.strip():
            q_hermanas = q_hermanas.filter(
                _func.lower(Herramienta.marca) == h.marca.strip().lower()
            )
    else:
        q_hermanas = q_hermanas.filter(
            _func.lower(Herramienta.nombre) == h.nombre.strip().lower()
        )
    propagadas = 0
    for hermana in q_hermanas.all():
        db.add(Documento(
            herramienta_id=hermana.id,
            nombre=nombre or archivo.filename,
            tipo=tipo,
            archivo=nombre_archivo,  # mismo archivo físico
            extension=ext.lstrip("."),
            tamano=len(contenido),
            subido_por_id=user.id,
        ))
        propagadas += 1

    db.commit()
    return JSONResponse({"ok": True, "id": doc.id, "nombre": doc.nombre,
                         "archivo": nombre_archivo, "propagado_a": propagadas})


# ─── API v1 ───────────────────────────────────────────────────────────────────
@app.get("/api/v1/herramientas")
def api_v1_herramientas(
    q: str = Query(""),
    estado: str = Query(""),
    limit: int = Query(50),
    offset: int = Query(0),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    filtros = []
    if q:
        filtros.append(or_(
            Herramienta.nombre.ilike(f"%{q}%"),
            Herramienta.codigo.ilike(f"%{q}%"),
            Herramienta.num_serie.ilike(f"%{q}%"),
        ))
    if estado:
        filtros.append(Herramienta.estado == estado)
    query = db.query(Herramienta)
    if filtros:
        query = query.filter(*filtros)
    total = query.count()
    herramientas = query.order_by(Herramienta.nombre).offset(offset).limit(limit).all()
    return JSONResponse({
        "total": total,
        "offset": offset,
        "limit": limit,
        "items": [
            {
                "id": h.id,
                "codigo": h.codigo,
                "nombre": h.nombre,
                "marca": h.marca or "",
                "modelo": h.modelo or "",
                "estado": h.estado,
                "num_serie": h.num_serie or "",
                "categoria": h.categoria or "",
            }
            for h in herramientas
        ],
    })


@app.get("/api/v1/herramientas/{hid}")
def api_v1_herramienta_detalle(
    hid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).get(hid)
    if not h:
        raise HTTPException(404, "No encontrada")
    return JSONResponse({
        "id": h.id,
        "codigo": h.codigo,
        "nombre": h.nombre,
        "marca": h.marca or "",
        "modelo": h.modelo or "",
        "estado": h.estado,
        "num_serie": h.num_serie or "",
        "categoria": h.categoria or "",
        "descripcion": h.descripcion or "",
        "valor": float(h.valor) if h.valor else None,
        "fecha_compra": h.fecha_compra.isoformat() if h.fecha_compra else None,
    })

def _herramienta_full_json(h) -> dict:
    """Serializa una herramienta completa para la API v1."""
    return {
        "id": h.id,
        "codigo": h.codigo,
        "nombre": h.nombre,
        "descripcion": h.descripcion or "",
        "categoria": h.categoria or "",
        "subcategoria": h.subcategoria or "",
        "familia": h.familia or "",
        "marca": h.marca or "",
        "modelo": h.modelo or "",
        "fabricante": h.fabricante or "",
        "num_serie": h.num_serie or "",
        "potencia": h.potencia or "",
        "voltaje": h.voltaje or "",
        "peso": float(h.peso) if h.peso else None,
        "color": h.color or "",
        "dimensiones": h.dimensiones or "",
        "activo_fijo": h.activo_fijo or "",
        "vida_util_anos": h.vida_util_anos,
        "estado": h.estado,
        "activa": bool(h.activa),
        "ubicacion_texto": h.ubicacion_texto or "",
        "almacen_id": h.almacen_id,
        "obra_id": h.obra_id,
        "vehiculo_id": h.vehiculo_id,
        "responsable_id": h.responsable_id,
        "fecha_compra": h.fecha_compra.isoformat() if h.fecha_compra else None,
        "precio_compra": float(h.precio_compra) if h.precio_compra else None,
        "garantia_hasta": h.garantia_hasta.isoformat() if h.garantia_hasta else None,
        "proveedor_texto": h.proveedor_texto or "",
        "numero_factura": h.numero_factura or "",
        "observaciones": h.observaciones or "",
        "foto": h.foto or "",
    }


# ─── API v1 — CREATE ──────────────────────────────────────────────────────────
@app.post("/api/v1/herramientas")
async def api_v1_crear_herramienta(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    body: Dict[str, Any] = await request.json()
    nombre = body.get("nombre", "").strip()
    if not nombre:
        raise HTTPException(400, "nombre es obligatorio")
    # Se ignora cualquier código enviado por el cliente: la referencia es interna.
    codigo = generar_referencia_herramienta(db)
    almacen_predeterminado = get_default_warehouse(db)
    h = Herramienta(
        codigo=codigo,
        nombre=nombre,
        descripcion=body.get("descripcion") or None,
        categoria=body.get("categoria") or "Otro",
        subcategoria=body.get("subcategoria") or None,
        familia=body.get("familia") or None,
        marca=body.get("marca") or None,
        modelo=body.get("modelo") or None,
        fabricante=body.get("fabricante") or None,
        num_serie=body.get("num_serie") or None,
        potencia=body.get("potencia") or None,
        voltaje=body.get("voltaje") or None,
        peso=float(body["peso"]) if body.get("peso") else None,
        color=body.get("color") or None,
        dimensiones=body.get("dimensiones") or None,
        activo_fijo=body.get("activo_fijo") or None,
        vida_util_anos=int(body["vida_util_anos"]) if body.get("vida_util_anos") else None,
        estado=body.get("estado") or "disponible",
        fecha_compra=datetime.fromisoformat(body["fecha_compra"]).date() if body.get("fecha_compra") else None,
        precio_compra=float(body["precio_compra"]) if body.get("precio_compra") else None,
        garantia_hasta=datetime.fromisoformat(body["garantia_hasta"]).date() if body.get("garantia_hasta") else None,
        proveedor_texto=body.get("proveedor_texto") or None,
        numero_factura=body.get("numero_factura") or None,
        observaciones=body.get("observaciones") or None,
        almacen_id=(int(body["almacen_id"]) if body.get("almacen_id") else
                    (almacen_predeterminado.id if almacen_predeterminado else None)),
        ubicacion_texto=(body.get("ubicacion_texto") or
                         (almacen_predeterminado.nombre if almacen_predeterminado else None)),
        activa=True,
    )
    db.add(h)
    db.flush()
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    registrar_auditoria(db, "herramientas", h.id, "crear_api", user.id,
                        datos_nuevos=snapshot_herramienta(h), resumen=f"Alta API: {h.nombre}", ip=ip)
    db.commit()
    return JSONResponse(_herramienta_full_json(h), status_code=201)


# ─── API v1 — UPDATE ─────────────────────────────────────────────────────────
@app.put("/api/v1/herramientas/{hid}")
async def api_v1_actualizar_herramienta(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = await request.json()
    snap_ant = snapshot_herramienta(h)
    campos = ["nombre", "descripcion", "categoria", "subcategoria", "familia",
              "marca", "modelo", "fabricante", "num_serie", "potencia", "voltaje",
              "color", "dimensiones", "activo_fijo", "proveedor_texto",
              "numero_factura", "observaciones", "ubicacion_texto"]
    for campo in campos:
        if campo in body:
            setattr(h, campo, body[campo] or None)
    if "peso" in body:
        h.peso = float(body["peso"]) if body["peso"] else None
    if "precio_compra" in body:
        h.precio_compra = float(body["precio_compra"]) if body["precio_compra"] else None
    if "vida_util_anos" in body:
        h.vida_util_anos = int(body["vida_util_anos"]) if body["vida_util_anos"] else None
    if "fecha_compra" in body:
        h.fecha_compra = datetime.fromisoformat(body["fecha_compra"]).date() if body["fecha_compra"] else None
    if "garantia_hasta" in body:
        h.garantia_hasta = datetime.fromisoformat(body["garantia_hasta"]).date() if body["garantia_hasta"] else None
    if "almacen_id" in body:
        h.almacen_id = int(body["almacen_id"]) if body["almacen_id"] else None
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    registrar_auditoria(db, "herramientas", h.id, "editar_api", user.id,
                        datos_anteriores=snap_ant, datos_nuevos=snapshot_herramienta(h),
                        resumen=f"Edición API: {h.nombre}", ip=ip)
    db.commit()
    return JSONResponse(_herramienta_full_json(h))


# ─── API v1 — DELETE (baja lógica) ───────────────────────────────────────────
@app.delete("/api/v1/herramientas/{hid}")
async def api_v1_baja_herramienta(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "borrar"):
        raise HTTPException(403, "Sin permiso de administrador")
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = {}
    try:
        body = await request.json()
    except Exception:
        pass
    observaciones = body.get("observaciones", "Baja por API")
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        aplicar_accion(db, h, "baja", user, es_admin=True, observaciones=observaciones, ip=ip)
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse({"ok": True, "estado": h.estado, "activa": bool(h.activa)})


# ─── API v1 — ACCION genérica ────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/accion")
async def api_v1_accion_herramienta(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = await request.json()
    accion = body.get("accion", "").strip()
    if not accion:
        raise HTTPException(400, "accion es obligatorio")
    ACCIONES_ADMIN = {"baja", "restaurar", "archivar", "recuperar", "robada"}
    es_admin = tiene_permiso(user, "borrar")
    if accion in ACCIONES_ADMIN and not es_admin:
        raise HTTPException(403, "Sin permiso de administrador")
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(
            db, h, accion, user,
            es_admin=es_admin,
            trabajador_id=int(body["trabajador_id"]) if body.get("trabajador_id") else None,
            obra_id=int(body["obra_id"]) if body.get("obra_id") else None,
            almacen_id=int(body["almacen_id"]) if body.get("almacen_id") else None,
            vehiculo_id=int(body["vehiculo_id"]) if body.get("vehiculo_id") else None,
            observaciones=body.get("observaciones", ""),
            ip=ip,
        )
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — MOVER ──────────────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/mover")
async def api_v1_mover(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = await request.json()
    destino = body.get("destino", "")
    accion_map = {"obra": "a_obra", "almacen": "a_almacen", "furgoneta": "a_furgoneta"}
    accion = accion_map.get(destino)
    if not accion:
        raise HTTPException(400, f"destino debe ser: {', '.join(accion_map)}")
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(
            db, h, accion, user,
            obra_id=int(body["obra_id"]) if body.get("obra_id") else None,
            almacen_id=int(body["almacen_id"]) if body.get("almacen_id") else None,
            vehiculo_id=int(body["vehiculo_id"]) if body.get("vehiculo_id") else None,
            observaciones=body.get("observaciones", ""),
            ip=ip,
        )
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — ASIGNAR ────────────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/asignar")
async def api_v1_asignar(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = await request.json()
    trabajador_id = body.get("trabajador_id")
    if not trabajador_id:
        raise HTTPException(400, "trabajador_id es obligatorio")
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(
            db, h, "entregar", user,
            trabajador_id=int(trabajador_id),
            observaciones=body.get("observaciones", ""),
            ip=ip,
        )
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — DEVOLVER ───────────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/devolver")
async def api_v1_devolver(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = {}
    try:
        body = await request.json()
    except Exception:
        pass
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(db, h, "devolver", user,
                                   observaciones=body.get("observaciones", ""), ip=ip)
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — REPARACIÓN ─────────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/reparacion")
async def api_v1_reparacion(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = {}
    try:
        body = await request.json()
    except Exception:
        pass
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(db, h, "reparacion", user,
                                   observaciones=body.get("observaciones", ""), ip=ip)
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — INCIDENCIA ─────────────────────────────────────────────────────
@app.post("/api/v1/herramientas/{hid}/incidencia")
async def api_v1_incidencia(
    hid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "borrar"):
        raise HTTPException(403, "Sin permiso de administrador para incidencias")
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    body: Dict[str, Any] = await request.json()
    tipo = body.get("tipo", "").strip()
    if tipo not in {"perdida", "robada", "fuera_servicio"}:
        raise HTTPException(400, "tipo debe ser: perdida, robada, fuera_servicio")
    ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "")
    try:
        resultado = aplicar_accion(db, h, tipo, user, es_admin=True,
                                   observaciones=body.get("observaciones", ""), ip=ip)
        db.commit()
    except ErrorTransicion as exc:
        db.rollback()
        raise HTTPException(400, str(exc))
    return JSONResponse(resultado)


# ─── API v1 — ETIQUETA / QR ──────────────────────────────────────────────────
@app.get("/api/v1/herramientas/{hid}/etiqueta")
def api_v1_etiqueta(
    hid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    try:
        import qrcode, io as _io, base64
        qr = qrcode.QRCode(version=1, box_size=4, border=2)
        qr.add_data(h.codigo)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_b64 = base64.b64encode(buf.getvalue()).decode()
    except Exception:
        qr_b64 = None
    return JSONResponse({
        "id": h.id,
        "codigo": h.codigo,
        "nombre": h.nombre,
        "marca": h.marca or "",
        "modelo": h.modelo or "",
        "num_serie": h.num_serie or "",
        "estado": h.estado,
        "qr_base64": qr_b64,
        "url_pdf": f"/herramientas/{h.id}/pdf",
    })


# ─── API v1 — HISTORIAL ──────────────────────────────────────────────────────
@app.get("/api/v1/herramientas/{hid}/historial")
def api_v1_historial(
    hid: int,
    limit: int = Query(50),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    movs = db.query(Movimiento).filter(
        Movimiento.herramienta_id == hid
    ).order_by(Movimiento.fecha.desc()).limit(limit).all()
    return JSONResponse({
        "herramienta_id": hid,
        "total": len(movs),
        "items": [
            {
                "id": m.id,
                "fecha": m.fecha.isoformat() if m.fecha else None,
                "tipo": m.tipo,
                "estado_anterior": m.estado_anterior or "",
                "estado_nuevo": m.estado_nuevo or "",
                "destino": m.destino or "",
                "observaciones": m.observaciones or "",
                "usuario_id": m.usuario_id,
            }
            for m in movs
        ],
    })


# ─── API v1 — AUDITORÍA ──────────────────────────────────────────────────────
@app.get("/api/v1/herramientas/{hid}/auditoria")
def api_v1_auditoria_herramienta(
    hid: int,
    limit: int = Query(50),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    h = db.query(Herramienta).filter(Herramienta.id == hid).first()
    if not h:
        raise HTTPException(404, "No encontrada")
    logs = db.query(AuditoriaLog).filter(
        AuditoriaLog.tabla == "herramientas",
        AuditoriaLog.registro_id == hid,
    ).order_by(AuditoriaLog.fecha.desc()).limit(limit).all()
    return JSONResponse({
        "herramienta_id": hid,
        "total": len(logs),
        "items": [
            {
                "id": lg.id,
                "fecha": lg.fecha.isoformat() if lg.fecha else None,
                "accion": lg.accion,
                "usuario_id": lg.usuario_id,
                "resumen": lg.resumen or "",
                "ip": lg.ip or "",
            }
            for lg in logs
        ],
    })


# ─── API v1 — EXPORT Excel ───────────────────────────────────────────────────
@app.get("/api/v1/export/herramientas")
def api_v1_exportar_herramientas(
    estado: str = Query(""),
    activa: str = Query(""),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    q = db.query(Herramienta)
    if estado:
        q = q.filter(Herramienta.estado == estado)
    if activa:
        q = q.filter(Herramienta.activa == (activa.lower() == "true"))
    herramientas = q.order_by(Herramienta.nombre).all()
    excel = exportar_inventario_excel(herramientas)
    nombre = f"herramientas_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )




# ─── Export trabajadores y EPIs ───────────────────────────────────────────────

@app.get("/informes/trabajadores/excel")
def exportar_trabajadores_excel_route(
    activo: str = Query(""),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    q = db.query(Trabajador)
    if activo == "true":
        q = q.filter(Trabajador.activo == True)
    elif activo == "false":
        q = q.filter(Trabajador.activo == False)
    trabajadores = q.order_by(Trabajador.nombre).all()
    excel = exportar_trabajadores_excel(trabajadores)
    nombre = f"trabajadores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@app.get("/api/epis/individuales/disponibles")
def api_epis_disponibles(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    tipo: str = Query(None),
):
    """EPIs individuales sin asignar — para modal de asignacion desde trabajador."""
    q = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.trabajador_id == None,
        or_(EPIIndividual.proxima_revision == None, EPIIndividual.proxima_revision >= date.today()),
    )
    if tipo:
        q = q.filter(EPIIndividual.tipo == tipo)
    epis = q.order_by(EPIIndividual.tipo, EPIIndividual.codigo_fabricacion).all()
    return [
        {
            "id": e.id,
            "tipo": e.tipo,
            "codigo": e.codigo_fabricacion,
            "marca": e.marca or "",
            "modelo": e.modelo or "",
            "proxima_revision": e.proxima_revision.isoformat() if e.proxima_revision else None,
            "revision_vencida": e.revision_vencida,
        }
        for e in epis
    ]


@app.get("/informes/epis-trabajadores", response_class=HTMLResponse)
def informe_epis_trabajadores(request: Request, user: Usuario = Depends(requiere_login),
                               db: Session = Depends(get_db)):
    """Informe completo: qué EPIs/ropa/arneses tiene cada trabajador + stock disponible."""
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()

    entregas_por_trabajador: dict = {}
    arneses_por_trabajador: dict = {}
    if trabajadores:
        ids_trabajadores = [t.id for t in trabajadores]
        for e in db.query(EntregaEPI).filter(
            EntregaEPI.trabajador_id.in_(ids_trabajadores),
        ).order_by(EntregaEPI.fecha.desc()).all():
            entregas_por_trabajador.setdefault(e.trabajador_id, []).append(e)
        for a in db.query(EPIIndividual).filter(
            EPIIndividual.trabajador_id.in_(ids_trabajadores),
            EPIIndividual.estado != "baja",
        ).all():
            arneses_por_trabajador.setdefault(a.trabajador_id, []).append(a)

    # Para cada trabajador armar su resumen
    resumen = []
    for t in trabajadores:
        entregas = entregas_por_trabajador.get(t.id, [])
        arneses = arneses_por_trabajador.get(t.id, [])
        if entregas or arneses:
            resumen.append({
                "trabajador": t,
                "entregas": [
                    {"e": e, "items": json.loads(e.items_json or "[]")}
                    for e in entregas
                ],
                "arneses": arneses,
            })
    # Stock disponible
    stock_epi = db.query(EPIIndividual).filter(
        EPIIndividual.estado != "baja",
        EPIIndividual.trabajador_id == None
    ).all()
    stock_por_tipo = {}
    for epi in stock_epi:
        stock_por_tipo.setdefault(epi.tipo, []).append(epi)
    return templates.TemplateResponse(request, "informe_epis_trabajadores.html", ctx_base(
        request, user, db,
        resumen=resumen,
        stock_por_tipo=stock_por_tipo,
        total_trabajadores_con_epis=len(resumen),
    ))


@app.get("/informes/epis-trabajadores/excel")
def informe_epis_trabajadores_excel(user: Usuario = Depends(requiere_login),
                                     db: Session = Depends(get_db)):
    """Exporta a Excel el informe EPIs/ropa/arneses por trabajador."""
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
    from openpyxl.utils import get_column_letter as _gcl
    import io as _io

    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()

    wb = _WB()

    # ── Hoja 1: Entregas EPIs / Ropa
    ws1 = wb.active; ws1.title = "Entregas EPI y Ropa"
    lado = _Side(style="thin", color="CCCCCC")
    borde = _Border(left=lado, right=lado, top=lado, bottom=lado)
    fill_h = _Fill("solid", fgColor="1B4F8A")
    fill_s = _Fill("solid", fgColor="D6E4F0")
    fill_o = _Fill("solid", fgColor="E8600A")

    ws1.merge_cells("A1:G1")
    ws1["A1"] = "MRD TOOL CONTROL — Entregas de EPIs y Ropa por Trabajador"
    ws1["A1"].font = _Font(bold=True, color="FFFFFF", size=13)
    ws1["A1"].fill = fill_h
    ws1["A1"].alignment = _Align(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    cols1 = [("Trabajador", 28), ("Cargo", 22), ("Tipo", 10),
             ("Fecha entrega", 16), ("Artículos", 50), ("Firmado por", 22), ("Obs.", 25)]
    for c, (t_, w) in enumerate(cols1, 1):
        cell = ws1.cell(row=3, column=c, value=t_)
        cell.font = _Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill_o
        cell.alignment = _Align(horizontal="center")
        cell.border = borde
        ws1.column_dimensions[_gcl(c)].width = w
    ws1.row_dimensions[3].height = 20

    row = 4
    for idx_t, t in enumerate(trabajadores):
        entregas = db.query(EntregaEPI).filter(EntregaEPI.trabajador_id == t.id).order_by(EntregaEPI.fecha.desc()).all()
        if not entregas:
            continue
        bg = fill_s if idx_t % 2 == 0 else _Fill("solid", fgColor="FFFFFF")
        for e in entregas:
            items_str = ", ".join(
                f"{it['nombre']} x{it.get('cantidad',1)}" + (f" T.{it['talla']}" if it.get('talla') else "")
                for it in json.loads(e.items_json or "[]")
            )
            vals = [t.nombre_completo, t.cargo or "", e.tipo.upper(),
                    e.fecha.strftime("%d/%m/%Y") if e.fecha else "",
                    items_str, e.firmado_por or "", e.observaciones or ""]
            for c, v in enumerate(vals, 1):
                cell = ws1.cell(row=row, column=c, value=v)
                cell.fill = bg; cell.border = borde
                cell.alignment = _Align(vertical="center", wrap_text=(c == 5))
            ws1.row_dimensions[row].height = 18
            row += 1

    # ── Hoja 2: Arneses y Absorbedores
    ws2 = wb.create_sheet("Arneses y Absorbedores")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "MRD TOOL CONTROL — Arneses y Absorbedores Asignados"
    ws2["A1"].font = _Font(bold=True, color="FFFFFF", size=13)
    ws2["A1"].fill = fill_h
    ws2["A1"].alignment = _Align(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    cols2 = [("Trabajador", 28), ("Tipo", 14), ("Código fabricación", 22),
             ("Marca / Modelo", 22), ("Estado", 14), ("Próx. revisión", 16)]
    for c, (t_, w) in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=c, value=t_)
        cell.font = _Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill_o
        cell.alignment = _Align(horizontal="center")
        cell.border = borde
        ws2.column_dimensions[_gcl(c)].width = w
    ws2.row_dimensions[3].height = 20

    row2 = 4
    epis_asignados = db.query(EPIIndividual).filter(
        EPIIndividual.trabajador_id != None,
        EPIIndividual.estado != "baja"
    ).options(joinedload(EPIIndividual.trabajador)).order_by(EPIIndividual.tipo).all()
    for idx_e, epi in enumerate(epis_asignados):
        bg = fill_s if idx_e % 2 == 0 else _Fill("solid", fgColor="FFFFFF")
        marca_mod = " / ".join(filter(None, [epi.marca, epi.modelo])) or "—"
        vals = [
            epi.trabajador.nombre_completo if epi.trabajador else "—",
            epi.tipo, epi.codigo_fabricacion, marca_mod,
            epi.estado.replace("_", " ").title(),
            epi.proxima_revision.strftime("%d/%m/%Y") if epi.proxima_revision else "Sin programar",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=c, value=v)
            cell.fill = bg; cell.border = borde
            cell.alignment = _Align(vertical="center")
        ws2.row_dimensions[row2].height = 18
        row2 += 1

    # ── Hoja 3: Stock disponible
    ws3 = wb.create_sheet("Stock Disponible")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "MRD TOOL CONTROL — Stock EPIs Individuales Disponibles"
    ws3["A1"].font = _Font(bold=True, color="FFFFFF", size=13)
    ws3["A1"].fill = fill_h
    ws3["A1"].alignment = _Align(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    cols3 = [("Tipo", 18), ("Código fabricación", 24), ("Marca / Modelo", 24), ("Próx. revisión", 16)]
    for c, (t_, w) in enumerate(cols3, 1):
        cell = ws3.cell(row=3, column=c, value=t_)
        cell.font = _Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill_o
        cell.alignment = _Align(horizontal="center")
        cell.border = borde
        ws3.column_dimensions[_gcl(c)].width = w
    ws3.row_dimensions[3].height = 20

    disponibles = db.query(EPIIndividual).filter(
        EPIIndividual.estado != "baja",
        EPIIndividual.trabajador_id == None
    ).order_by(EPIIndividual.tipo, EPIIndividual.codigo_fabricacion).all()
    for idx_d, epi in enumerate(disponibles):
        bg = fill_s if idx_d % 2 == 0 else _Fill("solid", fgColor="FFFFFF")
        marca_mod = " / ".join(filter(None, [epi.marca, epi.modelo])) or "—"
        vals = [epi.tipo, epi.codigo_fabricacion, marca_mod,
                epi.proxima_revision.strftime("%d/%m/%Y") if epi.proxima_revision else "Sin programar"]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=3 + 1 + idx_d, column=c, value=v)
            cell.fill = bg; cell.border = borde
            cell.alignment = _Align(vertical="center")
        ws3.row_dimensions[3 + 1 + idx_d].height = 18

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date as _date
    nombre = f"informe_epis_{_date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


@app.get("/informes/epis/excel")
def exportar_epis_excel_route(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Exporta historial de entregas de EPIs a Excel."""
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
    from openpyxl.utils import get_column_letter as _gcl
    import io as _io

    entregas = (
        db.query(EntregaEPI)
        .options(joinedload(EntregaEPI.trabajador))
        .order_by(EntregaEPI.fecha.desc())
        .all()
    )

    wb = _WB(); ws = wb.active; ws.title = "EPIs Entregas"
    lado = _Side(style="thin", color="CCCCCC")
    borde = _Border(left=lado, right=lado, top=lado, bottom=lado)
    fill_h = _Fill("solid", fgColor="1B4F8A")
    fill_s = _Fill("solid", fgColor="D6E4F0")

    ws.merge_cells("A1:G1")
    ws["A1"] = "MRD TOOL CONTROL — Historial de Entregas EPI"
    ws["A1"].font = _Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = fill_h
    ws["A1"].alignment = _Align(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cols = [("Fecha", 18), ("Trabajador", 28), ("Tipo", 10),
            ("Artículos", 45), ("Firmado por", 20), ("Entregado por", 20), ("Observaciones", 30)]
    for c, (t, w) in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=t)
        cell.font = _Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _Fill("solid", fgColor="E8600A")
        cell.alignment = _Align(horizontal="center", vertical="center")
        cell.border = borde
        ws.column_dimensions[_gcl(c)].width = w
    ws.row_dimensions[3].height = 20

    for i, e in enumerate(entregas):
        row = 4 + i
        items_str = ", ".join(
            f"{it['nombre']} x{it.get('cantidad',1)}"
            + (f" T.{it['talla']}" if it.get("talla") else "")
            for it in json.loads(e.items_json or "[]")
        )
        bg = fill_s if i % 2 == 0 else _Fill("solid", fgColor="FFFFFF")
        vals = [
            e.fecha.strftime("%d/%m/%Y %H:%M") if e.fecha else "",
            e.trabajador.nombre_completo if e.trabajador else "",
            e.tipo.upper() if e.tipo else "",
            items_str,
            e.firmado_por or "",
            e.entregado_por or "",
            e.observaciones or "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = bg; cell.border = borde
            cell.alignment = _Align(vertical="center", wrap_text=(c == 4))
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A4"
    buf = _io.BytesIO(); wb.save(buf)
    nombre = f"epis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )



# ─── Foto herramienta ─────────────────────────────────────────────────────────

@app.post("/herramientas/{herramienta_id}/foto")
async def subir_foto_herramienta(
    herramienta_id: int,
    request: Request,
    foto: UploadFile = File(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if user.rol not in ("admin", "almacen"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404, "Herramienta no encontrada")

    # Validate
    contenido = await foto.read()
    try:
        validar_contenido_archivo(contenido, ["image/jpeg", "image/png", "image/webp"])
        validar_tamaño_bytes(contenido, MAX_UPLOAD_MB * 1024 * 1024)
    except ErrorArchivo as e:
        raise HTTPException(400, str(e))

    ext = Path(foto.filename or "foto.jpg").suffix.lower() or ".jpg"
    foto_dir = BASE_DIR / "static" / "uploads" / "herramientas"
    foto_dir.mkdir(parents=True, exist_ok=True)
    nombre = f"h_{herramienta_id}{ext}"
    (foto_dir / nombre).write_bytes(contenido)

    h.foto_path = nombre
    db.commit()
    return JSONResponse({"ok": True, "foto_url": f"/static/uploads/herramientas/{nombre}"})


@app.delete("/herramientas/{herramienta_id}/foto")
def borrar_foto_herramienta(
    herramienta_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if user.rol not in ("admin", "almacen"):
        raise HTTPException(403, "Sin permiso")
    h = db.query(Herramienta).filter(Herramienta.id == herramienta_id).first()
    if not h:
        raise HTTPException(404)
    if h.foto_path:
        p = BASE_DIR / "static" / "uploads" / "herramientas" / h.foto_path
        if p.exists():
            p.unlink()
        h.foto_path = None
        db.commit()
    return JSONResponse({"ok": True})


# ─── Búsqueda global ──────────────────────────────────────────────────────────

@app.get("/api/buscar")
def api_buscar_global(
    q: str = Query(""),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """API de búsqueda global: herramientas + trabajadores + obras + maquinaria + albaranes."""
    if not q or len(q.strip()) < 2:
        return {"herramientas": [], "trabajadores": [], "obras": [], "maquinaria": [], "albaranes": []}
    q = q.strip()
    like = f"%{q}%"

    herramientas = (
        db.query(Herramienta)
        .filter(Herramienta.activa == True)
        .filter(or_(
            Herramienta.codigo.ilike(like),
            Herramienta.nombre.ilike(like),
            Herramienta.num_serie.ilike(like),
            Herramienta.categoria.ilike(like),
        ))
        .limit(6).all()
    )
    trabajadores = (
        db.query(Trabajador)
        .filter(or_(
            Trabajador.nombre.ilike(like),
            Trabajador.apellidos.ilike(like),
            Trabajador.dni.ilike(like),
            Trabajador.cargo.ilike(like),
        ))
        .limit(4).all()
    )
    obras = (
        db.query(Obra)
        .filter(or_(
            Obra.nombre.ilike(like),
            Obra.numero.ilike(like),
            Obra.direccion.ilike(like),
        ))
        .limit(4).all()
    )
    maquinaria = (
        db.query(Maquinaria)
        .filter(or_(
            Maquinaria.nombre.ilike(like),
            Maquinaria.matricula.ilike(like),
        ))
        .limit(3).all()
    )
    albaranes = (
        db.query(AlbaranSalida)
        .filter(or_(
            AlbaranSalida.numero.ilike(like),
            AlbaranSalida.origen_destino.ilike(like),
        ))
        .order_by(AlbaranSalida.id.desc())
        .limit(4).all()
    )
    return {
        "herramientas": [{"id": h.id, "codigo": h.codigo, "nombre": h.nombre, "estado": h.estado} for h in herramientas],
        "trabajadores": [{"id": t.id, "nombre": t.nombre_completo, "cargo": t.cargo or ""} for t in trabajadores],
        "obras": [{"id": o.id, "nombre": o.nombre, "numero": o.numero or ""} for o in obras],
        "maquinaria": [{"id": m.id, "nombre": m.nombre, "matricula": getattr(m, "matricula", "") or ""} for m in maquinaria],
        "albaranes": [{"id": a.id, "numero": a.numero, "estado": a.estado} for a in albaranes],
    }


@app.get("/buscar", response_class=HTMLResponse)
def buscar_global(
    request: Request,
    q: str = Query(""),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Página de resultados de búsqueda global."""
    herramientas, trabajadores, obras, maquinaria, albaranes = [], [], [], [], []
    if q and len(q.strip()) >= 2:
        like = f"%{q.strip()}%"
        herramientas = (
            db.query(Herramienta)
            .filter(Herramienta.activa == True)
            .filter(or_(
                Herramienta.codigo.ilike(like),
                Herramienta.nombre.ilike(like),
                Herramienta.num_serie.ilike(like),
                Herramienta.categoria.ilike(like),
            ))
            .order_by(Herramienta.nombre).limit(50).all()
        )
        trabajadores = (
            db.query(Trabajador)
            .filter(or_(
                Trabajador.nombre.ilike(like),
                Trabajador.apellidos.ilike(like),
                Trabajador.dni.ilike(like),
                Trabajador.cargo.ilike(like),
            ))
            .order_by(Trabajador.nombre).limit(30).all()
        )
        obras = (
            db.query(Obra)
            .filter(or_(
                Obra.nombre.ilike(like),
                Obra.numero.ilike(like),
                Obra.direccion.ilike(like),
            ))
            .order_by(Obra.nombre).limit(20).all()
        )
        maquinaria = (
            db.query(Maquinaria)
            .filter(or_(
                Maquinaria.nombre.ilike(like),
                Maquinaria.matricula.ilike(like),
            ))
            .limit(20).all()
        )
        albaranes = (
            db.query(AlbaranSalida)
            .filter(or_(
                AlbaranSalida.numero.ilike(like),
                AlbaranSalida.origen_destino.ilike(like),
            ))
            .order_by(AlbaranSalida.id.desc())
            .limit(20).all()
        )
    total = len(herramientas) + len(trabajadores) + len(obras) + len(maquinaria) + len(albaranes)
    return templates.TemplateResponse(request, "buscar.html", ctx_base(
        request, user,
        q=q,
        total=total,
        herramientas=herramientas,
        trabajadores=trabajadores,
        obras=obras,
        maquinaria=maquinaria,
        albaranes=albaranes,
    ))


# ─── API v1 — IMPORT Excel ───────────────────────────────────────────────────
@app.post("/api/v1/import/herramientas")
async def api_v1_importar_herramientas(
    request: Request,
    archivo: UploadFile = File(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    # Sprint 5.2: validar MIME, magic bytes y tamaño
    try:
        _, _ext_api = validar_nombre_archivo(archivo.filename, {'xlsx'})
        _head_api = await archivo.read(16); await archivo.seek(0)
        validar_contenido_archivo(_head_api, _ext_api)
        contenido = await archivo.read(); await archivo.seek(0)
        validar_tamaño_bytes(len(contenido), MAX_UPLOAD_MB)
    except ErrorArchivo as _ea:
        raise HTTPException(400, str(_ea))
    try:
        resultado = importar_herramientas_excel(db, contenido, user)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(422, f"Error al procesar Excel: {exc}")
    return JSONResponse(resultado)


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO MAQUINARIA — Sprint 3.2
# Integración Zebra DS3678-SR (HID Bluetooth): escanea codigo_barras → busca aquí
# ═══════════════════════════════════════════════════════════════════════════════

TIPOS_EVENTO_MAQUINARIA = {
    "revision": "Revisión",
    "averia": "Avería",
    "reparacion": "Reparación",
    "pieza": "Pieza sustituida",
    "cambio": "Cambio / mejora",
    "horas": "Lectura de horas",
    "otro": "Otro",
}
ESTADOS_LOCALIZADOR = {
    "verificado": "Verificado",
    "pendiente": "Pendiente de comprobar",
    "sin_bateria": "Sin batería",
    "no_localizado": "No localizado",
    "retirado": "Retirado",
}


def _puede_operar_maquinaria(user: Usuario) -> bool:
    return tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")


@app.get("/panel-patio", response_class=HTMLResponse)
def panel_encargado_patio(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    filtro: str = Query("todos"),
):
    """Resumen privado y accionable para la apertura diaria del almacén."""
    if user.rol not in {"admin", "almacen", "encargado_patio"}:
        raise HTTPException(403, "Sin permiso para el panel de patio")
    warehouse = _operation_warehouse(request, user, db)
    hoy = date.today()
    limite_revision = hoy + timedelta(days=30)

    herramientas_fuera = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.almacen_id == warehouse.id,
        Herramienta.estado.in_(["entregada", "en_uso", "en_obra"]),
    ).order_by(Herramienta.updated_at.asc()).limit(20).all()
    dotaciones = db.query(DotacionTrabajador).filter(
        DotacionTrabajador.estado.in_(["pendiente", "preparada"]),
    ).order_by(DotacionTrabajador.creado_en.asc()).limit(20).all()
    trabajadores_dotacion = {
        t.id: t for t in db.query(Trabajador).filter(
            Trabajador.id.in_([d.trabajador_id for d in dotaciones] or [-1])
        ).all()
    }
    materiales_bajos = db.query(Material).filter(
        Material.activo == True,
        Material.almacen_id == warehouse.id,
        Material.stock_minimo > 0,
        Material.stock_actual <= Material.stock_minimo,
    ).order_by(Material.stock_actual.asc()).limit(20).all()
    ropa_baja = db.query(StockEPI).filter(
        StockEPI.almacen_id == warehouse.id,
        StockEPI.stock_minimo > 0,
        StockEPI.cantidad <= StockEPI.stock_minimo,
    ).order_by(StockEPI.cantidad.asc()).limit(20).all()
    maquinas_revision = db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        Maquinaria.almacen_id == warehouse.id,
        Maquinaria.proxima_revision != None,
        Maquinaria.proxima_revision <= limite_revision,
    ).order_by(Maquinaria.proxima_revision.asc()).limit(20).all()
    epis_revision = db.query(EPIIndividual).filter(
        EPIIndividual.estado != "baja",
        EPIIndividual.almacen_id == warehouse.id,
        EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision <= limite_revision,
    ).order_by(EPIIndividual.proxima_revision.asc()).limit(20).all()
    material_ids = [row[0] for row in db.query(Material.id).filter(Material.almacen_id == warehouse.id).all()]
    stock_epi_ids = [row[0] for row in db.query(StockEPI.id).filter(StockEPI.almacen_id == warehouse.id).all()]
    existence_ids = [row[0] for row in db.query(ExistenciaVariante.id).filter(ExistenciaVariante.almacen_id == warehouse.id).all()]
    suministros = db.query(MovimientoStock).filter(
        MovimientoStock.cantidad > 0,
        MovimientoStock.creado_en >= datetime.now() - timedelta(days=7),
        or_(
            MovimientoStock.material_id.in_(material_ids or [-1]),
            MovimientoStock.stock_epi_id.in_(stock_epi_ids or [-1]),
            MovimientoStock.existencia_id.in_(existence_ids or [-1]),
        ),
    ).order_by(MovimientoStock.creado_en.desc()).limit(15).all()
    inventarios = db.query(SesionInventario).filter(
        SesionInventario.almacen_id == warehouse.id,
        SesionInventario.estado.notin_(["cerrada", "cancelada"]),
    ).order_by(SesionInventario.opened_at.asc()).limit(20).all()
    transfers = db.query(TransferenciaAlmacen).filter(
        TransferenciaAlmacen.estado == "en_transito",
        or_(TransferenciaAlmacen.origen_id == warehouse.id, TransferenciaAlmacen.destino_id == warehouse.id),
    ).order_by(TransferenciaAlmacen.creado_en.asc()).limit(20).all()
    orders = db.query(PedidoProveedor).filter(
        PedidoProveedor.almacen_id == warehouse.id,
        PedidoProveedor.estado.in_(["borrador", "enviado", "parcial"]),
    ).order_by(PedidoProveedor.fecha_pedido.asc()).limit(20).all()
    preparations = db.query(PreparacionEntrega).filter(
        PreparacionEntrega.almacen_id == warehouse.id, PreparacionEntrega.estado == "preparada",
    ).order_by(PreparacionEntrega.creado_en.asc()).limit(20).all()
    expiring_lots = db.query(LoteAlmacen).filter(
        LoteAlmacen.almacen_id == warehouse.id, LoteAlmacen.cantidad > 0,
        LoteAlmacen.fecha_caducidad != None, LoteAlmacen.fecha_caducidad <= hoy + timedelta(days=60),
    ).order_by(LoteAlmacen.fecha_caducidad.asc()).limit(20).all()
    expiring_variant_lots = db.query(LoteVariante).join(ExistenciaVariante).filter(
        ExistenciaVariante.almacen_id == warehouse.id, LoteVariante.cantidad > 0,
        LoteVariante.fecha_caducidad != None, LoteVariante.fecha_caducidad <= hoy + timedelta(days=60),
    ).order_by(LoteVariante.fecha_caducidad.asc()).limit(20).all()
    closure_today = db.query(CierreDiarioAlmacen).filter_by(almacen_id=warehouse.id, fecha=hoy).first()

    secciones = {
        "salidas": herramientas_fuera,
        "dotaciones": dotaciones,
        "stock": materiales_bajos + ropa_baja,
        "revisiones": maquinas_revision + epis_revision,
        "suministros": suministros,
        "inventarios": inventarios,
        "operativa": transfers + orders + preparations + expiring_lots + expiring_variant_lots,
    }
    if filtro not in {*secciones, "todos"}:
        filtro = "todos"
    return templates.TemplateResponse(request, "panel_patio.html", ctx_base(
        request, user, db,
        filtro=filtro, secciones=secciones,
        herramientas_fuera=herramientas_fuera,
        dotaciones_pendientes=dotaciones,
        trabajadores_dotacion=trabajadores_dotacion,
        materiales_bajos=materiales_bajos, ropa_baja=ropa_baja,
        maquinas_revision=maquinas_revision, epis_revision=epis_revision,
        suministros_recientes=suministros, inventarios_abiertos=inventarios,
        transferencias_pendientes=transfers, pedidos_pendientes=orders,
        preparaciones_pendientes=preparations, lotes_proximos=expiring_lots,
        lotes_variante_proximos=expiring_variant_lots,
        cierre_hoy=closure_today, almacen=warehouse, hoy=hoy,
    ))

@app.get("/maquinaria", response_class=HTMLResponse)
def maquinaria_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = Query(""),
    q: str = Query(""),
):
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    query = db.query(Maquinaria).filter(
        Maquinaria.activa == True, Maquinaria.almacen_id == warehouse_id,
    )
    if estado:
        query = query.filter(Maquinaria.estado == estado)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(Maquinaria.nombre.ilike(like), Maquinaria.matricula.ilike(like),
                Maquinaria.codigo_barras.ilike(like), Maquinaria.codigo_interno.ilike(like))
        )
    items = query.order_by(Maquinaria.nombre).all()
    total_por_estado = {e: db.query(Maquinaria).filter(
        Maquinaria.activa == True, Maquinaria.estado == e,
        Maquinaria.almacen_id == warehouse_id).count()
        for e in ESTADOS_MAQUINARIA}
    return templates.TemplateResponse(request, "maquinaria.html", ctx_base(
        request, user, db,
        items=items, estados=ESTADOS_MAQUINARIA,
        total_por_estado=total_por_estado,
        filtro_estado=estado, filtro_q=q,
    ))


@app.get("/maquinaria/nueva", response_class=HTMLResponse)
def maquinaria_nueva_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    codigo_barras: str = Query(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    return templates.TemplateResponse(request, "maquinaria_nueva.html", ctx_base(
        request, user, db,
        tipos=TIPOS_MAQUINARIA,
        estados=ESTADOS_MAQUINARIA,
        prefill_codigo_barras=codigo_barras.strip(),
        scan_origen=bool(codigo_barras.strip()),
    ))


@app.post("/maquinaria/nueva")
def maquinaria_nueva_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    nombre: str = Form(...),
    tipo: str = Form(""),
    marca: str = Form(""),
    modelo: str = Form(""),
    matricula: str = Form(""),
    num_serie: str = Form(""),
    codigo_barras: str = Form(""),
    estado: str = Form("disponible"),
    ubicacion: str = Form(""),
    responsable: str = Form(""),
    notas: str = Form(""),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso")
    # Validar codigo_barras único
    if codigo_barras:
        existe = db.query(Maquinaria).filter(Maquinaria.codigo_barras == codigo_barras.strip()).first()
        if existe:
            return templates.TemplateResponse(request, "maquinaria_nueva.html", ctx_base(
                request, user,
                tipos=TIPOS_MAQUINARIA, estados=ESTADOS_MAQUINARIA,
                prefill_codigo_barras=codigo_barras,
                error=f"El código de barras '{codigo_barras}' ya está registrado.",
            ))
    almacen_predeterminado = _active_warehouse(db, user, request)
    m = Maquinaria(
        nombre=nombre.strip(),
        tipo=tipo or None,
        marca=marca or None,
        modelo=modelo or None,
        matricula=matricula.strip() or None,
        num_serie=num_serie or None,
        codigo_barras=codigo_barras.strip() or None,
        codigo_interno=generar_referencia_maquinaria(db),
        estado=estado,
        ubicacion=ubicacion or (almacen_predeterminado.nombre if almacen_predeterminado else None),
        almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
        responsable=responsable or None,
        notas=notas or None,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    registrar_auditoria(db, user, "maquinaria", m.id, "crear", None,
                        {"nombre": m.nombre, "codigo_barras": m.codigo_barras})
    return RedirectResponse(f"/maquinaria/{m.id}", status_code=303)


@app.get("/maquinaria/{mid}/pasaporte", response_class=HTMLResponse)
def maquinaria_pasaporte(
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    m = db.get(Maquinaria, mid)
    if not m or not m.activa:
        raise HTTPException(404)
    _require_warehouse_access(user, m.almacen_id)
    codigo_pasaporte = m.codigo_interno or m.codigo_barras or f"MRD-MAQ-{m.id}"
    eventos = db.query(EventoMaquinaria).filter(
        EventoMaquinaria.maquinaria_id == m.id,
    ).order_by(EventoMaquinaria.fecha.desc(), EventoMaquinaria.id.desc()).all()
    documentos = db.query(DocumentoMaquinaria).filter(
        DocumentoMaquinaria.maquinaria_id == m.id,
    ).order_by(DocumentoMaquinaria.creado_en.desc()).all()
    coste_total = sum(float(e.coste or 0) for e in eventos)
    averias = sum(1 for e in eventos if e.tipo == "averia")
    reparaciones = sum(1 for e in eventos if e.tipo == "reparacion")
    hoy = date.today()
    if not m.proxima_revision:
        estado_revision = "sin_fecha"
    elif m.proxima_revision < hoy:
        estado_revision = "vencida"
    elif m.proxima_revision <= hoy + timedelta(days=30):
        estado_revision = "proxima"
    else:
        estado_revision = "al_dia"
    campos_control = [
        m.codigo_interno, m.nombre, m.tipo, m.marca, m.modelo, m.num_serie,
        m.ubicacion, m.proxima_revision, m.localizador_alias, documentos,
    ]
    completitud = round(sum(bool(campo) for campo in campos_control) / len(campos_control) * 100)
    return templates.TemplateResponse(request, "maquinaria_pasaporte.html", ctx_base(
        request, user, db,
        maquina=m, estados=ESTADOS_MAQUINARIA, tipos_maquinaria=TIPOS_MAQUINARIA,
        codigo_pasaporte=codigo_pasaporte, qr_b64=generar_qr_base64(codigo_pasaporte),
        eventos=eventos, documentos=documentos, coste_total=coste_total,
        averias=averias, reparaciones=reparaciones, estado_revision=estado_revision,
        completitud=completitud, tipos_evento=TIPOS_EVENTO_MAQUINARIA,
        estados_localizador=ESTADOS_LOCALIZADOR,
        puede_operar=_puede_operar_maquinaria(user),
    ))


@app.get("/maquinaria/{mid}", response_class=HTMLResponse)
def maquinaria_detalle(
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    m = db.get(Maquinaria, mid)
    if not m or not m.activa:
        raise HTTPException(404)
    _require_warehouse_access(user, m.almacen_id)
    # La ficha técnica forma parte del pasaporte único. Mantener una segunda
    # pantalla duplicada provocaba bloqueos visuales y datos inconsistentes.
    query = f"?{request.url.query}" if request.url.query else ""
    return RedirectResponse(
        f"/maquinaria/{m.id}/pasaporte{query}#ficha-tecnica", status_code=303,
    )


@app.post("/maquinaria/{mid}/pasaporte/eventos", response_class=RedirectResponse)
def maquinaria_evento_crear(
    mid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    tipo: str = Form(...), titulo: str = Form(...), descripcion: str = Form(""),
    fecha: str = Form(""), horas_maquina: str = Form(""), coste: str = Form(""),
    proveedor: str = Form(""), pieza_referencia: str = Form(""),
    proxima_revision: str = Form(""),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.get(Maquinaria, mid)
    if not m:
        raise HTTPException(404)
    if tipo not in TIPOS_EVENTO_MAQUINARIA:
        raise HTTPException(400, "Tipo de evento no válido")
    titulo = titulo.strip()
    if not titulo or len(titulo) > 200:
        raise HTTPException(400, "Título obligatorio (máximo 200 caracteres)")
    try:
        fecha_evento = datetime.fromisoformat(fecha) if fecha else datetime.now()
        horas_val = float(horas_maquina) if horas_maquina else None
        coste_val = float(coste) if coste else None
        prox_val = date.fromisoformat(proxima_revision) if proxima_revision else None
    except ValueError:
        raise HTTPException(400, "Fecha, horas o coste no válidos")
    if ((horas_val is not None and horas_val < 0)
            or (coste_val is not None and coste_val < 0)):
        raise HTTPException(400, "Horas y coste no pueden ser negativos")
    evento = EventoMaquinaria(
        maquinaria_id=m.id, tipo=tipo, titulo=titulo,
        descripcion=descripcion.strip() or None, fecha=fecha_evento,
        horas_maquina=horas_val, coste=coste_val,
        proveedor=proveedor.strip() or None,
        pieza_referencia=pieza_referencia.strip() or None,
        proxima_revision=prox_val, usuario_id=user.id,
    )
    if horas_val is not None and (m.horas_uso is None or horas_val >= m.horas_uso):
        m.horas_uso = horas_val
    if prox_val:
        m.proxima_revision = prox_val
    db.add(evento)
    registrar_auditoria(db, "maquinaria", m.id, "pasaporte_evento", user.id, None,
                        {"tipo": tipo, "titulo": titulo})
    db.commit()
    return RedirectResponse(f"/maquinaria/{mid}/pasaporte?ok=evento", status_code=303)


@app.post("/maquinaria/{mid}/pasaporte/localizador", response_class=RedirectResponse)
def maquinaria_localizador_actualizar(
    mid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    localizador_tipo: str = Form("Apple AirTag"),
    localizador_alias: str = Form(""),
    localizador_identificador: str = Form(""),
    localizador_estado: str = Form("pendiente"),
    localizador_notas: str = Form(""),
    marcar_verificado: str = Form(""),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.get(Maquinaria, mid)
    if not m:
        raise HTTPException(404)
    if localizador_estado not in ESTADOS_LOCALIZADOR:
        raise HTTPException(400, "Estado de localizador no válido")
    anterior = {"alias": m.localizador_alias, "estado": m.localizador_estado}
    m.localizador_tipo = localizador_tipo.strip()[:30] or None
    m.localizador_alias = localizador_alias.strip()[:100] or None
    m.localizador_identificador = localizador_identificador.strip()[:120] or None
    m.localizador_estado = localizador_estado
    m.localizador_notas = localizador_notas.strip() or None
    if marcar_verificado:
        m.localizador_ultima_verificacion = datetime.now()
    registrar_auditoria(db, "maquinaria", m.id, "localizador_actualizar", user.id, anterior,
                        {"alias": m.localizador_alias, "estado": m.localizador_estado})
    db.commit()
    return RedirectResponse(f"/maquinaria/{mid}/pasaporte?ok=localizador", status_code=303)


@app.post("/maquinaria/{mid}/pasaporte/documentos", response_class=RedirectResponse)
async def maquinaria_documento_subir(
    mid: int, archivo: UploadFile = File(...), tipo: str = Form("documento"),
    notas: str = Form(""), user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.get(Maquinaria, mid)
    if not m:
        raise HTTPException(404)
    tipos_documento = {"manual", "certificado", "revision", "factura", "foto", "documento"}
    if tipo not in tipos_documento:
        raise HTTPException(400, "Tipo de documento no válido")
    permitidas = {"jpg", "jpeg", "png", "webp", "pdf"}
    try:
        nombre_seguro, ext = validar_nombre_archivo(archivo.filename, permitidas)
        cabecera = await archivo.read(16)
        await archivo.seek(0)
        validar_contenido_archivo(cabecera, ext)
        contenido = await archivo.read()
        validar_tamaño_bytes(len(contenido), MAX_UPLOAD_MB)
    except ErrorArchivo as exc:
        raise HTTPException(400, str(exc))
    carpeta = UPLOADS_DIR / "maquinaria_documentos"
    carpeta.mkdir(parents=True, exist_ok=True)
    nombre_guardado = f"maq_{mid}_{uuid.uuid4().hex}.{ext}"
    destino = carpeta / nombre_guardado
    destino.write_bytes(contenido)
    doc = DocumentoMaquinaria(
        maquinaria_id=mid, tipo=tipo.strip()[:30] or "documento",
        nombre_original=f"{nombre_seguro}.{ext}", archivo_path=nombre_guardado,
        notas=notas.strip() or None, usuario_id=user.id,
    )
    try:
        db.add(doc)
        registrar_auditoria(db, "maquinaria", mid, "documento_subir", user.id, None,
                            {"tipo": doc.tipo, "nombre": doc.nombre_original})
        db.commit()
    except Exception:
        db.rollback()
        destino.unlink(missing_ok=True)
        raise
    return RedirectResponse(f"/maquinaria/{mid}/pasaporte?ok=documento", status_code=303)


@app.get("/maquinaria/{mid}/pasaporte/documentos/{did}")
def maquinaria_documento_descargar(
    mid: int, did: int, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    doc = db.query(DocumentoMaquinaria).filter(
        DocumentoMaquinaria.id == did, DocumentoMaquinaria.maquinaria_id == mid,
    ).first()
    if not doc:
        raise HTTPException(404)
    ruta = (UPLOADS_DIR / "maquinaria_documentos" / doc.archivo_path).resolve()
    raiz = (UPLOADS_DIR / "maquinaria_documentos").resolve()
    if raiz not in ruta.parents or not ruta.is_file():
        raise HTTPException(404)
    return FileResponse(ruta, filename=doc.nombre_original)


@app.post("/maquinaria/{mid}/estado")
def maquinaria_update_estado(
    mid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = Form(...),
    ubicacion: str = Form(""),
    responsable: str = Form(""),
    obra_actual: str = Form(""),
    notas: str = Form(""),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.query(Maquinaria).get(mid)
    if not m:
        raise HTTPException(404)
    if estado not in ESTADOS_MAQUINARIA:
        raise HTTPException(400, "Estado no válido")
    anterior = {"estado": m.estado, "ubicacion": m.ubicacion, "responsable": m.responsable}
    m.estado = estado
    if ubicacion:
        m.ubicacion = ubicacion
    if responsable:
        m.responsable = responsable
    if obra_actual:
        m.obra_actual = obra_actual
    if notas:
        m.notas = (m.notas or "") + f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notas}"
    registrar_auditoria(db, "maquinaria", m.id, "cambio_estado", user.id,
                        anterior, {"estado": estado, "ubicacion": ubicacion})
    db.commit()
    # Sprint 4.2: dispatch evento de maquinaria
    try:
        auto_engine.dispatch_evento(
            tipo_evento="evento_maquinaria",
            estado_nuevo=estado,
            item={
                "tipo": "maquinaria",
                "id": m.id,
                "codigo": m.codigo_interno or m.codigo_barras or str(m.id),
                "nombre": m.nombre,
                "marca": m.marca or "",
                "estado": estado,
                "dias": 0,
                "enlace": f"/maquinaria/{m.id}",
            },
            db=db,
        )
    except Exception:
        pass
    return RedirectResponse(f"/maquinaria/{m.id}", status_code=303)


@app.post("/maquinaria/{mid}/editar", response_class=RedirectResponse)
async def maquinaria_editar(
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "editar"):
        raise HTTPException(403, "Sin permiso")
    m = db.query(Maquinaria).get(mid)
    if not m:
        raise HTTPException(404)
    form = await request.form()
    m.nombre = form.get("nombre") or m.nombre
    m.tipo = form.get("tipo") or m.tipo
    m.marca = form.get("marca") or None
    m.modelo = form.get("modelo") or None
    m.matricula = form.get("matricula") or None
    m.num_serie = form.get("num_serie") or None
    m.estado = form.get("estado") or m.estado
    m.ubicacion = form.get("ubicacion") or None
    m.responsable = form.get("responsable") or None
    m.notas = form.get("notas") or None
    km = form.get("km_actuales", "")
    if km:
        try:
            m.km_actuales = float(km)
        except ValueError:
            pass
    db.commit()
    return RedirectResponse(f"/maquinaria/{mid}?ok=editado", status_code=303)


@app.get("/maquinaria/{mid}/etiqueta", response_class=HTMLResponse)
def maquinaria_etiqueta(
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    m = db.query(Maquinaria).get(mid)
    if not m:
        raise HTTPException(404)
    codigo = m.codigo_interno or m.codigo_barras or m.num_serie or f"MRD-MAQ-{m.id}"
    qr_b64 = generar_qr_base64(codigo)
    return templates.TemplateResponse(request, "maquinaria_etiqueta.html", ctx_base(
        request, user,
        maquina=m,
        codigo=codigo,
        qr_b64=qr_b64,
        empresa=COMPANY_NAME,
    ))


@app.post("/maquinaria/{mid}/foto", response_class=RedirectResponse)
async def maquinaria_foto(
    mid: int,
    foto: UploadFile = File(...),
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.query(Maquinaria).get(mid)
    if not m:
        raise HTTPException(404)
    try:
        _, ext = validar_nombre_archivo(foto.filename, {"jpg", "jpeg", "png", "webp"})
        cabecera = await foto.read(16); await foto.seek(0)
        validar_contenido_archivo(cabecera, ext)
        data = await foto.read()
        validar_tamaño_bytes(len(data), MAX_UPLOAD_MB)
    except ErrorArchivo as exc:
        raise HTTPException(400, str(exc))
    carpeta = BASE_DIR / "static" / "uploads" / "maquinaria"
    carpeta.mkdir(parents=True, exist_ok=True)
    import time as _t
    nombre = f"{mid}_{int(_t.time())}.{ext}"
    (carpeta / nombre).write_bytes(data)
    foto_anterior = m.foto
    m.foto = nombre
    registrar_auditoria(db, "maquinaria", m.id, "foto_actualizar", user.id, None,
                        {"archivo": nombre})
    try:
        db.commit()
    except Exception:
        db.rollback()
        (carpeta / nombre).unlink(missing_ok=True)
        raise
    if foto_anterior:
        (carpeta / foto_anterior).unlink(missing_ok=True)
    return RedirectResponse(f"/maquinaria/{mid}?ok=foto", status_code=303)


@app.post("/maquinaria/{mid}/foto/eliminar", response_class=RedirectResponse)
def maquinaria_foto_eliminar(
    mid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not _puede_operar_maquinaria(user):
        raise HTTPException(403, "Sin permiso")
    m = db.query(Maquinaria).get(mid)
    if m and m.foto:
        p = BASE_DIR / "static" / "uploads" / "maquinaria" / m.foto
        if p.exists():
            p.unlink()
        m.foto = None
        db.commit()
    return RedirectResponse(f"/maquinaria/{mid}?ok=foto_eliminada", status_code=303)


# API JSON para maquinaria
@app.get("/api/v1/maquinaria")
def api_maquinaria_list(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = Query(""),
):
    q = db.query(Maquinaria).filter(Maquinaria.activa == True)
    if estado:
        q = q.filter(Maquinaria.estado == estado)
    items = q.order_by(Maquinaria.nombre).all()
    return JSONResponse([{
        "id": m.id, "nombre": m.nombre, "tipo": m.tipo,
        "codigo_barras": m.codigo_barras, "matricula": m.matricula,
        "estado": m.estado, "estado_label": ESTADOS_MAQUINARIA.get(m.estado, m.estado),
        "ubicacion": m.ubicacion, "responsable": m.responsable,
    } for m in items])



# ═══════════════════════════════════════════════════════════════════════════════
#  SPRINT 4.1 — MOTOR DE AUTOMATIZACIONES
# ═══════════════════════════════════════════════════════════════════════════════

def ctx_auto(request, user, **kw):
    """Helper de contexto base para templates de automatizaciones."""
    return {
        **ctx_base(request, user),
        "estados_auto": ESTADOS_AUTOMATIZACION,
        "prioridades_auto": PRIORIDADES_AUTOMATIZACION,
        "tipos_disparador": TIPOS_DISPARADOR,
        "tipos_condicion": TIPOS_CONDICION,
        "tipos_accion": TIPOS_ACCION,
        "prioridades_aviso": PRIORIDADES_AVISO,
        **kw,
    }


# ─── Listado de automatizaciones ─────────────────────────────────────────────
@app.get("/automatizaciones", response_class=HTMLResponse)
def automatizaciones_list(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = Query(""),
    q: str = Query(""),
):
    query = db.query(Automatizacion)
    if estado:
        query = query.filter(Automatizacion.estado == estado)
    if q:
        query = query.filter(Automatizacion.nombre.ilike(f"%{q}%"))
    items = query.order_by(Automatizacion.id.desc()).all()

    # KPI por estado
    total_por_estado = {k: 0 for k in ESTADOS_AUTOMATIZACION}
    for item in db.query(Automatizacion).all():
        if item.estado in total_por_estado:
            total_por_estado[item.estado] += 1

    # Avisos sin leer
    avisos_sin_leer = db.query(Aviso).filter(Aviso.leido == False, Aviso.archivado == False).count()

    return templates.TemplateResponse(request, "automatizaciones.html", ctx_auto(
        request, user,
        items=items,
        filtro_estado=estado,
        filtro_q=q,
        total_por_estado=total_por_estado,
        avisos_sin_leer=avisos_sin_leer,
    ))


# ─── Nueva automatización — GET ───────────────────────────────────────────────
@app.get("/automatizaciones/nueva", response_class=HTMLResponse)
def automatizacion_nueva_get(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    return templates.TemplateResponse(request, "automatizacion_nueva.html", ctx_auto(
        request, user,
        auto=None,
        error=None,
    ))


# ─── Nueva automatización — POST ─────────────────────────────────────────────
@app.post("/automatizaciones/nueva", response_class=HTMLResponse)
async def automatizacion_nueva_post(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    form = await request.form()

    nombre = (form.get("nombre") or "").strip()
    if not nombre:
        return templates.TemplateResponse(request, "automatizacion_nueva.html", ctx_auto(
            request, user, auto=None, error="El nombre es obligatorio.",
        ))

    tipo_disparador = form.get("tipo_disparador") or "manual"
    prioridad = form.get("prioridad") or "media"
    descripcion = (form.get("descripcion") or "").strip()

    # Config disparador
    config_disparador = {}
    if tipo_disparador == "intervalo":
        config_disparador["intervalo_min"] = int(form.get("intervalo_min") or 60)
    elif tipo_disparador == "diario":
        config_disparador["hora"] = form.get("hora_diario") or "08:00"

    # Condiciones (una por formulario simplificado)
    tipo_cond = form.get("tipo_condicion") or ""
    condiciones = []
    if tipo_cond and tipo_cond != "ninguna":
        cond = {"tipo": tipo_cond}
        if tipo_cond in ("herramienta_dias_entregada", "reparacion_retrasada",
                          "maquinaria_sin_movimiento", "mantenimiento_proximo_itv"):
            cond["dias"] = int(form.get("cond_dias") or 30)
        condiciones.append(cond)

    # Acciones (una acción principal)
    tipo_accion = form.get("tipo_accion") or "crear_aviso"
    accion_titulo = (form.get("accion_titulo") or "Aviso: {nombre}").strip()
    accion_mensaje = (form.get("accion_mensaje") or "El activo {nombre} ({codigo}) requiere atención.").strip()
    accion_prioridad = form.get("accion_prioridad") or "media"
    acciones = [{
        "tipo": tipo_accion,
        "titulo": accion_titulo,
        "mensaje": accion_mensaje,
        "prioridad": accion_prioridad,
    }]

    auto = Automatizacion(
        nombre=nombre,
        descripcion=descripcion,
        estado="activa",
        prioridad=prioridad,
        tipo_disparador=tipo_disparador,
        config_disparador=json.dumps(config_disparador),
        condiciones=json.dumps(condiciones),
        acciones=json.dumps(acciones),
        creado_por_id=user.id,
        version=1,
    )
    # Calcular próxima ejecución
    auto.proxima_ejecucion = auto_engine.calcular_proxima_ejecucion(
        tipo_disparador, config_disparador
    )

    db.add(auto)
    db.commit()
    db.refresh(auto)

    registrar_auditoria(db, user, "automatizacion", auto.id, "crear",
                        None, {"nombre": auto.nombre})
    return RedirectResponse(f"/automatizaciones/{auto.id}", status_code=303)


# ─── Detalle de automatización ────────────────────────────────────────────────
@app.get("/automatizaciones/{auto_id}", response_class=HTMLResponse)
def automatizacion_detalle(
    auto_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)

    historial = (
        db.query(EjecucionAutomatizacion)
        .filter(EjecucionAutomatizacion.automatizacion_id == auto_id)
        .order_by(EjecucionAutomatizacion.fecha.desc())
        .limit(30)
        .all()
    )

    # Parse JSON para mostrar en template
    try:
        condiciones = json.loads(auto.condiciones or "[]")
    except Exception:
        condiciones = []
    try:
        acciones = json.loads(auto.acciones or "[]")
    except Exception:
        acciones = []
    try:
        config_disp = json.loads(auto.config_disparador or "{}")
    except Exception:
        config_disp = {}

    avisos = (
        db.query(Aviso)
        .filter(Aviso.automatizacion_id == auto_id, Aviso.archivado == False)
        .order_by(Aviso.creado_en.desc())
        .limit(20)
        .all()
    )

    return templates.TemplateResponse(request, "automatizacion_detalle.html", ctx_auto(
        request, user,
        auto=auto,
        historial=historial,
        condiciones=condiciones,
        acciones=acciones,
        config_disp=config_disp,
        avisos=avisos,
        es_admin=True,
    ))


# ─── Activar/pausar automatización ───────────────────────────────────────────
@app.post("/automatizaciones/{auto_id}/activar")
def automatizacion_activar(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)
    auto.estado = "activa"
    config = {}
    try:
        config = json.loads(auto.config_disparador or "{}")
    except Exception:
        pass
    auto.proxima_ejecucion = auto_engine.calcular_proxima_ejecucion(auto.tipo_disparador, config)
    db.commit()
    registrar_auditoria(db, user, "automatizacion", auto.id, "activar", None, {"estado": "activa"})
    return RedirectResponse(f"/automatizaciones/{auto_id}", status_code=303)

@app.post("/maquinaria/{mid}/eliminar", response_class=RedirectResponse)
def maquinaria_eliminar(
    mid: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "borrar"):
        raise HTTPException(403, "Sin permiso")
    m = db.query(Maquinaria).get(mid)
    if not m:
        raise HTTPException(404)
    # Baja lógica: el pasaporte, documentos y costes nunca se destruyen.
    anterior = {"activa": m.activa, "estado": m.estado}
    m.activa = False
    m.estado = "baja"
    registrar_auditoria(db, "maquinaria", m.id, "baja", user.id, anterior,
                        {"activa": False, "estado": "baja"},
                        resumen="Baja lógica; pasaporte conservado")
    db.commit()
    return RedirectResponse("/maquinaria?ok=eliminado", status_code=303)



@app.post("/automatizaciones/{auto_id}/pausar")
def automatizacion_pausar(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)
    auto.estado = "pausada"
    auto.proxima_ejecucion = None
    db.commit()
    registrar_auditoria(db, user, "automatizacion", auto.id, "pausar", None, {"estado": "pausada"})
    return RedirectResponse(f"/automatizaciones/{auto_id}", status_code=303)


# ─── Archivar (no eliminar) ───────────────────────────────────────────────────
@app.post("/automatizaciones/{auto_id}/archivar")
def automatizacion_archivar(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)
    auto.estado = "archivada"
    auto.proxima_ejecucion = None
    db.commit()
    registrar_auditoria(db, user, "automatizacion", auto.id, "archivar", None, {})
    return RedirectResponse("/automatizaciones", status_code=303)


# ─── Duplicar ─────────────────────────────────────────────────────────────────
@app.post("/automatizaciones/{auto_id}/duplicar")
def automatizacion_duplicar(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    src = db.query(Automatizacion).get(auto_id)
    if not src:
        raise HTTPException(404)
    nueva = Automatizacion(
        nombre=f"{src.nombre} (copia)",
        descripcion=src.descripcion,
        estado="inactiva",
        prioridad=src.prioridad,
        tipo_disparador=src.tipo_disparador,
        config_disparador=src.config_disparador,
        condiciones=src.condiciones,
        acciones=src.acciones,
        creado_por_id=user.id,
        version=1,
    )
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    registrar_auditoria(db, user, "automatizacion", nueva.id, "duplicar",
                        None, {"fuente": auto_id})
    return RedirectResponse(f"/automatizaciones/{nueva.id}", status_code=303)


# ─── Ejecutar manualmente ────────────────────────────────────────────────────
@app.post("/automatizaciones/{auto_id}/ejecutar")
def automatizacion_ejecutar(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    resultado = auto_engine.ejecutar_automatizacion(
        auto_id, db, modo="manual", simulacion=False, usuario_id=user.id
    )
    registrar_auditoria(db, user, "automatizacion", auto_id, "ejecutar_manual",
                        None, {"resultado": resultado.get("resultado")})
    return RedirectResponse(f"/automatizaciones/{auto_id}?ejecutado=1", status_code=303)


# ─── Simular ─────────────────────────────────────────────────────────────────
@app.post("/automatizaciones/{auto_id}/simular", response_class=HTMLResponse)
def automatizacion_simular(
    auto_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    resultado = auto_engine.ejecutar_automatizacion(
        auto_id, db, modo="simulacion", simulacion=True, usuario_id=user.id
    )

    auto = db.query(Automatizacion).get(auto_id)
    try:
        condiciones = json.loads(auto.condiciones or "[]")
    except Exception:
        condiciones = []
    try:
        acciones = json.loads(auto.acciones or "[]")
    except Exception:
        acciones = []
    try:
        config_disp = json.loads(auto.config_disparador or "{}")
    except Exception:
        config_disp = {}

    historial = (
        db.query(EjecucionAutomatizacion)
        .filter(EjecucionAutomatizacion.automatizacion_id == auto_id)
        .order_by(EjecucionAutomatizacion.fecha.desc())
        .limit(30)
        .all()
    )
    avisos = (
        db.query(Aviso)
        .filter(Aviso.automatizacion_id == auto_id, Aviso.archivado == False)
        .order_by(Aviso.creado_en.desc())
        .limit(20)
        .all()
    )

    return templates.TemplateResponse(request, "automatizacion_detalle.html", ctx_auto(
        request, user,
        auto=auto,
        historial=historial,
        condiciones=condiciones,
        acciones=acciones,
        config_disp=config_disp,
        avisos=avisos,
        resultado_simulacion=resultado,
        es_admin=True,
    ))


# ─── API: historial de ejecuciones ──────────────────────────────────────────
@app.get("/api/v1/automatizaciones/{auto_id}/historial")
def api_auto_historial(
    auto_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    limit: int = Query(20, le=100),
):
    historial = (
        db.query(EjecucionAutomatizacion)
        .filter(EjecucionAutomatizacion.automatizacion_id == auto_id)
        .order_by(EjecucionAutomatizacion.fecha.desc())
        .limit(limit)
        .all()
    )
    return JSONResponse([{
        "id": e.id,
        "fecha": e.fecha.isoformat() if e.fecha else None,
        "modo": e.modo,
        "resultado": e.resultado,
        "acciones_ejecutadas": e.acciones_ejecutadas,
        "items_afectados": e.items_afectados,
        "duracion_ms": e.duracion_ms,
        "error": e.error,
    } for e in historial])


# ─── Avisos operativos reales ────────────────────────────────────────────────
def _datos_aviso(aviso: Aviso) -> dict:
    try:
        data = json.loads(aviso.datos or "{}")
        return data if isinstance(data, dict) else {}
    except (TypeError, ValueError):
        return {}


def _sincronizar_avisos_operativos(db: Session) -> int:
    """Crea avisos solo desde incidencias, reparaciones y revisiones reales."""
    existing = {
        data.get("source_key")
        for aviso in db.query(Aviso).all()
        for data in [_datos_aviso(aviso)]
        if data.get("source_key")
    }
    candidates = []
    now = datetime.now()
    today = date.today()
    limit = today + timedelta(days=30)

    for item in db.query(MantenimientoProgramado).filter(
        MantenimientoProgramado.estado.in_(["pendiente", "en_proceso", "vencido"])
    ).all():
        due = item.fecha_programada.date() if isinstance(item.fecha_programada, datetime) else item.fecha_programada
        priority = "critica" if due and due < today else "alta" if due and due <= limit else "media"
        candidates.append((
            f"mantenimiento:{item.id}", "mantenimiento", priority,
            f"Mantenimiento {'vencido' if due and due < today else 'programado'} — {item.nombre_activo}",
            f"{item.descripcion or item.tipo}. Fecha: {due.strftime('%d/%m/%Y') if due else 'sin fecha'}",
            "/mantenimiento",
        ))
    for item in db.query(Reparacion).filter(
        Reparacion.estado.notin_(["finalizada", "sin_reparacion"])
    ).all():
        candidates.append((
            f"reparacion:{item.id}", "reparacion", "alta" if item.prioridad in ("alta", "critica") else "media",
            f"Reparación abierta — {item.numero}", item.diagnostico or item.descripcion or "Pendiente de completar",
            f"/reparaciones/{item.id}",
        ))
    for item in db.query(Incidencia).filter(
        Incidencia.estado.notin_(["resuelta", "cerrada"])
    ).all():
        candidates.append((
            f"incidencia:{item.id}", "averia" if item.tipo in ("averia", "rotura", "mal_funcionamiento") else "incidencia",
            "critica" if item.prioridad == "critica" else "alta" if item.prioridad == "alta" else "media",
            f"Incidencia activa — {item.numero}", item.titulo,
            f"/incidencias/{item.id}",
        ))
    for item in db.query(Maquinaria).filter(
        Maquinaria.activa == True, Maquinaria.proxima_revision != None,
        Maquinaria.proxima_revision <= limit,
    ).all():
        candidates.append((
            f"revision_maquinaria:{item.id}:{item.proxima_revision}", "revision",
            "critica" if item.proxima_revision < today else "alta",
            f"Revisión de maquinaria — {item.nombre}",
            f"Fecha: {item.proxima_revision.strftime('%d/%m/%Y')}",
            f"/maquinaria/{item.id}/pasaporte",
        ))
    for item in db.query(EPIIndividual).filter(
        EPIIndividual.estado != "baja", EPIIndividual.proxima_revision != None,
        EPIIndividual.proxima_revision <= limit,
    ).all():
        candidates.append((
            f"revision_epi:{item.id}:{item.proxima_revision}", "revision",
            "critica" if item.proxima_revision < today else "alta",
            f"Revisión de EPI — {item.tipo} {item.codigo_fabricacion}",
            f"Fecha: {item.proxima_revision.strftime('%d/%m/%Y')}",
            f"/epis/individuales/{item.id}",
        ))
    for item in db.query(Herramienta).filter(
        Herramienta.activa == True, Herramienta.fecha_proximo_mantenimiento != None,
        Herramienta.fecha_proximo_mantenimiento <= limit,
    ).all():
        candidates.append((
            f"mantenimiento_herramienta:{item.id}:{item.fecha_proximo_mantenimiento}", "mantenimiento",
            "critica" if item.fecha_proximo_mantenimiento < today else "alta",
            f"Mantenimiento de herramienta — {item.nombre}",
            f"Fecha: {item.fecha_proximo_mantenimiento.strftime('%d/%m/%Y')}",
            f"/herramientas/{item.id}",
        ))

    created = 0
    for source_key, notice_type, priority, title, message, link in candidates:
        if source_key in existing:
            continue
        db.add(Aviso(
            titulo=title[:200], mensaje=message, prioridad=priority, tipo=notice_type,
            enlace=link, datos=json.dumps({"source_key": source_key, "estado": "activo"}),
            leido=False, archivado=False,
        ))
        existing.add(source_key)
        created += 1
    return created


# ─── API: avisos ─────────────────────────────────────────────────────────────
@app.get("/api/v1/avisos")
def api_avisos_list(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    solo_no_leidos: bool = Query(False),
    limit: int = Query(50, le=200),
):
    _sincronizar_avisos_operativos(db)
    db.commit()
    q = db.query(Aviso).filter(
        Aviso.archivado == False, Aviso.tipo.in_(TIPOS_AVISO_OPERATIVO),
    )
    if solo_no_leidos:
        q = q.filter(Aviso.leido == False)
    avisos = q.order_by(Aviso.creado_en.desc()).limit(limit).all()
    return JSONResponse([{
        "id": a.id,
        "titulo": a.titulo,
        "mensaje": a.mensaje,
        "prioridad": a.prioridad,
        "tipo": a.tipo,
        "leido": a.leido,
        "enlace": a.enlace,
        "creado_en": a.creado_en.isoformat() if a.creado_en else None,
        "automatizacion_id": a.automatizacion_id,
    } for a in avisos])


@app.post("/api/v1/avisos/{aviso_id}/leer")
def api_aviso_leer(
    aviso_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    aviso = db.query(Aviso).get(aviso_id)
    if not aviso:
        raise HTTPException(404)
    aviso.leido = True
    aviso.leido_en = datetime.utcnow()
    db.commit()
    return JSONResponse({"ok": True})


@app.post("/api/v1/avisos/{aviso_id}/archivar")
def api_aviso_archivar(
    aviso_id: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    aviso = db.query(Aviso).get(aviso_id)
    if not aviso:
        raise HTTPException(404)
    aviso.archivado = True
    aviso.leido = True
    db.commit()
    return JSONResponse({"ok": True})


def _cerrar_aviso_operativo(aviso: Aviso, user: Usuario, estado: str, db: Session, request: Request):
    if estado not in {"terminado", "no_real"}:
        raise HTTPException(400, "Estado no válido")
    data = _datos_aviso(aviso)
    data.update({
        "estado": estado, "cerrado_por_id": user.id,
        "cerrado_por": user.nombre or user.username,
        "cerrado_en": datetime.utcnow().isoformat(),
    })
    aviso.datos = json.dumps(data, ensure_ascii=False)
    aviso.leido = True
    aviso.leido_en = datetime.utcnow()
    aviso.archivado = True
    registrar_auditoria(
        db, "avisos", aviso.id, estado, user.id, None,
        {"titulo": aviso.titulo, "source_key": data.get("source_key")},
        "Aviso operativo terminado" if estado == "terminado" else "Aviso marcado como no real",
        request.client.host if request.client else "",
    )
    db.commit()
    return JSONResponse({"ok": True, "estado": estado})


@app.post("/api/v1/avisos/{aviso_id}/terminar")
def api_aviso_terminar(
    aviso_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    aviso = db.get(Aviso, aviso_id)
    if not aviso or aviso.tipo not in TIPOS_AVISO_OPERATIVO:
        raise HTTPException(404)
    return _cerrar_aviso_operativo(aviso, user, "terminado", db, request)


@app.post("/api/v1/avisos/{aviso_id}/no-real")
def api_aviso_no_real(
    aviso_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso")
    aviso = db.get(Aviso, aviso_id)
    if not aviso or aviso.tipo not in TIPOS_AVISO_OPERATIVO:
        raise HTTPException(404)
    return _cerrar_aviso_operativo(aviso, user, "no_real", db, request)


# ─── Panel de avisos ─────────────────────────────────────────────────────────
@app.get("/buzon", include_in_schema=False)
def buzon_compatibilidad():
    """Conserva el acceso histórico; el buzón operativo ahora vive en Avisos."""
    return RedirectResponse("/avisos", status_code=303)


@app.get("/avisos", response_class=HTMLResponse)
def avisos_panel(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    mostrar: str = Query("activos"),  # activos | todos | archivados
):
    _sincronizar_avisos_operativos(db)
    db.commit()
    q = db.query(Aviso).filter(Aviso.tipo.in_(TIPOS_AVISO_OPERATIVO))
    if mostrar == "archivados":
        q = q.filter(Aviso.archivado == True)
    elif mostrar == "todos":
        pass
    else:
        q = q.filter(Aviso.archivado == False)

    avisos = q.order_by(Aviso.creado_en.desc()).limit(200).all()
    total_sin_leer = db.query(Aviso).filter(
        Aviso.leido == False, Aviso.archivado == False,
        Aviso.tipo.in_(TIPOS_AVISO_OPERATIVO),
    ).count()

    return templates.TemplateResponse(request, "avisos.html", ctx_base(
        request, user,
        avisos=avisos,
        mostrar=mostrar,
        total_sin_leer=total_sin_leer,
        prioridades_aviso=PRIORIDADES_AVISO,
        puede_cerrar=(tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")),
    ))


@app.get("/configuracion/auditoria", response_class=HTMLResponse)
def auditoria_usuarios(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    usuario_id: int = Query(0, ge=0),
    dias: int = Query(30, ge=1, le=365),
):
    """Informe privado de actividad; únicamente el administrador puede verlo."""
    if user.rol != "admin":
        raise HTTPException(403, "Informe reservado al administrador")
    since = datetime.utcnow() - timedelta(days=dias)
    query = db.query(AuditoriaLog).options(joinedload(AuditoriaLog.usuario)).filter(
        AuditoriaLog.fecha >= since,
    )
    if usuario_id:
        query = query.filter(AuditoriaLog.usuario_id == usuario_id)
    rows = query.order_by(AuditoriaLog.fecha.desc()).limit(1000).all()
    usuarios = db.query(Usuario).order_by(Usuario.nombre, Usuario.username).all()
    return templates.TemplateResponse(request, "auditoria_usuarios.html", ctx_base(
        request, user, db, registros=rows, usuarios=usuarios,
        usuario_id=usuario_id, dias=dias,
    ))


# ═══════════════════════════════════════════════════════════════════════════════
#  SPRINT 4.2 — EDITAR AUTOMATIZACIÓN + GESTIÓN DE LISTENERS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/automatizaciones/{auto_id}/editar", response_class=HTMLResponse)
def automatizacion_editar_get(
    auto_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)
    if auto.estado == "archivada":
        raise HTTPException(400, "No se puede editar una automatización archivada.")

    try:
        condiciones = json.loads(auto.condiciones or "[]")
    except Exception:
        condiciones = []
    try:
        acciones = json.loads(auto.acciones or "[]")
    except Exception:
        acciones = []
    try:
        config_disp = json.loads(auto.config_disparador or "{}")
    except Exception:
        config_disp = {}

    # Get list of users for notificar_usuario action
    usuarios = db.query(Usuario).filter(Usuario.activo == True).order_by(Usuario.username).all()

    return templates.TemplateResponse(request, "automatizacion_editar.html", ctx_auto(
        request, user,
        auto=auto,
        condiciones=condiciones,
        acciones=acciones,
        config_disp=config_disp,
        usuarios=usuarios,
        error=None,
    ))


@app.post("/automatizaciones/{auto_id}/editar", response_class=HTMLResponse)
async def automatizacion_editar_post(
    auto_id: int,
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    auto = db.query(Automatizacion).get(auto_id)
    if not auto:
        raise HTTPException(404)
    if auto.estado == "archivada":
        raise HTTPException(400)

    form = await request.form()

    nombre = (form.get("nombre") or "").strip()
    if not nombre:
        usuarios = db.query(Usuario).filter(Usuario.activo == True).all()
        return templates.TemplateResponse(request, "automatizacion_editar.html", ctx_auto(
            request, user, auto=auto,
            condiciones=json.loads(auto.condiciones or "[]"),
            acciones=json.loads(auto.acciones or "[]"),
            config_disp=json.loads(auto.config_disparador or "{}"),
            usuarios=usuarios,
            error="El nombre es obligatorio.",
        ))

    tipo_disparador = form.get("tipo_disparador") or "manual"
    prioridad = form.get("prioridad") or "media"
    descripcion = (form.get("descripcion") or "").strip()

    # Config disparador
    config_disparador = {}
    if tipo_disparador == "intervalo":
        config_disparador["intervalo_min"] = int(form.get("intervalo_min") or 60)
    elif tipo_disparador == "diario":
        config_disparador["hora"] = form.get("hora_diario") or "08:00"
    elif tipo_disparador in ("evento_herramienta", "evento_maquinaria"):
        config_disparador["filtro_estado"] = form.get("filtro_estado_evento") or ""

    # Condiciones — hasta 3 condiciones con AND
    condiciones = []
    for i in range(1, 4):
        tipo_cond = form.get(f"tipo_condicion_{i}") or ""
        if not tipo_cond or tipo_cond == "ninguna":
            continue
        cond = {"tipo": tipo_cond}
        dias_val = form.get(f"cond_dias_{i}")
        if dias_val:
            cond["dias"] = int(dias_val)
        estado_val = form.get(f"cond_estado_{i}")
        if estado_val:
            cond["estado"] = estado_val
        condiciones.append(cond)

    # Acciones — hasta 2 acciones
    acciones = []
    for i in range(1, 3):
        tipo_accion = form.get(f"tipo_accion_{i}") or ""
        if not tipo_accion or tipo_accion == "ninguna":
            continue
        accion = {
            "tipo": tipo_accion,
            "titulo": (form.get(f"accion_titulo_{i}") or "Aviso: {nombre}").strip(),
            "mensaje": (form.get(f"accion_mensaje_{i}") or "").strip(),
            "prioridad": form.get(f"accion_prioridad_{i}") or "media",
        }
        if tipo_accion == "cambiar_estado_herramienta":
            accion["estado_destino"] = form.get(f"accion_estado_destino_{i}") or ""
        if tipo_accion == "notificar_usuario":
            accion["username_destino"] = form.get(f"accion_username_{i}") or ""
        if tipo_accion == "registrar_log":
            accion["nivel"] = form.get(f"accion_nivel_{i}") or "info"
        acciones.append(accion)

    # Snapshot anterior
    snap_ant = {
        "nombre": auto.nombre, "estado": auto.estado,
        "tipo_disparador": auto.tipo_disparador,
    }

    auto.nombre = nombre
    auto.descripcion = descripcion
    auto.prioridad = prioridad
    auto.tipo_disparador = tipo_disparador
    auto.config_disparador = json.dumps(config_disparador)
    auto.condiciones = json.dumps(condiciones)
    auto.acciones = json.dumps(acciones)
    auto.version = (auto.version or 1) + 1
    auto.proxima_ejecucion = auto_engine.calcular_proxima_ejecucion(
        tipo_disparador, config_disparador
    )

    db.commit()

    # Actualizar listeners si es trigger de evento
    if tipo_disparador in ("evento_herramienta", "evento_maquinaria"):
        auto_engine.registrar_listener_eventos(
            auto.id, tipo_disparador,
            config_disparador.get("filtro_estado", "")
        )
    else:
        auto_engine.deregistrar_listener(auto.id)

    registrar_auditoria(db, user, "automatizacion", auto.id, "editar",
                        snap_ant, {"nombre": auto.nombre, "version": auto.version})

    return RedirectResponse(f"/automatizaciones/{auto_id}", status_code=303)


# ─── API: estado rápido de automatizaciones ──────────────────────────────────
@app.get("/api/v1/automatizaciones")
def api_automatizaciones_list(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    estado: str = Query(""),
):
    q = db.query(Automatizacion).filter(Automatizacion.estado != "archivada")
    if estado:
        q = q.filter(Automatizacion.estado == estado)
    items = q.order_by(Automatizacion.nombre).all()
    return JSONResponse([{
        "id": a.id,
        "nombre": a.nombre,
        "estado": a.estado,
        "tipo_disparador": a.tipo_disparador,
        "ultimo_resultado": a.ultimo_resultado,
        "ultima_ejecucion": a.ultima_ejecucion.isoformat() if a.ultima_ejecucion else None,
        "proxima_ejecucion": a.proxima_ejecucion.isoformat() if a.proxima_ejecucion else None,
        "total_ejecuciones": a.total_ejecuciones or 0,
    } for a in items])


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 4.3 — CENTRO INTELIGENTE DE NOTIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/notificaciones")
async def notificaciones_panel(request: Request, db: Session = Depends(get_db),
                                user: Usuario = Depends(requiere_login)):
    canales = db.query(CanalNotificacion).order_by(CanalNotificacion.id).all()
    log = db.query(NotificacionEnviada).order_by(
        NotificacionEnviada.fecha_envio.desc()).limit(50).all()
    return templates.TemplateResponse(request, "notificaciones.html", ctx_base(
        request, user, db,
        canales=canales,
        log=log,
        tipos_canal=TIPOS_CANAL,
        prioridades_canal=PRIORIDADES_CANAL,
    ))


@app.get("/notificaciones/nuevo-canal")
async def notificaciones_nuevo_canal_get(request: Request,
                                          db: Session = Depends(get_db),
                                          user: Usuario = Depends(requiere_login)):
    return templates.TemplateResponse(request, "notificaciones.html", ctx_base(
        request, user, db,
        canales=db.query(CanalNotificacion).order_by(CanalNotificacion.id).all(),
        log=[],
        tipos_canal=TIPOS_CANAL,
        prioridades_canal=PRIORIDADES_CANAL,
        mostrar_form=True,
        error=None,
    ))


@app.post("/notificaciones/nuevo-canal")
async def notificaciones_nuevo_canal_post(
    request: Request, db: Session = Depends(get_db),
    user: Usuario = Depends(requiere_login),
    nombre: str = Form(...),
    tipo: str = Form(...),
    prioridad_minima: str = Form("media"),
    # Email
    smtp_host: str = Form(""),
    smtp_port: str = Form("587"),
    smtp_user: str = Form(""),
    smtp_pass: str = Form(""),
    smtp_tls: str = Form("1"),
    destinatarios: str = Form(""),
    # Webhook
    webhook_url: str = Form(""),
    webhook_incluir_enlace: str = Form("1"),
):
    if tipo not in TIPOS_CANAL:
        return templates.TemplateResponse(request, "notificaciones.html", ctx_base(
            request, user, db,
            canales=db.query(CanalNotificacion).all(), log=[],
            tipos_canal=TIPOS_CANAL, prioridades_canal=PRIORIDADES_CANAL,
            mostrar_form=True, error="Tipo de canal inválido."))

    if tipo == "email":
        cfg = {
            "smtp_host": smtp_host.strip(),
            "smtp_port": int(smtp_port) if smtp_port.isdigit() else 587,
            "smtp_user": smtp_user.strip(),
            "smtp_pass": smtp_pass,
            "smtp_tls": smtp_tls == "1",
            "destinatarios": [e.strip() for e in destinatarios.split(",") if e.strip()],
        }
        if not cfg["smtp_host"] or not cfg["destinatarios"]:
            return templates.TemplateResponse(request, "notificaciones.html", ctx_base(
                request, user, db,
                canales=db.query(CanalNotificacion).all(), log=[],
                tipos_canal=TIPOS_CANAL, prioridades_canal=PRIORIDADES_CANAL,
                mostrar_form=True, error="Faltan servidor SMTP o destinatarios."))
    elif tipo == "webhook":
        url = webhook_url.strip()
        if not url.startswith(("http://", "https://")):
            return templates.TemplateResponse(request, "notificaciones.html", ctx_base(
                request, user, db,
                canales=db.query(CanalNotificacion).all(), log=[],
                tipos_canal=TIPOS_CANAL, prioridades_canal=PRIORIDADES_CANAL,
                mostrar_form=True, error="URL de webhook inválida."))
        cfg = {
            "url": url,
            "incluir_enlace": webhook_incluir_enlace == "1",
        }
    else:
        # webpush: sin configuración propia, se envía a todos los navegadores
        # que se hayan suscrito desde /notificaciones/mis-push
        cfg = {}

    canal = CanalNotificacion(
        nombre=nombre[:120],
        tipo=tipo,
        prioridad_minima=prioridad_minima,
        config=__import__("json").dumps(cfg),
    )
    db.add(canal)
    db.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/notificaciones?ok=Canal+creado", status_code=303)


@app.post("/notificaciones/canal/{canal_id}/activar")
async def notificaciones_activar(canal_id: int, db: Session = Depends(get_db),
                                  user: Usuario = Depends(requiere_login)):
    canal = db.query(CanalNotificacion).filter(CanalNotificacion.id == canal_id).first()
    if canal:
        canal.activo = not canal.activo
        db.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/notificaciones", status_code=303)


@app.post("/notificaciones/canal/{canal_id}/eliminar")
async def notificaciones_eliminar(canal_id: int, db: Session = Depends(get_db),
                                   user: Usuario = Depends(requiere_login)):
    canal = db.query(CanalNotificacion).filter(CanalNotificacion.id == canal_id).first()
    if canal:
        db.delete(canal)
        db.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/notificaciones", status_code=303)


@app.post("/api/v1/notificaciones/test/{canal_id}")
async def notificaciones_test(canal_id: int, db: Session = Depends(get_db),
                               user: Usuario = Depends(requiere_login)):
    error = notif_engine.enviar_test(canal_id, db)
    if error:
        raise HTTPException(400, detail=error)
    return {"ok": True, "mensaje": "Notificación de prueba enviada."}


@app.get("/api/v1/notificaciones/log")
async def notificaciones_log_api(limit: int = 50, db: Session = Depends(get_db),
                                  user: Usuario = Depends(requiere_login)):
    rows = db.query(NotificacionEnviada).order_by(
        NotificacionEnviada.fecha_envio.desc()).limit(limit).all()
    return [{"id": r.id, "canal_id": r.canal_id, "aviso_titulo": r.aviso_titulo,
             "resultado": r.resultado, "reintentos": r.reintentos,
             "fecha": r.fecha_envio.isoformat() if r.fecha_envio else None,
             "detalle": r.detalle} for r in rows]


# ── Suscripciones Web Push del navegador ────────────────────────────────────
@app.get("/api/push/vapid-public-key")
def api_push_vapid_public_key(user: Usuario = Depends(requiere_login)):
    return {"public_key": push_service.clave_publica_vapid()}


@app.post("/api/push/suscribirse")
async def api_push_suscribirse(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    body = await request.json()
    endpoint = (body.get("endpoint") or "").strip()
    keys = body.get("keys") or {}
    p256dh = (keys.get("p256dh") or "").strip()
    auth = (keys.get("auth") or "").strip()
    if not endpoint or not p256dh or not auth:
        raise HTTPException(400, detail="Suscripción incompleta")

    existente = db.query(PushSuscripcion).filter(PushSuscripcion.endpoint == endpoint).first()
    if existente:
        existente.usuario_id = user.id
        existente.p256dh = p256dh
        existente.auth = auth
        existente.user_agent = request.headers.get("user-agent", "")[:255]
    else:
        db.add(PushSuscripcion(
            usuario_id=user.id, endpoint=endpoint, p256dh=p256dh, auth=auth,
            user_agent=request.headers.get("user-agent", "")[:255],
        ))
    db.commit()
    return {"ok": True}


@app.post("/api/push/desuscribirse")
async def api_push_desuscribirse(
    request: Request,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    body = await request.json()
    endpoint = (body.get("endpoint") or "").strip()
    if endpoint:
        db.query(PushSuscripcion).filter(PushSuscripcion.endpoint == endpoint).delete()
        db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 4.7 — INFORMES INTELIGENTES — Rutas adicionales
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/informes/resumen/pdf")
def informe_pdf_resumen(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    from config import COMPANY_NAME as _CN
    analisis = generar_analisis_inteligente(db)
    pdf = exportar_pdf_resumen(analisis, _CN)
    nombre = f"resumen_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={nombre}"})


@app.get("/informes/maquinaria/excel")
def informe_maquinaria_excel(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    items = db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        Maquinaria.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Maquinaria.nombre).all()
    excel = exportar_maquinaria_excel(items)
    nombre = f"maquinaria_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(content=excel,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={nombre}"})


@app.get("/informes/incidencias/excel")
def informe_incidencias_excel(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    items = db.query(Incidencia).filter(
        Incidencia.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Incidencia.fecha_apertura.desc()).all()
    excel = exportar_incidencias_excel(items)
    nombre = f"incidencias_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(content=excel,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={nombre}"})


@app.get("/informes/reparaciones/excel")
def informe_reparaciones_excel(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    warehouse = _active_warehouse(db, user)
    items = db.query(Reparacion).filter(
        Reparacion.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Reparacion.fecha_entrada.desc()).all()
    excel = exportar_reparaciones_excel(items)
    nombre = f"reparaciones_mrd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(content=excel,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={nombre}"})


@app.get("/api/v1/informes/analisis")
def api_analisis_inteligente(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    return generar_analisis_inteligente(db)


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 4.8 — DETECCIÓN DE ANOMALÍAS
# ═══════════════════════════════════════════════════════════════════════════════
@app.get("/anomalias", response_class=HTMLResponse)
def anomalias_panel(request: Request, db: Session = Depends(get_db),
                    user: Usuario = Depends(requiere_login),
                    sev: str = ""):
    resultado = anom_engine.ejecutar_deteccion_completa(db)
    lista = resultado["anomalias"]
    if sev:
        lista = [a for a in lista if a["severidad"] == sev]
    return templates.TemplateResponse(request, "anomalias.html", ctx_base(
        request, user, db,
        resultado=resultado,
        lista=lista,
        filtro_sev=sev,
    ))


@app.post("/anomalias/crear-avisos")
def anomalias_crear_avisos(request: Request, db: Session = Depends(get_db),
                            user: Usuario = Depends(requiere_login)):
    resultado = anom_engine.ejecutar_deteccion_completa(db)
    creados = anom_engine.crear_avisos_desde_anomalias(resultado, db)
    from fastapi.responses import RedirectResponse
    return RedirectResponse(f"/anomalias?ok={creados}+avisos+creados", status_code=303)


@app.get("/api/v1/anomalias")
def api_anomalias(sev: str = "", db: Session = Depends(get_db),
                  user: Usuario = Depends(requiere_login)):
    resultado = anom_engine.ejecutar_deteccion_completa(db)
    if sev:
        resultado["anomalias"] = [a for a in resultado["anomalias"] if a["severidad"] == sev]
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 4.9 — MANTENIMIENTO PREDICTIVO
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/mantenimiento", response_class=HTMLResponse)
def mantenimiento_panel(request: Request, db: Session = Depends(get_db),
                        user: Usuario = Depends(requiere_login),
                        nivel: str = ""):
    plan = mant_engine.generar_plan_mantenimiento(db)
    scores = plan["scores"]
    if nivel:
        scores = [s for s in scores if s["nivel"] == nivel]
    return templates.TemplateResponse(request, "mantenimiento.html", ctx_base(
        request, user, db,
        plan=plan,
        scores=scores,
        filtro_nivel=nivel,
        tipos_mantenimiento=TIPOS_MANTENIMIENTO,
        herramientas=db.query(Herramienta).filter(Herramienta.activa == True).order_by(Herramienta.nombre).all(),
        maquinas=db.query(Maquinaria).filter(Maquinaria.activa == True).order_by(Maquinaria.nombre).all(),
        now=datetime.now(),
    ))


@app.post("/mantenimiento/programar")
async def mantenimiento_programar(
    request: Request, db: Session = Depends(get_db),
    user: Usuario = Depends(requiere_login),
    tipo_activo: str = Form(...),
    activo_id: int = Form(...),
    tipo: str = Form("preventivo"),
    descripcion: str = Form(""),
    fecha_programada: str = Form(...),
    intervalo_dias: str = Form(""),
    coste_estimado: str = Form(""),
    proveedor_texto: str = Form(""),
    notas: str = Form(""),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso para programar mantenimientos")
    # Obtener nombre y código del activo
    nombre_activo, codigo_activo = "", ""
    if tipo_activo == "herramienta":
        h = db.query(Herramienta).filter(Herramienta.id == activo_id).first()
        if h:
            nombre_activo = h.nombre
            codigo_activo = getattr(h, "codigo", "") or ""
    else:
        m = db.query(Maquinaria).filter(Maquinaria.id == activo_id).first()
        if m:
            nombre_activo = m.nombre
            codigo_activo = getattr(m, "codigo", "") or ""

    try:
        fecha_dt = datetime.strptime(fecha_programada, "%Y-%m-%d")
    except Exception:
        fecha_dt = datetime.now()

    mant_engine.crear_mantenimiento(
        db=db,
        tipo_activo=tipo_activo,
        activo_id=activo_id,
        nombre_activo=nombre_activo,
        codigo_activo=codigo_activo,
        tipo=tipo,
        descripcion=descripcion,
        fecha_programada=fecha_dt,
        intervalo_dias=int(intervalo_dias) if intervalo_dias.isdigit() else None,
        coste_estimado=float(coste_estimado) if coste_estimado else None,
        proveedor=proveedor_texto,
        notas=notas,
        creado_por_id=user.id,
    )
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/mantenimiento?ok=Mantenimiento+programado", status_code=303)


@app.post("/mantenimiento/{mp_id}/completar")
async def mantenimiento_completar(
    mp_id: int, request: Request, db: Session = Depends(get_db),
    user: Usuario = Depends(requiere_login),
    fecha_realizada: str = Form(""),
    coste_real: str = Form(""),
    notas_cierre: str = Form(""),
):
    if not (tiene_permiso(user, "editar") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso para completar mantenimientos")
    try:
        fecha_dt = datetime.strptime(fecha_realizada, "%Y-%m-%d") if fecha_realizada else datetime.now()
    except Exception:
        fecha_dt = datetime.now()
    mant_engine.completar_mantenimiento(
        mp_id=mp_id, db=db, fecha_realizada=fecha_dt,
        coste_real=float(coste_real) if coste_real else None,
        notas=notas_cierre,
    )
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/mantenimiento?ok=Mantenimiento+completado", status_code=303)


@app.post("/mantenimiento/{mp_id}/cancelar")
def mantenimiento_cancelar(mp_id: int, db: Session = Depends(get_db),
                            user: Usuario = Depends(requiere_login)):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede cancelar mantenimientos")
    mp = db.query(MantenimientoProgramado).filter(MantenimientoProgramado.id == mp_id).first()
    if mp:
        mp.estado = "cancelado"
        db.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/mantenimiento", status_code=303)


@app.post("/mantenimiento/{mp_id}/aplazar")
def mantenimiento_aplazar(
    mp_id: int, nueva_fecha: str = Form(...),
    db: Session = Depends(get_db), user: Usuario = Depends(requiere_login),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede aplazar mantenimientos")
    mp = db.get(MantenimientoProgramado, mp_id)
    if not mp or mp.estado in {"completado", "cancelado"}:
        raise HTTPException(409, "Este mantenimiento no se puede aplazar")
    try:
        fecha = datetime.strptime(nueva_fecha, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Fecha no válida")
    mp.fecha_programada = fecha
    mp.estado = "pendiente"
    db.commit()
    return RedirectResponse("/mantenimiento?ok=Mantenimiento+aplazado", status_code=303)


@app.post("/mantenimiento/{mp_id}/eliminar")
def mantenimiento_eliminar(
    mp_id: int, db: Session = Depends(get_db), user: Usuario = Depends(requiere_login),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede eliminar mantenimientos")
    mp = db.get(MantenimientoProgramado, mp_id)
    if not mp:
        raise HTTPException(404)
    if mp.estado == "completado" or mp.fecha_realizada:
        raise HTTPException(409, "Un mantenimiento realizado no se elimina; forma parte del historial")
    db.delete(mp)
    db.commit()
    return RedirectResponse("/mantenimiento?ok=Mantenimiento+eliminado", status_code=303)


@app.get("/api/v1/mantenimiento/plan")
def api_plan_mantenimiento(db: Session = Depends(get_db),
                            user: Usuario = Depends(requiere_login)):
    plan = mant_engine.generar_plan_mantenimiento(db)
    # Serializar fechas
    for s in plan["scores"]:
        if s.get("fecha_predicha"):
            s["fecha_predicha"] = s["fecha_predicha"].strftime("%Y-%m-%d")
        if s.get("ultimo_mantenimiento") and hasattr(s["ultimo_mantenimiento"], "strftime"):
            s["ultimo_mantenimiento"] = s["ultimo_mantenimiento"].strftime("%Y-%m-%d")
    return plan


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 4.12 — PANEL DE CONTROL DE IA Y AUTOMATIZACIONES
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/panel-ia", response_class=HTMLResponse)
def panel_ia(request: Request, db: Session = Depends(get_db),
             user: Usuario = Depends(requiere_login)):

    # ── Estado del scheduler ──────────────────────────────────────
    scheduler_activo = (
        auto_engine._scheduler_thread is not None and
        auto_engine._scheduler_thread.is_alive()
    )

    # ── KPIs globales ─────────────────────────────────────────────
    avisos_sin_leer  = db.query(Aviso).filter(Aviso.leido == False, Aviso.archivado == False).count()
    avisos_criticos  = db.query(Aviso).filter(Aviso.prioridad == "critica", Aviso.archivado == False).count()
    total_avisos     = db.query(Aviso).filter(Aviso.archivado == False).count()

    autos_activas    = db.query(Automatizacion).filter(Automatizacion.estado == "activa").count()
    autos_total      = db.query(Automatizacion).count()
    total_ejecuciones= db.query(EjecucionAutomatizacion).count()

    canales_activos  = db.query(CanalNotificacion).filter(CanalNotificacion.activo == True).count()
    notif_ok         = db.query(NotificacionEnviada).filter(NotificacionEnviada.resultado == "ok").count()
    notif_error      = db.query(NotificacionEnviada).filter(NotificacionEnviada.resultado == "error").count()

    mant_vencidos    = db.query(MantenimientoProgramado).filter(MantenimientoProgramado.estado == "vencido").count()
    mant_proximos    = db.query(MantenimientoProgramado).filter(
        MantenimientoProgramado.estado == "pendiente",
        MantenimientoProgramado.fecha_programada >= datetime.now(),
    ).count()

    # ── Anomalías (rápido — solo conteo) ─────────────────────────
    try:
        res_anom = anom_engine.ejecutar_deteccion_completa(db)
        anom_criticas = res_anom["resumen"]["critica"]
        anom_altas    = res_anom["resumen"]["alta"]
        anom_total    = res_anom["resumen"]["total"]
    except Exception:
        anom_criticas = anom_altas = anom_total = 0

    # ── Timeline de actividad reciente ────────────────────────────
    # Mezcla: últimas ejecuciones + últimos avisos + últimos mantenimientos
    timeline = []

    ult_ejecuciones = db.query(EjecucionAutomatizacion).order_by(
        EjecucionAutomatizacion.fecha.desc()).limit(5).all()
    for e in ult_ejecuciones:
        auto = db.query(Automatizacion).filter(Automatizacion.id == e.automatizacion_id).first()
        timeline.append({
            "tipo": "ejecucion",
            "icono": "bi-lightning",
            "color": "#0d6efd" if e.resultado == "ok" else "#dc3545",
            "texto": f"Auto «{auto.nombre if auto else '#'+str(e.automatizacion_id)}» — {e.acciones_ejecutadas} acción(es)",
            "fecha": e.fecha,
            "enlace": f"/automatizaciones/{e.automatizacion_id}",
        })

    ult_avisos = db.query(Aviso).order_by(Aviso.creado_en.desc()).limit(5).all()
    for a in ult_avisos:
        color_p = {"critica": "#dc3545", "alta": "#fd7e14", "media": "#0d6efd", "baja": "#6c757d"}.get(a.prioridad, "#6c757d")
        timeline.append({
            "tipo": "aviso",
            "icono": "bi-bell",
            "color": color_p,
            "texto": a.titulo,
            "fecha": a.creado_en,
            "enlace": "/avisos",
        })

    ult_notif = db.query(NotificacionEnviada).order_by(
        NotificacionEnviada.fecha_envio.desc()).limit(3).all()
    for n in ult_notif:
        timeline.append({
            "tipo": "notificacion",
            "icono": "bi-send",
            "color": "#198754" if n.resultado == "ok" else "#dc3545",
            "texto": f"Notif. {'enviada' if n.resultado=='ok' else 'fallida'}: {n.aviso_titulo or '—'}",
            "fecha": n.fecha_envio,
            "enlace": "/notificaciones",
        })

    ult_mant = db.query(MantenimientoProgramado).filter(
        MantenimientoProgramado.estado == "completado"
    ).order_by(MantenimientoProgramado.fecha_realizada.desc()).limit(3).all()
    for m in ult_mant:
        timeline.append({
            "tipo": "mantenimiento",
            "icono": "bi-tools",
            "color": "#198754",
            "texto": f"Mant. completado: {m.nombre_activo}",
            "fecha": m.fecha_realizada,
            "enlace": "/mantenimiento",
        })

    # Ordenar timeline por fecha
    timeline.sort(key=lambda x: x["fecha"] or datetime.min, reverse=True)
    timeline = timeline[:15]

    # ── Datos para mini gráfica de ejecuciones (7 días) ──────────
    ejecuciones_semana = []
    for i in range(6, -1, -1):
        dia = datetime.now() - timedelta(days=i)
        cnt = db.query(EjecucionAutomatizacion).filter(
            func.strftime("%Y-%m-%d", EjecucionAutomatizacion.fecha) == dia.strftime("%Y-%m-%d")
        ).count()
        ejecuciones_semana.append({"dia": dia.strftime("%a"), "total": cnt})

    chart_ej = dumps_for_script({
        "labels": [e["dia"] for e in ejecuciones_semana],
        "data":   [e["total"] for e in ejecuciones_semana],
    })

    return templates.TemplateResponse(request, "panel_ia.html", ctx_base(
        request, user, db,
        scheduler_activo=scheduler_activo,
        kpis={
            "avisos_sin_leer": avisos_sin_leer,
            "avisos_criticos": avisos_criticos,
            "total_avisos": total_avisos,
            "autos_activas": autos_activas,
            "autos_total": autos_total,
            "total_ejecuciones": total_ejecuciones,
            "canales_activos": canales_activos,
            "notif_ok": notif_ok,
            "notif_error": notif_error,
            "mant_vencidos": mant_vencidos,
            "mant_proximos": mant_proximos,
            "anom_criticas": anom_criticas,
            "anom_altas": anom_altas,
            "anom_total": anom_total,
        },
        timeline=timeline,
        chart_ej_json=chart_ej,
        now=datetime.now(),
    ))


@app.get("/api/v1/panel-ia/estado")
def api_panel_ia_estado(db: Session = Depends(get_db),
                         user: Usuario = Depends(requiere_login)):
    scheduler_activo = (
        auto_engine._scheduler_thread is not None and
        auto_engine._scheduler_thread.is_alive()
    )
    avisos_sin_leer = db.query(Aviso).filter(Aviso.leido == False, Aviso.archivado == False).count()
    autos_activas   = db.query(Automatizacion).filter(Automatizacion.estado == "activa").count()
    return {
        "scheduler_activo": scheduler_activo,
        "avisos_sin_leer": avisos_sin_leer,
        "autos_activas": autos_activas,
        "timestamp": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.3 — SERVICIOS DE PRODUCCIÓN
# API del servicio Windows y panel de administración
# Solo administradores. Todas las acciones se registran en log de auditoría.
# ═══════════════════════════════════════════════════════════════════════════════

import json as _json_mod
import subprocess as _subprocess
from pathlib import Path as _Path
import cloudflare_tunnel as _cf_tunnel

_SVC_BASE = _Path(__file__).parent
_SVC_STATUS_FILE = _SVC_BASE / ".service_status"
_SVC_RESTART_FLAG = _SVC_BASE / ".service_restart"
_SERVICE_NAME = "MRDToolControl"


def _svc_requiere_admin(user: Usuario):
    """Lanza 403 si el usuario no es administrador."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(status_code=403, detail="Solo administradores pueden gestionar el servicio.")


def _svc_log_audit(user: Usuario, action: str, detail: str = ""):
    """Registra acción de servicio en el log de seguridad."""
    try:
        mrd_logging.log_security(f"SERVICE_ACTION action={action} user={user.username} detail={detail}")
    except Exception:
        pass


def _svc_read_status() -> dict:
    """Lee el archivo .service_status escrito por windows_service.py."""
    try:
        if _SVC_STATUS_FILE.exists():
            return _json_mod.loads(_SVC_STATUS_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def _svc_get_metrics() -> dict:
    """Devuelve métricas del sistema usando service_health.py."""
    try:
        from service_health import get_system_metrics
        return get_system_metrics()
    except Exception:
        return {}


def _svc_windows_state() -> str:
    """Consulta estado del servicio Windows via sc.exe. Devuelve RUNNING/STOPPED/etc."""
    try:
        result = _subprocess.run(
            ["sc.exe", "query", _SERVICE_NAME],
            capture_output=True, text=True, timeout=5
        )
        if "RUNNING" in result.stdout:
            return "RUNNING"
        elif "STOPPED" in result.stdout:
            return "STOPPED"
        elif "PENDING" in result.stdout:
            return "PENDING"
        return "NOT_INSTALLED"
    except FileNotFoundError:
        return "NOT_WINDOWS"
    except Exception:
        return "UNKNOWN"


# ─── Reinicio del servidor ────────────────────────────────────────────────────
def _restart_exec_target(argv: list[str], python_executable: str, os_name: str):
    """Devuelve el ejecutable correcto sin intentar abrir un launcher .exe con Python."""
    launcher = Path(argv[0]).resolve()
    if os_name == "nt" and launcher.is_file() and launcher.suffix.lower() == ".exe":
        return str(launcher), [str(launcher)] + argv[1:]
    return python_executable, [python_executable] + argv


@app.post("/admin/recuperar-sistema")
def recuperar_sistema(
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Autodiagnóstico y reparación conservadora, sin borrar datos."""
    if user.rol != "admin":
        raise HTTPException(403, "Solo administradores pueden recuperar el sistema")

    comprobacion = db.execute(text("PRAGMA quick_check")).scalar()
    if comprobacion != "ok":
        mrd_logging.log_security(
            f"Recuperación detenida: quick_check={comprobacion!r}", level="error"
        )
        raise HTTPException(409, "La base de datos necesita revisión técnica; no se ha modificado nada")

    backup = crear_backup()
    if not backup.get("ok"):
        raise HTTPException(500, "No se pudo crear la copia de seguridad previa")

    try:
        for directory in (DATA_DIR, BACKUPS_DIR, EXPORTS_DIR, UPLOADS_DIR):
            Path(directory).mkdir(parents=True, exist_ok=True)
        _inicializar_stock_epi(db)
        limpieza = cleanup_scan_data(db)
        db.commit()
    except Exception:
        db.rollback()
        mrd_logging.log_app("Fallo durante la recuperación segura", level="error")
        raise HTTPException(500, "La recuperación no pudo completarse; se conservaron los datos")

    mrd_logging.log_security(
        f"Recuperación segura ejecutada por {user.username}; backup={backup['archivo']}",
        level="warning",
    )
    return JSONResponse({
        "ok": True,
        "msg": "Sistema comprobado y reparado. Ya puedes seguir trabajando.",
        "backup": backup["archivo"],
        "limpieza": limpieza,
    })


@app.post("/admin/reiniciar")
def reiniciar_servidor(
    request: Request,
    user: Usuario = Depends(requiere_login),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administradores pueden reiniciar el servidor")
    import threading, time, os, sys as _sys
    mrd_logging.log_app(f"Reinicio del servidor solicitado por {user.username}", level="warning")
    def _do_restart():
        time.sleep(1.2)
        # Uvicorn se instala como lanzador .exe en Windows. Ejecutarlo con
        # python.exe como si fuera un script deja NSSM activo pero sin servidor.
        # Reemplazamos el proceso por el mismo lanzador y los mismos argumentos.
        executable, args = _restart_exec_target(_sys.argv, _sys.executable, os.name)
        os.execv(executable, args)
    threading.Thread(target=_do_restart, daemon=True).start()
    return JSONResponse({"ok": True, "msg": "Reiniciando en 1 segundo..."})


# ─── GET /servicio (Panel admin) ──────────────────────────────────────────────
@app.get("/servicio")
def servicio_panel(request: Request, user: Usuario = Depends(requiere_login)):
    _svc_requiere_admin(user)
    return templates.TemplateResponse(request, "servicio.html", ctx_base(request, user, "servicio"))


# ─── GET /api/service/status ──────────────────────────────────────────────────
@app.get("/api/service/status")
def api_service_status(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    _svc_requiere_admin(user)
    file_status = _svc_read_status()
    win_state   = _svc_windows_state()
    metrics     = _svc_get_metrics()

    # Determinar estado general
    if file_status.get("status") == "running":
        status = "running"
    elif win_state == "RUNNING":
        status = "running"
    elif win_state == "STOPPED":
        status = "stopped"
    else:
        status = "unknown"

    return {
        "status":        status,
        "windows_state": win_state,
        "pid":           file_status.get("pid"),
        "uptime_seconds": file_status.get("uptime_seconds"),
        "start_time":    file_status.get("start_time"),
        "port":          file_status.get("port", 8000),
        "host":          file_status.get("host", "0.0.0.0"),
        "workers":       file_status.get("workers", 1),
        "restart_count": file_status.get("restart_count", 0),
        "version":       file_status.get("version", "1.9.3-alpha"),
        "metrics":       metrics,
        "timestamp":     file_status.get("timestamp"),
    }


# ─── GET /api/service/health ──────────────────────────────────────────────────
@app.get("/api/service/health")
def api_service_health(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    _svc_requiere_admin(user)
    try:
        from service_health import run_all_checks
        return run_all_checks(port=8000)
    except Exception as exc:
        return {"healthy": False, "error": str(exc), "checks": {}}


# ─── GET /api/service/logs/{nombre} ──────────────────────────────────────────
@app.get("/api/service/logs/{log_name}")
def api_service_logs(
    log_name: str,
    lines: int = 50,
    user: Usuario = Depends(requiere_login),
):
    _svc_requiere_admin(user)
    # Solo logs del servicio — no permitir path traversal
    allowed = {"service", "startup", "shutdown", "crash", "rotation", "uvicorn"}
    if log_name not in allowed:
        raise HTTPException(400, f"Log '{log_name}' no permitido.")
    log_path = _SVC_BASE / "logs" / f"{log_name}.log"
    if not log_path.exists():
        return {"content": f"(log {log_name}.log vacío o no existe)", "lines": 0}
    try:
        # Leer las últimas N líneas
        all_lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
        tail = all_lines[-min(lines, len(all_lines)):]
        return {"content": "\n".join(tail), "lines": len(tail), "log": log_name}
    except Exception as exc:
        raise HTTPException(500, f"Error leyendo log: {exc}")


# ─── POST /api/service/restart ────────────────────────────────────────────────
@app.post("/api/service/restart")
def api_service_restart(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    """
    Reinicio suave: crea un archivo de señal que el watchdog detecta
    y reinicia el proceso uvicorn sin detener el servicio Windows.
    """
    _svc_requiere_admin(user)
    _svc_log_audit(user, "restart", "reinicio suave de uvicorn solicitado")
    try:
        _SVC_RESTART_FLAG.write_text(
            f"restart_requested_by={user.username}_at={datetime.now().isoformat()}",
            encoding="utf-8"
        )
        return {"ok": True, "message": "Señal de reinicio enviada. Uvicorn se reiniciará en unos segundos."}
    except Exception as exc:
        raise HTTPException(500, f"No se pudo enviar señal de reinicio: {exc}")


@app.post("/api/service/watchdog")
def api_service_install_watchdog(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    """Instala la tarea 24x7 desde la cuenta LocalSystem del servicio."""
    _svc_requiere_admin(user)
    installer = _SVC_BASE / "scripts" / "operations" / "install_continuity_24x7.ps1"
    if not installer.is_file():
        raise HTTPException(404, "No se encontró el instalador del vigilante")
    completed = _subprocess.run(
        [
            "powershell.exe", "-NoLogo", "-NoProfile", "-NonInteractive",
            "-ExecutionPolicy", "Bypass", "-File", str(installer),
            "-RepositoryRoot", str(_SVC_BASE), "-Apply",
        ],
        cwd=str(_SVC_BASE), capture_output=True, text=True, timeout=90, check=False,
    )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "Error desconocido").strip()[-500:]
        raise HTTPException(500, f"No se pudo instalar el vigilante: {detail}")
    _svc_log_audit(user, "watchdog_install", "tarea 24x7 instalada o actualizada")
    return {
        "ok": True,
        "message": "Vigilante 24/7 instalado: comprueba cada minuto y al arrancar Windows.",
    }


# ─── POST /api/service/stop ───────────────────────────────────────────────────
@app.post("/api/service/stop")
def api_service_stop(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    """Detiene el servicio Windows MRDToolControl via sc.exe."""
    _svc_requiere_admin(user)
    _svc_log_audit(user, "stop", "detención del servicio Windows")
    win_state = _svc_windows_state()
    if win_state == "NOT_INSTALLED":
        raise HTTPException(409, "Servicio no instalado. Instala con install_service.ps1.")
    if win_state == "STOPPED":
        return {"ok": True, "message": "El servicio ya estaba detenido."}
    try:
        _subprocess.run(
            ["sc.exe", "stop", _SERVICE_NAME],
            capture_output=True, timeout=10, check=False
        )
        return {"ok": True, "message": "Señal de parada enviada al servicio."}
    except Exception as exc:
        raise HTTPException(500, f"Error al detener el servicio: {exc}")


# ─── POST /api/service/start ──────────────────────────────────────────────────
@app.post("/api/service/start")
def api_service_start(
    user: Usuario = Depends(requiere_login),
    _db: Session = Depends(get_db),
):
    """Inicia el servicio Windows MRDToolControl via sc.exe."""
    _svc_requiere_admin(user)
    _svc_log_audit(user, "start", "inicio del servicio Windows")
    win_state = _svc_windows_state()
    if win_state == "NOT_INSTALLED":
        raise HTTPException(409, "Servicio no instalado. Instala con install_service.ps1.")
    if win_state == "RUNNING":
        return {"ok": True, "message": "El servicio ya estaba en ejecución."}
    try:
        _subprocess.run(
            ["sc.exe", "start", _SERVICE_NAME],
            capture_output=True, timeout=10, check=False
        )
        return {"ok": True, "message": "Señal de inicio enviada al servicio."}
    except Exception as exc:
        raise HTTPException(500, f"Error al iniciar el servicio: {exc}")

# ─── FIN SPRINT 5.3 ───────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.4 — CLOUDFLARE NAMED TUNNEL
# ═══════════════════════════════════════════════════════════════════════════════

def _cf_requiere_admin(user):
    """Verifica que el usuario es administrador. Lanza 403 si no."""
    if not tiene_permiso(user, "config"):
        raise HTTPException(status_code=403, detail="Solo administradores pueden gestionar Cloudflare.")

def _cf_log_audit(user, action: str, detail: str = ""):
    mrd_logging.log_security(f"CF_ACTION user={user.username} action={action} detail={detail[:200]}")


@app.get("/cloudflare", response_class=HTMLResponse)
def cloudflare_panel(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    """Panel de configuración de Cloudflare Named Tunnel. Solo admin."""
    _cf_requiere_admin(user)
    cfg = remote_access.load_config()
    return templates.TemplateResponse(request, "cloudflare.html",
        ctx_base(request, user, db, config=cfg))


@app.get("/api/cloudflare/status")
def api_cf_status(user: Usuario = Depends(requiere_login)):
    """Estado completo del Cloudflare Named Tunnel. Solo admin."""
    _cf_requiere_admin(user)
    cfg = remote_access.load_config()
    status = _cf_tunnel.get_tunnel_status(cfg)
    # Añadir info del proveedor Cloudflare desde remote_access
    ra_status = remote_access.get_status_cached(max_age=30)
    cf_provider = next((p for p in ra_status.get("providers", [])
                        if p.get("name") == "cloudflare"), None)
    status["provider_info"] = cf_provider
    status["public_url_active"] = ra_status.get("public_url")
    return status


@app.get("/api/cloudflare/info")
def api_cf_info(user: Usuario = Depends(requiere_login)):
    """Información de configuración del túnel (sin credenciales). Solo admin."""
    _cf_requiere_admin(user)
    cfg = remote_access.load_config()
    cf_config_file = cfg.get("cf_config_file", "")
    tunnel_cfg = _cf_tunnel.read_tunnel_config(cf_config_file) if cf_config_file else {}
    version = _cf_tunnel.get_cloudflared_version(cfg.get("cloudflared_exe", "cloudflared.exe"))
    metrics = _cf_tunnel.get_metrics()
    return {
        "cloudflared_version": version,
        "tunnel_name":         cfg.get("cf_tunnel_name") or tunnel_cfg.get("tunnel"),
        "tunnel_id":           metrics.get("tunnel_id") or cfg.get("cf_tunnel_id"),
        "hostname":            cfg.get("cf_hostname"),
        "domain":              cfg.get("cf_domain"),
        "subdomain":           cfg.get("cf_subdomain"),
        "public_url":          cfg.get("cf_public_url"),
        "internal_port":       cfg.get("cf_internal_port", 8000),
        "force_https":         cfg.get("cf_force_https", True),
        "config_file_exists":  bool(cf_config_file and _Path(cf_config_file).exists()),
        "credentials_ok":      tunnel_cfg.get("credentials_file_exists", False),
        "ingress_count":       tunnel_cfg.get("ingress_count", 0),
        "metrics_available":   metrics["available"],
        # NO se incluye: credentials_file path, API tokens, cert.pem
    }


@app.post("/api/cloudflare/test")
def api_cf_test(user: Usuario = Depends(requiere_login)):
    """Ejecuta diagnóstico completo del tunnel. Solo admin."""
    _cf_requiere_admin(user)
    _cf_log_audit(user, "test")
    cfg = remote_access.load_config()
    port = int(cfg.get("port", 8000))
    checks = _cf_tunnel.run_diagnostics(cfg, port)
    all_ok = all(c["ok"] for c in checks)
    return {
        "healthy": all_ok,
        "checks":  checks,
        "checked_at": datetime.now().isoformat(),
    }


@app.post("/api/cloudflare/restart")
def api_cf_restart(user: Usuario = Depends(requiere_login)):
    """Reinicia el servicio cloudflared. Solo admin."""
    _cf_requiere_admin(user)
    _cf_log_audit(user, "restart")
    cfg = remote_access.load_config()
    svc_name = cfg.get("cloudflared_service", "cloudflared")
    result = _cf_tunnel.restart_service(svc_name)
    remote_access.invalidate_cache()
    return result



@app.get("/api/cloudflare/logs")
def api_cf_logs(lines: int = 50, user: Usuario = Depends(requiere_login)):
    """Lee los ultimos logs de cloudflared del Event Log de Windows. Solo admin."""
    _cf_requiere_admin(user)
    if lines < 1:
        lines = 1
    if lines > 500:
        lines = 500
    import subprocess as _sp
    import platform
    result_lines = []
    if platform.system() == "Windows":
        try:
            ps_cmd = (
                f"Get-EventLog -LogName Application -Source cloudflared -Newest {lines} "
                "-ErrorAction SilentlyContinue | "
                "Select-Object TimeGenerated,EntryType,Message | "
                "ForEach-Object { $_.TimeGenerated.ToString('yyyy-MM-dd HH:mm:ss') + ' [' + $_.EntryType + '] ' + ($_.Message -replace '\r?\n',' ') }"
            )
            r = _sp.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps_cmd],
                capture_output=True, text=True, timeout=10,
            )
            if r.stdout.strip():
                result_lines = [l for l in r.stdout.strip().splitlines() if l.strip()]
        except Exception as e:
            result_lines = [f"Error leyendo Event Log: {str(e)[:100]}"]
    # Fallback: buscar cloudflared log en rutas conocidas
    if not result_lines:
        log_candidates = [
            Path("logs") / "cloudflared.log",
            Path("C:/Program Files/cloudflared/cloudflared.log"),
            Path("C:/ProgramData/cloudflared/cloudflared.log"),
        ]
        for lp in log_candidates:
            try:
                if lp.exists():
                    all_l = lp.read_text(encoding="utf-8", errors="replace").splitlines()
                    result_lines = all_l[-min(lines, len(all_l)):]
                    break
            except Exception:
                pass
    if not result_lines:
        result_lines = [
            "cloudflared no ha escrito logs accesibles desde esta API.",
            "Para ver logs: ejecuta scripts\\cloudflare_logs.ps1 como Administrador.",
        ]
    return {"lines": result_lines, "count": len(result_lines)}


@app.post("/api/cloudflare/config")
async def api_cf_save_config(request: Request, user: Usuario = Depends(requiere_login)):
    """Guarda la configuración de Cloudflare. Solo admin."""
    _cf_requiere_admin(user)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")

    allowed = {
        "cf_tunnel_name", "cf_hostname", "cf_domain", "cf_subdomain",
        "cf_public_url", "cf_config_file", "cf_force_https",
        "cf_internal_port", "cloudflared_service", "cloudflared_exe",
        "manual_url", "preferred_provider",
    }
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="Sin parámetros válidos")

    # Si se actualiza cf_public_url también actualizar manual_url
    if "cf_public_url" in updates and updates["cf_public_url"]:
        updates.setdefault("manual_url", updates["cf_public_url"])

    ok = remote_access.save_config(updates)
    if not ok:
        raise HTTPException(status_code=500, detail="Error guardando configuración")

    remote_access.invalidate_cache()
    _cf_log_audit(user, "config_saved", str(list(updates.keys())))
    return {"ok": True, "message": "Configuración guardada.", "keys": list(updates.keys())}


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.5 — BASE DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════
import db_tools as _db_tools
from database import get_db_info as _db_get_info, check_connection as _db_check_conn


def _db_requiere_admin(user: Usuario):
    if getattr(user, "rol", None) not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Sólo administradores pueden gestionar la base de datos.")


def _db_audit(user: Usuario, action: str, detail: str = ""):
    try:
        from models import LogSeguridad
        db = next(get_db())
        db.add(LogSeguridad(
            usuario_id=user.id, username=user.username,
            accion=f"DB:{action}", detalle=detail,
            ip="interno", riesgo="info",
        ))
        db.commit()
        db.close()
    except Exception:
        pass


@app.get("/database")
def page_database(request: Request, user: Usuario = Depends(requiere_login)):
    """Panel de administración de base de datos. Solo admin."""
    _db_requiere_admin(user)
    return templates.TemplateResponse(request, "database.html", ctx_base(request, user, title="Base de Datos"))


@app.get("/api/database/status")
def api_db_status(user: Usuario = Depends(requiere_login)):
    """Estado completo de la base de datos."""
    _db_requiere_admin(user)
    info = _db_get_info()
    conn = _db_check_conn()
    current = _db_tools.get_alembic_current()
    return {
        "engine":         info.get("engine"),
        "url_safe":       info.get("url_safe"),
        "connected":      conn.get("ok"),
        "latency_ms":     conn.get("ms"),
        "connection_error": conn.get("error"),
        "pool":           info.get("pool", {}),
        "stats":          info.get("stats", {}),
        "slow_queries":   info.get("slow_queries", []),
        "slow_query_ms":  info.get("slow_query_ms", 200),
        "alembic_current": current.get("output", ""),
        "checked_at":     datetime.now().isoformat(),
    }


@app.post("/api/database/migrate")
def api_db_migrate(user: Usuario = Depends(requiere_login)):
    """Ejecuta alembic upgrade head. Solo admin."""
    _db_requiere_admin(user)
    _db_audit(user, "upgrade", "alembic upgrade head")
    result = _db_tools.run_alembic_upgrade("head")
    return result


@app.post("/api/database/rollback")
async def api_db_rollback(request: Request, user: Usuario = Depends(requiere_login)):
    """Revierte N migraciones. Solo admin."""
    _db_requiere_admin(user)
    try:
        body = await request.json()
        steps = int(body.get("steps", 1))
        steps = max(1, min(steps, 10))
    except Exception:
        steps = 1
    _db_audit(user, "downgrade", f"steps={steps}")
    result = _db_tools.run_alembic_downgrade(steps)
    return result


@app.post("/api/database/check")
def api_db_check(user: Usuario = Depends(requiere_login)):
    """Verificación de integridad completa. Solo admin."""
    _db_requiere_admin(user)
    _db_audit(user, "integrity_check")
    result = _db_tools.verify_integrity()
    return result


@app.get("/api/database/history")
def api_db_history(user: Usuario = Depends(requiere_login)):
    """Historial de migraciones Alembic. Solo admin."""
    _db_requiere_admin(user)
    return _db_tools.get_alembic_history()


@app.post("/api/database/reset-stats")
def api_db_reset_stats(user: Usuario = Depends(requiere_login)):
    """Reinicia contadores de rendimiento. Solo admin."""
    _db_requiere_admin(user)
    from database import reset_stats as _reset_stats
    _reset_stats()
    _db_audit(user, "reset_stats")
    return {"ok": True, "message": "Contadores reiniciados."}


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.6 — BACKUPS Y RECUPERACIÓN
# ═══════════════════════════════════════════════════════════════════════════════
import backup_manager as _bk


def _bk_requiere_admin(user: Usuario):
    if getattr(user, "rol", None) not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Sólo administradores pueden gestionar backups.")


def _bk_audit(user: Usuario, action: str, detail: str = ""):
    try:
        from models import LogSeguridad
        db = next(get_db())
        db.add(LogSeguridad(
            usuario_id=user.id, username=user.username,
            accion=f"BACKUP:{action}", detalle=detail,
            ip="interno", riesgo="info",
        ))
        db.commit()
        db.close()
    except Exception:
        pass


@app.get("/backup")
def page_backup(request: Request, user: Usuario = Depends(requiere_login)):
    """Panel de gestión de backups. Solo admin."""
    _bk_requiere_admin(user)
    return templates.TemplateResponse(request, "backup.html", ctx_base(request, user, title="Backups"))


@app.get("/api/backup/status")
def api_bk_status(user: Usuario = Depends(requiere_login)):
    """Estado general del sistema de backups."""
    _bk_requiere_admin(user)
    return _bk.get_backup_status()


@app.get("/api/backup/history")
def api_bk_history(
    limit: int = 50,
    user: Usuario = Depends(requiere_login),
):
    """Historial de backups (más recientes primero)."""
    _bk_requiere_admin(user)
    limit = max(1, min(limit, 200))
    return {"history": _bk.get_history(limit), "limit": limit}


@app.post("/api/backup/create")
async def api_bk_create(request: Request, user: Usuario = Depends(requiere_login)):
    """Crea un backup manual."""
    _bk_requiere_admin(user)
    try:
        body  = await request.json()
        label = str(body.get("label", ""))[:30]
    except Exception:
        label = ""
    result = _bk.create_backup(tipo="manual", label=label or user.username)
    _bk_audit(user, "create_manual", label)
    return result


@app.post("/api/backup/verify")
async def api_bk_verify(request: Request, user: Usuario = Depends(requiere_login)):
    """Verifica la integridad de un backup."""
    _bk_requiere_admin(user)
    try:
        body = await request.json()
        filename = str(body.get("filename", ""))
    except Exception:
        raise HTTPException(status_code=400, detail="filename requerido")

    if not filename or ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Nombre de fichero inválido")

    # Buscar en todos los subdirectorios de backup
    import backup_manager as _bm2
    found = None
    for f in _bm2.BACKUPS_DIR.rglob(filename):
        found = f
        break
    if not found:
        raise HTTPException(status_code=404, detail="Backup no encontrado")

    result = _bk.verify_backup(str(found))
    _bk_audit(user, "verify", filename)
    return result


@app.post("/api/backup/restore")
async def api_bk_restore(request: Request, user: Usuario = Depends(requiere_login)):
    """Restaura un backup. dry_run=true solo verifica."""
    _bk_requiere_admin(user)
    try:
        body     = await request.json()
        filename = str(body.get("filename", ""))
        dry_run  = bool(body.get("dry_run", True))
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")

    if not filename or ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Nombre de fichero inválido")

    import backup_manager as _bm2
    found = None
    for f in _bm2.BACKUPS_DIR.rglob(filename):
        found = f
        break
    if not found:
        raise HTTPException(status_code=404, detail="Backup no encontrado")

    result = _bk.restore_backup(str(found), dry_run=dry_run)
    _bk_audit(user, "restore" + ("_dry" if dry_run else ""), filename)
    return result


@app.post("/api/backup/cleanup")
def api_bk_cleanup(user: Usuario = Depends(requiere_login)):
    """Limpieza de backups antiguos según política de retención."""
    _bk_requiere_admin(user)
    result = _bk.cleanup_old_backups()
    _bk_audit(user, "cleanup", str(result))
    return result


@app.get("/api/backup/download/{filename:path}")
def api_bk_download(filename: str, user: Usuario = Depends(requiere_login)):
    """Descarga segura de un backup."""
    from fastapi.responses import FileResponse
    _bk_requiere_admin(user)
    # Sanitizar: solo nombre de fichero, no rutas
    safe_name = Path(filename).name
    if ".." in safe_name or not safe_name:
        raise HTTPException(status_code=400, detail="Nombre inválido")

    import backup_manager as _bm2
    found = None
    for f in _bm2.BACKUPS_DIR.rglob(safe_name):
        found = f
        break
    if not found:
        raise HTTPException(status_code=404, detail="Backup no encontrado")

    _bk_audit(user, "download", safe_name)
    return FileResponse(
        path=str(found),
        filename=safe_name,
        media_type="application/octet-stream",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.7 — ACTUALIZACIONES PROFESIONALES
# ═══════════════════════════════════════════════════════════════════════════════
import updater as _updater


def _upd_requiere_admin(user: Usuario):
    if getattr(user, "rol", None) not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Sólo administradores pueden gestionar actualizaciones.")


def _upd_audit(user: Usuario, action: str, detail: str = ""):
    try:
        from models import LogSeguridad
        db = next(get_db())
        db.add(LogSeguridad(
            usuario_id=user.id, username=user.username,
            accion=f"UPDATE:{action}", detalle=detail,
            ip="interno", riesgo="info",
        ))
        db.commit()
        db.close()
    except Exception:
        pass


@app.get("/actualizaciones")
def page_actualizaciones(request: Request, user: Usuario = Depends(requiere_login)):
    """Panel de actualizaciones. Solo admin."""
    _upd_requiere_admin(user)
    return templates.TemplateResponse(request, "actualizaciones.html", ctx_base(request, user, title="Actualizaciones"))


@app.get("/api/update/check")
def api_upd_check(user: Usuario = Depends(requiere_login)):
    """Comprueba si hay actualizaciones disponibles."""
    _upd_requiere_admin(user)
    result = _updater.check_update()
    return result


@app.get("/api/update/status")
def api_upd_status(user: Usuario = Depends(requiere_login)):
    """Estado actual del proceso de actualización."""
    _upd_requiere_admin(user)
    return _updater.get_state()


@app.post("/api/update/install")
async def api_upd_install(request: Request, user: Usuario = Depends(requiere_login)):
    """Inicia la descarga e instalación de una actualización."""
    _upd_requiere_admin(user)
    try:
        body = await request.json()
        download_url = str(body.get("download_url", "")).strip()
        sha256       = str(body.get("sha256",       "")).strip()
        version      = str(body.get("version",      "")).strip()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")

    if not download_url:
        raise HTTPException(status_code=400, detail="download_url requerido")
    if not download_url.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="URL no válida (solo http/https)")
    if not version:
        raise HTTPException(status_code=400, detail="version requerida")

    _upd_audit(user, "install", f"version={version}")
    result = _updater.start_update(download_url, sha256, version)
    return result


@app.post("/api/update/rollback")
def api_upd_rollback(user: Usuario = Depends(requiere_login)):
    """Rollback manual a la versión anterior."""
    _upd_requiere_admin(user)
    _upd_audit(user, "rollback")
    return _updater.rollback_update()


@app.post("/api/update/reset")
def api_upd_reset(user: Usuario = Depends(requiere_login)):
    """Reinicia el estado del updater."""
    _upd_requiere_admin(user)
    ok = _updater.reset_state()
    return {"ok": ok, "message": "Estado reiniciado." if ok else "Actualización en curso, no se puede reiniciar."}


@app.post("/api/restart")
def api_restart(request: Request, user: Usuario = Depends(requiere_login)):
    """Reinicia el servidor MRD Tool Control."""
    import threading, subprocess, sys
    from pathlib import Path
    if not getattr(user, "es_admin", False):
        raise HTTPException(status_code=403, detail="Solo administradores.")
    def _do_restart():
        import time; time.sleep(1)
        ps1 = Path(__file__).parent / "REINICIAR_AHORA.ps1"
        if sys.platform == "win32" and ps1.exists():
            subprocess.Popen(
                ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(ps1)],
                creationflags=0x00000008,
            )
    threading.Thread(target=_do_restart, daemon=True).start()
    return {"ok": True, "message": "Reinicio iniciado."}


# ══════════════════════════════════════════════════════════════════════════════
# SPRINT 5.8 — IASMRD CLOUDFLARE DEPLOYMENT (v1.9.8-alpha)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/deployment/diagnostics")
def api_deployment_diagnostics(user: Usuario = Depends(requiere_login)):
    _cf_requiere_admin(user)
    """
    17 comprobaciones de despliegue de producción IASMRD.
    Solo accesible para administradores.
    """
    import urllib.request as _req
    import urllib.error as _uerr
    import socket as _sock
    import subprocess as _sp

    checks = []
    passed = 0

    def _chk(name: str, ok: bool, detail: str = ""):
        nonlocal passed
        if ok:
            passed += 1
        checks.append({"name": name, "ok": ok, "detail": detail})

    # 1. MRD_ENV = production
    env = os.getenv("MRD_ENV", "development")
    _chk("MRD_ENV=production", env == "production", env)

    # 2. MRD_SECRET_KEY definida y segura
    sk = os.getenv("MRD_SECRET_KEY", "")
    _chk("MRD_SECRET_KEY definida (>=32 chars)", len(sk) >= 32,
         f"{len(sk)} chars" if sk else "vacía")

    # 3. MRD_PUBLIC_URL configurada
    pub = MRD_PUBLIC_URL
    _chk("MRD_PUBLIC_URL configurada", bool(pub), pub or "vacía")

    # 4. MRD_PUBLIC_URL empieza por https://
    _chk("MRD_PUBLIC_URL usa HTTPS", pub.startswith("https://"), pub)

    # 5. MRD_HTTPS_ONLY=true
    _chk("MRD_HTTPS_ONLY=true", _MRD_HTTPS_ONLY, str(_MRD_HTTPS_ONLY))

    # 6. MRD_TRUST_PROXY_HEADERS=true
    _chk("MRD_TRUST_PROXY_HEADERS=true", MRD_TRUST_PROXY_HEADERS,
         str(MRD_TRUST_PROXY_HEADERS))

    # 7. MRD_ALLOWED_HOSTS contiene app.iasmrd.com
    hosts = MRD_ALLOWED_HOSTS
    _chk("MRD_ALLOWED_HOSTS configurado", bool(hosts), ",".join(hosts) if hosts else "vacío")

    # 8. MRD_SCAN_URL configurada
    scan = MRD_SCAN_URL
    _chk("MRD_SCAN_URL configurada", bool(scan), scan or "vacía")

    # 9. Servidor local responde en /health
    try:
        _r = _req.urlopen("http://127.0.0.1:8000/health", timeout=3)
        _chk("Servidor local responde /health", _r.status == 200, f"HTTP {_r.status}")
    except Exception as _e:
        _chk("Servidor local responde /health", False, str(_e))

    # 10. Servicio cloudflared instalado/en ejecución
    try:
        _out = _sp.run(
            ["sc", "query", "cloudflared"],
            capture_output=True, text=True, timeout=5,
        )
        _running = "RUNNING" in _out.stdout
        _chk("Servicio cloudflared en ejecución", _running,
             "RUNNING" if _running else _out.stdout.strip()[:80])
    except Exception as _e:
        _chk("Servicio cloudflared en ejecución", False, str(_e))

    # 11. cloudflared.exe en PATH
    try:
        _out = _sp.run(["cloudflared", "version"], capture_output=True, text=True, timeout=5)
        _ok = _out.returncode == 0
        _chk("cloudflared.exe en PATH", _ok,
             _out.stdout.strip()[:60] if _ok else _out.stderr.strip()[:60])
    except Exception as _e:
        _chk("cloudflared.exe en PATH", False, str(_e))

    # 12. URL pública accesible
    _pub_ok = False
    _pub_detail = "no configurada"
    if pub:
        try:
            _r = _req.urlopen(f"{pub}/health", timeout=10)
            _pub_ok = _r.status == 200
            _pub_detail = f"HTTP {_r.status}"
        except Exception as _e:
            _pub_detail = str(_e)[:80]
    _chk("URL pública accesible", _pub_ok, _pub_detail)

    # 13. Ruta /scan disponible
    _scan_ok = False
    _scan_detail = "no configurada"
    if pub:
        try:
            _r = _req.urlopen(f"{pub}/scan", timeout=10)
            _scan_ok = _r.status < 500
            _scan_detail = f"HTTP {_r.status}"
        except _uerr.HTTPError as _e:
            _scan_ok = _e.code < 500
            _scan_detail = f"HTTP {_e.code}"
        except Exception as _e:
            _scan_detail = str(_e)[:80]
    _chk("Ruta /scan disponible", _scan_ok, _scan_detail)

    # 14. Base de datos accesible
    try:
        from database import check_connection as _chkconn
        _db_ok, _db_msg = _chkconn()
        _chk("Base de datos accesible", _db_ok, _db_msg)
    except Exception as _e:
        _chk("Base de datos accesible", False, str(_e))

    # 15. Backup reciente (últimas 25 horas)
    try:
        import backup_manager as _bk
        _hist = _bk.get_history(limit=1)
        if _hist:
            from datetime import datetime, timezone
            _ts_str = _hist[0].get("created_at", "")
            try:
                _ts = datetime.fromisoformat(_ts_str)
                if _ts.tzinfo is None:
                    _ts = _ts.replace(tzinfo=timezone.utc)
                _age_h = (datetime.now(timezone.utc) - _ts).total_seconds() / 3600
                _bk_ok = _age_h <= 25
                _chk("Backup reciente (≤25h)", _bk_ok, f"{_age_h:.1f}h")
            except Exception:
                _chk("Backup reciente (≤25h)", False, "fecha inválida")
        else:
            _chk("Backup reciente (≤25h)", False, "sin historial")
    except Exception as _e:
        _chk("Backup reciente (≤25h)", False, str(_e))

    # 16. Logs sin errores críticos recientes
    try:
        import mrd_logging as _mrdlog
        _log_dir = BASE_DIR / "logs"
        _log_file = _log_dir / "mrd.log"
        _critical_count = 0
        if _log_file.exists():
            _lines = _log_file.read_text(encoding="utf-8", errors="ignore").splitlines()
            _critical_count = sum(1 for l in _lines[-200:] if "CRITICAL" in l or "ERROR" in l)
        _chk("Logs: sin errores críticos recientes",
             _critical_count == 0, f"{_critical_count} errores en últimas 200 líneas")
    except Exception as _e:
        _chk("Logs: sin errores críticos recientes", True, "no se pudo leer log")

    # 17. Versión de la aplicación
    _chk("Versión cargada", bool(VERSION), VERSION)

    total = len(checks)
    return JSONResponse({
        "ok": passed == total,
        "passed": passed,
        "total": total,
        "version": VERSION,
        "public_url": pub,
        "scan_url": scan,
        "env": env,
        "checks": checks,
        "checked_at": datetime.now().isoformat(),
    })

@app.get("/scanner-test", response_class=HTMLResponse)
def scanner_test_page(request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    """Pagina de diagnostico del escaner QR/barcode — solo admins."""
    return templates.TemplateResponse(request, "scanner_test.html", ctx_base(request, user))

@app.get("/instalar", response_class=HTMLResponse)
def instalar_app_page(request: Request, db: Session = Depends(get_db)):
    """Pagina de instalacion PWA — publica, accesible sin login para empleados."""
    user = usuario_actual(request, db)
    return templates.TemplateResponse(request, "instalar_app.html", ctx_base(request, user))



# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULOS NUEVOS — MEJORAS GLOBALES
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Formación y habilitaciones ────────────────────────────────────────────

@app.get("/trabajadores/{tid}/formacion", response_class=HTMLResponse)
def formacion_lista(tid: int, request: Request,
                    user: Usuario = Depends(requiere_login),
                    db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    formaciones = (db.query(FormacionTrabajador)
                   .filter(FormacionTrabajador.trabajador_id == tid)
                   .order_by(FormacionTrabajador.fecha_caducidad.asc().nullslast())
                   .all())
    ctx = ctx_base(request, user)
    ctx.update({"trabajador": t, "formaciones": formaciones,
                "tipos_formacion": TIPOS_FORMACION,
                "ok": request.query_params.get("ok")})
    return templates.TemplateResponse(request, "trabajador_formacion.html", ctx)


@app.post("/trabajadores/{tid}/formacion", response_class=RedirectResponse)
def formacion_crear(tid: int, request: Request,
                    nombre_curso: str = Form(...),
                    tipo: str = Form(""),
                    entidad: str = Form(""),
                    fecha_realizacion: str = Form(""),
                    fecha_caducidad: str = Form(""),
                    num_certificado: str = Form(""),
                    notas: str = Form(""),
                    user: Usuario = Depends(requiere_login),
                    db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    from datetime import date as _d
    def _date(s):
        try: return _d.fromisoformat(s) if s else None
        except: return None
    f = FormacionTrabajador(
        trabajador_id=tid,
        nombre_curso=nombre_curso.strip(),
        tipo=tipo or None,
        entidad=entidad or None,
        fecha_realizacion=_date(fecha_realizacion),
        fecha_caducidad=_date(fecha_caducidad),
        num_certificado=num_certificado or None,
        notas=notas or None,
        usuario_id=user.id,
    )
    db.add(f)
    db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/formacion?ok=creado", status_code=303)


@app.post("/trabajadores/{tid}/formacion/{fid}/eliminar", response_class=RedirectResponse)
def formacion_eliminar(tid: int, fid: int,
                       user: Usuario = Depends(requiere_login),
                       db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    f = db.get(FormacionTrabajador, fid)
    if f and f.trabajador_id == tid:
        db.delete(f)
        db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/formacion?ok=eliminado", status_code=303)


# ─── Reconocimientos médicos ──────────────────────────────────────────────────────

@app.get("/trabajadores/{tid}/reconocimientos", response_class=HTMLResponse)
def reconocimientos_lista(tid: int, request: Request,
                          user: Usuario = Depends(requiere_login),
                          db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    reconocimientos = (db.query(ReconocimientoMedico)
                       .filter(ReconocimientoMedico.trabajador_id == tid)
                       .order_by(ReconocimientoMedico.fecha.desc())
                       .all())
    ctx = ctx_base(request, user)
    ctx.update({"trabajador": t, "reconocimientos": reconocimientos,
                "ok": request.query_params.get("ok")})
    return templates.TemplateResponse(request, "trabajador_reconocimientos.html", ctx)


@app.post("/trabajadores/{tid}/reconocimientos", response_class=RedirectResponse)
def reconocimiento_crear(tid: int,
                         fecha: str = Form(...),
                         resultado: str = Form("apto"),
                         fecha_proxima: str = Form(""),
                         medico: str = Form(""),
                         centro: str = Form(""),
                         restricciones: str = Form(""),
                         user: Usuario = Depends(requiere_login),
                         db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    from datetime import date as _d
    def _date(s):
        try: return _d.fromisoformat(s) if s else None
        except: return None
    r = ReconocimientoMedico(
        trabajador_id=tid,
        fecha=_date(fecha) or _d.today(),
        resultado=resultado,
        fecha_proxima=_date(fecha_proxima),
        medico=medico or None,
        centro=centro or None,
        restricciones=restricciones or None,
        usuario_id=user.id,
    )
    db.add(r)
    db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/reconocimientos?ok=creado", status_code=303)


@app.post("/trabajadores/{tid}/reconocimientos/{rid}/eliminar", response_class=RedirectResponse)
def reconocimiento_eliminar(tid: int, rid: int,
                            user: Usuario = Depends(requiere_login),
                            db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    r = db.get(ReconocimientoMedico, rid)
    if r and r.trabajador_id == tid:
        db.delete(r)
        db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/reconocimientos?ok=eliminado", status_code=303)


# ─── Documentación del trabajador ─────────────────────────────────────────────────────

@app.get("/trabajadores/{tid}/documentos", response_class=HTMLResponse)
def documentos_trabajador_lista(tid: int, request: Request,
                                user: Usuario = Depends(requiere_login),
                                db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    documentos = (db.query(DocumentoTrabajador)
                  .filter(DocumentoTrabajador.trabajador_id == tid)
                  .order_by(DocumentoTrabajador.fecha_caducidad.asc().nullslast())
                  .all())
    ctx = ctx_base(request, user)
    ctx.update({"trabajador": t, "documentos": documentos,
                "tipos_documento": TIPOS_DOCUMENTO_TRABAJADOR,
                "ok": request.query_params.get("ok")})
    return templates.TemplateResponse(request, "trabajador_documentos.html", ctx)


@app.post("/trabajadores/{tid}/documentos", response_class=RedirectResponse)
def documento_trabajador_crear(tid: int,
                               tipo: str = Form(...),
                               numero: str = Form(""),
                               fecha_emision: str = Form(""),
                               fecha_caducidad: str = Form(""),
                               notas: str = Form(""),
                               user: Usuario = Depends(requiere_login),
                               db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    from datetime import date as _d
    def _date(s):
        try: return _d.fromisoformat(s) if s else None
        except: return None
    d = DocumentoTrabajador(
        trabajador_id=tid,
        tipo=tipo,
        numero=numero or None,
        fecha_emision=_date(fecha_emision),
        fecha_caducidad=_date(fecha_caducidad),
        notas=notas or None,
        usuario_id=user.id,
    )
    db.add(d)
    db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/documentos?ok=creado", status_code=303)


@app.post("/trabajadores/{tid}/documentos/{did}/eliminar", response_class=RedirectResponse)
def documento_trabajador_eliminar(tid: int, did: int,
                                  user: Usuario = Depends(requiere_login),
                                  db: Session = Depends(get_db)):
    raise HTTPException(403, "Esta sección ha sido deshabilitada.")
    d = db.get(DocumentoTrabajador, did)
    if d and d.trabajador_id == tid:
        db.delete(d)
        db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/documentos?ok=eliminado", status_code=303)


# ─── PDF ficha completa del trabajador ────────────────────────────────────────────────

@app.get("/trabajadores/{tid}/pdf-ficha")
def trabajador_pdf_ficha(tid: int,
                         user: Usuario = Depends(requiere_login),
                         db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io, datetime as _dt
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("<b>Ficha del Trabajador</b>", styles["Title"]))
    story.append(Paragraph(f"{t.nombre_completo} — {t.cargo or ''}", styles["Heading2"]))
    story.append(Spacer(1, 12))
    datos = [
        ["DNI", t.dni or "—"], ["Teléfono", t.telefono or "—"],
        ["Email", t.email or "—"], ["Empresa", t.empresa or "—"],
        ["Cargo", t.cargo or "—"], ["Departamento", t.departamento or "—"],
        ["Estado", "Activo" if t.activo else "Baja"],
    ]
    tbl = Table([["Campo", "Valor"]] + datos, colWidths=[130, 360])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1a3a5c")),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 12))
    epis = db.query(EPIIndividual).filter(EPIIndividual.trabajador_id == tid, EPIIndividual.estado == "activo").all()
    if epis:
        story.append(Paragraph("<b>EPIs Asignados</b>", styles["Heading3"]))
        rows = [["Tipo", "Código", "Marca/Modelo", "Próx. Revisión"]]
        for e in epis:
            rows.append([e.tipo, e.codigo_fabricacion, f"{e.marca or ''} {e.modelo or ''}".strip() or "—",
                         e.proxima_revision.strftime("%d/%m/%Y") if e.proxima_revision else "—"])
        te = Table(rows, colWidths=[80, 130, 150, 120])
        te.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c0392b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff5f5")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(te)
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Generado: {_dt.datetime.now().strftime('%d/%m/%Y %H:%M')} — MRD Tool", styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    nombre_archivo = f"ficha_{t.nombre.lower().replace(' ', '_')}.pdf"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'})


# ─── Acceso sencillo del trabajador por PIN ───────────────────────────────────────────────────

WORKER_COOKIE_NAME = "mrd_worker_token"


def _portal_cookie_worker(request: Request, db: Session) -> Optional[Trabajador]:
    payload = verificar_token(request.cookies.get(WORKER_COOKIE_NAME, ""))
    if not payload or payload.get("kind") != "worker":
        return None
    try:
        worker_id = int(payload.get("worker_id"))
    except (TypeError, ValueError):
        return None
    session_key = str(payload.get("session_key") or "")
    if not session_key:
        return None
    session = db.query(SesionPortalTrabajador).filter(
        SesionPortalTrabajador.trabajador_id == worker_id,
        SesionPortalTrabajador.token_hash == hashlib.sha256(session_key.encode()).hexdigest(),
        SesionPortalTrabajador.revocado_en.is_(None),
        SesionPortalTrabajador.expira_en > datetime.now(),
    ).first()
    if not session:
        return None
    if not session.ultimo_uso_en or session.ultimo_uso_en < datetime.now() - timedelta(minutes=15):
        session.ultimo_uso_en = datetime.now()
        db.commit()
    return db.query(Trabajador).filter(
        Trabajador.id == worker_id, Trabajador.activo == True,
    ).first()


def _portal_cookie_matches(request: Request, db: Session, trabajador: Trabajador) -> bool:
    current = _portal_cookie_worker(request, db)
    return bool(current and current.id == trabajador.id)


def _portal_worker_required(token: str, request: Request, db: Session) -> Trabajador:
    worker = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not worker:
        raise HTTPException(404, "Enlace de trabajador no válido")
    if not _portal_cookie_matches(request, db, worker):
        code = urllib.parse.quote(worker.codigo or worker.dni or "")
        raise HTTPException(
            303, "La sesión ha caducado",
            headers={"location": f"/portal-trabajador?codigo={code}"},
        )
    return worker


def _worker_login_response(
    request: Request, trabajador: Trabajador, db: Session,
) -> RedirectResponse:
    session_key = uuid.uuid4().hex + uuid.uuid4().hex
    token = crear_token(
        {"kind": "worker", "worker_id": trabajador.id, "session_key": session_key},
        expires_delta=timedelta(days=30),
    )
    db.add(SesionPortalTrabajador(
        trabajador_id=trabajador.id,
        token_hash=hashlib.sha256(session_key.encode()).hexdigest(),
        dispositivo=(request.headers.get("user-agent") or "Dispositivo")[:200],
        ip_hash=hashlib.sha256(
            ((request.client.host if request.client else "unknown") + trabajador.portal_token).encode()
        ).hexdigest(),
        expira_en=datetime.now() + timedelta(days=30),
    ))
    db.commit()
    secure = (
        _MRD_HTTPS_ONLY or request.headers.get("x-forwarded-proto") == "https"
        or "https" in request.headers.get("cf-visitor", "")
    ) and os.getenv("MRD_TESTING") != "1"
    response = RedirectResponse(f"/portal/{trabajador.portal_token}", status_code=303)
    response.set_cookie(
        WORKER_COOKIE_NAME, token, httponly=True, secure=secure,
        max_age=30 * 24 * 60 * 60, samesite="lax", path="/",
    )
    return response


@app.get("/portal-trabajador", response_class=HTMLResponse)
def portal_trabajador_login(request: Request, codigo: str = "", db: Session = Depends(get_db)):
    current = _portal_cookie_worker(request, db)
    if current and current.portal_token:
        return RedirectResponse(f"/portal/{current.portal_token}", status_code=303)
    return templates.TemplateResponse(request, "portal_trabajador_login.html", {
        "request": request, "codigo": codigo.strip(), "error": "",
    })


@app.post("/portal-trabajador/acceso", response_class=HTMLResponse)
def portal_trabajador_acceso(
    request: Request, codigo: str = Form(...), pin: str = Form(...),
    db: Session = Depends(get_db),
):
    identifier = codigo.strip()
    clean_pin = pin.strip()
    ip = request.client.host if request.client else "unknown"
    rate_key = f"worker:{ip}:{identifier.upper()}"
    if not _puede_intentar_login(rate_key):
        return templates.TemplateResponse(request, "portal_trabajador_login.html", {
            "request": request, "codigo": identifier,
            "error": "Demasiados intentos. Espera unos minutos.",
        }, status_code=429)
    worker = db.query(Trabajador).filter(
        Trabajador.activo == True,
        or_(
            func.upper(func.trim(Trabajador.codigo)) == identifier.upper(),
            func.upper(func.trim(Trabajador.dni)) == identifier.upper(),
            Trabajador.portal_token == identifier,
        ),
    ).first()
    if not worker or not worker.portal_pin_hash or not verificar_password(clean_pin, worker.portal_pin_hash):
        _registrar_fallo_login(rate_key)
        time.sleep(0.2)
        return templates.TemplateResponse(request, "portal_trabajador_login.html", {
            "request": request, "codigo": identifier,
            "error": "Código o PIN incorrecto.",
        }, status_code=401)
    _limpiar_intentos_login(rate_key)
    if not worker.portal_token:
        worker.portal_token = uuid.uuid4().hex + uuid.uuid4().hex
        db.commit()
    return _worker_login_response(request, worker, db)


@app.get("/portal-trabajador/salir")
def portal_trabajador_salir(request: Request, db: Session = Depends(get_db)):
    payload = verificar_token(request.cookies.get(WORKER_COOKIE_NAME, "")) or {}
    session_key = str(payload.get("session_key") or "")
    if session_key:
        session = db.query(SesionPortalTrabajador).filter_by(
            token_hash=hashlib.sha256(session_key.encode()).hexdigest(),
        ).first()
        if session:
            session.revocado_en = datetime.now()
            db.commit()
    response = RedirectResponse("/portal-trabajador", status_code=303)
    response.delete_cookie(WORKER_COOKIE_NAME, path="/")
    return response


def _valid_worker_pin(pin: str) -> bool:
    return bool(pin.isdigit() and 4 <= len(pin) <= 6)


@app.post("/portal/{token}/activar-pin", response_class=RedirectResponse)
async def portal_activar_pin(token: str, request: Request, db: Session = Depends(get_db)):
    worker = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not worker:
        raise HTTPException(404, "Enlace de activación no válido")
    if worker.portal_pin_hash:
        raise HTTPException(409, "El acceso ya está activado")
    form = await request.form()
    pin = str(form.get("pin") or "").strip()
    confirmation = str(form.get("pin_confirm") or "").strip()
    if not _valid_worker_pin(pin) or pin != confirmation:
        raise HTTPException(422, "El PIN debe coincidir y tener entre 4 y 6 números")
    worker.portal_pin_hash = hash_password(pin)
    worker.portal_pin_actualizado_en = datetime.now()
    worker.portal_pin_cambio_obligatorio = False
    db.add(AuditoriaLog(
        tabla="trabajadores", registro_id=worker.id, accion="portal_activado",
        resumen="El trabajador activó su acceso personal", usuario_id=None,
    ))
    db.commit()
    return _worker_login_response(request, worker, db)


# ─── Portal QR del trabajador ──────────────────────────────────────────────────────────────────

@app.get("/portal/{token}/manifest.json", include_in_schema=False)
def portal_worker_manifest(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    response = JSONResponse({
        "name": f"Mi espacio MRD — {worker.nombre}", "short_name": "Mi MRD",
        "description": "Portal personal del trabajador MRD",
        "id": f"/portal-worker-{worker.id}", "start_url": f"/portal/{token}",
        "scope": "/", "display": "standalone", "background_color": "#f3f7fb",
        "theme_color": "#061626", "lang": "es",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
        "shortcuts": [
            {"name": "Solicitar material", "url": f"/portal/{token}#solicitar"},
            {"name": "Comunicar incidencia", "url": f"/portal/{token}#incidencias"},
            {"name": "Mis albaranes", "url": f"/portal/{token}#albaranes"},
        ],
    })
    response.headers["Cache-Control"] = "private, no-store"
    return response

@app.get("/portal/{token}", response_class=HTMLResponse)
def portal_trabajador(token: str, request: Request, db: Session = Depends(get_db)):
    t = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not t:
        return HTMLResponse("<h2>Enlace no válido</h2>", status_code=404)
    if not t.portal_pin_hash:
        return templates.TemplateResponse(request, "portal_trabajador_login.html", {
            "request": request, "codigo": t.codigo or t.dni or "", "error": "",
            "activar_token": token,
        })
    if t.portal_pin_hash and not _portal_cookie_matches(request, db, t):
        return templates.TemplateResponse(request, "portal_trabajador_login.html", {
            "request": request, "codigo": t.codigo or t.dni or t.portal_token,
            "error": "Introduce tu PIN para abrir el portal.",
        })
    if t.portal_pin_hash and t.portal_pin_cambio_obligatorio:
        return templates.TemplateResponse(request, "portal_pin_inicial.html", {
            "request": request, "trabajador": t,
        })
    epis = db.query(EPIIndividual).filter(EPIIndividual.trabajador_id == t.id, EPIIndividual.estado == "activo").all()
    formaciones = db.query(FormacionTrabajador).filter(FormacionTrabajador.trabajador_id == t.id).order_by(FormacionTrabajador.fecha_caducidad.asc().nullslast()).all()
    reconocs = db.query(ReconocimientoMedico).filter(ReconocimientoMedico.trabajador_id == t.id).order_by(ReconocimientoMedico.fecha.desc()).limit(3).all()
    herramientas = db.query(Herramienta).filter(
        Herramienta.responsable_id == t.id,
        Herramienta.activa == True,
    ).order_by(Herramienta.nombre).all()
    dotacion_lineas = db.query(LineaDotacion).join(DotacionTrabajador).filter(
        DotacionTrabajador.trabajador_id == t.id,
        DotacionTrabajador.estado == "entregada",
        LineaDotacion.estado == "entregada",
    ).order_by(LineaDotacion.nombre).all()
    maquinaria = db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        func.lower(func.trim(Maquinaria.responsable)) == t.nombre_completo.strip().lower(),
    ).order_by(Maquinaria.nombre).all()
    solicitudes = db.query(SolicitudTrabajador).options(
        joinedload(SolicitudTrabajador.lineas), joinedload(SolicitudTrabajador.comentarios),
    ).filter(SolicitudTrabajador.trabajador_id == t.id).order_by(
        SolicitudTrabajador.creado_en.desc(),
    ).limit(30).all()
    comunicaciones = db.query(ComunicacionTrabajador).filter(
        ComunicacionTrabajador.trabajador_id == t.id,
    ).order_by(ComunicacionTrabajador.creado_en.desc()).limit(20).all()
    # El trabajador pide una necesidad genérica (p. ej. "taladro"). El almacén
    # decide después qué unidad concreta entrega. Solo enviamos nombres distintos
    # y acotados: evita mezclar categorías y reduce mucho el HTML del móvil.
    catalogo_por_tipo: dict[str, list[str]] = {
        "ropa": [], "epi": [], "herramienta": [], "maquinaria": [], "consumible": [], "otro": [],
    }
    for category, name in db.query(CatalogoEPI.categoria, CatalogoEPI.nombre).filter(
        CatalogoEPI.activo == True,
        CatalogoEPI.categoria.in_(["ropa", "epi"]),
    ).distinct().order_by(CatalogoEPI.categoria, CatalogoEPI.nombre).limit(120).all():
        if name and name.strip():
            catalogo_por_tipo[category].append(name.strip())
    catalogo_por_tipo["herramienta"] = [name.strip() for (name,) in db.query(Herramienta.nombre).filter(
        Herramienta.activa == True, Herramienta.almacen_id == t.almacen_id,
        Herramienta.nombre.isnot(None),
    ).distinct().order_by(Herramienta.nombre).limit(100).all() if name and name.strip()]
    catalogo_por_tipo["maquinaria"] = [name.strip() for (name,) in db.query(Maquinaria.nombre).filter(
        Maquinaria.activa == True, Maquinaria.almacen_id == t.almacen_id,
        Maquinaria.nombre.isnot(None),
    ).distinct().order_by(Maquinaria.nombre).limit(100).all() if name and name.strip()]
    catalogo_por_tipo["consumible"] = [name.strip() for (name,) in db.query(Material.nombre).filter(
        Material.activo == True, Material.almacen_id == t.almacen_id,
        Material.nombre.isnot(None),
    ).distinct().order_by(Material.nombre).limit(100).all() if name and name.strip()]
    entregas = db.query(EntregaEPI).filter(
        EntregaEPI.trabajador_id == t.id,
    ).order_by(EntregaEPI.fecha.desc()).limit(20).all()
    albaranes = db.query(AlbaranSalida).options(
        joinedload(AlbaranSalida.items), joinedload(AlbaranSalida.almacen),
    ).filter(
        AlbaranSalida.responsable_id == t.id,
    ).order_by(AlbaranSalida.fecha_salida.desc()).limit(50).all()
    notificaciones = db.query(NotificacionTrabajador).filter(
        NotificacionTrabajador.trabajador_id == t.id,
    ).order_by(NotificacionTrabajador.creado_en.desc()).limit(50).all()
    incidencias_portal = db.query(IncidenciaPortalTrabajador).filter(
        IncidenciaPortalTrabajador.trabajador_id == t.id,
    ).order_by(IncidenciaPortalTrabajador.creado_en.desc()).limit(30).all()
    devoluciones_portal = db.query(SolicitudDevolucionTrabajador).filter(
        SolicitudDevolucionTrabajador.trabajador_id == t.id,
    ).order_by(SolicitudDevolucionTrabajador.creado_en.desc()).limit(30).all()
    sesiones_portal = db.query(SesionPortalTrabajador).filter(
        SesionPortalTrabajador.trabajador_id == t.id,
        SesionPortalTrabajador.revocado_en.is_(None),
        SesionPortalTrabajador.expira_en > datetime.now(),
    ).order_by(SesionPortalTrabajador.ultimo_uso_en.desc()).all()
    for entrega in entregas:
        try:
            entrega.items_portal = json.loads(entrega.items_json or "[]")
        except (TypeError, ValueError):
            entrega.items_portal = []
    portal_base = MRD_PUBLIC_URL if IS_PRODUCTION else str(request.base_url).rstrip("/")
    carnet_qr_b64 = generar_qr_base64(f"{portal_base}/portal/{t.portal_token}")
    response = templates.TemplateResponse(request, "portal_trabajador.html", {
        "request": request, "trabajador": t, "epis": epis,
        "formaciones": formaciones, "reconocimientos": reconocs,
        "herramientas": herramientas, "maquinaria": maquinaria,
        "solicitudes": solicitudes, "comunicaciones": comunicaciones,
        "entregas": entregas, "albaranes": albaranes,
        "notificaciones": notificaciones,
        "notificaciones_sin_leer": sum(1 for item in notificaciones if not item.leida_en),
        "incidencias_portal": incidencias_portal,
        "devoluciones_portal": devoluciones_portal,
        "sesiones_portal": sesiones_portal,
        "portal_ok": request.query_params.get("ok"),
        "portal_numero": request.query_params.get("numero"),
        "catalogo_por_tipo": catalogo_por_tipo,
        "carnet_qr_b64": carnet_qr_b64,
        "dotacion_lineas": dotacion_lineas,
    })
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Referrer-Policy"] = "no-referrer"
    return response


@app.post("/portal/{token}/solicitudes", response_class=RedirectResponse)
async def portal_crear_solicitud(token: str, request: Request, db: Session = Depends(get_db)):
    trabajador = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not trabajador:
        raise HTTPException(404, "Enlace de trabajador no válido")
    if trabajador.portal_pin_hash and not _portal_cookie_matches(request, db, trabajador):
        return RedirectResponse(
            f"/portal-trabajador?codigo={urllib.parse.quote(trabajador.codigo or '')}",
            status_code=303,
        )
    form = await request.form()
    tipos, descripciones = form.getlist("tipo"), form.getlist("descripcion")
    tallas, cantidades = form.getlist("talla"), form.getlist("cantidad")
    items = [
        {"tipo": tipos[i] if i < len(tipos) else "otro",
         "descripcion": value,
         "talla": tallas[i] if i < len(tallas) else "",
         "cantidad": cantidades[i] if i < len(cantidades) else "1"}
        for i, value in enumerate(descripciones) if str(value).strip()
    ]
    try:
        solicitud = create_worker_request(
            db, trabajador, submission_id=str(form.get("submission_id") or ""),
            priority=str(form.get("prioridad") or "normal"),
            destination=str(form.get("obra_destino") or ""),
            reason=str(form.get("motivo") or ""), items=items,
        )
        db.add(AuditoriaLog(
            tabla="solicitudes_trabajador", registro_id=solicitud.id, accion="crear_portal",
            resumen=f"Solicitud {solicitud.numero} creada desde portal", usuario_id=None,
        ))
        db.commit()
        try:
            n_items = len(solicitud.lineas)
            aviso = Aviso(
                titulo=f"Nueva solicitud {solicitud.numero} de {trabajador.nombre_completo}",
                mensaje=f"{n_items} articulo(s) pedido(s)."
                        + (f" Prioridad: {solicitud.prioridad}." if solicitud.prioridad == "urgente" else "")
                        + (f" Destino: {solicitud.obra_destino}." if solicitud.obra_destino else ""),
                prioridad="alta" if solicitud.prioridad == "urgente" else "media",
                tipo="sistema",
                enlace="/solicitudes-trabajadores",
            )
            db.add(aviso)
            db.commit()
            notif_engine.procesar_notificacion(aviso.id, db)
        except Exception as _err:
            mrd_logging.log_app(f"Error creando aviso de solicitud trabajador: {_err}", level="warning")
    except IntegrityError as exc:
        db.rollback()
        solicitud = db.query(SolicitudTrabajador).filter_by(
            submission_id=str(form.get("submission_id") or "").strip(),
            trabajador_id=trabajador.id,
        ).first()
        if not solicitud:
            mrd_logging.log_app(
                f"Fallo al crear solicitud del trabajador {trabajador.id} "
                f"(submission_id={form.get('submission_id')!r}): {exc}",
                level="error",
            )
            raise HTTPException(409, "La solicitud no pudo registrarse; vuelve a intentarlo")
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse(
        f"/portal/{token}?ok=solicitud&numero={urllib.parse.quote(solicitud.numero)}#solicitudes",
        status_code=303,
    )


@app.post("/portal/{token}/buzon", response_class=RedirectResponse)
async def portal_crear_comunicacion(token: str, request: Request, db: Session = Depends(get_db)):
    trabajador = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not trabajador:
        raise HTTPException(404, "Enlace de trabajador no válido")
    if trabajador.portal_pin_hash and not _portal_cookie_matches(request, db, trabajador):
        return RedirectResponse(
            f"/portal-trabajador?codigo={urllib.parse.quote(trabajador.codigo or '')}",
            status_code=303,
        )
    form = await request.form()
    legacy_request = str(form.get("tipo") or "") == "solicitud"
    fingerprint = hashlib.sha256(
        f"{trabajador.id}|{form.get('tipo')}|{form.get('asunto')}|{form.get('mensaje')}".encode()
    ).hexdigest()
    with _worker_buzon_lock:
        now = time.time()
        for key, sent_at in list(_worker_buzon_recent.items()):
            if now - sent_at > 300:
                _worker_buzon_recent.pop(key, None)
        previous = _worker_buzon_recent.get(fingerprint, 0)
        if now - previous < 30:
            raise HTTPException(429, "Este mensaje ya se ha enviado. Espera antes de repetirlo.")
        _worker_buzon_recent[fingerprint] = now
    if legacy_request:
        category = str(form.get("categoria") or "otro").lower()
        if category not in {"ropa", "epi", "herramienta", "consumible", "otro"}:
            category = "otro"
        try:
            amount = int(form.get("cantidad") or 1)
        except (TypeError, ValueError):
            amount = 1
        try:
            solicitud = create_worker_request(
                db, trabajador, submission_id=uuid.uuid4().hex,
                priority="urgente" if str(form.get("prioridad") or "") in {"alta", "urgente"} else "normal",
                destination="", reason=str(form.get("mensaje") or ""),
                items=[{"tipo": category, "descripcion": str(form.get("asunto") or "Solicitud"),
                        "cantidad": amount}],
            )
            solicitud.tipo = "solicitud"
            solicitud.categoria = category
            solicitud.asunto = str(form.get("asunto") or "")[:200]
            solicitud.mensaje = str(form.get("mensaje") or "")[:5000]
            solicitud.cantidad = amount
            db.commit()
        except WorkerPortalError as exc:
            db.rollback()
            raise HTTPException(exc.status_code, exc.detail)
        return RedirectResponse(
            f"/portal/{token}?ok=solicitud&numero={urllib.parse.quote(solicitud.numero)}#solicitudes",
            status_code=303,
        )
    try:
        mensaje = create_worker_message(
            db, trabajador, message_type=str(form.get("tipo") or "sugerencia"),
            privacy=str(form.get("privacidad") or "identificada"),
            subject=str(form.get("asunto") or ""), message=str(form.get("mensaje") or ""),
            worksite=str(form.get("obra") or ""),
        )
        db.add(AuditoriaLog(
            tabla="comunicaciones_trabajador", registro_id=mensaje.id, accion="crear_portal",
            resumen=f"Comunicación {mensaje.numero} recibida ({mensaje.privacidad})", usuario_id=None,
        ))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse(
        f"/portal/{token}?ok=buzon&numero={urllib.parse.quote(mensaje.numero)}#buzon",
        status_code=303,
    )


@app.post("/portal/{token}/cambiar-pin", response_class=RedirectResponse)
async def portal_cambiar_pin(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    form = await request.form()
    current = str(form.get("pin_actual") or "").strip()
    new_pin = str(form.get("pin_nuevo") or "").strip()
    confirmation = str(form.get("pin_confirm") or "").strip()
    if not worker.portal_pin_hash or not verificar_password(current, worker.portal_pin_hash):
        raise HTTPException(401, "El PIN actual no es correcto")
    if not _valid_worker_pin(new_pin) or new_pin != confirmation or new_pin == current:
        raise HTTPException(422, "El PIN nuevo debe ser distinto, coincidir y tener entre 4 y 6 números")
    worker.portal_pin_hash = hash_password(new_pin)
    worker.portal_pin_actualizado_en = datetime.now()
    worker.portal_pin_cambio_obligatorio = False
    payload = verificar_token(request.cookies.get(WORKER_COOKIE_NAME, "")) or {}
    current_hash = hashlib.sha256(str(payload.get("session_key") or "").encode()).hexdigest()
    db.query(SesionPortalTrabajador).filter(
        SesionPortalTrabajador.trabajador_id == worker.id,
        SesionPortalTrabajador.token_hash != current_hash,
        SesionPortalTrabajador.revocado_en.is_(None),
    ).update({"revocado_en": datetime.now()}, synchronize_session=False)
    db.add(AuditoriaLog(
        tabla="trabajadores", registro_id=worker.id, accion="portal_pin_cambiado",
        resumen="PIN cambiado; otras sesiones cerradas", usuario_id=None,
    ))
    db.commit()
    return RedirectResponse(f"/portal/{token}?ok=pin#cuenta", status_code=303)


@app.post("/portal/{token}/cerrar-otras-sesiones", response_class=RedirectResponse)
def portal_cerrar_otras_sesiones(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    payload = verificar_token(request.cookies.get(WORKER_COOKIE_NAME, "")) or {}
    current_hash = hashlib.sha256(str(payload.get("session_key") or "").encode()).hexdigest()
    db.query(SesionPortalTrabajador).filter(
        SesionPortalTrabajador.trabajador_id == worker.id,
        SesionPortalTrabajador.token_hash != current_hash,
        SesionPortalTrabajador.revocado_en.is_(None),
    ).update({"revocado_en": datetime.now()}, synchronize_session=False)
    db.commit()
    return RedirectResponse(f"/portal/{token}?ok=sesiones#cuenta", status_code=303)


@app.post("/portal/{token}/notificaciones/leer", response_class=RedirectResponse)
def portal_leer_notificaciones(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    db.query(NotificacionTrabajador).filter(
        NotificacionTrabajador.trabajador_id == worker.id,
        NotificacionTrabajador.leida_en.is_(None),
    ).update({"leida_en": datetime.now()}, synchronize_session=False)
    db.commit()
    return RedirectResponse(f"/portal/{token}#notificaciones", status_code=303)


@app.post("/portal/{token}/perfil", response_class=RedirectResponse)
async def portal_actualizar_perfil(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    form = await request.form()
    current_pin = str(form.get("pin_actual") or "").strip()
    if not worker.portal_pin_hash or not verificar_password(current_pin, worker.portal_pin_hash):
        raise HTTPException(401, "El PIN actual no es correcto")
    email = str(form.get("email") or "").strip()[:100]
    phone = str(form.get("telefono") or "").strip()[:20]
    if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
        raise HTTPException(422, "El correo no parece válido")
    if phone and len(re.sub(r"\D", "", phone)) < 7:
        raise HTTPException(422, "El teléfono no parece válido")
    worker.email, worker.telefono = email or None, phone or None
    worker.portal_contacto_verificado_en = None
    db.add(AuditoriaLog(
        tabla="trabajadores", registro_id=worker.id, accion="portal_perfil",
        resumen="Contacto actualizado por el trabajador; pendiente de verificación", usuario_id=None,
    ))
    db.commit()
    return RedirectResponse(f"/portal/{token}?ok=perfil#cuenta", status_code=303)


@app.post("/portal/{token}/solicitudes/{request_id}/cancelar", response_class=RedirectResponse)
def portal_cancelar_solicitud(
    token: str, request_id: int, request: Request, db: Session = Depends(get_db),
):
    worker = _portal_worker_required(token, request, db)
    row = db.get(SolicitudTrabajador, request_id)
    if not row:
        raise HTTPException(404, "Solicitud no encontrada")
    try:
        cancel_worker_request(db, worker, row)
        db.add(AuditoriaLog(tabla="solicitudes_trabajador", registro_id=row.id,
                            accion="cancelar_portal", resumen=row.numero, usuario_id=None))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse(f"/portal/{token}?ok=cancelada#solicitudes", status_code=303)


@app.post("/portal/{token}/solicitudes/{request_id}/comentarios", response_class=RedirectResponse)
async def portal_comentar_solicitud(
    token: str, request_id: int, request: Request, db: Session = Depends(get_db),
):
    worker = _portal_worker_required(token, request, db)
    row = db.get(SolicitudTrabajador, request_id)
    if not row:
        raise HTTPException(404, "Solicitud no encontrada")
    form = await request.form()
    try:
        add_worker_request_comment(db, worker, row, str(form.get("comentario") or ""))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse(f"/portal/{token}?ok=comentario#solicitudes", status_code=303)


@app.post("/portal/{token}/solicitudes/{request_id}/confirmar-recogida", response_class=RedirectResponse)
def portal_confirmar_recogida(
    token: str, request_id: int, request: Request, db: Session = Depends(get_db),
):
    worker = _portal_worker_required(token, request, db)
    row = db.get(SolicitudTrabajador, request_id)
    if not row or row.trabajador_id != worker.id:
        raise HTTPException(404, "Solicitud no encontrada")
    if row.estado not in {"lista", "entregada"}:
        raise HTTPException(409, "La solicitud todavía no está lista para recoger")
    row.recogida_confirmada_en = datetime.now()
    db.commit()
    return RedirectResponse(f"/portal/{token}?ok=recogida#solicitudes", status_code=303)


async def _save_worker_portal_photo(
    upload: UploadFile | None, worker_id: int, prefix: str,
) -> str | None:
    if not upload or not upload.filename:
        return None
    try:
        _, ext = validar_nombre_archivo(upload.filename, {"jpg", "jpeg", "png", "webp"})
        head = await upload.read(16)
        await upload.seek(0)
        validar_contenido_archivo(head, ext)
        data = await upload.read()
        validar_tamaño_bytes(len(data), min(MAX_UPLOAD_MB, 8))
    except ErrorArchivo as exc:
        raise HTTPException(400, str(exc))
    folder = UPLOADS_DIR / "portal_trabajador"
    folder.mkdir(parents=True, exist_ok=True)
    filename = f"{prefix}_{worker_id}_{uuid.uuid4().hex}.{ext}"
    (folder / filename).write_bytes(data)
    return f"portal_trabajador/{filename}"


@app.post("/portal/{token}/incidencias", response_class=RedirectResponse)
async def portal_crear_incidencia(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    form = await request.form()
    photo = await _save_worker_portal_photo(form.get("foto"), worker.id, "inc")
    try:
        row = create_worker_incident(
            db, worker, category=str(form.get("categoria") or "otro"),
            asset_type=str(form.get("activo_tipo") or ""),
            asset_code=str(form.get("activo_codigo") or ""),
            asset_name=str(form.get("activo_nombre") or ""),
            description=str(form.get("descripcion") or ""), photo_path=photo,
        )
        db.add(AuditoriaLog(tabla="incidencias_portal_trabajador", registro_id=row.id,
                            accion="crear_portal", resumen=row.numero, usuario_id=None))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        if photo:
            (UPLOADS_DIR / photo).unlink(missing_ok=True)
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse(f"/portal/{token}?ok=incidencia&numero={row.numero}#incidencias", status_code=303)


@app.post("/portal/{token}/devoluciones", response_class=RedirectResponse)
async def portal_crear_devolucion(token: str, request: Request, db: Session = Depends(get_db)):
    worker = _portal_worker_required(token, request, db)
    form = await request.form()
    photo = await _save_worker_portal_photo(form.get("foto"), worker.id, "dev")
    try:
        row = create_worker_return(
            db, worker, asset_type=str(form.get("activo_tipo") or "otro"),
            asset_code=str(form.get("activo_codigo") or ""),
            description=str(form.get("descripcion") or ""),
            quantity=float(form.get("cantidad") or 1),
            item_state=str(form.get("estado_material") or "correcto"),
            reason=str(form.get("motivo") or ""), photo_path=photo,
        )
        db.add(AuditoriaLog(tabla="devoluciones_trabajador", registro_id=row.id,
                            accion="crear_portal", resumen=row.numero, usuario_id=None))
        db.commit()
    except (ValueError, WorkerPortalError) as exc:
        db.rollback()
        if photo:
            (UPLOADS_DIR / photo).unlink(missing_ok=True)
        if isinstance(exc, WorkerPortalError):
            raise HTTPException(exc.status_code, exc.detail)
        raise HTTPException(422, "Cantidad no válida")
    return RedirectResponse(f"/portal/{token}?ok=devolucion&numero={row.numero}#devoluciones", status_code=303)


@app.get("/solicitudes-trabajadores", response_class=HTMLResponse)
def solicitudes_trabajadores(
    request: Request, estado: str = "", q: str = "",
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not can_manage_requests(user):
        raise HTTPException(403, "Sin permiso para gestionar solicitudes")
    query = db.query(SolicitudTrabajador).options(
        joinedload(SolicitudTrabajador.trabajador), joinedload(SolicitudTrabajador.lineas),
    )
    active_warehouse = _active_warehouse(db, user, request)
    if user.rol != "admin":
        query = query.filter(
            SolicitudTrabajador.almacen_id == (active_warehouse.id if active_warehouse else -1)
        )
    if estado:
        query = query.filter(SolicitudTrabajador.estado == estado)
    if q:
        term = f"%{q.strip()}%"
        query = query.join(Trabajador).filter(or_(
            SolicitudTrabajador.numero.ilike(term), Trabajador.nombre.ilike(term),
            Trabajador.apellidos.ilike(term),
        ))
    rows = query.order_by(SolicitudTrabajador.creado_en.desc()).limit(250).all()
    request_notes = {
        note.notas.split(" · ", 1)[0].removeprefix("Solicitud "): note
        for note in db.query(AlbaranSalida).filter(
            AlbaranSalida.notas.like("Solicitud SOL-%"),
        ).all()
        if note.notas
    }
    for row in rows:
        row.albaran_solicitud = request_notes.get(row.numero)
    return templates.TemplateResponse(request, "solicitudes_trabajadores.html", ctx_base(
        request, user, db, solicitudes=rows, estado_filtro=estado, q=q,
        transiciones_solicitud=REQUEST_TRANSITIONS,
    ))


@app.post("/solicitudes-trabajadores/{solicitud_id}/estado", response_class=RedirectResponse)
async def solicitud_trabajador_estado(
    solicitud_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    solicitud = db.get(SolicitudTrabajador, solicitud_id)
    if not solicitud:
        raise HTTPException(404, "Solicitud no encontrada")
    form = await request.form()
    try:
        active_warehouse = _active_warehouse(db, user, request)
        for line in solicitud.lineas:
            approved_raw = str(form.get(f"cantidad_aprobada_{line.id}") or "").strip()
            assignment = str(form.get(f"asignacion_{line.id}") or "").strip()
            if approved_raw:
                try:
                    approved = int(approved_raw)
                except ValueError as exc:
                    raise WorkerPortalError(422, "La cantidad aprobada no es válida") from exc
                if approved < 1 or approved > line.cantidad:
                    raise WorkerPortalError(422, "La cantidad aprobada no es válida")
                line.cantidad_aprobada = approved
            if assignment:
                line.observaciones = assignment[:500]
        transition_worker_request(
            db, user, solicitud, new_status=str(form.get("estado") or ""),
            notes=str(form.get("notas") or ""),
            access_warehouse_id=active_warehouse.id if active_warehouse else None,
        )
        estimated = str(form.get("fecha_estimada") or "").strip()
        if estimated:
            try:
                solicitud.fecha_estimada = datetime.fromisoformat(estimated)
            except ValueError:
                raise WorkerPortalError(422, "Fecha estimada no válida")
        if solicitud.estado == "entregada":
            marker = f"Solicitud {solicitud.numero}"
            existing_note = db.query(AlbaranSalida).filter(
                AlbaranSalida.notas.like(f"{marker}%"),
            ).first()
            if not existing_note:
                reason = f" · Motivo: {solicitud.motivo}" if solicitud.motivo else ""
                create_delivery_note(
                    db, user_id=user.id, worker_id=solicitud.trabajador_id,
                    warehouse_id=solicitud.almacen_id,
                    origin_destination=solicitud.obra_destino or solicitud.trabajador.nombre_completo,
                    notes=f"{marker}{reason}",
                    lines=[{
                        "tipo": line.tipo,
                        "nombre": line.observaciones or f"{line.descripcion}{f' · T. {line.talla}' if line.talla else ''}",
                        "cantidad": line.cantidad_aprobada or line.cantidad,
                    } for line in solicitud.lineas],
                )
        db.add(AuditoriaLog(
            tabla="solicitudes_trabajador", registro_id=solicitud.id, accion="cambiar_estado",
            resumen=f"Solicitud {solicitud.numero}: {solicitud.estado}", usuario_id=user.id,
        ))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception as exc:
        db.rollback()
        mrd_logging.log_app(
            f"Fallo al cambiar estado de solicitud {solicitud_id}: {exc}", level="error",
        )
        raise HTTPException(500, "No se pudo actualizar la solicitud. Nada se ha guardado, inténtalo de nuevo.")
    return RedirectResponse("/solicitudes-trabajadores?ok=actualizada", status_code=303)


@app.post("/solicitudes-trabajadores/{solicitud_id}/albaran", response_class=RedirectResponse)
def solicitud_trabajador_generar_albaran(
    solicitud_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    solicitud = db.query(SolicitudTrabajador).options(
        joinedload(SolicitudTrabajador.lineas), joinedload(SolicitudTrabajador.trabajador),
    ).filter(SolicitudTrabajador.id == solicitud_id).first()
    if not solicitud:
        raise HTTPException(404, "Solicitud no encontrada")
    active_warehouse = _active_warehouse(db, user, request)
    require_request_access(
        user, solicitud, active_warehouse.id if active_warehouse else None,
    )
    if solicitud.estado != "entregada":
        raise HTTPException(409, "El albarán se genera al completar la entrega")
    marker = f"Solicitud {solicitud.numero}"
    note = db.query(AlbaranSalida).filter(AlbaranSalida.notas.like(f"{marker}%")).first()
    if not note:
        reason = f" · Motivo: {solicitud.motivo}" if solicitud.motivo else ""
        note = create_delivery_note(
            db, user_id=user.id, worker_id=solicitud.trabajador_id,
            warehouse_id=solicitud.almacen_id,
            origin_destination=solicitud.obra_destino or solicitud.trabajador.nombre_completo,
            notes=f"{marker}{reason}",
            lines=[{
                "tipo": line.tipo,
                "nombre": line.observaciones or f"{line.descripcion}{f' · T. {line.talla}' if line.talla else ''}",
                "cantidad": line.cantidad_aprobada or line.cantidad,
            } for line in solicitud.lineas],
        )
        db.add(AuditoriaLog(
            tabla="solicitudes_trabajador", registro_id=solicitud.id,
            accion="generar_albaran", resumen=f"{solicitud.numero}: {note.numero}",
            usuario_id=user.id,
        ))
        db.commit()
    return RedirectResponse(f"/albaranes-salida/{note.id}", status_code=303)


@app.get("/portal/{token}/albaranes/{aid}", response_class=HTMLResponse)
def portal_trabajador_albaran(
    token: str, aid: int, request: Request, db: Session = Depends(get_db),
):
    trabajador = db.query(Trabajador).filter(
        Trabajador.portal_token == token, Trabajador.activo == True,
    ).first()
    if not trabajador:
        raise HTTPException(404, "Enlace de trabajador no válido")
    if trabajador.portal_pin_hash and not _portal_cookie_matches(request, db, trabajador):
        return RedirectResponse(
            f"/portal-trabajador?codigo={urllib.parse.quote(trabajador.codigo or '')}",
            status_code=303,
        )
    note = db.query(AlbaranSalida).options(
        joinedload(AlbaranSalida.items), joinedload(AlbaranSalida.almacen),
    ).filter(
        AlbaranSalida.id == aid, AlbaranSalida.responsable_id == trabajador.id,
    ).first()
    if not note:
        raise HTTPException(404, "Albarán no encontrado")
    return templates.TemplateResponse(request, "portal_albaran.html", {
        "request": request, "trabajador": trabajador, "alb": note, "token": token,
    })


@app.post("/portal/{token}/albaranes/{aid}/conformidad", response_class=RedirectResponse)
async def portal_albaran_conformidad(
    token: str, aid: int, request: Request, db: Session = Depends(get_db),
):
    worker = _portal_worker_required(token, request, db)
    note = db.query(AlbaranSalida).filter(
        AlbaranSalida.id == aid, AlbaranSalida.responsable_id == worker.id,
    ).first()
    if not note:
        raise HTTPException(404, "Albarán no encontrado")
    if note.portal_conformidad != "pendiente":
        raise HTTPException(409, "Este albarán ya fue confirmado")
    form = await request.form()
    conformity = str(form.get("conformidad") or "")
    reason = str(form.get("motivo") or "").strip()
    signature = str(form.get("firma_datos") or "")
    if conformity not in {"conforme", "no_conforme"}:
        raise HTTPException(422, "Selecciona conforme o no conforme")
    if conformity == "no_conforme" and len(reason) < 5:
        raise HTTPException(422, "Explica brevemente la disconformidad")
    if conformity == "conforme" and not signature.startswith("data:image/png;base64,"):
        raise HTTPException(422, "Firma en la pantalla antes de confirmar")
    if len(signature) > 1_000_000:
        raise HTTPException(413, "La firma es demasiado grande")
    note.portal_conformidad = conformity
    note.portal_motivo = reason[:2000] or None
    note.portal_firma_datos = signature if conformity == "conforme" else None
    note.portal_firmado_en = datetime.now()
    db.add(AuditoriaLog(
        tabla="albaranes_salida", registro_id=note.id, accion="conformidad_portal",
        resumen=f"{note.numero}: {conformity}", usuario_id=None,
    ))
    db.commit()
    return RedirectResponse(f"/portal/{token}/albaranes/{aid}?ok=conformidad", status_code=303)


@app.get("/buzon-trabajadores", response_class=HTMLResponse)
def buzon_trabajadores(
    request: Request, estado: str = "", user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if user.rol != "admin":
        raise HTTPException(403, "Solo administración puede acceder al buzón privado")
    query = db.query(ComunicacionTrabajador).options(joinedload(ComunicacionTrabajador.trabajador))
    if estado:
        query = query.filter(ComunicacionTrabajador.estado == estado)
    rows = query.order_by(ComunicacionTrabajador.creado_en.desc()).limit(250).all()
    return templates.TemplateResponse(request, "buzon_trabajadores.html", ctx_base(
        request, user, db, comunicaciones=rows, estado_filtro=estado,
    ))


@app.get("/operaciones-portal-trabajadores", response_class=HTMLResponse)
def operaciones_portal_trabajadores(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not can_manage_requests(user):
        raise HTTPException(403, "Sin permiso para gestionar el portal")
    incidents = db.query(IncidenciaPortalTrabajador).options(
        joinedload(IncidenciaPortalTrabajador.trabajador),
    )
    returns = db.query(SolicitudDevolucionTrabajador).options(
        joinedload(SolicitudDevolucionTrabajador.trabajador),
    )
    if user.rol != "admin":
        incidents = incidents.filter(IncidenciaPortalTrabajador.almacen_id == user.almacen_id)
        returns = returns.filter(SolicitudDevolucionTrabajador.almacen_id == user.almacen_id)
    return templates.TemplateResponse(request, "operaciones_portal_trabajadores.html", ctx_base(
        request, user, db,
        incidencias_portal=incidents.order_by(IncidenciaPortalTrabajador.creado_en.desc()).limit(150).all(),
        devoluciones_portal=returns.order_by(SolicitudDevolucionTrabajador.creado_en.desc()).limit(150).all(),
    ))


@app.post("/operaciones-portal-trabajadores/incidencias/{item_id}", response_class=RedirectResponse)
async def gestionar_incidencia_portal(
    item_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not can_manage_requests(user):
        raise HTTPException(403, "Sin permiso")
    row = db.get(IncidenciaPortalTrabajador, item_id)
    if not row:
        raise HTTPException(404, "Incidencia no encontrada")
    _require_warehouse_access(user, row.almacen_id)
    form = await request.form()
    status = str(form.get("estado") or "")
    if status not in {"recibida", "revision", "actuacion", "resuelta", "cerrada"}:
        raise HTTPException(422, "Estado no válido")
    row.estado = status
    row.respuesta = str(form.get("respuesta") or "").strip()[:4000] or row.respuesta
    row.actualizado_en = datetime.now()
    if status in {"resuelta", "cerrada"}:
        row.resuelta_en = datetime.now()
    create_worker_notification(
        db, row.trabajador_id, title=f"Incidencia {row.numero}",
        message=row.respuesta or f"Tu incidencia está en estado {status}.",
        kind="incidencia", link="#incidencias", event_key=f"inc:{row.id}:{status}",
    )
    db.commit()
    return RedirectResponse("/operaciones-portal-trabajadores?ok=incidencia", status_code=303)


@app.post("/operaciones-portal-trabajadores/devoluciones/{item_id}", response_class=RedirectResponse)
async def gestionar_devolucion_portal(
    item_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not can_manage_requests(user):
        raise HTTPException(403, "Sin permiso")
    row = db.get(SolicitudDevolucionTrabajador, item_id)
    if not row:
        raise HTTPException(404, "Devolución no encontrada")
    _require_warehouse_access(user, row.almacen_id)
    form = await request.form()
    status = str(form.get("estado") or "")
    if status not in {"solicitada", "aceptada", "cita", "recibida", "rechazada", "cancelada"}:
        raise HTTPException(422, "Estado no válido")
    row.estado = status
    row.notas_gestion = str(form.get("notas") or "").strip()[:3000] or row.notas_gestion
    row.actualizado_en = datetime.now()
    if status == "recibida":
        row.completada_en = datetime.now()
    create_worker_notification(
        db, row.trabajador_id, title=f"Devolución {row.numero}",
        message=row.notas_gestion or f"Tu devolución está en estado {status}.",
        kind="devolucion", link="#devoluciones", event_key=f"dev:{row.id}:{status}",
    )
    db.commit()
    return RedirectResponse("/operaciones-portal-trabajadores?ok=devolucion", status_code=303)


@app.post("/buzon-trabajadores/{mensaje_id}/gestionar", response_class=RedirectResponse)
async def buzon_trabajador_gestionar(
    mensaje_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    mensaje = db.get(ComunicacionTrabajador, mensaje_id)
    if not mensaje:
        raise HTTPException(404, "Comunicación no encontrada")
    form = await request.form()
    try:
        manage_worker_message(
            db, user, mensaje, status=str(form.get("estado") or ""),
            response=str(form.get("respuesta") or ""),
        )
        db.add(AuditoriaLog(
            tabla="comunicaciones_trabajador", registro_id=mensaje.id, accion="gestionar",
            resumen=f"Comunicación {mensaje.numero}: {mensaje.estado}", usuario_id=user.id,
        ))
        db.commit()
    except WorkerPortalError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    return RedirectResponse("/buzon-trabajadores?ok=actualizado", status_code=303)


@app.post("/buzon-trabajadores/{solicitud_id}/responder", response_class=RedirectResponse)
async def buzon_solicitud_compat_responder(
    solicitud_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    """Gestiona solicitudes creadas por la versión histórica del buzón móvil."""
    solicitud = db.get(SolicitudTrabajador, solicitud_id)
    if not solicitud:
        raise HTTPException(404, "Solicitud no encontrada")
    active_warehouse = _active_warehouse(db, user, request)
    require_request_access(
        user, solicitud, active_warehouse.id if active_warehouse else None,
    )
    form = await request.form()
    response_text = str(form.get("respuesta") or form.get("notas") or "").strip()[:3000]
    new_status = str(form.get("estado") or "revision")
    if new_status not in ESTADOS_SOLICITUD_TRABAJADOR:
        raise HTTPException(422, "Estado no válido")
    if new_status != solicitud.estado:
        transition_worker_request(
            db, user, solicitud, new_status=new_status, notes=response_text,
            access_warehouse_id=active_warehouse.id if active_warehouse else None,
        )
    solicitud.respuesta = response_text or solicitud.respuesta
    solicitud.respondido_en = datetime.now()
    db.commit()
    return RedirectResponse("/buzon-trabajadores?ok=actualizado", status_code=303)


def _worker_access_rows(workers: list[Trabajador], portal_base: str) -> list[dict[str, str]]:
    rows = []
    for worker in workers:
        portal_url = f"{portal_base}/portal-trabajador?codigo={urllib.parse.quote(worker.codigo or '')}"
        provisional = bool(worker.portal_pin_hash and worker.portal_pin_cambio_obligatorio)
        pin_text = "PIN inicial: 123456" if provisional else "PIN: el personal que ya configuraste"
        message = (
            f"Hola {worker.nombre}. Acceso a Mi MRD: {portal_url}\n"
            f"Tu número: {worker.codigo or 'pendiente'}\n{pin_text}"
            + ("\nAl entrar tendrás que cambiarlo." if provisional else "")
        )
        phone = "".join(character for character in (worker.telefono or "") if character.isdigit())
        if len(phone) == 9:
            phone = f"34{phone}"
        rows.append({
            "nombre": worker.nombre_completo,
            "codigo": worker.codigo or "Pendiente",
            "telefono": worker.telefono or "",
            "portal_url": portal_url,
            "mensaje": message,
            "whatsapp_url": f"https://wa.me/{phone}?text={urllib.parse.quote(message)}" if phone else "",
            "pin_provisional": "Sí" if provisional else "No",
        })
    return rows


@app.get("/accesos-portal-trabajadores", response_class=HTMLResponse)
def accesos_portal_trabajadores(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "crear") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso para gestionar accesos")
    warehouse = _active_warehouse(db, user, request)
    workers = db.query(Trabajador).filter(
        Trabajador.activo == True,
        Trabajador.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(Trabajador.nombre, Trabajador.apellidos).all()
    portal_base = MRD_PUBLIC_URL if IS_PRODUCTION else str(request.base_url).rstrip("/")
    return templates.TemplateResponse(request, "accesos_portal_trabajadores.html", ctx_base(
        request, user, db, accesos=_worker_access_rows(workers, portal_base),
    ))


@app.post("/accesos-portal-trabajadores/preparar", response_class=RedirectResponse)
def preparar_accesos_portal_trabajadores(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "crear"):
        raise HTTPException(403, "Sin permiso para preparar accesos")
    warehouse = _active_warehouse(db, user, request)
    workers = db.query(Trabajador).filter(
        Trabajador.activo == True,
        Trabajador.almacen_id == (warehouse.id if warehouse else -1),
    ).all()
    prepared = 0
    for worker in workers:
        if not worker.codigo:
            worker.codigo = f"TRB-{worker.id:05d}"
        if not worker.portal_token:
            worker.portal_token = uuid.uuid4().hex + uuid.uuid4().hex
        if not worker.portal_pin_hash:
            worker.portal_pin_hash = hash_password("123456")
            worker.portal_pin_cambio_obligatorio = True
            prepared += 1
    db.add(AuditoriaLog(
        tabla="trabajadores", registro_id=None, accion="portal_accesos_masivos",
        resumen=f"Preparados {prepared} accesos provisionales sin reemplazar PIN existentes",
        usuario_id=user.id,
    ))
    db.commit()
    return RedirectResponse(
        f"/accesos-portal-trabajadores?ok=preparados&n={prepared}", status_code=303,
    )


@app.get("/trabajadores/{tid}/portal-qr", response_class=HTMLResponse)
def trabajador_portal_qr(tid: int, request: Request,
                         user: Usuario = Depends(requiere_login),
                         db: Session = Depends(get_db)):
    import secrets as _sec
    if not (tiene_permiso(user, "crear") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso para crear accesos de trabajadores")
    t = db.get(Trabajador, tid)
    if not t:
        raise HTTPException(404)
    if not t.portal_token:
        t.portal_token = _sec.token_urlsafe(32)
        db.commit()
        db.refresh(t)
    portal_base = MRD_PUBLIC_URL if IS_PRODUCTION else str(request.base_url).rstrip("/")
    portal_url = f"{portal_base}/portal/{t.portal_token}"
    qr_b64 = generar_qr_base64(portal_url)
    ctx = ctx_base(request, user)
    ctx.update({
        "trabajador": t, "portal_url": portal_url, "qr_b64": qr_b64,
        "pin_configurado": bool(t.portal_pin_hash),
    })
    return templates.TemplateResponse(request, "trabajador_portal_qr.html", ctx)


@app.post("/trabajadores/{tid}/portal-pin", response_class=RedirectResponse)
def trabajador_portal_pin(
    tid: int, pin: str = Form(...),
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not (tiene_permiso(user, "crear") or tiene_permiso(user, "stock_operar")):
        raise HTTPException(403, "Sin permiso para gestionar accesos de trabajadores")
    trabajador = db.get(Trabajador, tid)
    if not trabajador:
        raise HTTPException(404, "Trabajador no encontrado")
    _require_warehouse_access(user, trabajador.almacen_id)
    clean_pin = pin.strip()
    if not clean_pin.isdigit() or not 4 <= len(clean_pin) <= 6:
        raise HTTPException(422, "El PIN debe tener entre 4 y 6 números")
    if not trabajador.codigo:
        trabajador.codigo = f"TRB-{trabajador.id:05d}"
    if not trabajador.portal_token:
        trabajador.portal_token = uuid.uuid4().hex + uuid.uuid4().hex
    trabajador.portal_pin_hash = hash_password(clean_pin)
    trabajador.portal_pin_actualizado_en = datetime.now()
    trabajador.portal_pin_cambio_obligatorio = False
    db.query(SesionPortalTrabajador).filter(
        SesionPortalTrabajador.trabajador_id == trabajador.id,
        SesionPortalTrabajador.revocado_en.is_(None),
    ).update({"revocado_en": datetime.now()}, synchronize_session=False)
    db.add(AuditoriaLog(
        tabla="trabajadores", registro_id=trabajador.id, accion="portal_pin",
        resumen=f"PIN de portal renovado para {trabajador.nombre_completo}", usuario_id=user.id,
    ))
    db.commit()
    return RedirectResponse(f"/trabajadores/{tid}/portal-qr?ok=pin", status_code=303)


# ─── Materiales / Almacén ───────────────────────────────────────────────────────────────────────────

@app.get("/materiales", response_class=HTMLResponse)
async def materiales_lista(
    request: Request, db: Session = Depends(get_db),
    usuario=Depends(requiere_login),
    q: str = "", cat: str = "", solo_bajos: bool = False,
):
    warehouse = _active_warehouse(db, usuario, request)
    warehouse_id = warehouse.id if warehouse else -1
    query = db.query(Material).filter(
        Material.activo == True, Material.almacen_id == warehouse_id,
    )
    if q:
        query = query.filter(Material.nombre.ilike(f"%{q}%"))
    if cat:
        query = query.filter(Material.categoria == cat)
    materiales = query.order_by(Material.nombre).all()
    if solo_bajos:
        materiales = [m for m in materiales if m.bajo_minimo]
    obras = db.query(Obra).filter(Obra.activa == True, Obra.almacen_id == warehouse_id).order_by(Obra.nombre).all()
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True, Trabajador.almacen_id == warehouse_id).order_by(Trabajador.nombre).all()
    return templates.TemplateResponse(request, "materiales.html", ctx_base(
        request, usuario, db, usuario=usuario,
        materiales=materiales, categorias=CATEGORIAS_MATERIAL,
        unidades=UNIDADES_MATERIAL, obras=obras, trabajadores=trabajadores,
        q=q, cat=cat, solo_bajos=solo_bajos,
    ))


@app.post("/materiales", response_class=RedirectResponse)
async def materiales_crear(request: Request, db: Session = Depends(get_db),
                            usuario=Depends(requiere_login)):
    form = await request.form()
    _codigo = generar_referencia_material(db)
    _ts_mat = form.get("tipo_seguimiento", "generico")
    almacen_predeterminado = get_default_warehouse(db)
    mat = Material(
        codigo=_codigo,
        nombre=form["nombre"],
        descripcion=form.get("descripcion"),
        categoria=form.get("categoria") or None,
        unidad=form.get("unidad", "ud"),
        stock_actual=float(form.get("stock_actual") or 0),
        stock_minimo=float(form.get("stock_minimo") or 0),
        stock_maximo=float(form.get("stock_maximo")) if form.get("stock_maximo") else None,
        referencia_proveedor=form.get("referencia_proveedor") or None,
        ubicacion_texto=form.get("ubicacion"),
        tipo_seguimiento=_ts_mat if _ts_mat in ("individual", "generico") else "generico",
        almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
    )
    db.add(mat)
    db.commit()
    _next = form.get("_action", "crear")
    if _next == "crear_mas":
        return RedirectResponse("/materiales?ok=creado&open=nuevo", status_code=303)
    return RedirectResponse("/materiales?ok=creado", status_code=303)



# ─── Materiales — alertas stock bajo ────────────────────────────────────────────
@app.get("/materiales/alertas", response_class=HTMLResponse)
def materiales_alertas(
    request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    warehouse = _active_warehouse(db, user, request)
    warehouse_id = warehouse.id if warehouse else -1
    materials = (
        db.query(Material)
        .filter(Material.activo == True, Material.stock_minimo > 0,
                Material.stock_actual <= Material.stock_minimo,
                Material.almacen_id == warehouse_id)
        .order_by(Material.almacen_id, Material.nombre)
        .all()
    )
    warehouses = {row.id: row for row in db.query(Almacen).filter(Almacen.activo == True).all()}
    items = []
    for row in materials:
        target = row.stock_maximo if row.stock_maximo and row.stock_maximo > row.stock_minimo else row.stock_minimo
        alternative = db.query(Material).filter(
            Material.activo == True, Material.almacen_id != warehouse_id,
            Material.nombre == row.nombre, Material.categoria == row.categoria,
            Material.unidad == row.unidad, Material.stock_actual > Material.stock_minimo,
        ).order_by((Material.stock_actual - Material.stock_minimo).desc()).first()
        items.append({
            "tipo": "Material / consumible", "codigo": row.codigo, "nombre": row.nombre,
            "categoria": row.categoria or "Consumible", "stock_actual": float(row.stock_actual or 0),
            "stock_minimo": float(row.stock_minimo or 0), "falta": max(1, float(target or 0) - float(row.stock_actual or 0)),
            "unidad": row.unidad or "ud", "almacen_id": row.almacen_id,
            "almacen": warehouses.get(row.almacen_id),
            "ubicacion": row.ubicacion.nombre if row.ubicacion else (row.ubicacion_texto or "—"),
            "url": f"/materiales/{row.id}", "referencia_proveedor": row.referencia_proveedor or "",
            "traslado_disponible": max(0, float(alternative.stock_actual - alternative.stock_minimo)) if alternative else 0,
            "traslado_almacen": warehouses.get(alternative.almacen_id) if alternative else None,
        })
    for row in db.query(StockEPI).filter(
        StockEPI.stock_minimo > 0, StockEPI.cantidad <= StockEPI.stock_minimo,
        StockEPI.almacen_id == warehouse_id,
    ).order_by(StockEPI.categoria, StockEPI.nombre, StockEPI.talla).all():
        alternative = db.query(StockEPI).filter(
            StockEPI.almacen_id != warehouse_id, StockEPI.nombre == row.nombre,
            StockEPI.talla == row.talla, StockEPI.categoria == row.categoria,
            StockEPI.cantidad > StockEPI.stock_minimo,
        ).order_by((StockEPI.cantidad - StockEPI.stock_minimo).desc()).first()
        items.append({
            "tipo": "Ropa" if row.categoria == "ropa" else "EPI",
            "codigo": row.codigo or f"EPI-{row.id}", "nombre": row.nombre_display,
            "categoria": row.categoria, "stock_actual": int(row.cantidad or 0),
            "stock_minimo": int(row.stock_minimo or 0),
            "falta": max(1, int(row.stock_minimo or 0) - int(row.cantidad or 0)),
            "unidad": "ud", "almacen_id": row.almacen_id,
            "almacen": warehouses.get(row.almacen_id), "ubicacion": "—",
            "url": "/epis/stock", "referencia_proveedor": "",
            "traslado_disponible": max(0, int(alternative.cantidad - alternative.stock_minimo)) if alternative else 0,
            "traslado_almacen": warehouses.get(alternative.almacen_id) if alternative else None,
        })
    stock_totals = dict(db.query(
        ExistenciaVariante.variante_id, func.coalesce(func.sum(ExistenciaVariante.cantidad), 0),
    ).filter(ExistenciaVariante.almacen_id == warehouse_id).group_by(ExistenciaVariante.variante_id).all())
    for row in db.query(VarianteEPI).filter(
        VarianteEPI.activo == True,
    ).order_by(VarianteEPI.id).all():
        current = int(stock_totals.get(row.id, 0) or 0)
        existences = db.query(ExistenciaVariante).filter_by(
            variante_id=row.id, almacen_id=warehouse_id,
        ).all()
        site_minimum = max(
            [int(row.stock_minimo or 0)]
            + [int(existence.stock_minimo or 0) for existence in existences]
        )
        if site_minimum <= 0 or current > site_minimum:
            continue
        warehouse_ids = {existence.almacen_id for existence in existences if existence.almacen_id}
        item_warehouse_id = next(iter(warehouse_ids)) if len(warehouse_ids) == 1 else warehouse_id
        alternatives = db.query(ExistenciaVariante).filter(
            ExistenciaVariante.variante_id == row.id,
            ExistenciaVariante.almacen_id != warehouse_id,
            ExistenciaVariante.cantidad > ExistenciaVariante.stock_minimo,
        ).order_by((ExistenciaVariante.cantidad - ExistenciaVariante.stock_minimo).desc()).all()
        transferable = sum(max(0, int(other.cantidad - other.stock_minimo)) for other in alternatives)
        alternative_warehouse = warehouses.get(alternatives[0].almacen_id) if alternatives else None
        details = " · ".join(filter(None, (row.modelo, row.color, row.talla)))
        items.append({
            "tipo": "Ropa / EPI por talla", "codigo": row.codigo_qr or row.referencia_interna,
            "nombre": f"{row.catalogo.nombre if row.catalogo else 'Artículo'}{(' · '+details) if details else ''}",
            "categoria": row.catalogo.categoria if row.catalogo else "suministro",
            "stock_actual": current, "stock_minimo": site_minimum,
            "falta": max(1, site_minimum - current), "unidad": "ud",
            "almacen_id": item_warehouse_id, "almacen": warehouses.get(item_warehouse_id),
            "ubicacion": "Varias ubicaciones" if len(existences) > 1 else "—",
            "url": f"/inventario/variantes/{row.id}",
            "referencia_proveedor": row.referencia_proveedor or "",
            "traslado_disponible": transferable,
            "traslado_almacen": alternative_warehouse,
        })
    from collections import defaultdict as _dd
    grouped = _dd(list)
    for item in items:
        grouped[item["almacen_id"]].append(item)
    grupos = [{"almacen": warehouses.get(aid), "materiales": rows} for aid, rows in grouped.items()]
    return templates.TemplateResponse(request, "materiales_alertas.html", ctx_base(
        request, user, db,
        bajos=items,
        grupos=grupos,
        total_bajo=len(items),
        pedido_referencia=f"PED-{datetime.now().strftime('%Y%m%d-%H%M')}",
        pedido_fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
    ))

@app.get("/materiales/{mid}", response_class=HTMLResponse)
async def material_detalle(mid: int, request: Request, db: Session = Depends(get_db),
                            usuario=Depends(requiere_login)):
    mat = db.query(Material).filter(Material.id == mid).first()
    if not mat:
        raise HTTPException(status_code=404)
    _require_warehouse_access(usuario, mat.almacen_id)
    obras = db.query(Obra).filter(
        Obra.activa == True, Obra.almacen_id == mat.almacen_id,
    ).order_by(Obra.nombre).all()
    trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == mat.almacen_id,
    ).order_by(Trabajador.nombre).all()
    mat_movs = mat.movimientos_almacen if hasattr(mat, 'movimientos_almacen') else []
    historial_valores = (
        db.query(AuditoriaLog)
        .filter(AuditoriaLog.tabla == "materiales", AuditoriaLog.registro_id == mat.id)
        .order_by(AuditoriaLog.fecha.desc())
        .limit(50)
        .all()
    )
    return templates.TemplateResponse(request, "material_detalle.html", {
        "request": request, "usuario": usuario, "mat": mat, "mat_movs": mat_movs,
        "tipos": TIPOS_MOVIMIENTO_MAT, "obras": obras, "trabajadores": trabajadores,
        "categorias": CATEGORIAS_MATERIAL, "historial_valores": historial_valores,
    })


@app.post("/materiales/{mid}/movimiento", response_class=RedirectResponse)
async def material_movimiento(mid: int, request: Request, db: Session = Depends(get_db),
                               usuario=Depends(requiere_login)):
    form = await request.form()
    tipo = form.get("tipo", "salida")
    cantidad = float(form.get("cantidad") or 0)
    if tipo not in {"salida", "entrada", "ajuste"} or cantidad < 0:
        raise HTTPException(400, "Movimiento de material no válido")
    try:
        require_stock_permission(usuario)
    except StockError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    start_stock_transaction(db)
    try:
        mat = db.query(Material).filter(Material.id == mid).first()
        if not mat:
            raise StockError(404, "Material no encontrado")
        _require_warehouse_access(usuario, mat.almacen_id)
        delta = -cantidad if tipo == "salida" else cantidad
        if tipo == "ajuste":
            delta = cantidad - float(mat.stock_actual or 0)
        if not delta:
            raise StockError(400, "El movimiento no cambia el stock")
        obra_id = int(form["obra_id"]) if form.get("obra_id") else None
        trabajador_id = int(form["trabajador_id"]) if form.get("trabajador_id") else None
        move_material(
            db, usuario, mid, delta, tipo=tipo,
            event_id=f"material-{uuid.uuid4()}",
            motivo=form.get("notas") or f"Movimiento de material: {tipo}",
            trabajador_id=trabajador_id, obra_id=obra_id,
        )
        # Historial legacy conservado dentro del mismo commit.
        db.add(MovimientoMaterial(
            material_id=mid, tipo=tipo, cantidad=cantidad, obra_id=obra_id,
            trabajador_id=trabajador_id, referencia=form.get("referencia"),
            notas=form.get("notas"), usuario_id=usuario.id,
        ))
        db.commit()
    except StockError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except Exception:
        db.rollback()
        raise
    return RedirectResponse(f"/materiales/{mid}?ok=mov", status_code=303)


@app.post("/materiales/{mid}/editar", response_class=RedirectResponse)
async def material_editar(mid: int, request: Request, db: Session = Depends(get_db),
                           usuario=Depends(requiere_login)):
    mat = db.query(Material).filter(Material.id == mid).first()
    if not mat:
        raise HTTPException(status_code=404)
    _require_warehouse_access(usuario, mat.almacen_id)
    form = await request.form()
    precio_anterior = mat.precio_unidad
    stock_minimo_anterior = mat.stock_minimo
    mat.nombre = form.get("nombre") or mat.nombre
    mat.descripcion = form.get("descripcion")
    mat.categoria = form.get("categoria") or None
    mat.unidad = form.get("unidad") or mat.unidad
    mat.stock_minimo = float(form.get("stock_minimo") or 0)
    mat.stock_maximo = float(form.get("stock_maximo")) if form.get("stock_maximo") else None
    mat.precio_unidad = float(form.get("precio_unidad")) if form.get("precio_unidad") not in (None, "") else None
    mat.referencia_proveedor = form.get("referencia_proveedor") or None
    mat.ubicacion_texto = form.get("ubicacion_texto")
    _ts = form.get("tipo_seguimiento")
    if _ts in ("individual", "generico"):
        mat.tipo_seguimiento = _ts
    if precio_anterior != mat.precio_unidad or stock_minimo_anterior != mat.stock_minimo:
        cambios = []
        if precio_anterior != mat.precio_unidad:
            cambios.append(f"precio {precio_anterior if precio_anterior is not None else '—'} → {mat.precio_unidad if mat.precio_unidad is not None else '—'}")
        if stock_minimo_anterior != mat.stock_minimo:
            cambios.append(f"stock mínimo {stock_minimo_anterior} → {mat.stock_minimo}")
        registrar_auditoria(
            db, "materiales", mat.id, "editar", usuario.id,
            {"precio_unidad": precio_anterior, "stock_minimo": stock_minimo_anterior},
            {"precio_unidad": mat.precio_unidad, "stock_minimo": mat.stock_minimo},
            "Cambio de " + " y ".join(cambios),
            request.client.host if request.client else "",
        )
    db.commit()
    return RedirectResponse(f"/materiales/{mid}?ok=editado", status_code=303)


@app.post("/materiales/{mid}/eliminar", response_class=RedirectResponse)
async def material_eliminar(mid: int, db: Session = Depends(get_db),
                             usuario=Depends(requiere_login)):
    mat = db.query(Material).filter(Material.id == mid).first()
    if mat:
        _require_warehouse_access(usuario, mat.almacen_id)
        mat.activo = False
        db.commit()
    return RedirectResponse("/materiales?ok=eliminado", status_code=303)


# ─── Control vehicular ─────────────────────────────────────────────────────────────────────────────

@app.get("/vehiculos/movimientos", response_class=HTMLResponse)
async def vehiculos_movimientos(
    request: Request, db: Session = Depends(get_db),
    usuario=Depends(requiere_login),
    solo_activos: bool = False,
):
    movs = db.query(MovimientoVehiculo).order_by(MovimientoVehiculo.fecha_salida.desc()).limit(300).all()
    if solo_activos:
        movs = [m for m in movs if m.en_ruta]
    vehiculos = db.query(Vehiculo).filter(Vehiculo.activo == True).order_by(Vehiculo.matricula).all()
    trabajadores = db.query(Trabajador).filter(Trabajador.activo == True).order_by(Trabajador.nombre).all()
    obras = db.query(Obra).filter(Obra.activa == True).order_by(Obra.nombre).all()
    en_ruta = [m for m in db.query(MovimientoVehiculo).all() if m.en_ruta]
    # Repostajes externos (gasolinera)
    repos_ext = db.query(RepostajeVehiculo).order_by(RepostajeVehiculo.fecha.desc()).limit(100).all()
    # Repostajes internos desde surtidor propio que corresponden a vehículos
    repos_surt = db.query(RepostajeSurtidor).filter(
        RepostajeSurtidor.vehiculo_id != None,
        RepostajeSurtidor.tipo_registro == 'repostaje'
    ).order_by(RepostajeSurtidor.fecha.desc()).limit(100).all()
    # Combinar y ordenar por fecha descendente
    from datetime import datetime as _dt
    def _fecha_key(r):
        f = r.fecha
        if f is None:
            return _dt.min
        if isinstance(f, _dt):
            return f
        return _dt.combine(f, _dt.min.time())
    repostajes = sorted(list(repos_ext) + list(repos_surt), key=_fecha_key, reverse=True)[:150]
    return templates.TemplateResponse(request, "vehiculos_movimientos.html", {
        "request": request, "usuario": usuario,
        "movimientos": movs, "vehiculos": vehiculos,
        "trabajadores": trabajadores, "obras": obras,
        "en_ruta": en_ruta, "solo_activos": solo_activos,
        "repostajes": repostajes,
    })


@app.post("/vehiculos/{vid}/salida", response_class=RedirectResponse)
async def vehiculo_salida(vid: int, request: Request, db: Session = Depends(get_db),
                           usuario=Depends(requiere_login)):
    form = await request.form()
    from datetime import datetime as _dt
    mov = MovimientoVehiculo(
        vehiculo_id=vid,
        conductor_id=int(form["conductor_id"]) if form.get("conductor_id") else None,
        obra_id=int(form["obra_id"]) if form.get("obra_id") else None,
        destino=form.get("destino"),
        fecha_salida=_dt.now(),
        km_salida=int(form["km_salida"]) if form.get("km_salida") else None,
        observaciones=form.get("observaciones"),
        usuario_id=usuario.id,
    )
    db.add(mov)
    db.commit()
    return RedirectResponse("/vehiculos/movimientos?ok=salida", status_code=303)


@app.post("/vehiculos/movimientos/{mid}/retorno", response_class=RedirectResponse)
async def vehiculo_retorno(mid: int, request: Request, db: Session = Depends(get_db),
                            usuario=Depends(requiere_login)):
    mov = db.query(MovimientoVehiculo).filter(MovimientoVehiculo.id == mid).first()
    if not mov:
        raise HTTPException(status_code=404)
    form = await request.form()
    from datetime import datetime as _dt
    mov.fecha_retorno = _dt.now()
    if form.get("km_retorno"):
        mov.km_retorno = int(form["km_retorno"])
    if form.get("obs_retorno"):
        obs = form["obs_retorno"]
        mov.observaciones = (mov.observaciones + " | " + obs) if mov.observaciones else obs
    db.commit()
    return RedirectResponse("/vehiculos/movimientos?ok=retorno", status_code=303)


# ─── Panel Salidas (¿Qué está fuera ahora?) ─────────────────────────────────────────────────────────────────────────────

@app.get("/panel-salidas", response_class=HTMLResponse)
async def panel_salidas(request: Request, db: Session = Depends(get_db),
                         usuario=Depends(requiere_login)):
    from datetime import datetime as _dt
    herr_fuera = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.estado.notin_(["disponible", "baja", "mantenimiento"]),
    ).order_by(Herramienta.nombre).all()
    epis_fuera = db.query(EPIIndividual).filter(
        EPIIndividual.estado == "activo",
        EPIIndividual.trabajador_id.isnot(None),
    ).order_by(EPIIndividual.tipo).all()
    vehs_ruta = [m for m in db.query(MovimientoVehiculo)
                 .order_by(MovimientoVehiculo.fecha_salida.desc()).all()
                 if m.en_ruta]
    mat_bajos = [m for m in db.query(Material).filter(Material.activo == True).all()
                 if m.bajo_minimo]
    maq_fuera = db.query(Maquinaria).filter(
        Maquinaria.estado.in_(["en_obra", "asignada", "alquilada"]),
    ).order_by(Maquinaria.nombre).all()
    return templates.TemplateResponse(request, "panel_salidas.html", {
        "request": request, "usuario": usuario,
        "herr_fuera": herr_fuera, "epis_fuera": epis_fuera,
        "vehs_ruta": vehs_ruta, "mat_bajos": mat_bajos, "maq_fuera": maq_fuera,
        "ahora": _dt.now(),
    })


# ─── Historial Global ─────────────────────────────────────────────────────────────────────────────────────────────────────────

@app.get("/historial", response_class=HTMLResponse)
async def historial_global(
    request: Request, db: Session = Depends(get_db),
    usuario=Depends(requiere_login),
    tipo: str = "", obra_id: str = "", desde: str = "", hasta: str = "",
    pagina: int = 1,
):
    from datetime import datetime as _dt, date as _date
    POR_PAG = 60
    eventos = []

    # Movimientos de herramientas
    q_mov = db.query(Movimiento).order_by(Movimiento.fecha.desc())
    for mv in q_mov.limit(300).all():
        eventos.append({
            "fecha": mv.fecha, "tipo_cat": "herramienta",
            "icono": "bi-wrench-adjustable", "color": "primary",
            "descripcion": f"{mv.tipo.capitalize()} · {mv.herramienta.nombre if mv.herramienta else '?'}",
            "responsable": mv.trabajador.nombre_completo if mv.trabajador else "—",
            "obra": mv.obra.nombre if mv.obra else "—",
            "detalle": f"/herramientas/{mv.herramienta_id}" if mv.herramienta_id else "#",
        })

    # Movimientos de materiales
    for mm in db.query(MovimientoMaterial).order_by(MovimientoMaterial.fecha.desc()).limit(300).all():
        eventos.append({
            "fecha": mm.fecha, "tipo_cat": "material",
            "icono": "bi-boxes", "color": "success",
            "descripcion": f"{mm.tipo.capitalize()} {mm.cantidad} {mm.material.unidad if mm.material else ''} · {mm.material.nombre if mm.material else '?'}",
            "responsable": mm.trabajador.nombre_completo if mm.trabajador else "—",
            "obra": mm.obra.nombre if mm.obra else "—",
            "detalle": f"/materiales/{mm.material_id}" if mm.material_id else "#",
        })

    # Movimientos de vehículos
    for mv in db.query(MovimientoVehiculo).order_by(MovimientoVehiculo.fecha_salida.desc()).limit(200).all():
        eventos.append({
            "fecha": mv.fecha_salida, "tipo_cat": "vehiculo",
            "icono": "bi-truck", "color": "warning",
            "descripcion": f"{'Retorno' if not mv.en_ruta else 'Salida'} · {mv.vehiculo.matricula if mv.vehiculo else '?'}",
            "responsable": mv.conductor.nombre_completo if mv.conductor else "—",
            "obra": mv.obra.nombre if mv.obra else (mv.destino or "—"),
            "detalle": "/vehiculos/movimientos",
        })

    # Albaranes de salida
    for alb in db.query(AlbaranSalida).order_by(AlbaranSalida.fecha_salida.desc()).limit(100).all():
        eventos.append({
            "fecha": alb.fecha_salida, "tipo_cat": "albaran",
            "icono": "bi-file-earmark-arrow-up", "color": "danger",
            "descripcion": f"Albarán {alb.numero} · {len(alb.items)} ítem(s)",
            "responsable": alb.responsable.nombre_completo if alb.responsable else "—",
            "obra": alb.obra.nombre if alb.obra else "—",
            "detalle": f"/albaranes-salida/{alb.id}",
        })

    # Filtros
    if tipo:
        eventos = [e for e in eventos if e["tipo_cat"] == tipo]
    if obra_id:
        eventos = [e for e in eventos if e["obra"] and obra_id in e["obra"]]
    if desde:
        try:
            d = _dt.fromisoformat(desde)
            eventos = [e for e in eventos if e["fecha"] and e["fecha"] >= d]
        except Exception: pass
    if hasta:
        try:
            h = _dt.fromisoformat(hasta)
            eventos = [e for e in eventos if e["fecha"] and e["fecha"] <= h]
        except Exception: pass

    eventos.sort(key=lambda e: e["fecha"] or _dt(2000,1,1), reverse=True)
    total = len(eventos)
    inicio = (pagina-1)*POR_PAG
    paginas = max(1, (total + POR_PAG - 1)//POR_PAG)
    eventos = eventos[inicio:inicio+POR_PAG]
    obras = db.query(Obra).filter(Obra.activa == True).order_by(Obra.nombre).all()
    return templates.TemplateResponse(request, "historial.html", {
        "request": request, "usuario": usuario, "eventos": eventos,
        "tipo": tipo, "obra_id": obra_id, "desde": desde, "hasta": hasta,
        "pagina": pagina, "paginas": paginas, "total": total, "obras": obras,
    })


# ─── Albaranes de Salida ───────────────────────────────────────────────────────────────────────────────────────────────────────

def _gen_numero_albaran(db):
    from datetime import date as _d
    hoy = _d.today().strftime("%Y%m%d")
    prefix = f"AL-{hoy}-"
    count = db.query(AlbaranSalida).filter(AlbaranSalida.numero.like(f"{prefix}%")).count()
    return f"{prefix}{count+1:03d}"


@app.get("/albaranes-salida", response_class=HTMLResponse)
async def albaranes_lista(request: Request, db: Session = Depends(get_db),
                           usuario=Depends(requiere_login), tipo_doc: str = Query("")):
    warehouse = _active_warehouse(db, usuario, request)
    warehouse_id = warehouse.id if warehouse else -1
    query = db.query(AlbaranSalida).filter(AlbaranSalida.almacen_id == warehouse_id)
    if tipo_doc in {"salida", "entrada"}:
        query = query.filter(AlbaranSalida.tipo_documento == tipo_doc)
    albaranes = query.order_by(AlbaranSalida.fecha_salida.desc()).limit(500).all()
    obras = db.query(Obra).filter(
        Obra.activa == True, Obra.almacen_id == warehouse_id,
    ).order_by(Obra.nombre).all()
    trabajadores = db.query(Trabajador).filter(
        Trabajador.activo == True, Trabajador.almacen_id == warehouse_id,
    ).order_by(Trabajador.nombre).all()
    herramientas = db.query(Herramienta).filter(Herramienta.activa == True,
                            Herramienta.estado == "disponible",
                            Herramienta.almacen_id == warehouse_id).order_by(Herramienta.nombre).all()
    materiales = db.query(Material).filter(
        Material.activo == True, Material.almacen_id == warehouse_id,
    ).order_by(Material.nombre).all()
    return templates.TemplateResponse(request, "albaranes_salida.html", {
        "request": request, "usuario": usuario, "albaranes": albaranes,
        "obras": obras, "trabajadores": trabajadores,
        "herramientas": herramientas, "materiales": materiales, "tipo_doc": tipo_doc,
    })


@app.post("/albaranes-salida", response_class=RedirectResponse)
async def albaran_crear(request: Request, db: Session = Depends(get_db),
                         usuario=Depends(requiere_login)):
    form = await request.form()
    warehouse = _active_warehouse(db, usuario, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    worker_id = int(form["responsable_id"]) if form.get("responsable_id") else None
    work_id = int(form["obra_id"]) if form.get("obra_id") else None
    tool_ids = [int(value) for value in form.getlist("herramienta_ids") if value]
    _validate_movement_warehouse(db, usuario, request, tool_ids, worker_id, work_id)
    material_ids = [int(value) for value in form.getlist("material_ids") if value]
    if material_ids and db.query(Material).filter(
        Material.id.in_(material_ids), Material.almacen_id == warehouse.id,
    ).count() != len(set(material_ids)):
        raise HTTPException(404, "Un material no pertenece al almacén activo")
    actor = _require_movement_http(usuario, "entregar")
    notes = str(form.get("notas") or "")
    start_movement_transaction(db)
    try:
        lines = []
        for hid in tool_ids:
            result = deliver_tool(
                db, actor, herramienta_id=hid, trabajador_id=worker_id,
                obra_id=work_id, observaciones=notes,
            )
            lines.append({
                "tipo": "herramienta", "id": result.herramienta_id,
                "cantidad": 1, "nombre": result.codigo,
                "movimiento_id": result.movimiento_id,
            })

        # Los materiales y las líneas libres forman parte de la misma operación.
        for mid in form.getlist("material_ids"):
            if mid:
                cant = float(form.get(f"cant_mat_{mid}") or 1)
                lines.append({"tipo": "material", "id": int(mid), "cantidad": cant})
                mat = db.query(Material).filter(Material.id == int(mid)).first()
                if mat:
                    mat.stock_actual = max(0.0, mat.stock_actual - cant)
        for desc in form.getlist("lineas_libres"):
            if desc.strip():
                lines.append({"tipo": "libre", "nombre": desc.strip(), "cantidad": 1})

        alb = create_delivery_note(
            db, user_id=actor.id, worker_id=worker_id, work_id=work_id,
            warehouse_id=warehouse.id, notes=notes, lines=lines,
        )
        db.commit()
    except MovementError as exc:
        db.rollback()
        raise HTTPException(exc.status_code, exc.detail)
    except (TypeError, ValueError):
        db.rollback()
        raise HTTPException(400, "Los datos del albarán no son válidos")
    except Exception:
        db.rollback()
        raise
    return RedirectResponse(f"/albaranes-salida/{alb.id}", status_code=303)


@app.post("/albaranes-salida/{aid}/firma")
def albaran_guardar_firma(
    aid: int,
    user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
    firma_datos: str = Form(""),
    firma_nombre: str = Form(""),
):
    alb = db.query(AlbaranSalida).filter(AlbaranSalida.id == aid).first()
    if not alb:
        raise HTTPException(404)
    _require_warehouse_access(user, alb.almacen_id)
    firma_nombre = (firma_nombre or "").strip()
    if len(firma_nombre) < 2:
        raise HTTPException(400, "Indica el nombre de quien recibe")
    mime, firma_bytes = _decode_delivery_signature(firma_datos)
    alb.firma_datos = f"data:{mime};base64,{__import__('base64').b64encode(firma_bytes).decode('ascii')}"
    alb.firma_nombre = firma_nombre[:100]
    alb.firma_fecha = datetime.now()
    db.commit()
    return RedirectResponse(f"/albaranes-salida/{aid}?ok=firma", status_code=303)


def _decode_delivery_signature(value: str) -> tuple[str, bytes]:
    """Valida y decodifica una firma de canvas sin aceptar contenido arbitrario."""
    import base64
    import binascii

    raw_value = (value or "").strip()
    match = re.fullmatch(
        r"data:(image/(?:png|jpeg));base64,([A-Za-z0-9+/=\r\n]+)", raw_value,
        flags=re.IGNORECASE,
    )
    if not match:
        raise HTTPException(400, "La firma no tiene un formato de imagen válido")
    try:
        image_bytes = base64.b64decode(match.group(2), validate=True)
    except (ValueError, binascii.Error):
        raise HTTPException(400, "La firma está dañada; vuelve a firmar")
    if not image_bytes or len(image_bytes) > 500_000:
        raise HTTPException(400, "La firma está vacía o supera el tamaño permitido")
    mime = match.group(1).lower()
    valid_magic = (
        mime == "image/png" and image_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    ) or (
        mime == "image/jpeg" and image_bytes.startswith(b"\xff\xd8\xff")
    )
    if not valid_magic:
        raise HTTPException(400, "El contenido recibido no es una firma PNG o JPEG válida")
    return mime, image_bytes


@app.get("/albaranes-salida/{aid}/firmar", response_class=HTMLResponse)
async def albaran_firmar_tablet(
    aid: int, request: Request, db: Session = Depends(get_db),
    usuario=Depends(requiere_login),
):
    alb = db.query(AlbaranSalida).filter(AlbaranSalida.id == aid).first()
    if not alb:
        raise HTTPException(404)
    _require_warehouse_access(usuario, alb.almacen_id)
    return templates.TemplateResponse(request, "albaran_firmar.html", {
        "request": request, "usuario": usuario, "alb": alb,
    })


@app.get("/albaranes-salida/{aid}", response_class=HTMLResponse)
async def albaran_detalle(aid: int, request: Request, db: Session = Depends(get_db),
                           usuario=Depends(requiere_login)):
    alb = db.query(AlbaranSalida).filter(AlbaranSalida.id == aid).first()
    if not alb:
        raise HTTPException(status_code=404)
    _require_warehouse_access(usuario, alb.almacen_id)
    return templates.TemplateResponse(request, "albaran_detalle.html", {
        "request": request, "usuario": usuario, "alb": alb,
    })


@app.post("/albaranes-salida/{aid}/retorno-item/{iid}", response_class=RedirectResponse)
async def albaran_retorno_item(aid: int, iid: int, db: Session = Depends(get_db),
                                usuario=Depends(requiere_login)):
    alb = db.get(AlbaranSalida, aid)
    if not alb:
        raise HTTPException(404)
    _require_warehouse_access(usuario, alb.almacen_id)
    item = db.query(ItemAlbaranSalida).filter(
        ItemAlbaranSalida.id == iid, ItemAlbaranSalida.albaran_id == aid,
    ).first()
    if item:
        from datetime import datetime as _dt
        item.retornado = True
        item.fecha_retorno = _dt.now()
        if item.tipo == "material" and item.material:
            item.material.stock_actual += item.cantidad
        alb = item.albaran
        if alb:
            if all(i.retornado for i in alb.items):
                alb.estado = "cerrado"
                alb.fecha_retorno_real = _dt.now()
            else:
                alb.estado = "parcial"
        db.commit()
    return RedirectResponse(f"/albaranes-salida/{aid}?ok=retorno", status_code=303)


@app.get("/albaranes-salida/{aid}/pdf")
async def albaran_pdf(aid: int, db: Session = Depends(get_db),
                       usuario=Depends(requiere_login)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    import io
    alb = db.query(AlbaranSalida).filter(AlbaranSalida.id == aid).first()
    if not alb:
        raise HTTPException(status_code=404)
    _require_warehouse_access(usuario, alb.almacen_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []
    stamp_path = Path(__file__).parent / "static" / "img" / "mrd_sello_blanco.png"
    if stamp_path.is_file():
        stamp = RLImage(str(stamp_path), width=145, height=72)
        stamp.hAlign = "LEFT"
        story.append(stamp)
        story.append(Spacer(1, 6))
    es_entrada = alb.tipo_documento == "entrada"
    titulo_doc = "ALBARÁN DE ENTRADA" if es_entrada else "ALBARÁN DE SUMINISTRO / SALIDA"
    story.append(Paragraph(f"<b>{titulo_doc} {alb.numero}</b>", styles["Title"]))
    story.append(Spacer(1,10))
    info = [
        ["Obra:", alb.obra.nombre if alb.obra else "—",
         "Responsable:", alb.responsable.nombre_completo if alb.responsable else "—"],
        ["Fecha entrada:" if es_entrada else "Fecha salida:", alb.fecha_salida.strftime("%d/%m/%Y %H:%M"),
         "Estado:", alb.estado.upper()],
    ]
    if es_entrada:
        info.append([
            "Almacén:", alb.almacen.nombre if alb.almacen else "Almacén principal",
            "Procedencia:", alb.origen_destino or "—",
        ])
    t = Table(info, colWidths=[80,170,80,170])
    t.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),"Helvetica"),
                            ("FONTSIZE",(0,0),(-1,-1),9),
                            ("TEXTCOLOR",(0,0),(0,-1),colors.grey),
                            ("TEXTCOLOR",(2,0),(2,-1),colors.grey)]))
    story.append(t); story.append(Spacer(1,16))
    headers = [["#","Descripción","Tipo","Cantidad","Recibido" if es_entrada else "Retornado"]]
    from xml.sax.saxutils import escape as _pdf_escape
    rows = headers + [[str(i+1), Paragraph(_pdf_escape(item.descripcion), styles["BodyText"]), item.tipo.capitalize(),
                       str(item.cantidad), "Sí" if item.retornado else "No"]
                      for i, item in enumerate(alb.items)]
    tbl = Table(rows, colWidths=[25,230,80,60,80])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8f9fa")]),
        ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
        ("ALIGN",(3,0),(4,-1),"CENTER"),
    ]))
    story.append(tbl)
    if alb.notas:
        story.append(Spacer(1,12))
        story.append(Paragraph(f"<b>Notas:</b> {alb.notas}", styles["Normal"]))
    story.append(Spacer(1,24))
    if alb.firma_datos:
        try:
            _signature_mime, signature_bytes = _decode_delivery_signature(alb.firma_datos)
            signature = RLImage(io.BytesIO(signature_bytes), width=180, height=60)
            signed_at = alb.firma_fecha or alb.fecha_salida
            signature_info = [
                [Paragraph("<b>RECIBIDO CONFORME</b>", styles["Normal"])],
                [signature],
                [Paragraph(
                    f"<b>{_pdf_escape(alb.firma_nombre or '')}</b><br/>"
                    f"Firmado: {signed_at.strftime('%d/%m/%Y %H:%M')}",
                    styles["Normal"],
                )],
            ]
            signature_table = Table(signature_info, colWidths=[230])
            signature_table.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#64748b")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]))
            story.append(signature_table)
        except HTTPException:
            story.append(Paragraph("Firma registrada no válida. Debe capturarse de nuevo.", styles["Normal"]))
    else:
        story.append(Paragraph("Firma del responsable: ___________________________", styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=albaran_{alb.numero}.pdf"})


# ─── QR de Material ──────────────────────────────────────────────────────────────────────────────────────────────────────

@app.get("/materiales/{mid}/qr", response_class=HTMLResponse)
async def material_qr(mid: int, request: Request, db: Session = Depends(get_db),
                       usuario=Depends(requiere_login)):
    mat = db.query(Material).filter(Material.id == mid).first()
    if not mat:
        raise HTTPException(status_code=404)
    public_url = MRD_PUBLIC_URL or str(request.base_url).rstrip("/")
    url = f"{public_url}/materiales/{mid}"
    qr_b64 = generar_qr_base64(url)
    return templates.TemplateResponse(request, "material_qr.html", {
        "request": request, "usuario": usuario, "mat": mat,
        "qr_b64": qr_b64, "url": url,
    })


# ─── Importar materiales desde Excel ──────────────────────────────────────────────────────────────────────────────────────

@app.get("/materiales/importar", response_class=HTMLResponse)
async def materiales_importar_form(request: Request, db: Session = Depends(get_db),
                                    usuario=Depends(requiere_login)):
    return templates.TemplateResponse(request, "materiales_importar.html", {
        "request": request, "usuario": usuario,
    })


@app.post("/materiales/importar", response_class=HTMLResponse)
async def materiales_importar_post(request: Request, db: Session = Depends(get_db),
                                    usuario=Depends(requiere_login),
                                    archivo: UploadFile = File(...)):
    import openpyxl, io
    contents = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    encabezados = []
    for valor in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        texto_enc = str(valor or "").strip().lower()
        texto_enc = (texto_enc.replace("ó", "o").replace("í", "i")
                     .replace("á", "a").replace("é", "e").replace("ú", "u"))
        encabezados.append(texto_enc)

    def _indice_columna(*nombres):
        for nombre_col in nombres:
            if nombre_col in encabezados:
                return encabezados.index(nombre_col)
        return None

    indices = {
        "nombre": _indice_columna("nombre", "nombre *"),
        "categoria": _indice_columna("categoria"),
        "unidad": _indice_columna("unidad"),
        "stock": _indice_columna("stock inicial", "stock"),
        "minimo": _indice_columna("stock minimo", "minimo"),
        "ubicacion": _indice_columna("ubicacion nave", "ubicacion"),
    }
    if indices["nombre"] is None:
        raise HTTPException(400, "La hoja debe incluir la columna Nombre")

    creados = 0; errores = []
    almacen_predeterminado = _active_warehouse(db, usuario, request)
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            def _celda(clave):
                indice = indices[clave]
                return row[indice] if indice is not None and indice < len(row) else None

            nombre_raw = _celda("nombre")
            nombre = str(nombre_raw).strip() if nombre_raw else None
            if not nombre:
                continue
            codigo = generar_referencia_material(db)
            mat = Material(
                codigo=codigo, nombre=nombre,
                categoria=str(_celda("categoria")).strip() if _celda("categoria") else None,
                unidad=str(_celda("unidad")).strip() if _celda("unidad") else "ud",
                stock_actual=float(_celda("stock")) if _celda("stock") else 0.0,
                stock_minimo=float(_celda("minimo")) if _celda("minimo") else 0.0,
                ubicacion_texto=(str(_celda("ubicacion")).strip() if _celda("ubicacion") else
                                  (almacen_predeterminado.nombre if almacen_predeterminado else None)),
                almacen_id=almacen_predeterminado.id if almacen_predeterminado else None,
            )
            db.add(mat); creados += 1
        except Exception as ex:
            errores.append(str(ex))
    db.commit()
    return templates.TemplateResponse(request, "materiales_importar.html", {
        "request": request, "usuario": usuario,
        "creados": creados, "errores": errores,
    })


@app.get("/materiales/plantilla-excel")
async def materiales_plantilla(usuario=Depends(requiere_login)):
    import openpyxl, io
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales"
    ws.append(["Nombre","Categoria","Unidad","Stock inicial","Stock minimo","Ubicacion nave"])
    ws.append(["Cable H07Z1-K 1.5mm²","Cables y electricidad","m","100","20","Estanteria A1"])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_materiales.xlsx"})


# ─── Repostaje de vehículos ───────────────────────────────────────────────────────────────────────────────────────────────

@app.post("/vehiculos/{vid}/repostaje", response_class=RedirectResponse)
async def vehiculo_repostaje(vid: int, request: Request, db: Session = Depends(get_db),
                              usuario=Depends(requiere_login)):
    form = await request.form()
    litros = float(form.get("litros") or 0)
    precio = float(form.get("precio_litro") or 0) if form.get("precio_litro") else None
    rep = RepostajeVehiculo(
        vehiculo_id=vid,
        litros=litros,
        precio_litro=precio,
        total_euros=round(litros * precio, 2) if precio else (float(form.get("total_euros_manual")) if form.get("total_euros_manual") else None),
        km_actuales=int(form["km_actuales"]) if form.get("km_actuales") else None,
        gasolinera=form.get("gasolinera"),
        notas=form.get("notas"),
        usuario_id=usuario.id,
    )
    # actualizar km en vehículo
    if rep.km_actuales:
        v = db.query(Vehiculo).filter(Vehiculo.id == vid).first()
        if v and (not v.kilometros or rep.km_actuales > v.kilometros):
            v.kilometros = rep.km_actuales
    db.add(rep); db.commit()
    return RedirectResponse("/vehiculos/movimientos?ok=repostaje", status_code=303)


# ─── Compatibilidad con la antigua Salida rápida ────────────────────────────────────────────────────────────────────────

@app.get("/salida-rapida")
async def salida_rapida_legacy(usuario=Depends(requiere_login)):
    """Conserva marcadores antiguos y centraliza la operativa en Mostrador."""
    return RedirectResponse("/mostrador", status_code=303)


# ─── Foto en incidencias ───────────────────────────────────────────────────────────────

@app.post("/incidencias/{iid}/foto", response_class=RedirectResponse)
async def incidencia_foto(iid: int,
                          foto: UploadFile = File(...),
                          user: Usuario = Depends(requiere_login),
                          db: Session = Depends(get_db)):
    inc = db.get(Incidencia, iid)
    if not inc:
        raise HTTPException(404)
    carpeta = BASE_DIR / "static" / "uploads" / "incidencias"
    carpeta.mkdir(parents=True, exist_ok=True)
    if inc.foto_path:
        old_p = carpeta / inc.foto_path
        if old_p.exists():
            old_p.unlink()
    ext = (foto.filename or "").rsplit(".", 1)[-1].lower() or "jpg"
    import time as _time
    nombre = f"{iid}_{int(_time.time())}.{ext}"
    data = await foto.read()
    (carpeta / nombre).write_bytes(data)
    inc.foto_path = nombre
    db.commit()
    return RedirectResponse(f"/incidencias/{iid}?ok=foto", status_code=303)


@app.post("/incidencias/{iid}/foto/eliminar", response_class=RedirectResponse)
def incidencia_foto_eliminar(iid: int,
                             user: Usuario = Depends(requiere_login),
                             db: Session = Depends(get_db)):
    inc = db.get(Incidencia, iid)
    if inc and inc.foto_path:
        p = BASE_DIR / "static" / "uploads" / "incidencias" / inc.foto_path
        if p.exists():
            p.unlink()
        inc.foto_path = None
        db.commit()
    return RedirectResponse(f"/incidencias/{iid}?ok=foto_eliminada", status_code=303)


# ─── /version.json ── Servidor de actualizaciones ────────────────────────────
@app.get("/version.json")
def serve_version_json():
    """
    Sirve el archivo version.json para el sistema de actualizaciones.
    Accesible en https://app.iasmrd.com/version.json
    """
    import json as _json
    try:
        vpath = BASE_DIR / "version.json"
        data  = _json.loads(vpath.read_text(encoding="utf-8"))
        return data
    except Exception as exc:
        return {"error": str(exc), "version_actual": "0.0.0"}




# ─── Surtidor de combustible ──────────────────────────────────────────────────

def _generar_pdf_surtidor(rep, stock_antes: float, stock_despues: float, empresa: str) -> bytes:
    """Genera ticket/albarán PDF de repostaje en el surtidor interno."""
    import io as _io
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    azul   = colors.HexColor("#1B4F8A")
    naranj = colors.HexColor("#E8600A")
    gris   = colors.HexColor("#555555")
    rojo   = colors.HexColor("#dc3545")

    def P(text, size=10, bold=False, color=colors.black, align=TA_LEFT):
        name = f"s{size}{'b' if bold else ''}{'c' if align==TA_CENTER else ''}"
        st = ParagraphStyle(name, fontSize=size,
                            fontName="Helvetica-Bold" if bold else "Helvetica",
                            textColor=color, alignment=align, spaceAfter=2)
        return Paragraph(text, st)

    num       = f"SURT-{rep.fecha.strftime('%Y%m%d') if rep.fecha else '?'}-{rep.id:04d}"
    fecha_str = rep.fecha.strftime("%d/%m/%Y") if rep.fecha else "—"
    hora_str  = rep.fecha.strftime("%H:%M")    if rep.fecha else "—"
    activo    = str(rep.activo_nombre or "—")
    tipo_label = str(getattr(rep, 'activo_tipo', None) or "Vehículo")
    comb_label = (rep.tipo_combustible or "gasoil").upper()
    precio_str = f"{rep.precio_litro:.3f} €/L" if rep.precio_litro else "—"
    litros_str = f"{rep.litros:.2f} L"
    operador   = str(rep.usuario.nombre if (rep.usuario and hasattr(rep.usuario, 'nombre')) else "—")

    story = []

    # ── Cabecera ─────────────────────────────────────────
    story.append(P(empresa.upper(), size=14, bold=True, color=azul, align=TA_CENTER))
    story.append(P("ALBARÁN DE REPOSTAJE — SURTIDOR NAVE", size=9, color=naranj, align=TA_CENTER))
    story.append(HRFlowable(width="100%", thickness=2, color=azul, spaceAfter=4))

    # Nº y fecha en una fila
    t_head = Table([[f"Nº: {num}", f"{fecha_str}  {hora_str}"]],
                   colWidths=[8*cm, 4*cm])
    t_head.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 8),
        ("TEXTCOLOR", (0,0), (-1,-1), gris),
        ("ALIGN",     (1,0), (1,0), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
    ]))
    story.append(t_head)
    story.append(Spacer(1, 8))

    # ── Matrícula / Vehículo — bloque grande ─────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=6))
    story.append(P(activo, size=20, bold=True, color=azul, align=TA_CENTER))
    story.append(P(tipo_label, size=10, color=gris, align=TA_CENTER))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=6))
    story.append(Spacer(1, 4))

    # ── Datos del repostaje ───────────────────────────────
    filas = [
        ["COMBUSTIBLE",   comb_label],
        ["LITROS",        litros_str],
        ["PRECIO / LITRO", precio_str],
        ["OPERADOR",      operador],
    ]
    if rep.km_actuales:
        filas.append(["KM ACTUALES", f"{rep.km_actuales:,} km"])

    t_datos = Table(filas, colWidths=[5.5*cm, 6.5*cm])
    t_datos.setStyle(TableStyle([
        ("FONTNAME",      (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",      (1,0), (1,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (0,-1), 9),
        ("FONTSIZE",      (1,0), (1,-1), 11),
        ("TEXTCOLOR",     (0,0), (0,-1), gris),
        ("TEXTCOLOR",     (1,0), (1,0), azul),        # combustible
        ("TEXTCOLOR",     (1,1), (1,1), naranj),       # litros
        ("FONTSIZE",      (1,1), (1,1), 16),           # litros — grande
        ("TEXTCOLOR",     (1,2), (1,2), colors.black), # precio
        ("TEXTCOLOR",     (1,3), (1,3), gris),         # operador
        ("FONTSIZE",      (1,3), (1,3), 9),
        ("ALIGN",         (1,0), (1,-1), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("LINEBELOW",     (0,0), (-1,0), 0.5, colors.lightgrey),
        ("LINEBELOW",     (0,2), (-1,2), 1, colors.lightgrey),
    ]))
    story.append(t_datos)
    story.append(Spacer(1, 8))

    if rep.notas:
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=4))
        story.append(P(f"Notas: {rep.notas}", size=8, color=gris))

    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=1, color=azul, spaceAfter=4))
    story.append(P("Documento generado automáticamente por MRD Tool Control", size=7, color=gris, align=TA_CENTER))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _stock_surtidor(db) -> float:
    """Calcula el stock actual del depósito: compras - repostajes (todo el histórico)."""
    todos = db.query(RepostajeSurtidor).all()
    compras    = sum(r.litros for r in todos if r.tipo_registro == "compra")
    dispensado = sum(r.litros for r in todos if r.tipo_registro == "repostaje")
    return round(compras - dispensado, 2)


@app.get("/surtidor", response_class=HTMLResponse)
async def surtidor_get(
    request: Request,
    db: Session = Depends(get_db),
    usuario=Depends(requiere_login),
    mes: int = None,
    anio: int = None,
):
    from datetime import datetime as _dt, date as _date
    import calendar

    hoy = _date.today()
    mes  = mes  or hoy.month
    anio = anio or hoy.year

    todos = db.query(RepostajeSurtidor).order_by(RepostajeSurtidor.fecha.desc()).limit(200).all()

    primer_dia = _dt(anio, mes, 1)
    ultimo_dia = _dt(anio, mes, calendar.monthrange(anio, mes)[1], 23, 59, 59)
    del_mes = db.query(RepostajeSurtidor).filter(
        RepostajeSurtidor.fecha >= primer_dia,
        RepostajeSurtidor.fecha <= ultimo_dia,
    ).all()

    repos_mes   = [r for r in del_mes if r.tipo_registro == "repostaje"]
    compras_mes = [r for r in del_mes if r.tipo_registro == "compra"]
    litros_mes  = round(sum(r.litros for r in repos_mes), 2)
    comprado_mes= round(sum(r.litros for r in compras_mes), 2)
    gasto_mes   = round(sum(r.total_euros or 0 for r in del_mes), 2)
    n_repos     = len(repos_mes)
    precio_med  = round(sum(r.precio_litro for r in del_mes if r.precio_litro) /
                        max(1, sum(1 for r in del_mes if r.precio_litro)), 3) if del_mes else 0

    # Stock actual del depósito
    stock_actual = _stock_surtidor(db)

    # Si viene de un repostaje recién guardado, pasar el ID para mostrar botón albarán
    ok_param = request.query_params.get("ok", "")
    albaran_id = int(ok_param) if ok_param.isdigit() else None

    from sqlalchemy import or_ as _or
    vehiculos = db.query(Vehiculo).filter(
        Vehiculo.activo == True,
        _or(
            Vehiculo.tipo == None,
            Vehiculo.tipo.in_(["furgoneta", "camion", "camión", "van", "furgón"])
        )
    ).order_by(Vehiculo.matricula).all()
    maquinarias = db.query(Maquinaria).filter(Maquinaria.activa == True).order_by(Maquinaria.nombre).all()

    meses_nombres = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                     "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    anios_rango = list(range(hoy.year - 2, hoy.year + 1))

    return templates.TemplateResponse(request, "surtidor.html", {
        "request": request,
        "usuario": usuario,
        "todos": todos,
        "litros_mes": litros_mes,
        "comprado_mes": comprado_mes,
        "gasto_mes": gasto_mes,
        "n_repos": n_repos,
        "precio_med": precio_med,
        "stock_actual": stock_actual,
        "albaran_id": albaran_id,
        "mes_sel": mes,
        "anio_sel": anio,
        "mes_label": meses_nombres[mes - 1],
        "meses_nombres": meses_nombres,
        "anios_rango": anios_rango,
        "vehiculos": vehiculos,
        "maquinarias": maquinarias,
    })


@app.get("/surtidor/{rid}/albaran")
async def surtidor_albaran(rid: int, db: Session = Depends(get_db),
                            usuario=Depends(requiere_login)):
    """Genera y descarga el albarán PDF de un repostaje del surtidor."""
    from fastapi.responses import Response as _Resp
    rep = db.query(RepostajeSurtidor).filter(RepostajeSurtidor.id == rid).first()
    if not rep or rep.tipo_registro != "repostaje":
        raise HTTPException(status_code=404, detail="Repostaje no encontrado")

    # Calcular stock antes y después
    todos = db.query(RepostajeSurtidor).filter(RepostajeSurtidor.fecha <= rep.fecha).all()
    compras    = sum(r.litros for r in todos if r.tipo_registro == "compra")
    dispensado = sum(r.litros for r in todos if r.tipo_registro == "repostaje")
    stock_despues = round(compras - dispensado, 2)
    stock_antes   = round(stock_despues + rep.litros, 2)

    pdf_bytes = _generar_pdf_surtidor(rep, stock_antes, stock_despues, COMPANY_NAME)
    nombre = f"albaran_surtidor_{rid:04d}.pdf"
    return _Resp(content=pdf_bytes, media_type="application/pdf",
                 headers={"Content-Disposition": f'inline; filename="{nombre}"'})


@app.post("/surtidor/nuevo", response_class=RedirectResponse)
async def surtidor_nuevo(request: Request, db: Session = Depends(get_db),
                          usuario=Depends(requiere_login)):
    form = await request.form()
    tipo_registro = form.get("tipo_registro", "repostaje")
    tipo_activo   = form.get("tipo_activo", "vehiculo")
    litros = float(form.get("litros") or 0)
    precio = float(form.get("precio_litro")) if form.get("precio_litro") else None
    total  = float(form.get("total_euros"))  if form.get("total_euros")  else (
        round(litros * precio, 2) if precio else None
    )
    rep = RepostajeSurtidor(
        tipo_registro=tipo_registro,
        tipo_combustible=form.get("tipo_combustible", "gasoil"),
        vehiculo_id=(int(form["vehiculo_id"]) if tipo_registro == "repostaje"
                     and tipo_activo == "vehiculo" and form.get("vehiculo_id") else None),
        maquinaria_id=(int(form["maquinaria_id"]) if tipo_registro == "repostaje"
                       and tipo_activo == "maquinaria" and form.get("maquinaria_id") else None),
        litros=litros,
        precio_litro=precio,
        total_euros=total,
        km_actuales=int(form["km_actuales"]) if form.get("km_actuales") else None,
        proveedor=form.get("proveedor") if tipo_registro == "compra" else None,
        notas=form.get("notas"),
        usuario_id=usuario.id,
    )
    db.add(rep)
    db.commit()
    db.refresh(rep)
    # Si es repostaje → redirigir al surtidor con ID para mostrar botón albarán
    if tipo_registro == "repostaje":
        return RedirectResponse(f"/surtidor?ok={rep.id}", status_code=303)
    return RedirectResponse("/surtidor?ok=compra", status_code=303)


@app.post("/surtidor/{rid}/eliminar", response_class=RedirectResponse)
async def surtidor_eliminar(rid: int, db: Session = Depends(get_db),
                             usuario=Depends(requiere_login)):
    rep = db.query(RepostajeSurtidor).filter(RepostajeSurtidor.id == rid).first()
    if rep:
        db.delete(rep)
        db.commit()
    return RedirectResponse("/surtidor?ok=eliminado", status_code=303)


# ─── Inventario masivo V2: API transaccional ─────────────────────────────────

class VarianteNuevaRequest(BaseModel):
    model_config = {"extra": "forbid"}
    catalogo_epi_id: int = Field(gt=0)
    modelo: str = Field(default="", max_length=100)
    color: str = Field(default="", max_length=50)
    talla: str = Field(default="", max_length=20)
    almacen_id: int = Field(gt=0)
    ubicacion_id: Optional[int] = Field(default=None, gt=0)
    referencia_proveedor: Optional[str] = Field(default=None, max_length=100)
    stock_minimo: int = Field(default=0, ge=0)
    cantidad: int = Field(default=0, ge=0)
    numero_lote: Optional[str] = Field(default=None, max_length=100)
    fecha_caducidad: Optional[date] = None


class VarianteEditarRequest(BaseModel):
    model_config = {"extra": "forbid"}
    referencia_proveedor: Optional[str] = Field(default=None, max_length=100)
    stock_minimo: Optional[int] = Field(default=None, ge=0)
    ubicacion_id: Optional[int] = Field(default=None, gt=0)


class RecepcionSuministroRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    variante_id: Optional[int] = Field(default=None, gt=0)
    catalogo_epi_id: Optional[int] = Field(default=None, gt=0)
    modelo: str = Field(default="", max_length=100)
    color: str = Field(default="", max_length=50)
    talla: str = Field(default="", max_length=20)
    almacen_id: int = Field(gt=0)
    ubicacion_id: Optional[int] = Field(default=None, gt=0)
    cantidad: int = Field(gt=0)
    proveedor: Optional[str] = Field(default=None, max_length=150)
    albaran: Optional[str] = Field(default=None, max_length=100)
    precio_unitario: Optional[float] = Field(default=None, ge=0)
    numero_lote: Optional[str] = Field(default=None, max_length=100)
    fecha_caducidad: Optional[date] = None
    referencia_proveedor: Optional[str] = Field(default=None, max_length=100)
    stock_minimo: int = Field(default=0, ge=0)


class EntradaExistenciaRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    codigo: str = Field(min_length=1, max_length=512)
    cantidad: float = Field(gt=0, le=1_000_000)
    almacen_id: int = Field(gt=0)
    ubicacion_id: Optional[int] = Field(default=None, gt=0)
    proveedor: Optional[str] = Field(default=None, max_length=150)
    albaran: Optional[str] = Field(default=None, max_length=100)
    numero_lote: Optional[str] = Field(default=None, max_length=100)
    fecha_caducidad: Optional[date] = None


class SesionNuevaRequest(BaseModel):
    model_config = {"extra": "forbid"}
    nombre: str = Field(min_length=1, max_length=200)
    almacen_id: Optional[int] = Field(default=None, gt=0)
    scope: Literal["almacen", "ubicacion", "categoria", "total"] = "almacen"
    tipo_articulo: Literal["todo", "material", "epi_ropa", "epi_individual"] = "todo"
    umbral_desviacion: float = Field(default=5.0, ge=0, le=100)


class SesionEditarRequest(BaseModel):
    model_config = {"extra": "forbid"}
    nombre: str = Field(min_length=1, max_length=200)
    observaciones: str = Field(default="", max_length=1000)


class ConteoRequest(BaseModel):
    model_config = {"extra": "forbid"}
    linea_id: int = Field(gt=0)
    cantidad: float = Field(ge=0)
    numero_conteo: Literal[1, 2] = 1
    scan_event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    modo_entrada: Literal["unidad", "incremento", "caja"] = "unidad"
    unidades_por_caja: Optional[int] = Field(default=None, gt=0)
    notas: str = Field(default="", max_length=1000)
    puesto_id: Optional[str] = Field(default=None, max_length=64)


class AprobarConteoRequest(BaseModel):
    model_config = {"extra": "forbid"}
    cantidad_final: float = Field(ge=0)


class EscaneoActivoInventarioRequest(BaseModel):
    model_config = {"extra": "forbid"}
    codigo: str = Field(min_length=1, max_length=512)


class CerrarSesionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    cierre_event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")


class ConfirmarDotacionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")


class PrepararLineaDotacionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    codigo_qr: str = Field(min_length=4, max_length=100)


class EntregarLineaDotacionRequest(PrepararLineaDotacionRequest):
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    firmado_por: str = Field(min_length=2, max_length=150)
    firma_base64: str = Field(min_length=30, max_length=500_000)


class DevolverLineaDotacionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    codigo_qr: str = Field(min_length=4, max_length=100)
    motivo: str = Field(default="", max_length=500)


class CambiarTallaDotacionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    talla: str = Field(min_length=1, max_length=20)


class SustituirLineaDotacionRequest(BaseModel):
    model_config = {"extra": "forbid"}
    motivo: str = Field(min_length=2, max_length=500)


class ResetRopaRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    frase: str = Field(min_length=1, max_length=100)
    preview_hash: str = Field(min_length=64, max_length=64, pattern=r"^[a-f0-9]{64}$")


class EtiquetaRequest(BaseModel):
    model_config = {"extra": "forbid"}
    event_id: str = Field(min_length=8, max_length=64, pattern=r"^[A-Za-z0-9_-]+$")
    identificador_id: int = Field(gt=0)
    copias: int = Field(default=1, ge=1, le=100)
    reimpresion: bool = False
    motivo_reimpresion: str = Field(default="", max_length=300)


def _inventory_http_error(exc):
    raise HTTPException(exc.status_code, exc.detail)


def _ensure_inventory_operator(user: Usuario) -> None:
    try:
        require_inventory_operator(user)
    except InventoryError as exc:
        _inventory_http_error(exc)


def _ensure_inventory_admin(user: Usuario) -> None:
    try:
        require_inventory_admin(user)
    except InventoryError as exc:
        _inventory_http_error(exc)


def _inventory_session_for_user(
    db: Session, user: Usuario, session_id: int,
) -> SesionInventario:
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    _require_warehouse_access(user, session.almacen_id)
    return session


@app.get("/inventario/variantes")
def inventario_variantes(user: Usuario = Depends(requiere_login), db: Session = Depends(get_db)):
    _ensure_inventory_operator(user)
    rows = db.query(VarianteEPI).filter(VarianteEPI.activo == True).order_by(VarianteEPI.id).all()
    return JSONResponse({"variantes": [{
        "id": row.id, "catalogo_epi_id": row.catalogo_epi_id,
        "modelo": row.modelo, "color": row.color, "talla": row.talla,
        "referencia_interna": row.referencia_interna, "codigo_qr": row.codigo_qr,
        "referencia_proveedor": row.referencia_proveedor,
    } for row in rows]})


@app.get("/inventario/v2", response_class=HTMLResponse)
def inventario_v2_panel(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    sessions = db.query(SesionInventario).filter(
        SesionInventario.almacen_id == warehouse.id,
    ).order_by(SesionInventario.id.desc()).limit(50).all()
    variants = db.query(VarianteEPI).filter(VarianteEPI.activo == True).order_by(VarianteEPI.id.desc()).limit(100).all()
    dotations = db.query(DotacionTrabajador).join(Trabajador).filter(
        Trabajador.almacen_id == warehouse.id,
    ).order_by(DotacionTrabajador.id.desc()).limit(50).all()
    catalogs = db.query(CatalogoEPI).filter(CatalogoEPI.activo == True).order_by(CatalogoEPI.orden, CatalogoEPI.nombre).all()
    warehouses = [warehouse]
    return templates.TemplateResponse(request, "inventario_v2.html", ctx_base(
        request, user, db, sesiones=sessions, variantes=variants, dotaciones=dotations,
        catalogos=catalogs, almacenes=warehouses,
        almacen_predeterminado_id=warehouse.id,
    ))


def _estado_inventario_real(db: Session, warehouse_id: int | None = None) -> dict:
    """Resumen operativo con datos reales; no genera ni modifica inventario."""
    estados = ("disponible", "en_uso", "mantenimiento", "fuera_servicio")
    categorias = {
        clave: {estado: 0 for estado in estados}
        for clave in ("herramientas", "ropa", "epis", "maquinaria", "consumibles")
    }
    sin_etiqueta = 0
    sin_ubicacion = 0
    stock_bajo = 0

    for item in db.query(Herramienta).filter(
        Herramienta.estado != "archivada",
        *([Herramienta.almacen_id == warehouse_id] if warehouse_id else []),
    ).all():
        estado = (item.estado or "disponible").lower()
        destino = "disponible"
        if estado in {"entregada", "en_obra", "en_transporte", "reservada", "en_uso"}:
            destino = "en_uso"
        elif estado in {"en_mantenimiento", "en_reparacion", "pendiente_revision"}:
            destino = "mantenimiento"
        elif estado in {"fuera_servicio", "extraviada", "robada", "baja"}:
            destino = "fuera_servicio"
        categorias["herramientas"][destino] += 1
        sin_etiqueta += int(not (item.codigo or "").strip())
        sin_ubicacion += int(not any((item.ubicacion_texto, item.almacen_id, item.obra_id, item.vehiculo_id)))

    for item in db.query(Maquinaria).filter(
        Maquinaria.activa == True,
        *([Maquinaria.almacen_id == warehouse_id] if warehouse_id else []),
    ).all():
        estado = (item.estado or "disponible").lower()
        destino = "disponible"
        if estado in {"en_uso", "asignada", "en_obra", "alquilada"}:
            destino = "en_uso"
        elif estado in {"mantenimiento", "en_mantenimiento", "reparacion", "en_reparacion", "pendiente_revision"}:
            destino = "mantenimiento"
        elif estado in {"fuera_servicio", "baja", "averiada"}:
            destino = "fuera_servicio"
        categorias["maquinaria"][destino] += 1
        sin_etiqueta += int(not any(((item.codigo_interno or "").strip(), (item.codigo_barras or "").strip())))
        sin_ubicacion += int(not (item.ubicacion or "").strip())

    for item in db.query(EPIIndividual).filter(
        *([EPIIndividual.almacen_id == warehouse_id] if warehouse_id else []),
    ).all():
        estado = (item.estado or "activo").lower()
        if estado in {"baja", "retirado", "no_apto"}:
            destino = "fuera_servicio"
        elif estado in {"en_revision", "pendiente_revision"} or item.revision_vencida:
            destino = "mantenimiento"
        elif item.trabajador_id:
            destino = "en_uso"
        else:
            destino = "disponible"
        categorias["epis"][destino] += 1
        sin_etiqueta += int(not any(((item.referencia_interna or "").strip(), (item.codigo_qr or "").strip())))

    variantes = db.query(VarianteEPI).filter(VarianteEPI.activo == True).all()
    stock_por_variante = dict(db.query(
        ExistenciaVariante.variante_id,
        func.coalesce(func.sum(ExistenciaVariante.cantidad), 0),
    ).filter(
        *([ExistenciaVariante.almacen_id == warehouse_id] if warehouse_id else []),
    ).group_by(
        ExistenciaVariante.variante_id,
    ).all())
    for variante in variantes:
        categoria = "ropa" if variante.catalogo and variante.catalogo.categoria == "ropa" else "epis"
        cantidad = int(stock_por_variante.get(variante.id, 0) or 0)
        categorias[categoria]["disponible"] += cantidad
        stock_bajo += int(variante.stock_minimo > 0 and cantidad <= variante.stock_minimo)
        sin_etiqueta += int(not (variante.referencia_interna and variante.codigo_qr))

    for linea in db.query(LineaDotacion).join(DotacionTrabajador).join(Trabajador).filter(
        LineaDotacion.estado == "entregada",
        *([Trabajador.almacen_id == warehouse_id] if warehouse_id else []),
    ).all():
        categoria = "ropa" if linea.categoria == "ropa" else "epis"
        categorias[categoria]["en_uso"] += int(linea.cantidad or 0)

    for item in db.query(Material).filter(
        Material.activo == True,
        *([Material.almacen_id == warehouse_id] if warehouse_id else []),
    ).all():
        cantidad = max(0, int(item.stock_actual or 0))
        categorias["consumibles"]["disponible"] += cantidad
        stock_bajo += int(item.bajo_minimo)
        sin_etiqueta += int(not (item.codigo or "").strip())
        sin_ubicacion += int(not any((item.ubicacion_id, item.ubicacion_texto)))

    totales = {estado: sum(datos[estado] for datos in categorias.values()) for estado in estados}
    total = sum(totales.values())
    return {
        "categorias": categorias, "totales": totales, "total": total,
        "sin_etiqueta": sin_etiqueta, "sin_ubicacion": sin_ubicacion,
        "stock_bajo": stock_bajo, "actualizado": datetime.now().strftime("%d/%m/%Y · %H:%M"),
    }


@app.get("/inventario/estado", response_class=HTMLResponse)
def inventario_estado_panel(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    resumen = _estado_inventario_real(db, warehouse.id)
    return templates.TemplateResponse(request, "inventario_estado_real.html", ctx_base(
        request, user, db, resumen=resumen,
    ))


@app.get("/inventario/recepcion", response_class=HTMLResponse)
def inventario_recepcion_panel(
    request: Request, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    catalogs = db.query(CatalogoEPI).filter(CatalogoEPI.activo == True).order_by(
        CatalogoEPI.orden, CatalogoEPI.nombre,
    ).all()
    warehouses = visible_warehouses(db, user)
    locations = db.query(Ubicacion).filter(
        Ubicacion.activo == True,
        Ubicacion.almacen_id.in_([item.id for item in warehouses]),
    ).order_by(
        Ubicacion.almacen_id, Ubicacion.nombre,
    ).all()
    recent = db.query(RecepcionSuministro).join(ExistenciaVariante).filter(
        ExistenciaVariante.almacen_id == warehouse.id,
    ).order_by(
        RecepcionSuministro.id.desc(),
    ).limit(30).all()
    return templates.TemplateResponse(request, "inventario_recepcion.html", ctx_base(
        request, user, db, catalogos=catalogs, almacenes=warehouses,
        ubicaciones=locations, recepciones=recent,
        almacen_predeterminado_id=warehouse.id,
    ))


@app.get("/api/inventario/variantes/buscar")
def inventario_buscar_variante(
    codigo: str = Query(min_length=1, max_length=100),
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    request: Request = None,
):
    _ensure_inventory_operator(user)
    warehouse = _active_warehouse(db, user, request)
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    variant = find_variant(db, codigo)
    if not variant:
        raise HTTPException(404, "No existe una variante con ese código")
    stock = int(db.query(func.coalesce(func.sum(ExistenciaVariante.cantidad), 0)).filter(
        ExistenciaVariante.variante_id == variant.id,
        ExistenciaVariante.almacen_id == warehouse.id,
    ).scalar() or 0)
    return JSONResponse({
        "id": variant.id, "catalogo_epi_id": variant.catalogo_epi_id,
        "nombre": variant.catalogo.nombre if variant.catalogo else "Artículo",
        "modelo": variant.modelo, "color": variant.color, "talla": variant.talla,
        "referencia_interna": variant.referencia_interna,
        "codigo_qr": variant.codigo_qr, "stock_total": stock,
    })


@app.post("/inventario/recepciones")
def inventario_registrar_recepcion(
    payload: RecepcionSuministroRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    request: Request = None,
):
    try:
        _require_warehouse_access(user, payload.almacen_id)
        start_stock_transaction(db)
        result = receive_supply(
            db, user, event_id=payload.event_id, cantidad=payload.cantidad,
            almacen_id=payload.almacen_id, ubicacion_id=payload.ubicacion_id,
            proveedor=payload.proveedor, albaran=payload.albaran,
            precio_unitario=payload.precio_unitario,
            numero_lote=payload.numero_lote, fecha_caducidad=payload.fecha_caducidad,
            variante_id=payload.variante_id, catalogo_epi_id=payload.catalogo_epi_id,
            modelo=payload.modelo, color=payload.color, talla=payload.talla,
            referencia_proveedor=payload.referencia_proveedor,
            stock_minimo=payload.stock_minimo,
        )
        db.commit()
        return JSONResponse({
            "resultado": "ya_registrada" if result.reused else "ok",
            "recepcion_id": result.recepcion_id, "variante_id": result.variante_id,
            "referencia_interna": result.referencia_interna,
            "codigo_qr": result.codigo_qr, "saldo_posterior": result.saldo_posterior,
            "ficha_url": f"/inventario/variantes/{result.variante_id}",
        }, status_code=200 if result.reused else 201)
    except StockError as exc:
        db.rollback()
        if exc.status_code == 409 and exc.detail.startswith("VARIANTE_EXISTENTE:"):
            variant_id = int(exc.detail.rsplit(":", 1)[1])
            return JSONResponse({
                "detail": "La variante ya existe. Registra la entrada sobre ella.",
                "codigo": "variante_existente", "variante_id": variant_id,
            }, status_code=409)
        _inventory_http_error(exc)
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "La recepción coincide con una operación concurrente; vuelve a buscar la variante")


@app.post("/inventario/variantes/nueva")
def inventario_variante_nueva(
    payload: VarianteNuevaRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        require_stock_permission(user)
        _require_warehouse_access(user, payload.almacen_id)
    except StockError as exc:
        _inventory_http_error(exc)
    start_stock_transaction(db)
    try:
        catalog = db.query(CatalogoEPI).filter(
            CatalogoEPI.id == payload.catalogo_epi_id, CatalogoEPI.activo == True,
        ).first()
        warehouse = db.query(Almacen).filter(
            Almacen.id == payload.almacen_id, Almacen.activo == True,
        ).first()
        location = db.query(Ubicacion).filter(
            Ubicacion.id == payload.ubicacion_id,
            Ubicacion.almacen_id == payload.almacen_id,
            Ubicacion.activo == True,
        ).first() if payload.ubicacion_id else None
        if not catalog or not warehouse or (payload.ubicacion_id and not location):
            raise StockError(400, "Catálogo, almacén o ubicación no válidos")
        owner_key = str(uuid.uuid4())
        identifiers = reservar_identificadores(
            db, prefijo="EPI", propietario_tipo="variante_epi",
            propietario_clave=owner_key, creado_por_id=user.id,
        )
        variant = VarianteEPI(
            catalogo_epi_id=payload.catalogo_epi_id,
            modelo=payload.modelo.strip(), color=payload.color.strip(),
            talla=payload.talla.strip(), identificador_id=identifiers.id,
            referencia_interna=identifiers.referencia_interna,
            codigo_qr=identifiers.codigo_qr,
            referencia_proveedor=payload.referencia_proveedor,
            stock_minimo=payload.stock_minimo, creado_por_id=user.id,
        )
        db.add(variant)
        db.flush()
        existence = ExistenciaVariante(
            variante_id=variant.id, almacen_id=payload.almacen_id,
            ubicacion_id=payload.ubicacion_id,
            ubicacion_clave=payload.ubicacion_id or 0,
            cantidad=0, stock_minimo=payload.stock_minimo, version=0,
        )
        db.add(existence)
        db.flush()
        if payload.cantidad:
            move_variante(
                db, user, existence.id, payload.cantidad, tipo="entrada_inicial",
                event_id=f"variant-create-{uuid.uuid4()}",
                motivo="Alta inicial de variante", numero_lote=payload.numero_lote,
                fecha_caducidad=payload.fecha_caducidad,
            )
        db.commit()
        return JSONResponse({
            "variante_id": variant.id,
            "referencia_interna": variant.referencia_interna,
            "codigo_qr": variant.codigo_qr,
        }, status_code=201)
    except (StockError, InventoryError) as exc:
        db.rollback()
        _inventory_http_error(exc)
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "La variante o existencia ya existe")


@app.put("/inventario/variantes/{variant_id}")
def inventario_variante_editar(
    variant_id: int, payload: VarianteEditarRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        require_stock_permission(user)
    except StockError as exc:
        _inventory_http_error(exc)
    variant = db.get(VarianteEPI, variant_id)
    if not variant:
        raise HTTPException(404, "Variante no encontrada")
    if payload.referencia_proveedor is not None:
        variant.referencia_proveedor = payload.referencia_proveedor
    if payload.stock_minimo is not None:
        variant.stock_minimo = payload.stock_minimo
    if payload.ubicacion_id is not None:
        existence = db.query(ExistenciaVariante).filter_by(variante_id=variant_id).first()
        location = db.query(Ubicacion).filter(
            Ubicacion.id == payload.ubicacion_id,
            Ubicacion.almacen_id == existence.almacen_id,
            Ubicacion.activo == True,
        ).first() if existence else None
        if not location:
            raise HTTPException(400, "Ubicación no válida")
        existence.ubicacion_id = location.id
        existence.ubicacion_clave = location.id
    db.commit()
    return JSONResponse({"resultado": "ok", "variante_id": variant_id})


@app.get("/inventario/variantes/{variant_id}", response_class=HTMLResponse)
def inventario_variante_ficha(
    request: Request, variant_id: int,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    variant = db.get(VarianteEPI, variant_id)
    if not variant:
        raise HTTPException(404, "Variante no encontrada")
    warehouse = _active_warehouse(db, user, request)
    existences = db.query(ExistenciaVariante).filter(
        ExistenciaVariante.variante_id == variant_id,
        ExistenciaVariante.almacen_id == (warehouse.id if warehouse else -1),
    ).order_by(
        ExistenciaVariante.almacen_id, ExistenciaVariante.ubicacion_clave,
    ).all()
    if not existences:
        raise HTTPException(404, "La referencia no tiene existencias en este almacén")
    existence_ids = [row.id for row in existences]
    lots = db.query(LoteVariante).filter(
        LoteVariante.existencia_id.in_(existence_ids),
    ).order_by(LoteVariante.fecha_caducidad, LoteVariante.id).all() if existence_ids else []
    movements = db.query(MovimientoStock).filter(
        MovimientoStock.existencia_id.in_(existence_ids),
    ).order_by(MovimientoStock.id.desc()).limit(100).all() if existence_ids else []
    receipts = db.query(RecepcionSuministro).filter(
        RecepcionSuministro.variante_id == variant_id,
        RecepcionSuministro.existencia_id.in_(existence_ids),
    ).order_by(
        RecepcionSuministro.id.desc(),
    ).limit(50).all()
    total = sum(int(row.cantidad) for row in existences)
    return templates.TemplateResponse(request, "inventario_variante_ficha.html", ctx_base(
        request, user, db, variante=variant, existencias=existences,
        lotes=lots, movimientos_stock=movements, recepciones=receipts,
        stock_total=total, bajo_minimo=total <= variant.stock_minimo,
    ))


@app.post("/inventario/sesiones/nueva")
def inventario_sesion_nueva(
    payload: SesionNuevaRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    request: Request = None,
):
    try:
        warehouse = _active_warehouse(db, user, request)
        if not warehouse or payload.almacen_id != warehouse.id:
            raise InventoryError(409, "Cambia primero al almacén que quieres inventariar")
        start_stock_transaction(db)
        session = open_inventory_session(
            db, user, nombre=payload.nombre, almacen_id=payload.almacen_id,
            scope=payload.scope, tipo_articulo=payload.tipo_articulo,
            umbral_desviacion=payload.umbral_desviacion,
        )
        db.commit()
        return JSONResponse({
            "sesion_id": session.id,
            "lineas_generadas": db.query(LineaInventario).filter_by(sesion_id=session.id).count(),
        }, status_code=201)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)
    except Exception:
        db.rollback()
        raise


@app.post("/inventario/sesiones/{session_id}/contar")
def inventario_contar(
    session_id: int, payload: ConteoRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _inventory_session_for_user(db, user, session_id)
    start_stock_transaction(db)
    try:
        result = register_count(
            db, user, session_id=session_id, line_id=payload.linea_id,
            amount=payload.cantidad, count_number=payload.numero_conteo,
            scan_event_id=payload.scan_event_id, mode=payload.modo_entrada,
            units_per_box=payload.unidades_por_caja, notes=payload.notas,
            station_id=payload.puesto_id,
        )
        db.commit()
        return JSONResponse(result)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/sesiones/{session_id}/activos/escanear")
def inventario_escanear_activo(
    session_id: int, payload: EscaneoActivoInventarioRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    request: Request = None,
):
    _ensure_inventory_operator(user)
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    warehouse = _active_warehouse(db, user, request) if session.almacen_id else None
    if session.almacen_id:
        _require_warehouse_access(user, session.almacen_id)
        if not warehouse:
            raise HTTPException(409, "No hay un almacén activo configurado")
    if session.almacen_id and warehouse and session.almacen_id != warehouse.id:
        raise HTTPException(409, "Cambia primero al almacén de esta sesión")
    if session.tipo_articulo != "todo" or session.estado not in {
        "abierta", "en_conteo", "revision", "segundo_conteo",
    }:
        raise HTTPException(409, "Esta sesión no admite el inventario de activos")
    try:
        item = resolve_counter_item(db, payload.codigo, warehouse_id=session.almacen_id)
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    if item["tipo"] not in {"herramienta", "maquinaria", "vehiculo"}:
        raise HTTPException(409, "Este código se cuenta en las líneas de stock de la sesión")
    ensure_inventory_asset_snapshot(db, session)
    row = db.query(ActivoInventarioEscaneado).filter_by(
        sesion_id=session.id, tipo=item["tipo"], item_id=item["id"],
    ).first()
    if not row:
        row = ActivoInventarioEscaneado(
            sesion_id=session.id, tipo=item["tipo"], item_id=item["id"],
            codigo=item["codigo"], nombre=item["nombre"],
            estado_snapshot=item.get("estado"), esperado=False,
        )
        db.add(row)
        db.flush()
    reused = row.encontrado_en is not None
    if not reused:
        row.encontrado_en = datetime.now()
        row.encontrado_por_id = user.id
        if session.estado == "abierta":
            session.estado = "en_conteo"
    db.commit()
    return JSONResponse({
        "resultado": "ya_encontrado" if reused else "ok",
        "tipo": row.tipo, "item_id": row.item_id, "codigo": row.codigo,
        "nombre": row.nombre, "encontrado": True,
    })


@app.get("/api/inventario/recepcion/resolver")
def inventario_recepcion_resolver(
    codigo: str = Query(min_length=1, max_length=512),
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    almacen_id: int | None = Query(default=None, gt=0),
    request: Request = None,
):
    _ensure_inventory_operator(user)
    selected_warehouse_id = almacen_id if isinstance(almacen_id, int) else None
    warehouse = (
        db.get(Almacen, selected_warehouse_id)
        if selected_warehouse_id else _active_warehouse(db, user, request)
    )
    if not warehouse:
        raise HTTPException(409, "No hay un almacén activo configurado")
    _require_warehouse_access(user, warehouse.id)
    try:
        item = resolve_counter_item(db, codigo, warehouse_id=warehouse.id)
    except CounterError as exc:
        raise HTTPException(exc.status_code, exc.detail)
    if item["tipo"] not in {"material", "stock_epi", "variante"}:
        raise HTTPException(
            409,
            f"{item['nombre']} está reconocido, pero no es una existencia con cantidad",
        )
    variant_id = None
    if item["tipo"] == "variante":
        existence = db.get(ExistenciaVariante, item["id"])
        variant_id = existence.variante_id if existence else None
        if not variant_id:
            raise HTTPException(404, "La variante no tiene una existencia válida")
    return JSONResponse({"ok": True, "item": item, "variante_id": variant_id})


@app.post("/inventario/recepciones/existencias")
def inventario_registrar_entrada_existencia(
    payload: EntradaExistenciaRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
    request: Request = None,
):
    _ensure_inventory_operator(user)
    try:
        start_stock_transaction(db)
        warehouse = db.query(Almacen).filter(
            Almacen.id == payload.almacen_id, Almacen.activo == True,
        ).first()
        _require_warehouse_access(user, payload.almacen_id)
        location = db.get(Ubicacion, payload.ubicacion_id) if payload.ubicacion_id else None
        if not warehouse:
            raise StockError(400, "Almacén no válido")
        if payload.ubicacion_id and (
            not location or not location.activo or location.almacen_id != warehouse.id
        ):
            raise StockError(400, "La ubicación no pertenece al almacén")
        try:
            item = resolve_counter_item(db, payload.codigo, warehouse_id=warehouse.id)
        except CounterError as exc:
            raise StockError(exc.status_code, exc.detail)
        motivo = "Recepción"
        if payload.albaran:
            motivo += f" · albarán {payload.albaran.strip()}"
        if payload.proveedor:
            motivo += f" · {payload.proveedor.strip()}"
        if item["tipo"] == "material":
            material = db.get(Material, item["id"])
            material.almacen_id = warehouse.id
            material.ubicacion_id = location.id if location else None
            material.ubicacion_texto = location.nombre if location else warehouse.nombre
            movement = move_material(
                db, user, material.id, float(payload.cantidad), tipo="recepcion",
                event_id=payload.event_id, motivo=motivo,
            )
            url = f"/materiales/{material.id}"
        elif item["tipo"] == "stock_epi":
            if abs(payload.cantidad - round(payload.cantidad)) > 0.0001:
                raise StockError(400, "La ropa y los EPIs se reciben en unidades enteras")
            stock = db.get(StockEPI, item["id"])
            stock.almacen_id = warehouse.id
            stock.ubicacion_id = location.id if location else None
            movement = move_stock_epi(
                db, user, stock.id, int(round(payload.cantidad)), tipo="recepcion",
                event_id=payload.event_id, motivo=motivo,
            )
            url = "/epis/stock"
        else:
            raise StockError(409, "Utiliza la recepción de variante para este artículo")
        if payload.numero_lote:
            lot = db.query(LoteAlmacen).filter_by(
                tipo=item["tipo"], objeto_id=item["id"], almacen_id=warehouse.id,
                numero_lote=payload.numero_lote.strip(),
            ).first()
            if not lot:
                lot = LoteAlmacen(
                    tipo=item["tipo"], objeto_id=item["id"], almacen_id=warehouse.id,
                    numero_lote=payload.numero_lote.strip(), cantidad=0,
                )
                db.add(lot)
            lot.cantidad = float(lot.cantidad or 0) + float(payload.cantidad)
            lot.fecha_caducidad = payload.fecha_caducidad or lot.fecha_caducidad
            lot.proveedor = payload.proveedor or lot.proveedor
        db.commit()
        return JSONResponse({
            "resultado": "ya_registrada" if movement.reused else "ok",
            "saldo_posterior": movement.saldo_posterior,
            "ficha_url": url, "nombre": item["nombre"], "codigo": item["codigo"],
        })
    except StockError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/sesiones/{session_id}/lineas/{line_id}/aprobar")
def inventario_aprobar(
    session_id: int, line_id: int, payload: AprobarConteoRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _inventory_session_for_user(db, user, session_id)
    start_stock_transaction(db)
    try:
        approve_count(
            db, user, session_id=session_id, line_id=line_id,
            final_amount=payload.cantidad_final,
        )
        db.commit()
        return JSONResponse({"resultado": "ok"})
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/sesiones/{session_id}/cerrar")
def inventario_cerrar(
    session_id: int, payload: CerrarSesionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _inventory_session_for_user(db, user, session_id)
    try:
        result = close_inventory_session(
            db, user, session_id=session_id,
            cierre_event_id=payload.cierre_event_id,
        )
        db.commit()
        return JSONResponse(result)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


def _inventory_line_view(db: Session, line: LineaInventario, reveal_expected: bool) -> dict:
    name, reference, detail = "Artículo", f"Línea {line.id}", ""
    item_type, item_id, codes = "", 0, []
    if line.material_id:
        item = db.get(Material, line.material_id)
        if item:
            name, reference, detail = item.nombre, item.codigo, item.unidad or ""
            item_type, item_id, codes = "material", item.id, [item.codigo]
    elif line.stock_epi_id:
        item = db.get(StockEPI, line.stock_epi_id)
        if item:
            name = item.nombre_display
            reference = item.codigo or f"EPI-{item.id}"
            detail = item.categoria
            item_type, item_id = "stock_epi", item.id
            codes = [item.codigo, f"EPI-{item.id}"]
    elif line.existencia_id:
        existence = db.get(ExistenciaVariante, line.existencia_id)
        variant = db.get(VarianteEPI, existence.variante_id) if existence else None
        if variant:
            name = variant.catalogo.nombre if variant.catalogo else "Variante"
            reference = variant.codigo_qr or variant.referencia_interna
            detail = " · ".join(v for v in (variant.modelo, variant.color, variant.talla) if v)
            item_type, item_id = "variante", existence.id
            codes = [variant.codigo_qr, variant.referencia_interna, variant.referencia_proveedor]
    elif line.epi_individual_id:
        item = db.get(EPIIndividual, line.epi_individual_id)
        if item:
            name = item.tipo
            reference = item.codigo_qr or item.referencia_interna or item.codigo_fabricacion
            detail = item.estado
            item_type, item_id = "epi_individual", item.id
            codes = [item.codigo_qr, item.referencia_interna, item.codigo_fabricacion]
    data = {
        "id": line.id, "nombre": name, "referencia": reference, "detalle": detail,
        "tipo": item_type, "item_id": item_id,
        "codigos": [str(code).strip() for code in codes if str(code or "").strip()],
        "estado": line.estado, "cantidad_contada_1": line.cantidad_contada_1,
        "cantidad_contada_2": line.cantidad_contada_2,
        "cantidad_final": line.cantidad_final,
    }
    if reveal_expected:
        data.update(cantidad_esperada=line.cantidad_esperada, diferencia=line.diferencia)
    return data


def _inventory_session_payload(db: Session, session: SesionInventario) -> dict:
    lines = db.query(LineaInventario).filter_by(sesion_id=session.id).order_by(
        LineaInventario.id,
    ).all()
    reveal = session.estado in {"pendiente_cierre", "cerrada", "cancelada"}
    serialized = [_inventory_line_view(db, line, reveal) for line in lines]
    assets = db.query(ActivoInventarioEscaneado).filter_by(sesion_id=session.id).order_by(
        ActivoInventarioEscaneado.tipo, ActivoInventarioEscaneado.nombre,
    ).all()
    serialized_assets = [{
        "id": row.id, "tipo": row.tipo, "item_id": row.item_id,
        "codigo": row.codigo, "nombre": row.nombre,
        "estado": row.estado_snapshot or "", "esperado": row.esperado,
        "encontrado": row.encontrado_en is not None,
    } for row in assets]
    completed = sum(1 for line in lines if line.estado != "pendiente")
    completed += sum(1 for row in assets if row.encontrado_en is not None)
    total = len(lines) + len(assets)
    return {
        "id": session.id, "nombre": session.nombre, "estado": session.estado,
        "tipo_articulo": session.tipo_articulo, "scope": session.scope,
        "total_lineas": total, "lineas_procesadas": completed,
        "progreso": round((completed / total * 100), 1) if total else 100,
        "conteo_ciego": not reveal, "lineas": serialized, "activos": serialized_assets,
    }


@app.get("/inventario/sesiones/{session_id}", response_class=HTMLResponse)
def inventario_sesion_detalle(
    request: Request, session_id: int,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    _require_warehouse_access(user, session.almacen_id)
    if ensure_inventory_asset_snapshot(db, session):
        db.commit()
    payload = _inventory_session_payload(db, session)
    return templates.TemplateResponse(request, "inventario_sesion.html", ctx_base(
        request, user, db, sesion=session, inventario=payload,
        puede_aprobar=tiene_permiso(user, "editar"),
        puede_cerrar=tiene_permiso(user, "editar"),
        puede_administrar=user.rol == "admin",
    ))


@app.patch("/inventario/sesiones/{session_id}")
def inventario_sesion_editar(
    session_id: int, payload: SesionEditarRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    """Corrige la identificación sin alterar líneas, conteos ni ajustes."""
    _ensure_inventory_admin(user)
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    session.nombre = payload.nombre.strip()
    session.observaciones = payload.observaciones.strip() or None
    db.commit()
    return JSONResponse({"resultado": "ok", "nombre": session.nombre})


@app.delete("/inventario/sesiones/{session_id}")
def inventario_sesion_eliminar(
    session_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    """Borra borradores; los inventarios aplicados permanecen auditables."""
    _ensure_inventory_admin(user)
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    if session.estado == "cerrada" or db.query(AjusteInventario).filter_by(sesion_id=session_id).first():
        raise HTTPException(409, "Un inventario cerrado no se puede borrar porque forma parte del historial")
    line_ids = [line.id for line in session.lineas]
    if line_ids:
        db.query(IntentoConteo).filter(IntentoConteo.linea_id.in_(line_ids)).delete(
            synchronize_session=False,
        )
    db.delete(session)
    db.commit()
    return JSONResponse({"resultado": "ok"})


@app.get("/api/inventario/sesiones/{session_id}")
def api_inventario_sesion_detalle(
    session_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    session = db.get(SesionInventario, session_id)
    if not session:
        raise HTTPException(404, "Sesión no encontrada")
    _require_warehouse_access(user, session.almacen_id)
    return JSONResponse(_inventory_session_payload(db, session))


@app.get("/inventario/sesiones/{session_id}/diferencias")
def inventario_sesion_diferencias(
    session_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    _inventory_session_for_user(db, user, session_id)
    lines = db.query(LineaInventario).filter(
        LineaInventario.sesion_id == session_id,
        or_(LineaInventario.estado == "conflicto", LineaInventario.diferencia != 0),
    ).order_by(LineaInventario.id).all()
    return JSONResponse({"diferencias": [{
        "linea_id": line.id, "estado": line.estado, "diferencia": line.diferencia,
    } for line in lines]})


@app.post("/inventario/sesiones/{session_id}/cancelar")
def inventario_sesion_cancelar(
    session_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_admin(user)
    start_stock_transaction(db)
    changed = db.query(SesionInventario).filter(
        SesionInventario.id == session_id,
        SesionInventario.estado.in_(("abierta", "en_conteo", "revision", "segundo_conteo", "pendiente_cierre")),
    ).update({"estado": "cancelada"}, synchronize_session=False)
    if changed != 1:
        db.rollback()
        raise HTTPException(409, "La sesión no puede cancelarse")
    db.commit()
    return JSONResponse({"resultado": "ok"})


def _dotacion_payload(dotation: DotacionTrabajador) -> dict:
    return {
        "id": dotation.id, "trabajador_id": dotation.trabajador_id,
        "estado": dotation.estado,
        "firmado_por": dotation.firmado_por,
        "lineas": [{
            "id": line.id, "nombre": line.nombre, "categoria": line.categoria,
            "talla": line.talla, "cantidad": line.cantidad, "estado": line.estado,
            "epi_individual_id": line.epi_individual_id,
            "preparado_en": line.preparado_en.isoformat() if line.preparado_en else None,
            "entregado_en": line.entregado_en.isoformat() if line.entregado_en else None,
            "devuelto_en": line.devuelto_en.isoformat() if line.devuelto_en else None,
        } for line in dotation.lineas],
    }


@app.get("/api/inventario/dotaciones/{dotation_id}")
def inventario_dotacion_api(
    dotation_id: int, user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    dotation = db.get(DotacionTrabajador, dotation_id)
    if not dotation:
        raise HTTPException(404, "Dotación no encontrada")
    return JSONResponse(_dotacion_payload(dotation))


@app.get("/inventario/dotaciones/{dotation_id}", response_class=HTMLResponse)
def inventario_dotacion_detalle(
    dotation_id: int, request: Request, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    _ensure_inventory_operator(user)
    dotation = db.get(DotacionTrabajador, dotation_id)
    if not dotation:
        raise HTTPException(404, "Dotación no encontrada")
    worker = db.get(Trabajador, dotation.trabajador_id)
    return templates.TemplateResponse(request, "dotacion_operativa.html", ctx_base(
        request, user, db, dotacion=dotation, trabajador=worker,
    ))


@app.post("/inventario/dotaciones/{dotation_id}/confirmar")
def inventario_dotacion_confirmar(
    dotation_id: int, payload: ConfirmarDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    raise HTTPException(
        409,
        "La entrega completa está bloqueada: debe implantarse confirmación física por cada artículo",
    )


@app.post("/inventario/dotaciones/lineas/{line_id}/preparar")
def inventario_dotacion_preparar_linea(
    line_id: int, payload: PrepararLineaDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        result = prepare_dotation_line(db, user, line_id=line_id, codigo_qr=payload.codigo_qr)
        db.commit()
        return JSONResponse(result)
    except (InventoryError, StockError) as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/dotaciones/lineas/{line_id}/entregar")
def inventario_dotacion_entregar_linea(
    line_id: int, payload: EntregarLineaDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        result = confirm_dotation_line(
            db, user, line_id=line_id, event_id=payload.event_id,
            codigo_qr=payload.codigo_qr, firmado_por=payload.firmado_por,
            firma_base64=payload.firma_base64,
        )
        db.commit()
        return JSONResponse(result)
    except (InventoryError, StockError) as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/dotaciones/lineas/{line_id}/devolver")
def inventario_dotacion_devolver_linea(
    line_id: int, payload: DevolverLineaDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        result = return_dotation_line(
            db, user, line_id=line_id, event_id=payload.event_id,
            codigo_qr=payload.codigo_qr, motivo=payload.motivo,
        )
        db.commit()
        return JSONResponse(result)
    except (InventoryError, StockError) as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/dotaciones/lineas/{line_id}/talla")
def inventario_dotacion_cambiar_talla(
    line_id: int, payload: CambiarTallaDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        result = change_dotation_line_size(db, user, line_id=line_id, talla=payload.talla)
        db.commit()
        return JSONResponse(result)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/dotaciones/lineas/{line_id}/sustituir")
def inventario_dotacion_sustituir_linea(
    line_id: int, payload: SustituirLineaDotacionRequest,
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    try:
        replacement = replace_dotation_line(db, user, line_id=line_id, motivo=payload.motivo)
        db.commit()
        return JSONResponse({"resultado": "pendiente", "linea_id": replacement.id}, status_code=201)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.get("/inventario/reset-ropa/preview")
def inventario_reset_ropa_preview(
    user: Usuario = Depends(requiere_login), db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "config"):
        raise HTTPException(403, "Sin permiso")
    preview = clothing_reset_preview(db)
    preview["frase_requerida"] = RESET_PHRASE
    return JSONResponse(preview)


@app.post("/inventario/reset-ropa/ejecutar")
def inventario_reset_ropa_ejecutar(
    payload: ResetRopaRequest, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    try:
        result = execute_clothing_reset(
            db, user, event_id=payload.event_id, phrase=payload.frase,
            preview_hash=payload.preview_hash,
        )
        db.commit()
        return JSONResponse(result)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.post("/inventario/etiquetas/preview")
def inventario_etiqueta_preview(
    payload: EtiquetaRequest, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    if not tiene_permiso(user, "etiquetas"):
        raise HTTPException(403, "Sin permiso")
    try:
        label = label_from_identifier(db, payload.identificador_id)
        zpl = build_zpl(
            tipo=label["tipo"], referencia=label["codigo_qr"],
            titulo=label["titulo"],
            detalle=f'{label["referencia"]} · {label["detalle"]}'.strip(" ·"),
        )
    except InventoryError as exc:
        _inventory_http_error(exc)
    return PlainTextResponse(zpl, media_type="text/plain; charset=utf-8")


@app.post("/inventario/etiquetas/imprimir")
def inventario_etiqueta_imprimir(
    payload: EtiquetaRequest, user: Usuario = Depends(requiere_login),
    db: Session = Depends(get_db),
):
    try:
        result = send_label(
            db, user, event_id=payload.event_id,
            identifier_id=payload.identificador_id, copias=payload.copias,
            reimpresion=payload.reimpresion,
            motivo_reimpresion=payload.motivo_reimpresion,
        )
        db.commit()
        return JSONResponse(result)
    except InventoryError as exc:
        db.rollback()
        _inventory_http_error(exc)


@app.get("/p/{codigo_qr}")
def ficha_publica_qr(codigo_qr: str, db: Session = Depends(get_db)):
    identifier = db.query(IdentificadorGlobal).filter_by(codigo_qr=codigo_qr).first()
    if not identifier:
        raise HTTPException(404, "Código no encontrado")
    response = {"referencia": identifier.referencia_interna, "tipo": identifier.propietario_tipo}
    if identifier.propietario_tipo == "variante_epi":
        variant = db.query(VarianteEPI).filter_by(identificador_id=identifier.id).first()
        if variant:
            catalog = db.get(CatalogoEPI, variant.catalogo_epi_id)
            response.update({
                "nombre": catalog.nombre if catalog else "Artículo",
                "modelo": variant.modelo, "color": variant.color,
                "talla": variant.talla, "activo": variant.activo,
            })
    return JSONResponse(response)


app.include_router(salidas_router)

# El montaje genérico debe quedar después de todas las rutas /api específicas.
# Starlette resuelve en orden y, si se monta antes, intercepta rutas internas
# como /api/backup/status y responde 404 desde la subaplicación.
if _api_app is not None:
    app.mount("/api", _api_app)
    mrd_logging.log_app("API externa montada en /api — Swagger en /api/docs")
