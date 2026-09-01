import asyncio
import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

import identificadores
import main
from auth import hash_password
from identificadores import (
    PREFIJO_HERRAMIENTA, PREFIJO_MAQUINARIA, PREFIJO_MATERIAL,
    asegurar_referencias_maquinaria,
    generar_referencia_herramienta,
    generar_referencia_maquinaria,
    generar_referencia_material,
)
from models import Herramienta, Maquinaria, Material, Usuario
from reports import generar_plantilla_importacion, importar_herramientas_excel


ROOT = Path(__file__).resolve().parents[1]


def _usuario_admin(db, sufijo="ref"):
    usuario = Usuario(
        username=f"admin-{sufijo}",
        password_hash=hash_password("TestAdmin@2026!"),
        nombre="Administrador",
        rol="admin",
        activo=True,
        must_change_password=False,
    )
    db.add(usuario)
    db.flush()
    return usuario


def test_referencia_tiene_prefijo_fuerte_y_columna_unique(db):
    referencia = generar_referencia_herramienta(db)

    assert referencia.startswith(PREFIJO_HERRAMIENTA)
    assert len(referencia) == len(PREFIJO_HERRAMIENTA) + 32
    assert Herramienta.__table__.c.codigo.unique is True
    assert Herramienta.__table__.c.codigo.nullable is False


def test_generador_reintenta_si_referencia_ya_existe_en_otro_inventario(db, monkeypatch):
    repetido = "ABCDEF123456ABCDEF123456ABCDEF12"
    nuevo = "654321FEDCBA654321FEDCBA654321FE"
    db.add(Material(codigo=f"{PREFIJO_HERRAMIENTA}{repetido}", nombre="Reserva global"))
    db.flush()
    valores = iter([repetido.lower(), nuevo.lower()])
    monkeypatch.setattr(identificadores, "token_hex", lambda _n: next(valores))

    referencia = generar_referencia_herramienta(db)

    assert referencia == f"{PREFIJO_HERRAMIENTA}{nuevo}"


def test_cada_familia_tiene_prefijo_reservado(db):
    assert generar_referencia_herramienta(db).startswith(PREFIJO_HERRAMIENTA)
    assert generar_referencia_material(db).startswith(PREFIJO_MATERIAL)
    assert generar_referencia_maquinaria(db).startswith(PREFIJO_MAQUINARIA)


def test_maquinaria_historica_recibe_referencia_sin_cambiar_la_existente(db):
    antigua = Maquinaria(nombre="Alimak histórico", codigo_barras="FAB-001")
    existente = Maquinaria(nombre="Maquinillo", codigo_interno="MRD-MAQ-CONSERVAR")
    db.add_all([antigua, existente])
    db.flush()

    creadas = asegurar_referencias_maquinaria(db)

    assert creadas == 1
    assert antigua.codigo_interno.startswith(PREFIJO_MAQUINARIA)
    assert existente.codigo_interno == "MRD-MAQ-CONSERVAR"


def test_formulario_no_permite_escribir_referencia():
    html = (ROOT / "templates" / "nueva_herramienta.html").read_text(encoding="utf-8")

    assert 'name="codigo"' not in html
    assert "Solo el programa puede crearla" in html
    assert "Automática al guardar" in html

    material_nuevo = (ROOT / "templates" / "materiales.html").read_text(encoding="utf-8")
    material_editar = (ROOT / "templates" / "material_detalle.html").read_text(encoding="utf-8")
    maquinaria = (ROOT / "templates" / "maquinaria_nueva.html").read_text(encoding="utf-8")
    assert 'name="codigo"' not in material_nuevo
    assert 'name="codigo"' not in material_editar
    assert 'name="codigo_interno"' not in maquinaria


def test_plantilla_excel_no_solicita_codigo_y_importacion_lo_genera(db, monkeypatch):
    plantilla = load_workbook(io.BytesIO(generar_plantilla_importacion()), data_only=True)
    encabezados = [c.value for c in plantilla.active[3]]
    assert not any("Código" in str(v) for v in encabezados if v)

    wb = Workbook()
    ws = wb.active
    ws.append(["Código", "Nombre *", "Categoría"])
    ws.append(["CODIGO-IMPUESTO-POR-USUARIO", "Taladro de prueba", "Herramienta"])
    buf = io.BytesIO()
    wb.save(buf)
    monkeypatch.setattr("identificadores.token_hex", lambda _n: "112233aabbcc112233aabbcc112233aa")

    resultado = importar_herramientas_excel(db, buf.getvalue(), _usuario_admin(db, "excel"))
    creada = db.query(Herramienta).filter(Herramienta.nombre == "Taladro de prueba").one()

    assert resultado["creadas"] == 1
    assert creada.codigo == "MRD-HTA-112233AABBCC112233AABBCC112233AA"
    assert creada.codigo != "CODIGO-IMPUESTO-POR-USUARIO"


class _JsonRequest:
    headers = {}
    client = None

    async def json(self):
        return {"codigo": "CODIGO-EXTERNO", "nombre": "Herramienta API"}


def test_api_ignora_codigo_externo_y_asigna_referencia_interna(db, monkeypatch):
    monkeypatch.setattr(main, "generar_referencia_herramienta", lambda _db: "MRD-HTA-A1B2C3D4E5F6A1B2C3D4E5F6A1B2C3D4")
    usuario = _usuario_admin(db, "api")

    respuesta = asyncio.run(main.api_v1_crear_herramienta(
        request=_JsonRequest(), user=usuario, db=db,
    ))
    creada = db.query(Herramienta).filter(Herramienta.nombre == "Herramienta API").one()

    assert respuesta.status_code == 201
    assert creada.codigo == "MRD-HTA-A1B2C3D4E5F6A1B2C3D4E5F6A1B2C3D4"
    assert creada.codigo != "CODIGO-EXTERNO"


def test_qr_de_herramienta_se_construye_desde_referencia_unica():
    codigo = (ROOT / "main.py").read_text(encoding="utf-8")

    assert "generar_qr_base64(h.codigo)" in codigo
    assert "generar_qr_bytes(h.codigo)" in codigo
