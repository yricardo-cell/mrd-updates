"""
Exportación de informes Excel - MRD TOOL CONTROL
"""
import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


MRD_BLUE = "1B4F8A"
MRD_ORANGE = "E8600A"
LIGHT_BLUE = "D6E4F0"
LIGHT_ORANGE = "FAD7C0"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
DARK = "1A1A2E"


def _borde():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _header_fill(color):
    return PatternFill("solid", fgColor=color)


def exportar_inventario_excel(herramientas: List) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Encabezado principal
    ws.merge_cells("A1:L1")
    ws["A1"] = f"MRD TOOL CONTROL — Inventario de Herramientas"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = _header_fill(MRD_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Columnas
    cols = [
        ("Código", 20), ("Nombre", 35), ("Categoría", 18),
        ("Marca", 18), ("Modelo", 18), ("N° Serie", 18),
        ("Estado", 15), ("Ubicación", 25), ("Responsable", 25),
        ("Precio", 12), ("Fecha Compra", 15), ("Observaciones", 30),
    ]
    row_h = 4
    for c_idx, (titulo, ancho) in enumerate(cols, start=1):
        cell = ws.cell(row=row_h, column=c_idx, value=titulo)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = _header_fill(MRD_ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _borde()
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[row_h].height = 22

    # Datos
    ESTADOS_COLORES = {
        "disponible": "2ECC71", "entregada": "E74C3C",
        "en_obra": "E67E22", "en_furgoneta": "9B59B6",
        "en_reparacion": "F39C12", "perdida": "7F8C8D", "baja": "BDC3C7",
    }
    for i, h in enumerate(herramientas):
        row = row_h + 1 + i
        fill_bg = _header_fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        estado = getattr(h, 'estado', '')
        estado_color = ESTADOS_COLORES.get(estado, "AAAAAA")

        valores = [
            getattr(h, 'codigo', ''),
            getattr(h, 'nombre', ''),
            getattr(h, 'categoria', ''),
            getattr(h, 'marca', ''),
            getattr(h, 'modelo', ''),
            getattr(h, 'num_serie', ''),
            estado.replace("_", " ").title(),
            getattr(h, 'ubicacion', ''),
            (h.responsable.nombre_completo if hasattr(h, 'responsable') and h.responsable else ""),
            getattr(h, 'precio', ''),
            (h.fecha_compra.strftime("%d/%m/%Y") if hasattr(h, 'fecha_compra') and h.fecha_compra else ""),
            getattr(h, 'observaciones', ''),
        ]
        for c_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = _borde()
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx == 12))
            if c_idx == 7:
                cell.fill = _header_fill(estado_color)
                cell.font = Font(color=WHITE, bold=True)
            else:
                cell.fill = fill_bg
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{row_h}:L{row_h + len(herramientas)}"

    # Hoja estadísticas
    ws2 = wb.create_sheet("Estadísticas")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 15

    ws2["A1"] = "Estadísticas de Inventario"
    ws2["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws2["A1"].fill = _header_fill(MRD_BLUE)
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    estados = {}
    categorias = {}
    for h in herramientas:
        est = h.estado if hasattr(h, 'estado') else h.get('estado', 'desconocido')
        cat = h.categoria if hasattr(h, 'categoria') else h.get('categoria', 'Sin categoría')
        estados[est] = estados.get(est, 0) + 1
        categorias[cat] = categorias.get(cat, 0) + 1

    ws2["A3"] = "Estado"
    ws2["B3"] = "Cantidad"
    ws2["C3"] = "Porcentaje"
    for cell in [ws2["A3"], ws2["B3"], ws2["C3"]]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _header_fill(MRD_ORANGE)

    total = len(herramientas)
    for r, (est, cnt) in enumerate(sorted(estados.items()), start=4):
        ws2.cell(row=r, column=1, value=est.replace("_", " ").title())
        ws2.cell(row=r, column=2, value=cnt)
        pct = f"{cnt/total*100:.1f}%" if total else "0%"
        ws2.cell(row=r, column=3, value=pct)

    ws2["A3"].border = _borde()

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_movimientos_excel(movimientos: List) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.merge_cells("A1:H1")
    ws["A1"] = "MRD TOOL CONTROL — Historial de Movimientos"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = _header_fill(MRD_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = [
        ("Fecha", 18), ("Herramienta", 35), ("Código", 20),
        ("Tipo", 18), ("Estado Ant.", 15), ("Estado Nuevo", 15),
        ("Trabajador / Destino", 28), ("Registrado por", 20),
    ]
    row_h = 3
    for c_idx, (titulo, ancho) in enumerate(cols, start=1):
        cell = ws.cell(row=row_h, column=c_idx, value=titulo)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _header_fill(MRD_ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde()
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho

    TIPO_COLORES = {
        "entrega": "E74C3C", "devolucion": "2ECC71", "alta": "3498DB",
        "traslado": "9B59B6", "reparacion": "F39C12", "perdida": "7F8C8D",
        "baja": "BDC3C7",
    }

    for i, m in enumerate(movimientos):
        row = row_h + 1 + i
        fill_bg = _header_fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        tipo = m.tipo if hasattr(m, 'tipo') else m.get('tipo', '')

        fecha = m.fecha if hasattr(m, 'fecha') else m.get('fecha')
        fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if fecha else ""

        herramienta_nombre = ""
        herramienta_codigo = ""
        if hasattr(m, 'herramienta') and m.herramienta:
            herramienta_nombre = m.herramienta.nombre
            herramienta_codigo = m.herramienta.codigo

        trabajador_str = ""
        if hasattr(m, 'trabajador') and m.trabajador:
            trabajador_str = m.trabajador.nombre_completo

        usuario_str = ""
        if hasattr(m, 'usuario') and m.usuario:
            usuario_str = m.usuario.nombre

        valores = [
            fecha_str,
            herramienta_nombre,
            herramienta_codigo,
            tipo.replace("_", " ").title(),
            (m.estado_anterior or "").replace("_", " ").title() if hasattr(m, 'estado_anterior') else "",
            (m.estado_nuevo or "").replace("_", " ").title() if hasattr(m, 'estado_nuevo') else "",
            trabajador_str or (m.destino if hasattr(m, 'destino') and m.destino else ""),
            usuario_str,
        ]
        for c_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = _borde()
            cell.alignment = Alignment(vertical="center")
            if c_idx == 4:
                color_tipo = TIPO_COLORES.get(tipo, "AAAAAA")
                cell.fill = _header_fill(color_tipo)
                cell.font = Font(color=WHITE, bold=True)
            else:
                cell.fill = fill_bg
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{row_h}:H{row_h + len(movimientos)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── Plantilla de importación ─────────────────────────────────────────────────
COLUMNAS_IMPORTACION = [
    ("nombre",          "Nombre *",             35),
    ("categoria",       "Categoría",            18),
    ("subcategoria",    "Subcategoría",         18),
    ("familia",         "Familia",              15),
    ("marca",           "Marca",                18),
    ("modelo",          "Modelo",               18),
    ("fabricante",      "Fabricante",           18),
    ("num_serie",       "N° Serie",             18),
    ("potencia",        "Potencia",             12),
    ("voltaje",         "Voltaje",              12),
    ("peso",            "Peso (kg)",            12),
    ("color",           "Color",                12),
    ("dimensiones",     "Dimensiones",          18),
    ("activo_fijo",     "Nº Activo Fijo",       18),
    ("vida_util_anos",  "Vida Útil (años)",     15),
    ("estado",          "Estado",               15),
    ("fecha_compra",    "Fecha Compra",         15),
    ("precio_compra",   "Precio Compra (€)",    15),
    ("proveedor_texto", "Proveedor",            25),
    ("numero_factura",  "Nº Factura",           18),
    ("observaciones",   "Observaciones",        30),
    ("ubicacion_texto", "Ubicación",            25),
]


def generar_plantilla_importacion() -> bytes:
    """Genera un Excel vacío con las columnas correctas para importar herramientas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Herramientas"

    ncols = len(COLUMNAS_IMPORTACION)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "MRD TOOL CONTROL — Plantilla de Importación de Herramientas"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = _header_fill(MRD_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Complete los campos (los marcados con * son obligatorios). "
        "La referencia y el QR los asigna automáticamente MRD. Estado por defecto: disponible."
    )
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    for c_idx, (campo, titulo, ancho) in enumerate(COLUMNAS_IMPORTACION, start=1):
        cell = ws.cell(row=3, column=c_idx, value=titulo)
        color = MRD_ORANGE if titulo.endswith("*") else "4A6FA5"
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = _header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _borde()
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[3].height = 22

    # Fila de ejemplo
    ejemplo = {
        "nombre": "Taladro percutor Hilti TE-700",
        "categoria": "Maquinaria eléctrica",
        "subcategoria": "Percutores",
        "familia": "",
        "marca": "Hilti",
        "modelo": "TE-700",
        "fabricante": "Hilti AG",
        "num_serie": "SN12345678",
        "potencia": "1000W",
        "voltaje": "230V",
        "peso": "5.5",
        "color": "Rojo",
        "dimensiones": "40x20x15 cm",
        "activo_fijo": "AF-2024-001",
        "vida_util_anos": "10",
        "estado": "disponible",
        "fecha_compra": "2024-01-15",
        "precio_compra": "850.00",
        "proveedor_texto": "Hilti Store",
        "numero_factura": "FAC-2024-001",
        "observaciones": "Herramienta de ejemplo — eliminar antes de importar",
        "ubicacion_texto": "Almacén Principal",
    }
    for c_idx, (campo, _, _w) in enumerate(COLUMNAS_IMPORTACION, start=1):
        cell = ws.cell(row=4, column=c_idx, value=ejemplo.get(campo, ""))
        cell.fill = _header_fill(LIGHT_BLUE)
        cell.border = _borde()
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[4].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Importar desde Excel ──────────────────────────────────────────────────────
def importar_herramientas_excel(db, contenido: bytes, user) -> dict:
    """
    Parsea un archivo .xlsx de herramientas y las inserta en la base de datos.
    Retorna dict con: creadas, omitidas y errores. Las referencias las genera MRD.
    """
    from openpyxl import load_workbook
    from models import Herramienta
    from datetime import datetime as _dt

    wb2 = load_workbook(filename=io.BytesIO(contenido), read_only=True, data_only=True)
    ws2 = wb2.active

    # Detectar la fila de encabezado. La referencia ya no se admite desde Excel:
    # el programa crea una distinta para cada herramienta importada.
    header_row = None
    campo_to_col: dict = {}
    for row_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        for col_idx, cell_val in enumerate(row):
            if not cell_val or not isinstance(cell_val, str):
                continue
            normalized = (cell_val.lower().strip()
                          .replace(" *", "").replace("nº ", "").replace("n° ", "")
                          .replace("(kg)", "").replace("(€)", "").replace("(años)", "")
                          .strip())
            for campo, titulo, _ in COLUMNAS_IMPORTACION:
                t_norm = (titulo.lower()
                          .replace(" *", "").replace("nº ", "").replace("n° ", "")
                          .replace("(kg)", "").replace("(€)", "").replace("(años)", "")
                          .strip())
                if normalized == t_norm or normalized == campo:
                    campo_to_col[campo] = col_idx
        if "nombre" in campo_to_col:
            header_row = row_idx
            break

    if not header_row or "nombre" not in campo_to_col:
        raise ValueError("No se encontró encabezado con columna 'Nombre'")

    creadas = 0
    omitidas = 0
    errores = []
    from identificadores import generar_referencia_herramienta

    for data_row in ws2.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v for v in data_row if v is not None):
            continue

        def _val(campo):
            idx = campo_to_col.get(campo)
            if idx is None or idx >= len(data_row):
                return None
            v = data_row[idx]
            if v is None:
                return None
            return str(v).strip() if not isinstance(v, str) else v.strip()

        nombre = _val("nombre")
        if not nombre:
            errores.append("Fila sin nombre omitida")
            continue

        try:
            codigo = generar_referencia_herramienta(db)
            peso_raw = _val("peso")
            precio_raw = _val("precio_compra")
            vida_raw = _val("vida_util_anos")
            fecha_raw = _val("fecha_compra")

            h = Herramienta(
                codigo=codigo,
                nombre=nombre,
                categoria=_val("categoria") or "Otro",
                subcategoria=_val("subcategoria") or None,
                familia=_val("familia") or None,
                marca=_val("marca") or None,
                modelo=_val("modelo") or None,
                fabricante=_val("fabricante") or None,
                num_serie=_val("num_serie") or None,
                potencia=_val("potencia") or None,
                voltaje=_val("voltaje") or None,
                peso=float(peso_raw) if peso_raw else None,
                color=_val("color") or None,
                dimensiones=_val("dimensiones") or None,
                activo_fijo=_val("activo_fijo") or None,
                vida_util_anos=int(float(vida_raw)) if vida_raw else None,
                estado=_val("estado") or "disponible",
                fecha_compra=_dt.strptime(fecha_raw, "%Y-%m-%d").date() if fecha_raw and len(fecha_raw) == 10 else None,
                precio_compra=float(precio_raw) if precio_raw else None,
                proveedor_texto=_val("proveedor_texto") or None,
                numero_factura=_val("numero_factura") or None,
                observaciones=_val("observaciones") or None,
                ubicacion_texto=_val("ubicacion_texto") or None,
                activa=True,
            )
            db.add(h)
            db.flush()
            creadas += 1
        except Exception as exc:
            errores.append(f"Error al importar '{nombre}': {exc}")

    return {
        "ok": True,
        "creadas": creadas,
        "omitidas": omitidas,
        "errores": errores,
        "total_procesadas": creadas + omitidas + len(errores),
    }



# ═══════════════════════════════════════════════════════════════════
# SPRINT 4.7 — INFORMES INTELIGENTES
# ═══════════════════════════════════════════════════════════════════

from datetime import timedelta


def generar_analisis_inteligente(db) -> dict:
    """
    Analiza la BD y devuelve KPIs, tendencias, alertas e insights narrativos.
    No usa ML — estadística descriptiva pura.
    """
    from models import Herramienta, Maquinaria, Movimiento, Incidencia, Reparacion, Material
    from sqlalchemy import func, case

    ahora = datetime.now()
    hace_30 = ahora - timedelta(days=30)
    hace_60 = ahora - timedelta(days=60)
    hace_7  = ahora - timedelta(days=7)
    hace_90 = ahora - timedelta(days=90)

    # ── Herramientas ──────────────────────────────────────────────
    total_h = db.query(Herramienta).filter(Herramienta.activa == True).count()
    estados_h = {}
    for row in db.query(Herramienta.estado, func.count()).filter(Herramienta.activa == True).group_by(Herramienta.estado).all():
        estados_h[row[0]] = row[1]

    # Herramientas entregadas/en_obra hace más de 30 días sin devolver
    h_mucho_tiempo = db.query(Herramienta).filter(
        Herramienta.activa == True,
        Herramienta.estado.in_(["entregada", "en_obra"]),
        Herramienta.ultima_actualizacion <= hace_30,
    ).count() if hasattr(Herramienta, 'ultima_actualizacion') else 0

    # ── Maquinaria ────────────────────────────────────────────────
    total_m = db.query(Maquinaria).filter(Maquinaria.activa == True).count()
    estados_m = {}
    for row in db.query(Maquinaria.estado, func.count()).filter(Maquinaria.activa == True).group_by(Maquinaria.estado).all():
        estados_m[row[0]] = row[1]

    # ── Movimientos por mes (últimos 6 meses) ─────────────────────
    mov_por_mes = []
    for i in range(5, -1, -1):
        fecha_ref = ahora - timedelta(days=30 * i)
        mes_str = fecha_ref.strftime("%Y-%m")
        count = db.query(Movimiento).filter(
            func.strftime("%Y-%m", Movimiento.fecha) == mes_str
        ).count()
        mov_por_mes.append({"mes": fecha_ref.strftime("%b %Y"), "total": count})

    # Tendencia movimientos: comparar último mes vs anterior
    mov_mes_actual = mov_por_mes[-1]["total"]
    mov_mes_anterior = mov_por_mes[-2]["total"]
    tendencia_mov = "sube" if mov_mes_actual > mov_mes_anterior else ("baja" if mov_mes_actual < mov_mes_anterior else "igual")
    delta_mov = mov_mes_actual - mov_mes_anterior

    # ── Incidencias ───────────────────────────────────────────────
    total_inc = db.query(Incidencia).count()
    inc_abiertas = db.query(Incidencia).filter(Incidencia.estado == "abierta").count()
    inc_este_mes = db.query(Incidencia).filter(Incidencia.fecha_apertura >= hace_30).count()
    inc_mes_pasado = db.query(Incidencia).filter(
        Incidencia.fecha_apertura >= hace_60,
        Incidencia.fecha_apertura < hace_30,
    ).count()

    # ── Reparaciones ──────────────────────────────────────────────
    total_rep = db.query(Reparacion).count()
    rep_abiertas = db.query(Reparacion).filter(Reparacion.estado.in_(["pendiente", "en_proceso"])).count()
    rep_retrasadas = db.query(Reparacion).filter(
        Reparacion.estado.in_(["pendiente", "en_proceso"]),
        Reparacion.fecha_entrada <= hace_30,
    ).count()

    # ── Materiales con stock bajo ─────────────────────────────────
    mat_stock_bajo = db.query(Material).filter(
        Material.activo == True,
        Material.stock_actual <= Material.stock_minimo,
    ).count()

    # ── Insights narrativos ───────────────────────────────────────
    insights = []
    alertas = []

    pct_disponible = round(estados_h.get("disponible", 0) / total_h * 100) if total_h else 0
    pct_uso = 100 - pct_disponible

    if pct_uso >= 80:
        alertas.append(f"Alta utilización: el {pct_uso}% del inventario de herramientas está fuera.")
    elif pct_uso >= 60:
        insights.append(f"Buena actividad: {pct_uso}% del inventario en uso.")
    else:
        insights.append(f"Baja utilización: solo el {pct_uso}% del inventario está en uso.")

    if h_mucho_tiempo > 0:
        alertas.append(f"{h_mucho_tiempo} herramienta(s) llevan más de 30 días entregadas sin devolución.")

    if rep_retrasadas > 0:
        alertas.append(f"{rep_retrasadas} reparación(es) llevan más de 30 días sin cerrar.")

    if mat_stock_bajo > 0:
        alertas.append(f"{mat_stock_bajo} material(es) con stock por debajo del mínimo.")

    if inc_abiertas > 5:
        alertas.append(f"{inc_abiertas} incidencias abiertas pendientes de resolución.")

    if tendencia_mov == "sube" and delta_mov > 0:
        insights.append(f"Actividad creciente: +{delta_mov} movimientos respecto al mes anterior.")
    elif tendencia_mov == "baja" and delta_mov < 0:
        insights.append(f"Actividad decreciente: {delta_mov} movimientos respecto al mes anterior.")

    maq_operativas = estados_m.get("operativa", 0) + estados_m.get("disponible", 0)
    if total_m > 0:
        pct_maq = round(maq_operativas / total_m * 100)
        if pct_maq < 50:
            alertas.append(f"Solo el {pct_maq}% de la maquinaria está operativa.")
        else:
            insights.append(f"{pct_maq}% de la maquinaria operativa.")

    return {
        "generado_en": ahora.strftime("%d/%m/%Y %H:%M"),
        "herramientas": {
            "total": total_h,
            "estados": estados_h,
            "pct_disponible": pct_disponible,
            "pct_uso": pct_uso,
            "mucho_tiempo": h_mucho_tiempo,
        },
        "maquinaria": {
            "total": total_m,
            "estados": estados_m,
            "operativas": maq_operativas,
        },
        "movimientos": {
            "por_mes": mov_por_mes,
            "mes_actual": mov_mes_actual,
            "tendencia": tendencia_mov,
            "delta": delta_mov,
        },
        "incidencias": {
            "total": total_inc,
            "abiertas": inc_abiertas,
            "este_mes": inc_este_mes,
            "mes_anterior": inc_mes_pasado,
        },
        "reparaciones": {
            "total": total_rep,
            "abiertas": rep_abiertas,
            "retrasadas": rep_retrasadas,
        },
        "materiales": {
            "stock_bajo": mat_stock_bajo,
        },
        "insights": insights,
        "alertas": alertas,
    }


def exportar_trabajadores_excel(trabajadores: list) -> bytes:
    """Exporta lista de trabajadores a Excel con formato MRD."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores"

    ws.merge_cells("A1:J1")
    ws["A1"] = "MRD TOOL CONTROL — Listado de Trabajadores"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = _header_fill(MRD_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    cols = [
        ("Código", 14), ("Nombre", 22), ("Apellidos", 22), ("DNI", 14),
        ("Cargo", 20), ("Departamento", 18), ("Teléfono", 16),
        ("Email", 28), ("Empresa", 20), ("Estado", 12),
    ]
    row_h = 4
    for c_idx, (titulo, ancho) in enumerate(cols, start=1):
        cell = ws.cell(row=row_h, column=c_idx, value=titulo)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = _header_fill(MRD_ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borde()
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[row_h].height = 22

    for i, t in enumerate(trabajadores):
        row = row_h + 1 + i
        bg = _header_fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        valores = [
            getattr(t, "codigo", "") or "",
            getattr(t, "nombre", "") or "",
            getattr(t, "apellidos", "") or "",
            getattr(t, "dni", "") or "",
            getattr(t, "cargo", "") or "",
            getattr(t, "departamento", "") or "",
            getattr(t, "telefono", "") or "",
            getattr(t, "email", "") or "",
            getattr(t, "empresa", "") or "",
            "Activo" if getattr(t, "activo", True) else "Inactivo",
        ]
        for c_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill = bg
            cell.border = _borde()
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_maquinaria_excel(maquinaria_list: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Maquinaria"

    headers = ["Código", "Nombre", "Tipo", "Marca", "Modelo", "Matrícula",
               "Estado", "Obra Actual", "Año", "Próx. ITV", "Activa"]
    col_w   = [14, 28, 18, 16, 16, 14, 16, 20, 8, 14, 8]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _header_fill(MRD_BLUE)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()
        ws.column_dimensions[get_column_letter(i)].width = col_w[i - 1]
    ws.row_dimensions[1].height = 28

    for row_n, m in enumerate(maquinaria_list, 2):
        vals = [
            getattr(m, "codigo", ""),
            getattr(m, "nombre", ""),
            getattr(m, "tipo", ""),
            getattr(m, "marca", "") or "",
            getattr(m, "modelo", "") or "",
            getattr(m, "matricula", "") or "",
            getattr(m, "estado", ""),
            getattr(m, "obra_actual_texto", "") or "",
            getattr(m, "anio_fabricacion", "") or "",
            getattr(m, "proxima_itv", "").strftime("%d/%m/%Y") if getattr(m, "proxima_itv", None) else "",
            "Sí" if getattr(m, "activa", True) else "No",
        ]
        fill = PatternFill("solid", fgColor=GRAY if row_n % 2 == 0 else WHITE)
        for col_n, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col_n, value=val)
            c.fill = fill
            c.border = _borde()
            c.alignment = Alignment(vertical="center")
            c.font = Font(size=9)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_incidencias_excel(incidencias: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidencias"

    headers = ["ID", "Título", "Descripción", "Estado", "Prioridad",
               "Herramienta", "Fecha apertura", "Fecha cierre", "Reportado por"]
    col_w   = [6, 30, 40, 14, 12, 20, 16, 16, 20]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _header_fill(MRD_ORANGE)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()
        ws.column_dimensions[get_column_letter(i)].width = col_w[i - 1]
    ws.row_dimensions[1].height = 28

    for row_n, inc in enumerate(incidencias, 2):
        h_nombre = ""
        if getattr(inc, "herramienta", None):
            h_nombre = inc.herramienta.nombre
        reporter = ""
        if getattr(inc, "reportado_por", None):
            reporter = inc.reportado_por.nombre or inc.reportado_por.username
        vals = [
            inc.id,
            getattr(inc, "titulo", ""),
            getattr(inc, "descripcion", "") or "",
            getattr(inc, "estado", ""),
            getattr(inc, "prioridad", "") or "",
            h_nombre,
            inc.fecha_apertura.strftime("%d/%m/%Y") if getattr(inc, "fecha_apertura", None) else "",
            inc.fecha_cierre.strftime("%d/%m/%Y") if getattr(inc, "fecha_cierre", None) else "",
            reporter,
        ]
        fill = PatternFill("solid", fgColor=GRAY if row_n % 2 == 0 else WHITE)
        for col_n, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col_n, value=val)
            c.fill = fill
            c.border = _borde()
            c.alignment = Alignment(vertical="center", wrap_text=(col_n == 3))
            c.font = Font(size=9)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_reparaciones_excel(reparaciones: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reparaciones"

    headers = ["ID", "Herramienta", "Estado", "Taller / Proveedor",
               "Fecha entrada", "Fecha salida", "Coste (€)", "Descripción", "Notas"]
    col_w   = [6, 24, 14, 22, 14, 14, 10, 35, 25]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _header_fill(MRD_BLUE)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borde()
        ws.column_dimensions[get_column_letter(i)].width = col_w[i - 1]
    ws.row_dimensions[1].height = 28

    for row_n, rep in enumerate(reparaciones, 2):
        h_nombre = ""
        if getattr(rep, "herramienta", None):
            h_nombre = rep.herramienta.nombre
        vals = [
            rep.id,
            h_nombre,
            getattr(rep, "estado", ""),
            getattr(rep, "taller", "") or getattr(rep, "proveedor_texto", "") or "",
            rep.fecha_entrada.strftime("%d/%m/%Y") if getattr(rep, "fecha_entrada", None) else "",
            rep.fecha_salida.strftime("%d/%m/%Y") if getattr(rep, "fecha_salida", None) else "",
            float(rep.coste or 0) if getattr(rep, "coste", None) is not None else "",
            getattr(rep, "descripcion", "") or "",
            getattr(rep, "notas", "") or "",
        ]
        fill = PatternFill("solid", fgColor=GRAY if row_n % 2 == 0 else WHITE)
        for col_n, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col_n, value=val)
            c.fill = fill
            c.border = _borde()
            c.alignment = Alignment(vertical="center", wrap_text=(col_n in (8, 9)))
            c.font = Font(size=9)
            if col_n == 7 and isinstance(val, float):
                c.number_format = "#,##0.00 €"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_inventario_pdf(herramientas: list, company_name: str = "MRD ESTRUCTURAS") -> bytes:
    """Listado completo de herramientas en PDF apaisado, con la misma
    identidad visual que exportar_pdf_resumen()."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.3*cm, rightMargin=1.3*cm,
                             topMargin=1.3*cm, bottomMargin=1.3*cm)
    azul = colors.HexColor("#1B4F8A")
    gris = colors.HexColor("#6C757D")
    ESTADOS_COLORES = {
        "disponible": colors.HexColor("#2ECC71"), "entregada": colors.HexColor("#E74C3C"),
        "en_obra": colors.HexColor("#E67E22"), "en_furgoneta": colors.HexColor("#9B59B6"),
        "en_reparacion": colors.HexColor("#F39C12"), "perdida": colors.HexColor("#7F8C8D"),
        "baja": colors.HexColor("#BDC3C7"),
    }

    h1 = ParagraphStyle("h1", fontSize=16, textColor=azul, spaceAfter=4,
                         fontName="Helvetica-Bold", alignment=TA_LEFT)
    small = ParagraphStyle("s", fontSize=8, textColor=gris, spaceAfter=2, leading=11)

    story = [
        Paragraph(company_name, h1),
        Paragraph(f"Inventario de herramientas — {len(herramientas)} registros — "
                  f"{datetime.now().strftime('%d/%m/%Y %H:%M')}", small),
        HRFlowable(width="100%", thickness=2, color=azul, spaceAfter=8),
    ]

    encabezado = ["Código", "Nombre", "Categoría", "Estado", "Ubicación", "Responsable"]
    data = [encabezado]
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DEE2E6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i, h in enumerate(herramientas, start=1):
        estado = getattr(h, "estado", "") or ""
        data.append([
            getattr(h, "codigo", "") or "",
            getattr(h, "nombre", "") or "",
            getattr(h, "categoria", "") or "",
            estado.replace("_", " ").title(),
            getattr(h, "ubicacion_texto", "") or "",
            (h.responsable.nombre_completo if getattr(h, "responsable", None) else ""),
        ])
        estilo.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F9FA") if i % 2 else colors.white))
        color_estado = ESTADOS_COLORES.get(estado)
        if color_estado:
            estilo.append(("TEXTCOLOR", (3, i), (3, i), color_estado))
            estilo.append(("FONTNAME", (3, i), (3, i), "Helvetica-Bold"))

    tabla = Table(data, colWidths=[2.6*cm, 6*cm, 3.2*cm, 3*cm, 4*cm, 5*cm], repeatRows=1)
    tabla.setStyle(TableStyle(estilo))
    story.append(tabla)
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.5, color=gris))
    story.append(Paragraph("Generado automáticamente por MRD TOOL CONTROL", small))

    doc.build(story)
    return buf.getvalue()


def exportar_pdf_resumen(analisis: dict, company_name: str = "MRD ESTRUCTURAS") -> bytes:
    """PDF ejecutivo de 1 página con KPIs e insights. Usa reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    azul = colors.HexColor("#1B4F8A")
    naranja = colors.HexColor("#E8600A")
    rojo = colors.HexColor("#DC3545")
    verde = colors.HexColor("#198754")
    gris = colors.HexColor("#6C757D")

    h1 = ParagraphStyle("h1", fontSize=18, textColor=azul, spaceAfter=4,
                         fontName="Helvetica-Bold", alignment=TA_LEFT)
    h2 = ParagraphStyle("h2", fontSize=11, textColor=azul, spaceAfter=3,
                         spaceBefore=10, fontName="Helvetica-Bold")
    normal = ParagraphStyle("n", fontSize=9, textColor=colors.black, spaceAfter=3, leading=13)
    small = ParagraphStyle("s", fontSize=8, textColor=gris, spaceAfter=2, leading=11)
    alert_style = ParagraphStyle("a", fontSize=9, textColor=rojo, spaceAfter=3,
                                  leading=13, fontName="Helvetica-Bold")
    ok_style = ParagraphStyle("ok", fontSize=9, textColor=verde, spaceAfter=3, leading=13)

    story = []

    # Header
    story.append(Paragraph(f"{company_name}", h1))
    story.append(Paragraph(f"Informe de Control de Activos — {analisis['generado_en']}", small))
    story.append(HRFlowable(width="100%", thickness=2, color=azul, spaceAfter=10))

    # KPI tabla herramientas + maquinaria
    h = analisis["herramientas"]
    m = analisis["maquinaria"]
    inc = analisis["incidencias"]
    rep = analisis["reparaciones"]

    kpi_data = [
        ["HERRAMIENTAS", "", "MAQUINARIA", ""],
        [f"{h['total']}", "Total", f"{m['total']}", "Total"],
        [f"{h['estados'].get('disponible',0)}", "Disponibles",
         f"{m['operativas']}", "Operativas"],
        [f"{h['pct_uso']}%", "En uso",
         f"{m['estados'].get('en_reparacion',0)}", "En reparación"],
        ["", "", "", ""],
        ["INCIDENCIAS", "", "REPARACIONES", ""],
        [f"{inc['abiertas']}", "Abiertas", f"{rep['abiertas']}", "Abiertas"],
        [f"{inc['este_mes']}", "Este mes", f"{rep['retrasadas']}", "Retrasadas"],
    ]
    kpi_table = Table(kpi_data, colWidths=[2.2*cm, 4.5*cm, 2.2*cm, 4.5*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 5), (-1, 5), azul),
        ("TEXTCOLOR", (0, 5), (-1, 5), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (0, 4), "Helvetica-Bold"),
        ("FONTNAME", (2, 1), (2, 4), "Helvetica-Bold"),
        ("FONTNAME", (0, 6), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 6), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (0, 4), 14),
        ("FONTSIZE", (2, 1), (2, 4), 14),
        ("FONTSIZE", (0, 6), (0, -1), 14),
        ("FONTSIZE", (2, 6), (2, -1), 14),
        ("TEXTCOLOR", (0, 1), (0, 4), azul),
        ("TEXTCOLOR", (2, 1), (2, 4), naranja),
        ("TEXTCOLOR", (0, 6), (0, -1), rojo),
        ("TEXTCOLOR", (2, 6), (2, -1), rojo),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUND", (0, 1), (-1, 4), [colors.HexColor("#F8F9FA"), colors.white]),
        ("ROWBACKGROUND", (0, 6), (-1, -1), [colors.HexColor("#F8F9FA"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 14))

    # Movimientos por mes
    story.append(Paragraph("Movimientos — últimos 6 meses", h2))
    mov = analisis["movimientos"]["por_mes"]
    if mov:
        max_val = max(r["total"] for r in mov) or 1
        bar_data = [["Mes", "Movimientos", "Gráfica"]]
        for r in mov:
            barra = "█" * int(r["total"] / max_val * 20) if r["total"] else ""
            bar_data.append([r["mes"], str(r["total"]), barra])
        bar_table = Table(bar_data, colWidths=[3*cm, 3*cm, 10.5*cm])
        bar_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9ECEF")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TEXTCOLOR", (2, 1), (2, -1), azul),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
            ("ROWHEIGHT", (0, 0), (-1, -1), 16),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ]))
        story.append(bar_table)

    story.append(Spacer(1, 14))

    # Alertas e insights
    alertas = analisis.get("alertas", [])
    insights = analisis.get("insights", [])

    if alertas:
        story.append(Paragraph("⚠ Alertas detectadas", h2))
        for a in alertas:
            story.append(Paragraph(f"• {a}", alert_style))
        story.append(Spacer(1, 6))

    if insights:
        story.append(Paragraph("✓ Observaciones", h2))
        for i in insights:
            story.append(Paragraph(f"• {i}", ok_style))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=gris))
    story.append(Paragraph(
        f"Generado automáticamente por MRD TOOL CONTROL · {analisis['generado_en']}",
        small))

    doc.build(story)
    return buf.getvalue()


# ─── Importación de Trabajadores ─────────────────────────────────────────────

def generar_plantilla_trabajadores() -> bytes:
    """Genera Excel de plantilla para importar trabajadores."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores"

    headers = [
        "nombre*", "apellidos", "dni", "codigo",
        "telefono", "email", "empresa", "cargo",
        "departamento", "observaciones", "activo",
    ]
    header_fill = PatternFill("solid", fgColor="1B4F8A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 8

    # Fila de ejemplo
    ejemplo = ["Juan", "García López", "12345678A", "T-001",
               "600123456", "juan@mrdestructuras.com", "MRD Estructuras",
               "Operario", "Obra", "", "si"]
    for col, v in enumerate(ejemplo, 1):
        ws.cell(row=2, column=col, value=v)

    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("nombre*", "Obligatorio. Nombre del trabajador."),
        ("apellidos", "Opcional."),
        ("dni", "Clave de actualización: si existe un trabajador con este DNI, se actualiza."),
        ("codigo", "Clave alternativa de actualización si no hay DNI."),
        ("telefono", "Opcional."),
        ("email", "Opcional."),
        ("empresa", "Por defecto: MRD Estructuras."),
        ("cargo", "Opcional."),
        ("departamento", "Opcional."),
        ("observaciones", "Opcional."),
        ("activo", "si/no o 1/0. Por defecto: si."),
    ]
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 60
    ws2.cell(row=1, column=1, value="Campo").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Descripción").font = Font(bold=True)
    for i, (campo, desc) in enumerate(instrucciones, 2):
        ws2.cell(row=i, column=1, value=campo)
        ws2.cell(row=i, column=2, value=desc)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_trabajadores_excel(contenido: bytes, db) -> dict:
    """
    Lee Excel y realiza upsert de trabajadores.
    Clave de actualización: DNI (si existe) o codigo.
    No elimina registros existentes.
    Devuelve resumen {creados, actualizados, errores, filas_procesadas}.
    """
    from openpyxl import load_workbook
    from models import Trabajador

    wb = load_workbook(filename=io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"creados": 0, "actualizados": 0, "errores": ["Archivo vacío"], "filas_procesadas": 0}

    # Normalizar cabecera
    header_raw = [str(c).strip().lower().rstrip("*") if c else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header_raw)}

    def get(row, name, default=""):
        idx = col.get(name)
        if idx is None:
            return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    creados = 0
    actualizados = 0
    errores = []

    for i, row in enumerate(rows[1:], start=2):
        nombre = get(row, "nombre")
        if not nombre:
            continue  # Fila vacía

        try:
            dni    = get(row, "dni") or None
            codigo = get(row, "codigo") or None
            activo_raw = get(row, "activo", "si").lower()
            activo = activo_raw not in ("no", "0", "false", "n")

            # Buscar existente por DNI o código
            existente = None
            if dni:
                existente = db.query(Trabajador).filter(Trabajador.dni == dni).first()
            if not existente and codigo:
                existente = db.query(Trabajador).filter(Trabajador.codigo == codigo).first()

            if existente:
                existente.nombre      = nombre
                existente.apellidos   = get(row, "apellidos") or existente.apellidos
                existente.telefono    = get(row, "telefono") or existente.telefono
                existente.email       = get(row, "email") or existente.email
                existente.empresa     = get(row, "empresa") or existente.empresa or "MRD Estructuras"
                existente.cargo       = get(row, "cargo") or existente.cargo
                existente.departamento = get(row, "departamento") or existente.departamento
                existente.observaciones = get(row, "observaciones") or existente.observaciones
                if dni:
                    existente.dni = dni
                if codigo:
                    existente.codigo = codigo
                existente.activo = activo
                actualizados += 1
            else:
                t = Trabajador(
                    nombre=nombre,
                    apellidos=get(row, "apellidos") or None,
                    dni=dni,
                    codigo=codigo,
                    telefono=get(row, "telefono") or None,
                    email=get(row, "email") or None,
                    empresa=get(row, "empresa") or "MRD Estructuras",
                    cargo=get(row, "cargo") or None,
                    departamento=get(row, "departamento") or None,
                    observaciones=get(row, "observaciones") or None,
                    activo=activo,
                )
                db.add(t)
                creados += 1

        except Exception as e:
            errores.append(f"Fila {i}: {e}")

    db.commit()
    return {
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "filas_procesadas": creados + actualizados + len(errores),
    }
